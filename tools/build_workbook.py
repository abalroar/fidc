# -*- coding: utf-8 -*-
"""Monta o workbook analítico a partir de data/Solfácil/*.csv.

Paleta única: laranja, preto e tons de cinza. Nenhum verde, vermelho ou azul.
Toda tabela e ListObject nomeado tbl_*. Todo gráfico e nativo.
"""
import csv, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, AreaChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "data", "solfacil_claude")
OUTDIR = os.path.join(ROOT, "outputs", "solfacil")

# ---------------------------------------------------------------- paleta
LARANJA      = "EC7000"
LARANJA_ESC  = "A85000"
LARANJA_CLA  = "F7C89A"
PRETO        = "000000"
CINZA_ESC    = "323436"
CINZA_MED    = "6E6E6E"
CINZA_CLARO  = "BFBFBF"
CINZA_FUNDO  = "F2F2F2"
BRANCO       = "FFFFFF"

F_TITULO  = Font(name="Calibri", size=15, bold=True, color=PRETO)
F_LEITURA = Font(name="Calibri", size=10.5, color=CINZA_ESC)
F_META    = Font(name="Calibri", size=9, italic=True, color=CINZA_MED)
F_CAB     = Font(name="Calibri", size=10, bold=True, color=BRANCO)
F_CORPO   = Font(name="Calibri", size=10, color=PRETO)
F_SUB     = Font(name="Calibri", size=11, bold=True, color=PRETO)

FILL_CAB  = PatternFill("solid", fgColor=CINZA_ESC)
FILL_FAIXA = PatternFill("solid", fgColor=LARANJA)
BORDA_BASE = Border(bottom=Side(style="thin", color=CINZA_CLARO))




# Rótulos de exibição: as chaves das colunas são ASCII por estabilidade de código;
# o cabeçalho do Excel mostra o texto em português corrente.
ROTULOS = {
    "veiculo_id": "Veículo", "tipo": "Tipo", "nome_comercial": "Nome comercial",
    "nome_oficial": "Nome oficial", "numero_emissao": "Nº da emissão",
    "cnpj_ou_emissora": "CNPJ / emissora", "securitizadora": "Securitizadora",
    "data_inicio_ou_emissao": "Início ou emissão", "data_bookbuilding": "Bookbuilding",
    "data_encerramento_oferta": "Encerramento da oferta", "situacao": "Situação",
    "administrador": "Administrador", "gestor": "Gestor", "custodiante": "Custodiante",
    "agente_fiduciario": "Agente fiduciário", "escriturador": "Escriturador", "auditor": "Auditor",
    "agencia_rating": "Agência de rating", "coordenador_lider": "Coordenador líder",
    "demais_coordenadores": "Demais coordenadores", "participantes_especiais": "Participantes especiais",
    "pl_ou_saldo_Rmi": "PL ou saldo (R$ mi)", "carteira_Rmi": "Carteira (R$ mi)",
    "data_base": "Data-base", "fonte_id": "Fonte", "status": "Status",
    "camada": "Camada", "serie": "Série", "isin": "ISIN", "data_emissao": "Data de emissão",
    "data_vencimento": "Vencimento", "prazo_meses": "Prazo (meses)",
    "quantidade_ofertada_lote_base": "Qtd. ofertada (lote base)", "quantidade_subscrita": "Qtd. subscrita",
    "valor_nominal_unitario_R": "Valor nominal unitário (R$)",
    "montante_ofertado_Rmi": "Montante ofertado (R$ mi)", "montante_subscrito_Rmi": "Montante subscrito (R$ mi)",
    "montante_reportado_deck_Rmi": "Montante reportado (R$ mi)", "pct_da_emissao": "% da emissão",
    "indexador": "Indexador", "taxa_teto_lamina": "Taxa - perna indexada da lâmina",
    "taxa_piso_lamina": "Taxa - piso da lâmina", "taxa_contratada": "Taxa contratada",
    "perna_que_prevaleceu": "Perna que prevaleceu", "duration_dias": "Duration (dias)",
    "pu_atual_R": "PU (R$)", "rating_agencia": "Agência", "rating_nota": "Nota",
    "colocacao": "Colocação", "retida_pelo_originador": "Retida pelo originador",
    "instrumento": "Instrumento", "ordem_cronologica": "Ordem",
    "montante_captado_Rmi": "Montante captado (R$ mi)", "familia_indexador": "Família do indexador",
    "preco_taxa_contratada": "Preço (taxa contratada)", "ano": "Ano",
    "numero_de_tranches": "Nº de tranches", "spread_numerico": "Spread (% a.a.)",
    "ordem": "Ordem", "data": "Data", "documento": "Documento", "tipo_de_fonte": "Tipo de fonte",
    "url_ou_origem": "URL ou origem", "data_de_acesso": "Data de acesso",
    "trecho_pagina": "Trecho / página", "termo": "Termo", "definicao": "Definição",
    "metrica": "Métrica", "formula": "Fórmula", "qualificador": "Qualificador",
    "dimensao": "Dimensão", "como_funciona_no_FIDC": "Como funciona no FIDC",
    "como_funciona_no_CRI": "Como funciona no CRI", "vantagem_real": "Vantagem real",
    "evidencia": "Evidência", "o_que_falta_para_confirmar": "O que falta para confirmar",
    "prioridade": "Prioridade", "o_que_falta": "O que falta",
    "pergunta_que_responderia": "Pergunta que responderia", "a_quem_pedir": "A quem pedir",
    "aba_afetada": "Aba afetada", "indicador": "Indicador", "valor": "Valor", "unidade": "Unidade",
    "leitura": "Leitura", "codigo_negociacao": "Código de negociação", "pu_R": "PU (R$)",
    "quantidade_em_circulacao": "Qtd. em circulação", "saldo_da_serie_R": "Saldo da série (R$)",
    "total_distribuido_aos_investidores_R": "Total distribuído aos investidores (R$)",
}



# Vocabulário de tokens usados nos nomes de coluna. As chaves das colunas são ASCII por
# estabilidade de código; o cabeçalho do Excel é montado a partir daqui, em português corrente.
TOKENS = {
    "serie": "série", "series": "séries", "criterio": "critério", "criterios": "critérios",
    "patrimonio": "patrimônio", "razao": "razão", "razoes": "razões", "amortizacao": "amortização",
    "subordinacao": "subordinação", "inadimplencia": "inadimplência", "adimplencia": "adimplência",
    "competencia": "competência", "periodo": "período", "indice": "índice", "indices": "índices",
    "maximo": "máximo", "maxima": "máxima", "minimo": "mínimo", "minima": "mínima",
    "media": "média", "medio": "médio", "publica": "pública", "publico": "público",
    "publicas": "públicas", "publicos": "públicos", "unica": "única", "unico": "único",
    "ultima": "última", "ultimo": "último", "numero": "número", "calculo": "cálculo",
    "formula": "fórmula", "metrica": "métrica", "historico": "histórico", "senior": "sênior",
    "junior": "júnior", "premio": "prêmio", "preco": "preço", "informacao": "informação",
    "distribuicao": "distribuição", "classificacao": "classificação", "concentracao": "concentração",
    "operacao": "operação", "operacoes": "operações", "veiculo": "veículo", "veiculos": "veículos",
    "liquido": "líquido", "liquida": "líquida", "liquidos": "líquidos", "analise": "análise",
    "credito": "crédito", "debenture": "debênture", "nao": "não", "sao": "são",
    "emissao": "emissão", "ocorrencia": "ocorrência", "consequencia": "consequência",
    "decisao": "decisão", "descricao": "descrição", "observacao": "observação",
    "redacao": "redação", "restricao": "restrição", "vedacoes": "vedações", "variacao": "variação",
    "cessao": "cessão", "coobrigacao": "coobrigação", "constituicao": "constituição",
    "carencia": "carência", "balao": "balão", "vagao": "vagão", "quorum": "quórum",
    "ja": "já", "fisicas": "físicas", "instituicoes": "instituições", "ambito": "âmbito",
    "mudanca": "mudança", "discricionaria": "discricionária", "juros": "juros",
    "posicao": "posição", "condicao": "condição", "gatilho": "gatilho", "evidencia": "evidência",
    "referencia": "referência", "dimensao": "dimensão", "definicao": "definição",
    "comparacao": "comparação", "situacao": "situação", "aquisicao": "aquisição",
    "avaliacao": "avaliação", "integralizacao": "integralização", "atras": "atraso",
    "ocorreu": "ocorreu", "titulares": "titulares", "atuais": "atuais",
}
SIGLAS = {
    "pct": "%", "Rmi": "(R$ mi)", "R": "(R$)", "PF": "PF", "PJ": "PJ", "UF": "UF",
    "ANBIMA": "ANBIMA", "CCB": "CCB", "MTM": "MTM", "VNU": "VNU", "CRI": "CRI", "FIDC": "FIDC",
    "PDD": "PDD", "WAM": "WAM", "CDI": "CDI", "bps": "bps", "ISIN": "ISIN", "CNPJ": "CNPJ",
    "id": "", "A": "A", "B": "B", "pp": "p.p.", "d": "d", "n": "nº", "top10": "top 10",
    "90d": "90d", "98pct": "98%", "181": "181", "15d": "15d",
}


def rotulo(c):
    if c in ROTULOS:
        return ROTULOS[c]
    partes = c.split("_")
    saida = []
    for i, t in enumerate(partes):
        if t in SIGLAS:
            v = SIGLAS[t]
            if v:
                saida.append(v)
            continue
        chave = t.lower()
        saida.append(TOKENS.get(chave, t))
    texto = " ".join(saida).strip()
    return texto[:1].upper() + texto[1:] if texto else c


def ler(nome):
    with open(os.path.join(DATA, nome), encoding="utf-8") as fh:
        r = list(csv.reader(fh))
    return r[0], r[1:]


def numerico(v):
    if not isinstance(v, str):
        return v
    if v in ("n/d", "n/a", "", None):
        return v
    t = v.replace(".", "").replace(",", ".") if (v.count(",") == 1 and v.replace(",", "").replace(".", "").replace("-", "").isdigit()) else v
    try:
        f = float(t)
        return int(f) if f == int(f) and abs(f) < 1e15 and "." not in t else f
    except ValueError:
        return v


def aba(wb, nome, titulo, leitura, meta):
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False
    ws["A1"] = titulo; ws["A1"].font = F_TITULO
    ws["A2"] = leitura; ws["A2"].font = F_LEITURA
    ws["A3"] = meta; ws["A3"].font = F_META
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(vertical="top", wrap_text=False)
    return ws


def tabela(ws, cols, rows, nome_tabela, linha0=5, larguras=None, wrap_cols=()):
    """Escreve uma tabela nativa começando em linha0. Devolve (primeira_linha_dados, ultima_linha)."""
    for j, c in enumerate(cols, start=1):
        cel = ws.cell(row=linha0, column=j, value=rotulo(c))
        cel.font = F_CAB; cel.fill = FILL_CAB
        cel.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[linha0].height = 30
    for i, r in enumerate(rows, start=linha0 + 1):
        for j, v in enumerate(r, start=1):
            cel = ws.cell(row=i, column=j, value=numerico(v))
            cel.font = F_CORPO
            cel.border = BORDA_BASE
            if cols[j - 1] in wrap_cols:
                cel.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cel.alignment = Alignment(vertical="top")
    fim = linha0 + len(rows)
    ref = f"A{linha0}:{get_column_letter(len(cols))}{fim}"
    t = Table(displayName=nome_tabela, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                      showColumnStripes=False, showFirstColumn=False, showLastColumn=False)
    ws.add_table(t)
    if larguras:
        for letra, w in larguras.items():
            ws.column_dimensions[letra].width = w
    else:
        for j, c in enumerate(cols, start=1):
            largura = max(12, min(40, len(c) + 4))
            if rows:
                largura = max(largura, min(46, max(len(str(r[j - 1])) for r in rows[:60]) + 2))
            ws.column_dimensions[get_column_letter(j)].width = largura
    ws.freeze_panes = ws.cell(row=linha0 + 1, column=1)
    return linha0 + 1, fim


def nota(ws, linha, texto):
    ws.cell(row=linha, column=1, value=texto).font = F_LEITURA
    ws.cell(row=linha, column=1).alignment = Alignment(vertical="top", wrap_text=False)


def estilo_serie(s, cor, linha=False):
    gp = GraphicalProperties(solidFill=cor)
    gp.line = LineProperties(solidFill=cor, w=22000) if linha else LineProperties(noFill=True)
    s.graphicalProperties = gp


# ================================================================ construcao
wb = Workbook()
wb.remove(wb.active)
DB = "Data-base: FIDCs em 31/07/2026 | CRIs na última competência por operação | escopo público até 22/08/2026"

# ---------------------------------------------------------------- 00_Painel
c, r = ler("00_painel.csv")
ws = aba(wb, "00_Painel", "Painel do programa Solfácil",
         "Sete FIDCs financiam a carteira enquanto ela e originada; seis operações de CRI compram pools fechados e alongam o funding. O quadro abaixo só traz números detalhados nas abas seguintes.",
         DB + " | Fontes: ver 17_Fontes")
ini, fim = tabela(ws, c, r, "tbl_00_Painel", larguras={"A": 42, "B": 14, "C": 24, "D": 74, "E": 14, "F": 30},
                  wrap_cols=("leitura",))

linha = fim + 3
ws.cell(row=linha, column=1, value="MAPA DO PROGRAMA").font = F_SUB
linha += 1
mapa = [
    ("1. ORIGINAÇÃO", "Solfácil e cerca de 4 mil integradores | CCB pré-fixada de PF e PJ | destinação a sistema fotovoltaico"),
    ("2. WAREHOUSE", "FIDCs I a VII | compram e financiam a carteira durante a originação | revolvência e reinvestimento"),
    ("3. TAKE-OUT", "Kanastra 1ª a 4ª e VERT 174ª e 177a | cessão definitiva sem coobrigação | novo patrimônio separado por emissão"),
    ("4. INVESTIDORES", "Super Sênior até Subordinado Jr. | séries públicas ao mercado e série privada retida pela originadora"),
]
for nome_etapa, desc in mapa:
    cel = ws.cell(row=linha, column=1, value=nome_etapa)
    cel.font = Font(name="Calibri", size=10, bold=True, color=PRETO)
    cel.fill = FILL_FAIXA
    cel.alignment = Alignment(vertical="center", horizontal="center")
    ws.cell(row=linha, column=2, value=desc).font = F_CORPO
    ws.row_dimensions[linha].height = 18
    linha += 1

linha += 2
ws.cell(row=linha, column=1, value="ÍNDICE DAS ABAS").font = F_SUB
linha += 1
INDICE = [
    ("01_Veiculos", "Um a um: os 7 FIDCs e as 6 operações de CRI, com prestadores e coordenadores"),
    ("02_Series", "As 34 séries de CRI e as 34 classes de cotas de FIDC, com taxa, prazo e rating"),
    ("03_Elegibilidade", "Critério a critério, com a redação literal, e o que mudou entre operações"),
    ("04_Concentracao", "Cap contratual por devedor contra o limite ANBIMA"),
    ("05_Prazos_WAM", "O descasamento de prazo na mesma escala de meses"),
    ("06_Waterfall", "A ordem de pagamentos nos dois regimes"),
    ("06b_Waterfall_Visual", "Os dois regimes lado a lado, degrau a degrau"),
    ("07_Subordinada", "Quando a subordinada pode sair e o que já saiu"),
    ("08_PDD", "Provisão por faixa de atraso e o efeito vagão"),
    ("09_Eventos", "Gatilhos de desalavancagem, resgate e recompra"),
    ("09b_Garantias", "Onde esta a garantia real - e onde não esta"),
    ("10_Subscritores", "Quem comprou na emissão, por tipo de investidor"),
    ("11_Matriz_FIDC_CRI", "Quais fundos cederam e quais poderiam ceder para cada CRI"),
    ("11b_Cessoes", "Uma linha por cessão documentada"),
    ("12_Custo_Captacao", "Taxa por série e evolução do spread por camada"),
    ("13_Cronograma", "A curva de amortização projetada e o que já foi pago"),
    ("14_Antes_Depois", "O que mudou nos FIDCs cedentes depois do take-out"),
    ("15_FIDC_vs_CRI", "Onde o CRI ganha, onde não ganha e o que não da para afirmar"),
    ("16_Conflitos", "Divergências entre fontes e a decisão adotada em cada uma"),
    ("17_Fontes", "O inventário completo, incluindo o que não foi localizado"),
    ("18_Metodologia", "A fórmula e o qualificador de cada métrica"),
    ("19_Glossario", "O jargão em português direto"),
    ("20_Lacunas", "O que falta e a quem pedir"),
    ("21_Funding_por_Tranche", "As 70 tranches do programa: valor captado e preço de cada uma"),
    ("22_Custo_Senior", "Como o custo da camada sênior evoluiu de 2020 a 2026"),
    ("23_Debenture", "A debênture de 2022, quarto tipo de veículo do programa"),
    ("24_Timeline", "Linha do tempo de todas as emissões e o muro de vencimentos"),
    ("25_De_Para_FIDC_CRI", "Qual fundo cedeu para qual operação e o que segue vigente"),
    ("26_Ranking_Risco", "Do mandato mais restritivo ao mais permissivo"),
    ("27_Waterfall_Comparado", "As duas cascatas documentadas, degrau a degrau"),
    ("28_Resgate_Subordinada", "Quando a subordinada pode virar caixa da originadora"),
    ("29_Preco_de_Aquisicao", "Como cada veículo precifica o crédito que compra"),
    ("30_Mismatch_e_Prazo", "Descasamento e prazo sugerido de aprovação da sênior"),
    ("31_Historico_Saques", "O que já saiu de subordinada, e onde"),
    ("32_Pontos_em_Aberto", "As dez perguntas para a diligência"),
]
for nm, desc in INDICE:
    cel = ws.cell(row=linha, column=1, value=nm)
    cel.font = Font(name="Calibri", size=10, color=PRETO, underline="single")
    cel.hyperlink = f"#'{nm if nm != '13_Cronograma' else '13_Cronograma_Pagamentos'}'!A1"
    ws.cell(row=linha, column=2, value=desc).font = F_CORPO
    linha += 1

# ---------------------------------------------------------------- 01_Veiculos
c, r = ler("01_veiculos.csv")
ws = aba(wb, "01_Veiculos", "Veículos do programa",
         "Uma linha por veículo. Os FIDCs II e IV são os únicos cujo nome oficial e CNPJ constam de documento primario - eles aparecem nomeados nos Prospectos das duas primeiras emissões de CRI.",
         DB)
tabela(ws, c, r, "tbl_01_Veiculos", wrap_cols=("demais_coordenadores", "participantes_especiais", "agente_fiduciario", "escriturador", "coordenador_lider", "nome_oficial", "administrador"))

# ---------------------------------------------------------------- 02_Series
c, r = ler("02_series.csv")
ws = aba(wb, "02_Series", "Séries de CRI e classes de cotas de FIDC",
         "34 séries de CRI (5+5+6+7+6+5) e 34 classes de cotas de FIDC. São dimensões diferentes e não se somam. Lote base, quantidade subscrita e valor reportado ficam em colunas separadas: teto de oferta nunca e colocação realizada.",
         DB + " | Fontes: lâminas, comunicados, prospectos e ANX-DECK")
tabela(ws, c, r, "tbl_02_Series", wrap_cols=("taxa_teto_lamina", "taxa_contratada", "perna_que_prevaleceu", "rating_nota", "quantidade_ofertada_lote_base", "fonte_id"))

# ---------------------------------------------------------------- 03_Elegibilidade
c, r = ler("03_elegibilidade.csv")
ws = aba(wb, "03_Elegibilidade", "Critérios de elegibilidade, critério a critério",
         "O cap por devedor endureceu de CRI-III para CRI-V e passou a ser escalonado pela integralização: 0,15% no início, 0,07% com o pool maduro - 30% mais granular que o limite fixo de 0,10% de CRI-III. CRI-V também introduziu a vedação a parcela balão e o enquadramento do PJ na Resolucao CMN 5.118, que CRI-III não trazia.",
         DB + " | Redação literal das lâminas de CRI-I, CRI-III e CRI-V")
ini, fim = tabela(ws, c, r, "tbl_03_Elegibilidade",
                  wrap_cols=("cap_individual_pct_patrimonio_separado", "taxa_retorno_minima_pro_forma",
                             "quem_atesta_elegibilidade", "vedacoes_expressas", "redação_literal",
                             "adimplencia_na_cessao", "amortizacao_mensal_sem_balao", "enquadramento_PJ"))
linha = fim + 3
ws.cell(row=linha, column=1, value="LINHAS DERIVADAS: O QUE APERTA E QUANTO").font = F_SUB
c2, r2 = ler("03b_elegibilidade_deltas.csv")
tabela(ws, c2, r2, "tbl_03b_Deltas", linha0=linha + 1,
       wrap_cols=("o_que_muda", "leitura_de_credito"))
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 70
nota(ws, linha + 1 + len(r2) + 3,
     "Nota: em CRI-V os Critérios de Elegibilidade são verificados sobre dados enviados pela Gestora do Cedente Fundo - evidência textual de que o cedente é um FIDC, e não apenas a originadora direta.")

# ---------------------------------------------------------------- 04_Concentracao
import solfacil_criterios as C_
c, r = ler("04_concentracao.csv")
ws = aba(wb, "04_Concentracao", "Concentração: limite contratual contra limite de mercado",
         C_.NOTA_CONCENTRACAO, DB)
ini, fim = tabela(ws, c, r, "tbl_04_Concentracao",
                  wrap_cols=("cap_individual_escalonado", "classificacao_ANBIMA"))
ws.column_dimensions["C"].width = 60
ws.column_dimensions["M"].width = 60

ch = BarChart(); ch.type = "bar"; ch.style = None
ch.title = "Cap individual por devedor, em % do Patrimônio Separado (escala logarítmica)"
ch.y_axis.title = "% do Patrimônio Separado"
dados = Reference(ws, min_col=2, min_row=ini - 1, max_row=ini + 5)
cats = Reference(ws, min_col=1, min_row=ini, max_row=ini + 5)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
estilo_serie(ch.series[0], LARANJA)
ch.legend = None; ch.height = 7.5; ch.width = 17
ws.add_chart(ch, f"A{fim + 3}")
nota(ws, fim + 20, "O limite ANBIMA de 20% por devedor não aparece no gráfico porque está duas ordens de grandeza acima da escala contratual: seria uma barra 80 a 285 vezes maior que a maior das seis.")

# ---------------------------------------------------------------- 05_Prazos_WAM
c, r = ler("05_prazos_wam.csv")
ws = aba(wb, "05_Prazos_WAM", "Descasamento de prazo, tudo em meses",
         "O ativo pode ir a 3.845 dias (126 meses) por recebível, com média ponderada de até 2.000 dias (66 meses). O passivo tem vencimento legal de 58 a 144 meses, mas duration muito menor - e a duration varia por operação, não pelo programa: 38 a 43 meses em CRI-I, 59 a 119 em CRI-III e apenas 22 a 23 em CRI-V.",
         DB + " | Duration aproximada, sujeita a redução por amortização extraordinária")
ini, fim = tabela(ws, c, r, "tbl_05_Prazos")

# Grafico: series publicas de CRI com duration conhecida, na mesma escala de meses
sub = [(x[0], x[2], x[3], x[5], x[8], x[11], x[13]) for x in r
       if x[1] == "CRI" and x[13] not in ("n/d", "n/a", "")]
base = fim + 3
ws.cell(row=base, column=1, value="COMPARAÇÃO NA MESMA ESCALA DE MESES").font = F_SUB
hdr = ["Série", "WAM contratual max (meses)", "Prazo max do recebível (meses)",
       "Duration (meses)", "Prazo legal da série (meses)"]
linhas_g = [[f"{s[0]} {s[1]} {s[2]}", numerico(s[3]), numerico(s[4]), numerico(s[6]), numerico(s[5])] for s in sub]
gi, gf = tabela(ws, hdr, linhas_g, "tbl_05_Comparacao", linha0=base + 1)

ch = BarChart(); ch.type = "bar"; ch.grouping = "clustered"
ch.title = "Prazo do ativo contra prazo do passivo, por série de CRI"
ch.x_axis.title = "Meses"
dados = Reference(ws, min_col=2, max_col=5, min_row=gi - 1, max_row=gf)
cats = Reference(ws, min_col=1, min_row=gi, max_row=gf)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
for s, cor in zip(ch.series, [CINZA_ESC, PRETO, LARANJA, CINZA_CLARO]):
    estilo_serie(s, cor)
ch.height = 16; ch.width = 26
ws.add_chart(ch, f"H{base + 1}")

# ---------------------------------------------------------------- 06_Waterfall
import solfacil_estrutura as E_
c, r = ler("06_waterfall.csv")
ws = aba(wb, "06_Waterfall", "Ordem de pagamentos",
         E_.NOTA_WATERFALL, DB + " | Única ordem integralmente documentada: CRI-II (2o Aditamento ao Termo de Securitização)")
tabela(ws, c, r, "tbl_06_Waterfall",
       wrap_cols=("regime", "gatilho_de_mudanca_para_sequencial", "quem_recebe_juros_antes_de_principal",
                  "subordinado_jr_prioridade", "cash_sweep", "condicionalidade", "reserva_de_juros",
                  "reserva_para_resgate_antecipado", "senior_prioridade", "mezanino_prioridade",
                  "subordinado_prioridade", "nome_contratual"))

# ---------------------------------------------------------------- 06b_Waterfall_Visual
c, r = ler("06b_waterfall_degraus.csv")
ws = aba(wb, "06b_Waterfall_Visual", "Os dois regimes, degrau a degrau",
         "A esquerda, o regime pró-rata condicionado: cada camada só recebe se as coberturas acima estiverem enquadradas, e recebe até um saldo alvo. A direita, o sequencial: some a condição de cobertura e some o alvo - a camada de cima é paga até 98% antes de a próxima receber principal.",
         DB + " | Fonte: ANX-TS2-K2, cláusulas 6.5.1 e 6.5.2")
pro = [x for x in r if x[1] == "Pró-rata"]
seq = [x for x in r if x[1] == "Sequencial"]
ws.cell(row=5, column=1, value="REGIME PRÓ-RATA CONDICIONADO").font = F_SUB
ws.cell(row=5, column=5, value="REGIME SEQUENCIAL").font = F_SUB
INTENS = {"Despesa": CINZA_CLARO, "SS": PRETO, "S": CINZA_ESC, "M": CINZA_MED, "Sub": LARANJA_ESC, "Jr": LARANJA, "Fim": CINZA_FUNDO}


def bloco(ws, linha, col, ordem, item, desc, cor, cor_texto=BRANCO):
    a = ws.cell(row=linha, column=col, value=f"{ordem}  {item}")
    a.fill = PatternFill("solid", fgColor=cor)
    a.font = Font(name="Calibri", size=9, bold=True, color=cor_texto)
    a.alignment = Alignment(vertical="center", horizontal="center")
    b = ws.cell(row=linha, column=col + 1, value=desc)
    b.font = Font(name="Calibri", size=9, color=PRETO)
    b.alignment = Alignment(vertical="center")


def cor_do_degrau(desc):
    d = desc.lower()
    if "despesa" in d or "investimentos permitidos" in d:
        return INTENS["Despesa"], PRETO
    if "prêmio final" in d:
        return INTENS["Jr"], PRETO
    if "5ª série" in d:
        return INTENS["Jr"], PRETO
    if "4ª série" in d:
        return INTENS["Sub"], BRANCO
    if "3ª série" in d:
        return INTENS["M"], BRANCO
    if "2ª série" in d:
        return INTENS["S"], BRANCO
    if "1ª série" in d:
        return INTENS["SS"], BRANCO
    if "extraordinária" in d:
        return INTENS["Despesa"], PRETO
    return CINZA_FUNDO, PRETO


for i, x in enumerate(pro):
    cor, ct = cor_do_degrau(x[4])
    bloco(ws, 6 + i, 1, x[2], x[3], x[4], cor, ct)
for i, x in enumerate(seq):
    cor, ct = cor_do_degrau(x[4])
    bloco(ws, 6 + i, 5, x[2], x[3], x[4], cor, ct)
ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 88
ws.column_dimensions["C"].width = 3
ws.column_dimensions["E"].width = 14; ws.column_dimensions["F"].width = 88
nota(ws, 6 + max(len(pro), len(seq)) + 2,
     "Legenda de intensidade: preto = Super Sênior, cinza escuro = Sênior, cinza médio = Mezanino, laranja escuro = Subordinado, laranja = Subordinado Jr., cinza claro = despesas e sobras. A cor indica a camada, não risco nem severidade.")

# ---------------------------------------------------------------- 07_Subordinada
c, r = ler("07_subordinada.csv")
ws = aba(wb, "07_Subordinada", "Saque da subordinada: quando pode sair",
         E_.NOTA_SUBORDINADA, DB)
tabela(ws, c, r, "tbl_07_Subordinada",
       wrap_cols=("testes_exigidos", "indices_de_cobertura", "trava_temporal", "vedacoes_pos_evento",
                  "impacto_na_senior", "saque_permitido", "quem_solicita", "quórum"))

# ---------------------------------------------------------------- 08_PDD
c, r = ler("08_pdd.csv")
ws = aba(wb, "08_PDD", "Provisão por faixa de atraso",
         E_.NOTA_PDD, DB + " | Curva inicial: CRI-I, CRI-II, FIDC IV e V | Curva posterior: CRI-III a CRI-VI, FIDC VI e VII")
ini, fim = tabela(ws, c, r, "tbl_08_PDD",
                  wrap_cols=("base_de_incidencia", "efeito_vagao", "tratamento_do_dia_181", "curva", "métrica"))

base = fim + 3
ws.cell(row=base, column=1, value="AS DUAS CURVAS DE PROVISÃO, LADO A LADO").font = F_SUB
hdr = ["Faixa de atraso", "Curva inicial (CRI-I e CRI-II)", "Curva posterior (CRI-III a CRI-VI)"]
faixas = ["Até 15 dias", "16 a 30", "31 a 60", "61 a 90", "91 a 120", "121 a 150", "151 a 180", "Acima de 180"]
ini_v = [0.0, 1.0, 3.0, 10.0, 30.0, 50.0, 70.0, 100.0]
pos_v = [0.0, 1.5, 5.0, 10.0, 37.0, 58.0, 78.0, 100.0]
linhas_g = [[f, i, p] for f, i, p in zip(faixas, ini_v, pos_v)]
gi, gf = tabela(ws, hdr, linhas_g, "tbl_08_Curvas", linha0=base + 1)
ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"
ch.title = "Percentual provisionado por faixa de atraso"
ch.y_axis.title = "% do valor presente do recebível"; ch.x_axis.title = "Faixa de atraso"
dados = Reference(ws, min_col=2, max_col=3, min_row=gi - 1, max_row=gf)
cats = Reference(ws, min_col=1, min_row=gi, max_row=gf)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
estilo_serie(ch.series[0], CINZA_CLARO); estilo_serie(ch.series[1], LARANJA)
ch.height = 9; ch.width = 20
ws.add_chart(ch, f"F{base + 1}")
nota(ws, gf + 3, "A curva posterior reconhece 7 p.p. a mais na faixa de 91 a 120 dias e 8 p.p. a mais entre 121 e 180 - reconhecimento mais conservador, não piora do ativo.")

# ---------------------------------------------------------------- 09_Eventos
c, r = ler("09_eventos.csv")
ws = aba(wb, "09_Eventos", "Eventos e gatilhos",
         "Os gatilhos numéricos só estão integralmente documentados em CRI-II: Índice de Atraso de Estoque limitado a 15%, rebaixamento de 2 níveis de rating e desenquadramento das Razões de Cobertura em 2 datas consecutivas ou 4 alternadas em 12 meses. Vencimento antecipado e 'não aplicável' por escrito em CRI-III e CRI-V - registrado como documentado, não como lacuna.",
         DB)
tabela(ws, c, r, "tbl_09_Eventos",
       wrap_cols=("descricao_do_gatilho", "parametro_numerico", "consequencia_automatica",
                  "quorum_de_dispensa", "prazo_de_cura", "status"))

# ---------------------------------------------------------------- 09b_Garantias
c, r = ler("09b_garantias.csv")
ws = aba(wb, "09b_Garantias", "Garantias: onde estão e onde não estão",
         E_.NOTA_GARANTIAS, DB)
tabela(ws, c, r, "tbl_09b_Garantias",
       wrap_cols=("garantia_no_ambito_do_veiculo", "garantia_sobre_os_direitos_creditorios",
                  "onde_e_contratada", "coobrigacao_do_cedente", "redação_literal", "status"))

# ---------------------------------------------------------------- 10_Subscritores
import solfacil_mercado as M_
c, r = ler("10c_concentracao_subscritores.csv")
ws = aba(wb, "10_Subscritores", "Quem comprou na emissão",
         "55,6% da série Super Sênior de CRI-I foi para uma única instituição financeira ligada ao emissor ou ao consórcio - 200.000 de 360.000 CRI, um só subscritor - enquanto 989 pessoas físicas ficaram com 42,2%. Categoria não é titular: as colunas separam a maior categoria do maior titular único.",
         "Data-base: 23/02/2024 (Anúncio de Encerramento de CRI-I) | Posição corrente: n/d em todas as operações")
ini, fim = tabela(ws, c, r, "tbl_10_Concentracao",
                  wrap_cols=("maior_categoria", "maior_titular_unico_categoria"))
linha = fim + 3
ws.cell(row=linha, column=1, value="TABELA LITERAL DO FORMULÁRIO CVM - CRI-I").font = F_SUB
c2, r2 = ler("10_subscritores_longa.csv")
gi, gf = tabela(ws, c2, r2, "tbl_10_Longa", linha0=linha + 1, wrap_cols=("tipo_de_investidor",))
ws.column_dimensions["D"].width = 74
linha = gf + 3
ws.cell(row=linha, column=1, value="DISTRIBUIÇÃO INICIAL AGREGADA POR OPERAÇÃO").font = F_SUB
c3, r3 = ler("10b_subscritores_agregado.csv")
gi2, gf2 = tabela(ws, c3, r3, "tbl_10_Agregado", linha0=linha + 1, wrap_cols=("fonte_da_posicao",))
nota(ws, gf2 + 3, M_.NOTA_SUBSCRITORES)

# ---------------------------------------------------------------- 11_Matriz_FIDC_CRI
c, r = ler("11_matriz_fidc_cri.csv")
ws = aba(wb, "11_Matriz_FIDC_CRI", "Matriz de cessão: 7 FIDCs por 6 CRIs",
         M_.NOTA_MATRIZ, DB)
CRIS = ["CRI-I", "CRI-II", "CRI-III", "CRI-IV", "CRI-V", "CRI-VI"]
FIDCS = ["FIDC-I", "FIDC-II", "FIDC-III", "FIDC-IV", "FIDC-V", "FIDC-VI", "FIDC-VII"]
estado = {(x[0], x[1]): x[2] for x in r}
hdr = ["FIDC"] + CRIS
linhas_m = [[f] + [estado.get((f, k), "n/d") for k in CRIS] for f in FIDCS]
ini, fim = tabela(ws, hdr, linhas_m, "tbl_11_Matriz")
for i in range(ini, fim + 1):
    for j in range(2, 8):
        cel = ws.cell(row=i, column=j)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cel.value == "Cedeu":
            cel.font = Font(name="Calibri", size=10, bold=True, color=PRETO)
    ws.row_dimensions[i].height = 30
for j in range(2, 8):
    ws.column_dimensions[get_column_letter(j)].width = 22
linha = fim + 3
ws.cell(row=linha, column=1, value="DETALHE POR CELULA").font = F_SUB
c2, r2 = ler("11_matriz_fidc_cri.csv")
gi, gf = tabela(ws, c2, r2, "tbl_11_Detalhe", linha0=linha + 1,
                wrap_cols=("criterio_que_bloqueia_ou_evidencia",))
ws.column_dimensions["D"].width = 90

# ---------------------------------------------------------------- 11b_Cessoes
c, r = ler("11b_cessoes.csv")
ws = aba(wb, "11b_Cessoes", "Cessões documentadas",
         "Uma linha por cessão. Volume cedido, percentual do pool e preço por lote são n/d nas seis operações - o ledger de cessões não é público e é o item 4 da lista de lacunas.",
         DB)
tabela(ws, c, r, "tbl_11b_Cessoes", wrap_cols=("pct_do_pool_do_CRI", "fidc_cedente"))

# ---------------------------------------------------------------- 12_Custo_Captacao
c, r = ler("12_custo_captacao.csv")
ws = aba(wb, "12_Custo_Captacao", "Custo de captação",
         M_.NOTA_CUSTO, DB)
ini, fim = tabela(ws, c, r, "tbl_12_Custo", wrap_cols=("observação", "status"))
ws.column_dimensions["G"].width = 90

linha = fim + 3
ws.cell(row=linha, column=1, value="EVOLUÇÃO DO SPREAD POR CAMADA AO LONGO DAS SEIS OPERAÇÕES").font = F_SUB
c2, r2 = ler("12b_spread_por_camada.csv")
gi, gf = tabela(ws, c2, r2, "tbl_12b_Spread", linha0=linha + 1, wrap_cols=("unidade", "fonte_id"))

# Grafico apenas das camadas comparaveis em DI+ (mezanino e subordinado)
base = gf + 3
ws.cell(row=base, column=1, value="SÉRIES COMPARÁVEIS EM DI + SPREAD").font = F_SUB
hdr = ["Operação", "Mezanino (DI + %)", "Subordinado (DI + %)"]
ops = ["CRI-II", "CRI-III", "CRI-IV", "CRI-V", "CRI-VI"]
mez = [6.00, 5.75, 5.50, 5.50, 5.50]
sub_ = [10.00, 10.00, 10.00, 8.00, 8.00]
linhas_g = [[o, m, s] for o, m, s in zip(ops, mez, sub_)]
gi2, gf2 = tabela(ws, hdr, linhas_g, "tbl_12_Evolucao", linha0=base + 1)
ch = LineChart()
ch.title = "Spread contratado sobre 100% do DI, por camada"
ch.y_axis.title = "% a.a. sobre o DI"; ch.x_axis.title = "Operação"
dados = Reference(ws, min_col=2, max_col=3, min_row=gi2 - 1, max_row=gf2)
cats = Reference(ws, min_col=1, min_row=gi2, max_row=gf2)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
for s, cor in zip(ch.series, [LARANJA, CINZA_ESC]):
    s.graphicalProperties = GraphicalProperties()
    s.graphicalProperties.line = LineProperties(solidFill=cor, w=28000)
    s.marker = Marker(symbol="circle", size=6)
    s.smooth = False
ch.height = 9; ch.width = 19
ws.add_chart(ch, f"F{base + 1}")
nota(ws, gf2 + 3, "CRI-I não entra no gráfico: suas quatro séries públicas são pré-fixadas e não há spread sobre DI a comparar. A queda de spread é consistente com melhora de percepção de risco, mas o dado público não permite atribuir causa - mudaram também o indexador, a camada e o ciclo de juros.")

# ---------------------------------------------------------------- 13_Cronograma_Pagamentos
c, r = ler("13_cronograma_pagamentos.csv")
ws = aba(wb, "13_Cronograma_Pagamentos", "Cronograma de pagamentos: projetado e realizado",
         "Só a 1ª série de CRI-II tem cronograma contratual de principal - as outras quatro amortizam até um Saldo Devedor Target, então o Anexo I delas e calendario de datas, não curva de principal. Toda linha traz a coluna status: Projetado vem do Anexo I, Realizado vem dos informes mensais.",
         "Fonte do projetado: ANX-TS2-K2, Anexo I | Fonte do realizado: ANX-DECK, anexo A3")
ini, fim = tabela(ws, c, r, "tbl_13_Cronograma")

# Curva de morte da 1a serie (unica com percentual contratual)
s1 = [x for x in r if x[1] == "1ª"]
base = fim + 3
ws.cell(row=base, column=1, value="CURVA DE AMORTIZAÇÃO PROJETADA - CRI-II, 1a SÉRIE (SUPER SÊNIOR)").font = F_SUB
hdr = ["Pagamento", "Data", "Saldo remanescente (% do VNU)"]
linhas_g = [[numerico(x[3]), x[4], numerico(x[8])] for x in s1]
gi, gf = tabela(ws, hdr, linhas_g, "tbl_13_Curva", linha0=base + 1)
ch = AreaChart(); ch.grouping = "standard"
ch.title = "Saldo remanescente da 1ª série de CRI-II ao longo dos 60 pagamentos"
ch.y_axis.title = "% do Valor Nominal Unitário"; ch.x_axis.title = "Número do pagamento"
dados = Reference(ws, min_col=3, min_row=gi - 1, max_row=gf)
cats = Reference(ws, min_col=1, min_row=gi, max_row=gf)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
estilo_serie(ch.series[0], LARANJA)
ch.legend = None; ch.height = 10; ch.width = 24
ws.add_chart(ch, f"F{base + 1}")
nota(ws, gf + 3, "O ponto de 98% amortizado - que dispara o resgate compulsório da série - e atingido no 59o pagamento, em 08/05/2029, um pagamento antes do vencimento. As demais séries não têm percentual contratual e por isso não aparecem no gráfico.")

linha = gf + 5
ws.cell(row=linha, column=1, value="AMORTIZAÇÃO REALIZADA POR CAMADA (AGREGADA)").font = F_SUB
c2, r2 = ler("13b_amortizacao_realizada.csv")
gi2, gf2 = tabela(ws, c2, r2, "tbl_13b_Realizado", linha0=linha + 1)
nota(ws, gf2 + 3, "O realizado público é agregado por camada - primeira e última ocorrência, meses com pagamento e total. A série mês a mês por camada não consta dos informes disponíveis, então a curva realizada não pode ser desenhada sem os informes mensais completos.")

# ---------------------------------------------------------------- 14_Antes_Depois
import solfacil_sintese as Z_
c, r = ler("14_antes_depois.csv")
ws = aba(wb, "14_Antes_Depois", "Antes e depois do take-out",
         "Janela de duas competências nos FIDCs VI e VII, em torno do take-out da VERT 177ª de 31/07/2026. As competências t-3 a t+3 pedidas não existem: o evento e a última competência disponível.",
         "Data-base: 30/06/2026 e 31/07/2026 | Fonte: CVM Informe Mensal FIDC, via ANX-DECK")
ini, fim = tabela(ws, c, r, "tbl_14_AntesDepois", wrap_cols=("evento",))
ws.column_dimensions["I"].width = 60

base = fim + 3
ws.cell(row=base, column=1, value="PL E CARTEIRA NAS DUAS COMPETÊNCIAS").font = F_SUB
hdr = ["Competência", "PL FIDC VI (R$ mi)", "Carteira FIDC VI (R$ mi)", "PL FIDC VII (R$ mi)", "Carteira FIDC VII (R$ mi)"]
linhas_g = [["30/06/2026 (t-1)", 437.4, 399.3, 564.7, 544.1],
            ["31/07/2026 (t=0, take-out)", 211.1, 147.7, 619.6, 446.1]]
gi, gf = tabela(ws, hdr, linhas_g, "tbl_14_Serie", linha0=base + 1)
ch = LineChart()
ch.title = "PL e carteira dos dois FIDCs cedentes, com marcador na competência do take-out"
ch.y_axis.title = "R$ mi"
dados = Reference(ws, min_col=2, max_col=5, min_row=gi - 1, max_row=gf)
cats = Reference(ws, min_col=1, min_row=gi, max_row=gf)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
for s, cor in zip(ch.series, [PRETO, CINZA_MED, LARANJA_ESC, LARANJA]):
    s.graphicalProperties = GraphicalProperties()
    s.graphicalProperties.line = LineProperties(solidFill=cor, w=28000)
    s.marker = Marker(symbol="circle", size=7)
    s.smooth = False
ch.height = 10; ch.width = 22
ws.add_chart(ch, f"H{base + 1}")
nota(ws, gf + 3, Z_.NOTA_ANTES_DEPOIS)

# ---------------------------------------------------------------- 15_FIDC_vs_CRI
c, r = ler("15_fidc_vs_cri.csv")
ws = aba(wb, "15_FIDC_vs_CRI", "Onde o CRI ganha, onde não ganha, e o que não da para afirmar",
         "Doze dimensões, com veredito explícito. Em quatro delas a vantagem é do FIDC, em cinco do CRI e em três o dado público não sustenta veredito. A última coluna diz o que falta para fechar cada uma.",
         DB)
ini, fim = tabela(ws, c, r, "tbl_15_Veredito",
                  wrap_cols=("como_funciona_no_FIDC", "como_funciona_no_CRI", "evidência", "o_que_falta_para_confirmar"))
for letra, w in {"A": 34, "B": 62, "C": 62, "D": 16, "E": 62, "F": 52}.items():
    ws.column_dimensions[letra].width = w

# ---------------------------------------------------------------- 16_Conflitos
c, r = ler("16_conflitos.csv")
ws = aba(wb, "16_Conflitos", "Divergências entre fontes e a decisão adotada",
         "Dez casos testados. Em cinco não havia divergência real, só perímetro diferente - o mais importante é o total de séries: os anúncios contam só as séries públicas, e somando a série privada de cada operação chega-se exatamente a 34.",
         DB)
ini, fim = tabela(ws, c, r, "tbl_16_Conflitos",
                  wrap_cols=("valor_fonte_A", "valor_fonte_B", "decisao_adotada", "justificativa",
                             "campo_em_conflito", "fonte_A", "fonte_B"))
for letra, w in {"A": 6, "B": 34, "C": 46, "D": 30, "E": 13, "F": 46, "G": 30, "H": 13, "I": 46, "J": 84, "K": 10}.items():
    ws.column_dimensions[letra].width = w

# ---------------------------------------------------------------- 17_Fontes
c, r = ler("17_fontes.csv")
ws = aba(wb, "17_Fontes", "Inventário de fontes",
         "Doze documentos obtidos e cinco buscas registradas como não localizadas. Ausência confirmada é informação: as duas últimas linhas de busca sustentam a afirmação de que o universo esta completo, com a limitação declarada.",
         "Data de acesso: 22/08/2026")
ini, fim = tabela(ws, c, r, "tbl_17_Fontes",
                  wrap_cols=("documento", "tipo_de_fonte", "url_ou_origem", "trecho_pagina", "status"))
for letra, w in {"A": 20, "B": 62, "C": 44, "D": 46, "E": 13, "F": 13, "G": 74, "H": 26}.items():
    ws.column_dimensions[letra].width = w

# ---------------------------------------------------------------- 18_Metodologia
c, r = ler("18_metodologia.csv")
ws = aba(wb, "18_Metodologia", "Fórmula e qualificador de cada métrica",
         "Cada métrica calculada traz a fórmula e a ressalva que a limita. Quatro métricas pedidas ficaram sem cálculo por falta de insumo público - estão registradas como n/d, com o insumo que falta nomeado.",
         DB)
ini, fim = tabela(ws, c, r, "tbl_18_Metodologia", wrap_cols=("fórmula", "qualificador"))
for letra, w in {"A": 34, "B": 74, "C": 88, "D": 40}.items():
    ws.column_dimensions[letra].width = w

# ---------------------------------------------------------------- 19_Glossario
c, r = ler("19_glossario.csv")
ws = aba(wb, "19_Glossario", "Glossário",
         "Vinte e dois termos em português direto, sem definir jargão com jargão.", DB)
ini, fim = tabela(ws, c, r, "tbl_19_Glossario", wrap_cols=("definição",))
ws.column_dimensions["A"].width = 38; ws.column_dimensions["B"].width = 108

# ---------------------------------------------------------------- 20_Lacunas
c, r = ler("20_lacunas.csv")
ws = aba(wb, "20_Lacunas", "O que falta e a quem pedir",
         "Dez itens, em ordem de prioridade. Os quatro primeiros bloqueiam, respectivamente, a confirmação do universo, a ordem de pagamentos de cinco das seis operações, o volume efetivamente colocado e a economia da cessão.",
         DB)
ini, fim = tabela(ws, c, r, "tbl_20_Lacunas",
                  wrap_cols=("o_que_falta", "pergunta_que_responderia", "a_quem_pedir", "aba_afetada"))
for letra, w in {"A": 11, "B": 72, "C": 66, "D": 34, "E": 34}.items():
    ws.column_dimensions[letra].width = w


# ---------------------------------------------------------------- 21_Funding_por_Tranche
import solfacil_funding as FU_
c, r = ler("21_funding_por_tranche.csv")
ws = aba(wb, "21_Funding_por_Tranche", "Preço e valor captado, tranche a tranche",
         "Todas as 70 tranches do programa em ordem cronológica: 34 classes de cotas de FIDC, 34 séries de CRI e 2 séries de debênture. "
         "A família do indexador está na tabela porque é ela que define o que se compara com o quê - preço em IPCA+ não se compara com preço em DI+ sem a curva de juros.",
         DB + " | Fontes: escrituras, prospectos, lâminas, comunicados e Termos de Securitização")
ini, fim = tabela(ws, c, r, "tbl_21_Funding",
                  wrap_cols=("preco_taxa_contratada", "camada", "serie", "fonte_id", "instrumento"))
for letra, w in {"A": 8, "B": 14, "C": 20, "D": 12, "E": 18, "F": 26, "G": 16,
                 "H": 16, "I": 54, "J": 20, "K": 14, "L": 30, "M": 14}.items():
    ws.column_dimensions[letra].width = w

base = fim + 3
ws.cell(row=base, column=1, value="CAPTAÇÃO POR ANO E POR INSTRUMENTO").font = F_SUB
c2, r2 = ler("25_captacao_por_ano.csv")
gi, gf = tabela(ws, c2, r2, "tbl_21_PorAno", linha0=base + 1)

anos = sorted({x[0] for x in r2})
mapa = {(x[0], x[1]): float(x[2]) for x in r2}
base2 = gf + 3
ws.cell(row=base2, column=1, value="MATRIZ ANO × INSTRUMENTO (R$ mi)").font = F_SUB
hdr = ["Ano", "FIDC (warehouse)", "CRI (take-out)", "Debênture"]
linhas_g = [[a, mapa.get((a, "FIDC (warehouse)"), 0.0), mapa.get((a, "CRI (take-out)"), 0.0),
             mapa.get((a, "Debênture"), 0.0)] for a in anos]
gi2, gf2 = tabela(ws, hdr, linhas_g, "tbl_21_Matriz", linha0=base2 + 1)
ch = BarChart(); ch.type = "col"; ch.grouping = "stacked"; ch.overlap = 100
ch.title = "Captação por ano, por instrumento"
ch.y_axis.title = "R$ mi"; ch.x_axis.title = "Ano de emissão da tranche"
dados = Reference(ws, min_col=2, max_col=4, min_row=gi2 - 1, max_row=gf2)
cats = Reference(ws, min_col=1, min_row=gi2, max_row=gf2)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
for sr, cor in zip(ch.series, [CINZA_ESC, LARANJA, CINZA_CLARO]):
    estilo_serie(sr, cor)
ch.height = 10; ch.width = 22
ws.add_chart(ch, f"F{base2 + 1}")
nota(ws, gf2 + 3,
     "O valor por ano é emissão de tranche, não saldo em aberto: os FIDCs reabrem classes ao longo do tempo, "
     "então a soma das cotas de um mesmo fundo em anos diferentes não é o tamanho do fundo hoje. "
     "A leitura do gráfico é a virada de 2024: até 2023 o funding do programa é integralmente FIDC e debênture; "
     "de 2024 em diante o CRI passa a responder pela maior parte da captação anual.")

# ---------------------------------------------------------------- 22_Custo_Senior
c, r = ler("22_custo_senior_timeline.csv")
ws = aba(wb, "22_Custo_Senior", "Evolução do custo da camada sênior",
         FU_.NOTA_CUSTO_SENIOR,
         DB + " | Taxa contratada na data de emissão de cada veículo")
ini, fim = tabela(ws, c, r, "tbl_22_CustoSenior", wrap_cols=("taxa_contratada", "fonte_id"))
ws.column_dimensions["H"].width = 40

base = fim + 3
ws.cell(row=base, column=1, value="FAMÍLIA DI+ : A ÚNICA SÉRIE LONGA E COMPARÁVEL").font = F_SUB
di = [x for x in r if x[6] == "DI+"]
hdr = ["Veículo", "Data", "Spread sobre o DI (% a.a.)"]
linhas_g = [[f"{x[2]} ({x[3]})", x[1], float(x[8])] for x in di]
gi, gf = tabela(ws, hdr, linhas_g, "tbl_22_DI", linha0=base + 1)
ch = LineChart()
ch.title = "Spread da camada sênior sobre o DI, por data de emissão"
ch.y_axis.title = "% a.a. sobre o DI"; ch.x_axis.title = "Veículo, em ordem de emissão"
dados = Reference(ws, min_col=3, min_row=gi - 1, max_row=gf)
cats = Reference(ws, min_col=1, min_row=gi, max_row=gf)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
for sr in ch.series:
    sr.graphicalProperties = GraphicalProperties()
    sr.graphicalProperties.line = LineProperties(solidFill=LARANJA, w=28000)
    sr.marker = Marker(symbol="circle", size=7)
    sr.smooth = False
ch.legend = None; ch.height = 9; ch.width = 20
ws.add_chart(ch, f"F{base + 1}")

base2 = gf + 3
ws.cell(row=base2, column=1, value="FAMÍLIA IPCA+ E FAMÍLIA PRÉ, LADO A LADO").font = F_SUB
outros = [x for x in r if x[6] in ("IPCA+", "Pré")]
hdr2 = ["Veículo", "Data", "Família", "Taxa ou spread (% a.a.)"]
linhas_g2 = [[f"{x[2]} ({x[3]})", x[1], x[6], float(x[8])] for x in outros]
gi2, gf2 = tabela(ws, hdr2, linhas_g2, "tbl_22_Outros", linha0=base2 + 1)
nota(ws, gf2 + 3,
     "Os números de famílias diferentes não se somam nem se comparam: 7,22% em IPCA+ e 14,81% em pré-fixado "
     "descrevem custos que só ficam na mesma régua depois de convertidos pela curva DI e pela inflação implícita "
     "da respectiva data-base - conversão que não foi feita por falta desses insumos, e que está registrada como lacuna.")

# ---------------------------------------------------------------- 23_Debenture
c, r = ler("23_debenture_posicao.csv")
ws = aba(wb, "23_Debenture", "Debênture Amazônia Solar, o veículo fora do perímetro original",
         FU_.NOTA_DEBENTURE,
         "Data-base: 31/07/2026 | Fontes: escritura de 18/02/2022 e relatório mensal do agente fiduciário")
ini, fim = tabela(ws, c, r, "tbl_23_Posicao")
linha = fim + 3
ws.cell(row=linha, column=1, value="CARTEIRA E CAIXA NA DATA-BASE").font = F_SUB
c2, r2 = ler("24_debenture_carteira.csv")
gi, gf = tabela(ws, c2, r2, "tbl_23_Carteira", linha0=linha + 1)
ws.column_dimensions["A"].width = 54
nota(ws, gf + 3,
     "A escritura fixa o vencimento da 1ª série em 18/02/2033 e o da 2ª em 18/08/2035; o relatório do agente "
     "fiduciário exibe 18/08/2035 para as duas. A divergência está registrada em 16_Conflitos e o valor da "
     "escritura foi o adotado, por ser o documento constitutivo.")


# ---------------------------------------------------------------- 24 a 32: comitê de crédito
import solfacil_comite as K_

c, r = ler("26_timeline_consolidada.csv")
ws = aba(wb, "24_Timeline", "Linha do tempo consolidada",
         "Todos os eventos de emissão do programa em ordem cronológica, dos sete FIDCs, das seis operações de CRI e da debênture. "
         + K_.NOTA_TIMELINE, DB)
ini, fim = tabela(ws, c, r, "tbl_24_Timeline", wrap_cols=("evento", "situacao"))
linha = fim + 3
ws.cell(row=linha, column=1, value="VENCIMENTOS LEGAIS POR ANO").font = F_SUB
c2, r2 = ler("26b_vencimentos_por_ano.csv")
gi, gf = tabela(ws, c2, r2, "tbl_24_Vencimentos", linha0=linha + 1)
ch = BarChart(); ch.type = "col"
ch.title = "Valor nominal com vencimento legal no ano - CRI e debênture"
ch.y_axis.title = "R$ mi"; ch.x_axis.title = "Ano de vencimento"
dados = Reference(ws, min_col=2, min_row=gi - 1, max_row=gf)
cats = Reference(ws, min_col=1, min_row=gi, max_row=gf)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
estilo_serie(ch.series[0], LARANJA)
ch.legend = None; ch.height = 9; ch.width = 20
ws.add_chart(ch, f"E{linha + 1}")

c, r = ler("27_depara_fidc_cri.csv")
ws = aba(wb, "25_De_Para_FIDC_CRI", "De-para: qual fundo cedeu para qual operação",
         K_.NOTA_DEPARA, DB)
ini, fim = tabela(ws, c, r, "tbl_25_DePara", wrap_cols=("qualidade_da_evidencia", "cedentes_documentados"))
ws.column_dimensions["G"].width = 90
linha = fim + 3
ws.cell(row=linha, column=1, value="SITUAÇÃO DE CADA FUNDO HOJE").font = F_SUB
c2, r2 = ler("27b_fidc_status.csv")
tabela(ws, c2, r2, "tbl_25_Status", linha0=linha + 1, wrap_cols=("nome_oficial", "cedeu_para", "situacao_hoje"))

c, r = ler("28_ranking_permissividade.csv")
ws = aba(wb, "26_Ranking_Risco", "Ranking de permissividade dos mandatos",
         K_.NOTA_RANKING, DB + " | Inferido: agregação de parâmetros documentados; fórmula em 18_Metodologia")
ini, fim = tabela(ws, c, r, "tbl_26_Ranking", wrap_cols=("leitura", "fonte_id", "status"))
ws.column_dimensions["N"].width = 96
ch = BarChart(); ch.type = "bar"
ch.title = "Índice de permissividade do mandato (0 = mais restritivo)"
ch.y_axis.title = "Índice"
dados = Reference(ws, min_col=12, min_row=ini - 1, max_row=fim)
cats = Reference(ws, min_col=2, min_row=ini, max_row=fim)
ch.add_data(dados, titles_from_data=True); ch.set_categories(cats)
estilo_serie(ch.series[0], LARANJA)
ch.legend = None; ch.height = 11; ch.width = 20
ws.add_chart(ch, f"A{fim + 3}")

c, r = ler("29_waterfall_comparado.csv")
ws = aba(wb, "27_Waterfall_Comparado", "As duas cascatas integralmente documentadas, lado a lado",
         K_.NOTA_WATERFALL_CMP, DB + " | CRI-II: 2º Aditamento ao Termo | CRI-VI: Termo da 177ª emissão")
ini, fim = tabela(ws, c, r, "tbl_27_WfCmp", wrap_cols=("CRI-II (2ª Kanastra)", "CRI-VI (177ª VERT)", "divergencia"))
for letra, w in {"A": 40, "B": 8, "C": 62, "D": 62, "E": 52}.items():
    ws.column_dimensions[letra].width = w
linha = fim + 3
ws.cell(row=linha, column=1, value="RAZÕES DE COBERTURA CONTRATUAIS").font = F_SUB
c2, r2 = ler("29b_razoes_de_cobertura.csv")
tabela(ws, c2, r2, "tbl_27_Coberturas", linha0=linha + 1)

c, r = ler("30_resgate_subordinada.csv")
ws = aba(wb, "28_Resgate_Subordinada", "Em que condições a subordinada sai antes do fim",
         K_.NOTA_SUBORDINADA_CAIXA, DB)
ini, fim = tabela(ws, c, r, "tbl_28_Resgate",
                  wrap_cols=("como_a_subordinada_recebe", "gatilho_ou_teste", "trava_temporal",
                             "pode_ser_fonte_de_caixa_do_originador", "evidencia_documental", "status"))
for letra, w in {"C": 46, "D": 54, "F": 40, "G": 54, "H": 52, "J": 34}.items():
    ws.column_dimensions[letra].width = w

c, r = ler("31_preco_de_aquisicao.csv")
ws = aba(wb, "29_Preco_de_Aquisicao", "Como cada veículo precifica o crédito que compra",
         K_.NOTA_PRECO, DB)
ini, fim = tabela(ws, c, r, "tbl_29_Preco",
                  wrap_cols=("mecanismo_de_preco", "teto_contratual", "preco_efetivamente_praticado", "status"))
for letra, w in {"C": 54, "D": 60, "E": 54, "F": 34, "H": 40}.items():
    ws.column_dimensions[letra].width = w

c, r = ler("32_mismatch_prazo.csv")
ws = aba(wb, "30_Mismatch_e_Prazo", "Descasamento de prazo e prazo sugerido de aprovação",
         K_.NOTA_MISMATCH, DB)
ini, fim = tabela(ws, c, r, "tbl_30_Mismatch", wrap_cols=("racional", "regime_muda_em", "prazo_sugerido_de_exposicao"))
ws.column_dimensions["L"].width = 96
nota(ws, fim + 3, K_.PRAZO_SUGERIDO_NOTA)

c, r = ler("33_historico_saques_subordinada.csv")
ws = aba(wb, "31_Historico_Saques", "Histórico de amortização de subordinada",
         K_.NOTA_HISTORICO, DB)
ini, fim = tabela(ws, c, r, "tbl_31_Historico", wrap_cols=("respeitou_o_regulamento",))
ws.column_dimensions["I"].width = 84

c, r = ler("34_pontos_em_aberto.csv")
ws = aba(wb, "32_Pontos_em_Aberto", "O que permanece sem resposta nos documentos disponíveis",
         "Dez perguntas em ordem de prioridade, com o documento que fecharia cada uma e o impacto na decisão de crédito.", DB)
ini, fim = tabela(ws, c, r, "tbl_32_Aberto",
                  wrap_cols=("pergunta_do_comite", "o_que_falta", "onde_obter", "impacto_na_decisao"))
for letra, w in {"A": 11, "B": 52, "C": 56, "D": 40, "E": 60}.items():
    ws.column_dimensions[letra].width = w

# ---------------------------------------------------------------- salvar
os.makedirs(OUTDIR, exist_ok=True)
hoje = "20260822"
caminho = os.path.join(OUTDIR, f"Solfacil_CRI_FIDC_{hoje}_claude.xlsx")
wb.save(caminho)
print(f"Workbook salvo: {caminho}")
print(f"Abas: {len(wb.sheetnames)}")
for n in wb.sheetnames:
    print("  -", n)
