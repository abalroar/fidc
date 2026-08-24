# -*- coding: utf-8 -*-
"""Workbook dos três quadros de comitê, com tabelas nativas e gráficos nativos."""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, AreaChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import Marker

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUTDIR = os.path.join(ROOT, "outputs", "solfacil")
sys.path.insert(0, HERE)
import solfacil_comite_tabelas as T

NAVY, NAVY2 = "1F3864", "2E5496"
LARANJA, LARANJA_CLA = "EC7000", "F7C89A"
ND_FILL = "FCE4D6"
PRETO, CINZA_ESC, CINZA_MED, CINZA_CLARO, BRANCO = "000000", "323436", "6E6E6E", "D9D9D9", "FFFFFF"

F_TIT = Font(name="Calibri", size=14, bold=True, color=PRETO)
F_SUB = Font(name="Calibri", size=10, color=CINZA_ESC)
F_META = Font(name="Calibri", size=8, italic=True, color=CINZA_MED)
F_CAB = Font(name="Calibri", size=9.5, bold=True, color=BRANCO)
F_ATR = Font(name="Calibri", size=9.5, bold=True, color=PRETO)
F_VAL = Font(name="Calibri", size=9.5, color=PRETO)
F_ND = Font(name="Calibri", size=9.5, color=CINZA_ESC, italic=True)
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_ATR = PatternFill("solid", fgColor="EAEFF7")
FILL_ND = PatternFill("solid", fgColor=ND_FILL)
BORDA = Border(bottom=Side(style="thin", color=CINZA_CLARO),
               right=Side(style="thin", color=CINZA_CLARO))

wb = Workbook(); wb.remove(wb.active)


def aba(nome, titulo, sub, meta):
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False
    ws["A1"] = titulo; ws["A1"].font = F_TIT
    ws["A2"] = sub; ws["A2"].font = F_SUB
    ws["A3"] = meta; ws["A3"].font = F_META
    ws.row_dimensions[1].height = 20
    return ws


def matriz(ws, cols, linhas, nome_tab, l0=5, larg=None, num_dir=True):
    """Tabela atributo x veículo, com destaque nas células a preencher."""
    for j, c in enumerate(cols, start=1):
        cel = ws.cell(row=l0, column=j, value=c)
        cel.font = F_CAB; cel.fill = FILL_NAVY
        cel.alignment = Alignment(horizontal="center" if j > 1 else "left",
                                  vertical="center", wrap_text=True)
    ws.row_dimensions[l0].height = 26
    for i, linha in enumerate(linhas, start=l0 + 1):
        for j, v in enumerate(linha, start=1):
            cel = ws.cell(row=i, column=j, value=v)
            cel.border = BORDA
            if j == 1:
                cel.font = F_ATR; cel.fill = FILL_ATR
                cel.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                if v == "n/d":
                    cel.font = F_ND; cel.fill = FILL_ND
                else:
                    cel.font = F_VAL
                cel.alignment = Alignment(horizontal="center" if num_dir else "left",
                                          vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 20
    fim = l0 + len(linhas)
    t = Table(displayName=nome_tab, ref=f"A{l0}:{get_column_letter(len(cols))}{fim}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                      showColumnStripes=False)
    ws.add_table(t)
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = (larg or [28] + [17] * (len(cols) - 1))[j - 1]
    ws.freeze_panes = ws.cell(row=l0 + 1, column=2)
    return l0 + 1, fim


def nota(ws, linha, texto):
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = Font(name="Calibri", size=9, color=CINZA_ESC)
    c.alignment = Alignment(vertical="top", wrap_text=False)


def estilo(serie, cor, linha=False, largura=22000):
    gp = GraphicalProperties(solidFill=cor)
    gp.line = LineProperties(solidFill=cor, w=largura) if linha else LineProperties(noFill=True)
    serie.graphicalProperties = gp


# ═══════════════════════════════ 1A · FIDCs
ws = aba("1A_FIDCs", "Quadro 1A · Perfil dos FIDCs Solfácil",
         "Uma linha por atributo do cadastro e do informe mensal da CVM. FIDC VIII em discussão, coluna reservada.",
         "Data-base 31/07/2026 · células em destaque: campo do CVM Fundos.NET não obtido nesta compilação")
ini, fim = matriz(ws, T.FIDC_COLS, T.FIDC_LINHAS, "tbl_1A_FIDCs",
                  larg=[30, 17, 17, 17, 17, 17, 17, 17, 17])
nota(ws, fim + 2, T.FIDC_NOTA)

ch = BarChart(); ch.type = "col"; ch.grouping = "stacked"; ch.overlap = 100
ch.title = "Composição do PL por classe de cota (% do PL)"
ch.y_axis.title = "% do PL"
for r, cor, nome in [(fim - 9, NAVY, "Sênior 1"), (fim - 8, LARANJA, "Mezanino"), (fim - 7, LARANJA_CLA, "Sub Júnior")]:
    pass
ws2 = ws
base = fim + 4
ws.cell(row=base, column=1, value="DADOS DO GRÁFICO (% do PL)").font = F_ATR
hdr = ["Classe"] + T.FIDC_COLS[1:8]
dados_g = [["Sênior 1", 63.6, 72.9, 66.8, 100.0, 79.1, 66.4, 74.0],
           ["Mezanino", 32.5, 14.9, 21.0, 0.0, 13.5, 18.7, 20.8],
           ["Sub Júnior", 3.9, 12.2, 12.2, 0.0, 7.4, 14.9, 5.2]]
gi, gf = matriz(ws, hdr, dados_g, "tbl_1A_Grafico", l0=base + 1,
                larg=[16] + [13] * 7)
ch = BarChart(); ch.type = "col"; ch.grouping = "stacked"; ch.overlap = 100
ch.title = "Composição do PL por classe de cota"
ch.y_axis.title = "% do PL"
dados = Reference(ws, min_col=2, max_col=8, min_row=gi - 1, max_row=gf)
cats = Reference(ws, min_col=1, min_row=gi, max_row=gf)
ch.add_data(dados, titles_from_data=True, from_rows=True)
ch.set_categories(Reference(ws, min_col=2, max_col=8, min_row=gi - 1))
for s, cor in zip(ch.series, [NAVY, LARANJA, LARANJA_CLA]):
    estilo(s, cor)
ch.height, ch.width = 8, 18
ws.add_chart(ch, f"J{base + 1}")

# ═══════════════════════════════ 1B · CRIs
ws = aba("1B_CRIs", "Quadro 1B · Perfil das operações de CRI Solfácil",
         "Mesma estrutura do quadro dos FIDCs, com abertura por série e pela securitizadora.",
         "Data-base 21/08/2026 · células em destaque: campo do Informe Mensal CRI não obtido nesta compilação")
ini, fim = matriz(ws, T.CRI_COLS, T.CRI_LINHAS, "tbl_1B_CRIs",
                  larg=[32, 22, 24, 22, 24, 22, 22])
nota(ws, fim + 2, T.CRI_NOTA)

base = fim + 4
ws.cell(row=base, column=1, value="DADOS DO GRÁFICO (% da emissão por série)").font = F_ATR
hdr = ["Série"] + T.CRI_COLS[1:]
dados_g = [["1ª", 59.7, 65.0, 49.0, 43.3, 22.1, 15.5],
           ["2ª", 14.9, 18.0, 16.0, 21.7, 47.9, 69.5],
           ["3ª", 17.9, 10.0, 18.0, 12.0, 15.0, 8.0],
           ["4ª", 5.0, 4.0, 10.0, 6.0, 8.0, 4.0],
           ["5ª", 2.5, 3.0, 4.0, 10.0, 4.0, 3.0],
           ["6ª", 0.0, 0.0, 3.0, 4.0, 3.0, 0.0],
           ["7ª", 0.0, 0.0, 0.0, 3.0, 0.0, 0.0]]
gi, gf = matriz(ws, hdr, dados_g, "tbl_1B_Grafico", l0=base + 1, larg=[12] + [13] * 6)
ch = BarChart(); ch.type = "col"; ch.grouping = "stacked"; ch.overlap = 100
ch.title = "Composição da emissão por série"
ch.y_axis.title = "% do volume emitido"
ch.add_data(Reference(ws, min_col=2, max_col=7, min_row=gi - 1, max_row=gf),
            titles_from_data=True, from_rows=True)
ch.set_categories(Reference(ws, min_col=2, max_col=7, min_row=gi - 1))
for s, cor in zip(ch.series, [NAVY, NAVY2, LARANJA, LARANJA_CLA, CINZA_MED, CINZA_CLARO, "9DC3E6"]):
    estilo(s, cor)
ch.height, ch.width = 8, 18
ws.add_chart(ch, f"I{base + 1}")

# ═══════════════════════════════ 2 · Emissões
ws = aba("2_Emissoes", "Quadro 2 · Emissões, remuneração-alvo e subordinação mínima",
         "Primeira e última emissão, volume acumulado, abertura por data e a forma como o índice de subordinação aparece em cada veículo.",
         "Fonte: prospectos, lâminas, comunicados de bookbuilding e termos de securitização")
ini, fim = matriz(ws, T.EMI_FIDC_COLS, T.EMI_FIDC, "tbl_2A_EmiFIDC",
                  larg=[13, 14, 14, 17, 62, 62, 42], num_dir=False)
for i in range(ini, fim + 1):
    ws.row_dimensions[i].height = 46
nota(ws, fim + 2, "Volume emitido soma as emissões registradas de cada fundo; não é saldo em aberto.")

base = fim + 4
ws.cell(row=base, column=1, value="OPERAÇÕES DE CRI").font = F_ATR
gi, gf = matriz(ws, T.EMI_CRI_COLS, T.EMI_CRI, "tbl_2B_EmiCRI", l0=base + 1,
                larg=[22, 14, 17, 66, 66, 52], num_dir=False)
for i in range(gi, gf + 1):
    ws.row_dimensions[i].height = 46
nota(ws, gf + 2, T.NOTA_SUBORDINACAO)

# ═══════════════════════════════ 3 · Cronograma e runoff
ws = aba("3_Cronograma_Runoff", "Quadro 3 · Cronograma de amortização e runoff da exposição",
         "Trimestre a trimestre do 2T26 ao 3T38. Vencimento legal é documentado; PMT contratual só existe para a 1ª série de CRI-II.",
         "Fonte: datas de vencimento das 34 séries de CRI e 2 de debênture; Anexo I do 2º aditamento ao TS da 2ª emissão Kanastra")
ini, fim = matriz(ws, T.CRONO_COLS, T.CRONOGRAMA, "tbl_3_Cronograma",
                  larg=[12, 19, 25, 28, 19, 46], num_dir=True)
nota(ws, fim + 2, T.NOTA_CRONOGRAMA)

# gráfico 1: PMT por trimestre
ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"
ch.title = "Vencimento legal por trimestre (R$ mm)"
ch.y_axis.title = "R$ mm"; ch.x_axis.title = "Trimestre"
ch.add_data(Reference(ws, min_col=3, min_row=ini - 1, max_row=fim), titles_from_data=True)
ch.set_categories(Reference(ws, min_col=1, min_row=ini, max_row=fim))
estilo(ch.series[0], LARANJA)
ch.legend = None; ch.height, ch.width = 8, 30
ws.add_chart(ch, f"H5")

# gráfico 2: runoff do saldo
ch2 = AreaChart(); ch2.grouping = "standard"
ch2.title = "Runoff do saldo nominal por vencimento legal (R$ mm)"
ch2.y_axis.title = "R$ mm"; ch2.x_axis.title = "Trimestre"
ch2.add_data(Reference(ws, min_col=5, min_row=ini - 1, max_row=fim), titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=1, min_row=ini, max_row=fim))
estilo(ch2.series[0], NAVY)
ch2.legend = None; ch2.height, ch2.width = 8, 30
ws.add_chart(ch2, f"H22")

# runoff por instrumento
base = fim + 5
ws.cell(row=base, column=1, value="RUNOFF POR INSTRUMENTO — SALDO NOMINAL AO FIM DO TRIMESTRE (R$ mm)").font = F_ATR
VEIC = ["CRI-I", "CRI-II", "CRI-III", "CRI-IV", "CRI-V", "CRI-VI", "DEB-I"]
ROT = {"CRI-I": "CRI I", "CRI-II": "CRI II", "CRI-III": "CRI III", "CRI-IV": "CRI IV",
       "CRI-V": "CRI V", "CRI-VI": "CRI VI", "DEB-I": "Debênture"}
TOT = {"CRI-I": 603.0, "CRI-II": 750.0, "CRI-III": 750.0, "CRI-IV": 450.0,
       "CRI-V": 470.6, "CRI-VI": 647.059, "DEB-I": 60.0}
saldo = dict(TOT)
linhas_rf = []
for rot, sal_ini, venc_tri, pmt, det in T.RUNOFF:
    for k, v in det.items():
        saldo[k] = round(saldo.get(k, 0.0) - v, 3)
    linhas_rf.append([rot] + [round(max(saldo.get(v, 0.0), 0.0), 1) for v in VEIC])
gi, gf = matriz(ws, ["Trimestre"] + [ROT[v] for v in VEIC], linhas_rf, "tbl_3_Runoff",
                l0=base + 1, larg=[12] + [13] * 7)
ch3 = AreaChart(); ch3.grouping = "stacked"
ch3.title = "Exposição total por instrumento (R$ mm)"
ch3.y_axis.title = "R$ mm"; ch3.x_axis.title = "Trimestre"
ch3.add_data(Reference(ws, min_col=2, max_col=8, min_row=gi - 1, max_row=gf), titles_from_data=True)
ch3.set_categories(Reference(ws, min_col=1, min_row=gi, max_row=gf))
for s, cor in zip(ch3.series, [NAVY, NAVY2, LARANJA, LARANJA_CLA, CINZA_MED, CINZA_CLARO, "9DC3E6"]):
    estilo(s, cor)
ch3.height, ch3.width = 9, 30
ws.add_chart(ch3, f"J{base + 1}")
nota(ws, gf + 2,
     "Este runoff assume que cada série permanece integralmente em aberto até o vencimento legal. "
     "Onde o cronograma real existe — 1ª série de CRI-II — a amortização é linear e antecipa R$ 327,8 mm "
     "entre o 2T26 e o 2T29, contra a barra única de R$ 487,5 mm que o vencimento legal projeta para o 2T29. "
     "A diferença mede o quanto a leitura por vencimento legal superestima a exposição.")

# ═══════════════════════════════ Fontes
ws = aba("4_Fontes_e_Lacunas", "Quadro 4 · Fontes e campos a obter no CVM Fundos.NET",
         "O que sustenta cada linha dos quadros anteriores e o que precisa ser preenchido.",
         "Compilação de 24/08/2026")
LAC = [
    ("Rentabilidade YTD por classe de cota", "FIDCs I a VII", "CVM — Informe Mensal FIDC, campo de rentabilidade da cota", "n/d — a obter"),
    ("Rentabilidade YTD por série", "CRIs I a VI", "CVM — Informe Mensal CRI", "n/d — a obter"),
    ("Over 90 / carteira por operação de CRI", "CRIs I a VI", "CVM — Informe Mensal CRI", "n/d — a obter"),
    ("Saldo devedor atual por série", "CRIs I a VI", "CVM — Informe Mensal CRI; B3", "n/d — a obter"),
    ("Cronograma de PMT por série", "CRI-I, III, IV, V e VI", "Anexo I dos termos de securitização — Fundos.NET", "n/d — a obter"),
    ("Cronograma de amortização das cotas", "FIDCs I a VII", "Regulamento e suplementos — Fundos.NET", "n/d — a obter"),
    ("Subordinação mínima", "CRI-I, III, IV e V", "Termos de securitização — Fundos.NET", "n/d — a obter"),
    ("Volume e data da cota sub júnior", "FIDCs IV, V, VI e VII", "CVM — registro de ofertas", "n/d — a obter"),
    ("Remuneração das cotas sênior", "FIDC IV (séries A, B e C)", "Suplementos de classe — Fundos.NET", "n/d — a obter"),
    ("Dados completos", "FIDC VIII", "Em discussão; sem registro na CVM até a data-base", "coluna reservada"),
    ("PL, composição por cota, over 90", "FIDCs I a VII", "CVM — Informe Mensal FIDC, 31/07/2026", "obtido"),
    ("Volume, séries, taxas, vencimentos, ISIN", "CRIs I a VI", "Prospectos, lâminas, comunicados e termos", "obtido"),
    ("Cronograma de PMT", "CRI-II, 1ª série", "Anexo I do 2º aditamento ao TS da 2ª emissão Kanastra", "obtido"),
    ("Razões de cobertura", "CRI-II e CRI-VI", "Termos de securitização", "obtido"),
]
ini, fim = matriz(ws, ["Campo", "Veículos", "Onde obter", "Status"], LAC, "tbl_4_Lacunas",
                  larg=[44, 30, 62, 20], num_dir=False)
nota(ws, fim + 2,
     "Esta compilação não teve acesso à rede: nenhum campo foi consultado diretamente no CVM Fundos.NET. "
     "Os campos marcados como obtidos vieram dos documentos em PDF do acervo e da consolidação de informes "
     "já disponível. Os marcados n/d estão destacados nos quadros para preenchimento.")

os.makedirs(OUTDIR, exist_ok=True)
caminho = os.path.join(OUTDIR, "Solfacil_Comite_Quadros_20260824_claude.xlsx")
wb.save(caminho)
print("Workbook salvo:", caminho)
print("Abas:", wb.sheetnames)
