from __future__ import annotations

import hashlib
import io
import json
from typing import Any, Dict, List

import streamlit as st
from openpyxl import load_workbook


st.set_page_config(page_title="Upload de Modelo FIDC", page_icon="📄", layout="wide")

st.title("Upload do modelo financeiro (.xlsm)")
st.markdown(
    """
Envie o seu modelo financeiro em formato **.xlsm**. O app extrai a lista de abas,
intervalos nomeados e todas as fórmulas encontradas, gerando um arquivo JSON para
análise posterior.
"""
)


def _extract_model_pack(file_bytes: bytes) -> Dict[str, Any]:
    workbook = load_workbook(io.BytesIO(file_bytes), keep_vba=True, data_only=False)

    output: Dict[str, Any] = {
        "sheets": workbook.sheetnames,
        "named_ranges": [],
        "cells": {},
    }

    defined_names = workbook.defined_names
    if hasattr(defined_names, "definedName"):
        defined_name_items = defined_names.definedName
    elif hasattr(defined_names, "defined_names"):
        defined_name_items = defined_names.defined_names
    else:
        defined_name_items = []

    for defined_name in defined_name_items:
        output["named_ranges"].append(
            {
                "name": defined_name.name,
                "refers_to": defined_name.attr_text or "",
            }
        )

    for worksheet in workbook.worksheets:
        sheet_cells: List[Dict[str, str]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formula_value = cell.value
                    if not isinstance(formula_value, str):
                        formula_value = str(formula_value)
                    sheet_cells.append(
                        {
                            "addr": cell.coordinate,
                            "formula": formula_value,
                        }
                    )
        output["cells"][worksheet.title] = sheet_cells

    return output


left, right = st.columns([2, 1], gap="large")

with left:
    st.subheader("1. Envie seu arquivo")
    uploaded_file = st.file_uploader(
        "Selecione um arquivo .xlsm",
        type=["xlsm"],
        accept_multiple_files=False,
        help="O arquivo é processado localmente para gerar o JSON de fórmulas.",
    )

    if uploaded_file is None:
        st.info("Aguardando o upload do arquivo .xlsm.")
    else:
        file_bytes = uploaded_file.getvalue()
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        file_size_mb = len(file_bytes) / (1024 * 1024)

        st.success("Arquivo recebido com sucesso!")
        st.write(
            {
                "Nome": uploaded_file.name,
                "Tamanho (MB)": f"{file_size_mb:.2f}",
                "SHA256": file_hash,
            }
        )

        with st.spinner("Extraindo fórmulas e intervalos nomeados..."):
            model_pack = _extract_model_pack(file_bytes)

        st.subheader("Resumo do modelo")
        st.write(
            {
                "Total de abas": len(model_pack["sheets"]),
                "Total de intervalos nomeados": len(model_pack["named_ranges"]),
                "Total de fórmulas": sum(
                    len(cells) for cells in model_pack["cells"].values()
                ),
            }
        )

        st.caption("Primeiras abas detectadas")
        st.write(model_pack["sheets"][:10])

        json_payload = json.dumps(model_pack, ensure_ascii=False, indent=2)

        st.download_button(
            "Baixar JSON de fórmulas",
            data=json_payload,
            file_name="model_pack_formulas.json",
            mime="application/json",
        )

        st.download_button(
            "Baixar o arquivo enviado",
            data=file_bytes,
            file_name=uploaded_file.name,
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
        )

with right:
    st.subheader("2. Próximos passos")
    st.markdown(
        """
- Confirmar quais abas e cálculos do modelo precisam ser replicados.
- Mapear entradas, saídas e premissas.
- Transformar a lógica em componentes da nova plataforma.
"""
    )

    st.subheader("3. Checklist")
    st.checkbox("O arquivo contém macros (.xlsm)", value=True, disabled=True)
    st.checkbox("Existe documentação ou notas no arquivo")
    st.checkbox("Há exemplos de saída esperada")

st.divider()

st.caption(
    "Se quiser incluir comentários adicionais, descreva no chat quais partes do modelo são mais críticas."
)
