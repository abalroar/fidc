"""Published Office bundle for the audited FIDC industry revision.

The Streamlit request path only reads an immutable, prebuilt bundle.  It never
starts Node or silently serves a stale/legacy deck.  The bundle is produced by
``scripts/build_fidc_revision_artifacts.mjs`` from the same editorial payload
used by the Industry Data page and is accepted only when payload and file
hashes match its manifest.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
import hashlib
from io import BytesIO
import json
import os
from pathlib import Path
import posixpath
import re
import shutil
from typing import Callable, Iterable
import unicodedata
import zipfile
from xml.etree import ElementTree

from services.industry_taxonomy_review import (
    assert_taxonomy_review_ledger_matches_audit,
    taxonomy_review_audit_digest,
    taxonomy_review_ledger_digest,
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATA_DIR = ROOT / "data" / "industry_study"
PAYLOAD_NAME = "artifact_payload.json"
BUNDLE_MANIFEST_NAME = "industry_export_bundle.json"
MATERIALIZED_PPTX_NAME = "industry_executive_revised.pptx"
MATERIALIZED_XLSX_NAME = "industry_data_revised.xlsx"
MATERIALIZED_PORTFOLIO_XLSX_NAME = "carteira_101_flagships.xlsx"
MATERIALIZED_TOP100_XLSX_NAME = "top100_fidcs_middle_market.xlsx"
MATERIALIZED_HTML_NAME = "provider_flows_explorer.html"
BUNDLE_SCHEMA = "fidc_revision_export_bundle_v5"
PAYLOAD_SCHEMA = "fidc_revision_artifact_payload_v10"
TOP100_PLUS2_ADDITIONAL_CNPJS = {"44302112000172", "61669748000176"}
ISSUANCE_TAXONOMY_TABLE_DIMENSIONS: tuple[tuple[int, int], ...] = ((6, 8),)
STRUCTURAL_MVP_SLIDE_SEQUENCE: tuple[tuple[str, ...], ...] = (
    ("risco estrutural", "financeiro", "carteira i"),
    ("risco estrutural", "adquirencia", "carteira i"),
    ("risco estrutural", "agro / revenda", "carteira i"),
    ("risco estrutural", "risco corporativo", "carteira i"),
    ("risco estrutural", "consignado inss e fgts", "carteira i"),
    ("risco estrutural", "factoring", "carteira i"),
)
TYPE_RANKING_SLIDE_SEQUENCE: tuple[tuple[str, ...], ...] = (
    ("fomento mercantil", "crescimento marginal em seis meses", "jun/26", "top 15"),
    ("fomento mercantil", "crescimento marginal em seis meses", "dez/25", "top 15"),
    ("agro, industria e comercio", "maior salto absoluto", "jun/26", "top 15"),
    ("agro, industria e comercio", "maior salto absoluto", "dez/25", "top 15"),
    ("financeiro", "maior bloco", "ainda crescendo", "jun/26", "top 15"),
    ("financeiro", "maior bloco", "ainda crescendo", "dez/25", "top 15"),
    ("outros", "unico bloco que encolheu", "jun/26", "top 15"),
    ("outros", "unico bloco que encolheu", "dez/25", "top 15"),
)
CURRENT_TOP15_SLIDE_SEQUENCE: tuple[tuple[str, ...], ...] = (
    (
        "top 15",
        "ibba esteve em 8 das 15 maiores ofertas do semestre",
        "liderou 5 delas",
        "jan–jun/26",
    ),
    (
        "top 15",
        "as 15 maiores ofertas de 2025 mantem a base anual de comparacao",
        "2025fy",
    ),
)
HISTORICAL_TOP15_SLIDE_SEQUENCE: tuple[tuple[str, ...], ...] = (
    (
        "top 15",
        "historico",
        "maiores ofertas de 2024",
        "1/2",
        "agencia",
        "rating",
    ),
    (
        "top 15",
        "historico",
        "maiores ofertas de 2024",
        "2/2",
        "agencia",
        "rating",
    ),
    (
        "top 15",
        "historico",
        "maiores ofertas de 2023",
        "1/2",
        "agencia",
        "rating",
    ),
    (
        "top 15",
        "historico",
        "maiores ofertas de 2023",
        "2/2",
        "agencia",
        "rating",
    ),
)
HISTORICAL_TOP15_TABLE_DIMENSIONS: tuple[tuple[int, int], ...] = (
    (9, 12),
    (8, 12),
    (9, 12),
    (8, 12),
)
EXPECTED_SLIDE_SEQUENCE: tuple[tuple[str, ...], ...] = (
    ("industria de fidcs", "ago-26", "dados de referencia: jun-26"),
    ("escala da industria", "r$ 821,0 bi", "r$ 13,780 tri"),
    ("emissoes", "fidcs seguem ganhando escala nas emissoes", "1s26 ytd yoy"),
    ("saldo e tipos de fidcs", "financeiros dominam saldo e novas emissoes"),
    ("emissoes por categoria anbima", "emissoes por setor", "total emitido"),
    ("abrir", "outros", "revela que 63% do mercado e credito financeiro"),
    ("adquirencia e r$ 99 bi", "33 cnpjs reclassificados"),
    ("financeiro explicou 70% do crescimento da carteira",),
    ("ranking", "top 20 fidcs"),
    *TYPE_RANKING_SLIDE_SEQUENCE,
    *STRUCTURAL_MVP_SLIDE_SEQUENCE,
    ("emissoes crescem 15% no semestre",),
    ("22 ofertas concentram 42% de todo o volume",),
    ("garantia firme", "yoy ytd", "melhores esforcos repr."),
    *CURRENT_TOP15_SLIDE_SEQUENCE,
    *HISTORICAL_TOP15_SLIDE_SEQUENCE,
    ("o que muda a leitura do mercado",),
    ("qi lidera administracao", "btg lidera gestao e custodia"),
    ("prestadores", "ranking e concentracao"),
    ("quase todo o volume vai para o investidor profissional",),
    ("distribuicao por numero de cotistas",),
)
EXPECTED_SLIDES = len(EXPECTED_SLIDE_SEQUENCE)


def _contract_slide_numbers(
    sequence: tuple[tuple[str, ...], ...],
) -> tuple[int, ...]:
    """Resolve ordinal positions from the shared editorial contract."""

    return tuple(EXPECTED_SLIDE_SEQUENCE.index(tokens) + 1 for tokens in sequence)


CURRENT_TOP15_SLIDE_NUMBERS = _contract_slide_numbers(
    CURRENT_TOP15_SLIDE_SEQUENCE
)
HISTORICAL_TOP15_SLIDE_NUMBERS = _contract_slide_numbers(
    HISTORICAL_TOP15_SLIDE_SEQUENCE
)
if len(STRUCTURAL_MVP_SLIDE_SEQUENCE) != 6:
    raise RuntimeError("capítulo MVP de risco estrutural deve conter seis slides")
if len(HISTORICAL_TOP15_TABLE_DIMENSIONS) != len(
    HISTORICAL_TOP15_SLIDE_SEQUENCE
) or any(
    sum(rows - 1 for rows, _ in year_dimensions) != 15
    for year_dimensions in (
        HISTORICAL_TOP15_TABLE_DIMENSIONS[:2],
        HISTORICAL_TOP15_TABLE_DIMENSIONS[2:],
    )
):
    raise RuntimeError("contrato das tabelas históricas Top 15 está inconsistente")
BLOCKED_PPTX_AUDIENCE_COPY: tuple[str, ...] = (
    "clique para inserir",
    "click to add",
    "atualizar para",
    "copilot",
    "claude code",
    "prompt antigo",
)
BLOCKED_PUBLISHED_TEXT_MARKERS: tuple[str, ...] = (
    "√",
    "\ufffd",
)
REQUIRED_WORKBOOK_SHEETS = {
    "QA Inadimplência",
    "Base por fundo-CNPJ",
    "Base competência-CNPJ",
    "Checks revisão",
    "Concentração de monoestruturas",
    "Market share por subtipo",
    "Top 20 FIDCs",
    "Top 20 Outros",
    "Curadoria Top 20",
    "Curadoria flagship",
    "Carteira 1 curadoria",
    "Carteira 1 vs flagships",
    "Risco estrutural ativos",
    "Risco estrutural taxonomia",
    "Carteira 1 evolução",
    "Taxonomia por nível",
    "Comparativos históricos",
    "Ranking prestadores",
    "Inadimplência por recebível",
    "Histórico inad. coorte",
    "Reconciliação Tabelas I-II",
    "Ranking independentes",
    "FIDCs por banco",
    "Detalhe coorte bancos",
    "Atribuição prestadores",
    "Fluxos prestadores",
    "Migração CBSF",
    "Taxonomia adquirência",
    "Adquirência reclass.",
    "Curadoria Cartão",
    "Top 20 por Tipo ANBIMA",
    "Auditoria Top 20 Tipo",
    "Curadoria Outros Top 100",
    "Dispersão inadimplência",
    "Ofertas encerradas",
    "Comparativo renda fixa",
    "Regime de colocação",
    "Histograma ofertas",
    "Crédito Privado Ampliado",
    "Originadores 2026",
    "Top 15 ofertas",
    "Auditoria emissões",
    "Remuneração-alvo",
    "Cobertura emissões",
    "Curadoria perfis",
    "Validação emissões",
    "Emissões por categoria",
    "Público-alvo ofertas",
    "Principais conclusões",
    "Curadoria Atlântico",
    "Série Atlântico",
    "Cedentes · Leia-me",
    "Cedentes · Top 437",
    "Cedentes · Cobertura",
    "Taxonomia · de-para",
    "Taxonomia · Outros",
    "Taxonomia · impacto",
    "Universo elegível",
    "FICs excluídos",
    "Decisões do ledger",
}
REVISION_EMISSION_AUDIT_REQUIRED_HEADERS = frozenset(
    {
        "Preço unitário por tipo de cota",
        "Fonte preço",
        "Remuneração-alvo por tipo de cota",
        "Fonte remuneração",
    }
)
REVISION_EMISSION_COVERAGE_TARGET_LABEL = "Remuneração-alvo"
REQUIRED_PORTFOLIO_WORKBOOK_SHEETS = {
    "Leia-me",
    "Carteira 101",
    "Casos 99",
    "Nomes editáveis",
    "Flagships",
    "Cobertura e lacunas",
    "Dicionário",
    "Fontes manuais",
    "Preços por cota",
    "Auditoria documental",
    "Evidências documentais",
    "Cobertura varredura",
    "Dicionário de campos",
}
PORTFOLIO_WORKBOOK_MINIMUM_HEADERS = frozenset(
    {
        "CNPJ",
        "Nome completo do fundo (CVM)",
        "PL atual",
        "Sub / PL atual",
        "Mínimo Jr literal",
        "Mínimo Jr calculado*",
        "Mínimo Jr ajustado*",
        "Suporte total*",
        "Suporte Jr + Mezanino*",
        "Índice estrutural usado",
        "Folga / falta",
        "Preço por cota · leitura",
        "Originador*",
        "Cedente*",
        "Sacado / devedor*",
        "Tipo de recebível*",
        "Categoria de risco atual",
        "Categoria de risco proposta",
        "Subtipo de risco diagnosticado",
        "Middle Market · status",
        "Fonte documental",
        "Status do preenchimento",
    }
)
PORTFOLIO_WORKBOOK_EXPECTED_ROWS = {
    "Carteira 101": 101,
    "Casos 99": 99,
    "Flagships": 47,
}


class RevisionExportUnavailable(RuntimeError):
    """Raised when the published revision bundle is missing or inconsistent."""


@dataclass(frozen=True)
class RevisionExportStatus:
    payload_path: str
    payload_exists: bool
    payload_schema: str
    latest_complete: str
    bundle_manifest_path: str
    bundle_exists: bool
    bundle_id: str
    bundle_valid: bool
    validation_error: str
    pptx_path: str
    pptx_exists: bool
    xlsx_path: str
    xlsx_exists: bool
    portfolio_xlsx_path: str
    portfolio_xlsx_exists: bool
    top100_xlsx_path: str
    top100_xlsx_exists: bool
    html_path: str
    html_exists: bool
    artifact_runtime_available: bool

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class _ValidatedBundle:
    manifest: dict[str, object]
    pptx_path: Path
    pptx_bytes: bytes
    xlsx_path: Path
    xlsx_bytes: bytes
    portfolio_xlsx_path: Path
    portfolio_xlsx_bytes: bytes
    top100_xlsx_path: Path
    top100_xlsx_bytes: bytes
    html_path: Path
    html_bytes: bytes


def revision_dir(data_dir: Path = DEFAULT_DATA_DIR) -> Path:
    return Path(data_dir).resolve() / "generated_revision"


def revision_payload_path(data_dir: Path = DEFAULT_DATA_DIR) -> Path:
    return revision_dir(data_dir) / PAYLOAD_NAME


def revision_bundle_manifest_path(data_dir: Path = DEFAULT_DATA_DIR) -> Path:
    configured = os.environ.get("FIDC_EXPORT_MANIFEST", "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    return revision_dir(data_dir) / BUNDLE_MANIFEST_NAME


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _valid_zip(payload: bytes, required_member: str) -> bool:
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            return required_member in archive.namelist()
    except (OSError, zipfile.BadZipFile):
        return False


def _chart_members(archive: zipfile.ZipFile) -> list[str]:
    return [
        name
        for name in archive.namelist()
        if "/charts/chart" in name and name.endswith(".xml")
    ]


def _validate_no_mojibake_text(value: str, artifact_label: str) -> None:
    marker = next(
        (
            candidate
            for candidate in BLOCKED_PUBLISHED_TEXT_MARKERS
            if candidate in value
        ),
        None,
    )
    if marker is None:
        match = re.search(r"(?:Ã|Â)[\u0080-\u00bf]", value)
        marker = match.group(0) if match is not None else None
    if marker is not None:
        raise RevisionExportUnavailable(
            f"{artifact_label} contém texto corrompido ou ilegível: {marker}"
        )


def _validate_no_mojibake_office_archive(
    archive: zipfile.ZipFile,
    artifact_label: str,
) -> None:
    for name in archive.namelist():
        if not name.endswith((".xml", ".rels")):
            continue
        raw = archive.read(name)
        try:
            raw.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise RevisionExportUnavailable(
                f"{artifact_label} contém parte OOXML fora de UTF-8: {name}"
            ) from exc
        try:
            root = ElementTree.fromstring(raw)
        except ElementTree.ParseError as exc:
            raise RevisionExportUnavailable(
                f"{artifact_label} contém parte OOXML inválida: {name}"
            ) from exc
        values: list[str] = []
        for element in root.iter():
            values.extend(
                value
                for value in (
                    element.text,
                    element.tail,
                    *element.attrib.values(),
                )
                if value
            )
        try:
            _validate_no_mojibake_text(" ".join(values), artifact_label)
        except RevisionExportUnavailable as exc:
            raise RevisionExportUnavailable(f"{exc} em {name}") from exc


def _normalize_office_text(value: str) -> str:
    """Return case- and accent-insensitive Office text with compact spacing."""

    normalized = unicodedata.normalize("NFKD", value.casefold())
    return " ".join(
        "".join(
            character
            for character in normalized
            if not unicodedata.combining(character)
        ).split()
    )


def _normalized_slide_text(payload: bytes) -> str:
    """Return normalized text from one slide or notes XML part."""

    try:
        root = ElementTree.fromstring(payload)
    except ElementTree.ParseError:
        return ""
    visible = " ".join(
        node.text or ""
        for node in root.iter()
        if node.tag.endswith("}t")
    )
    return _normalize_office_text(visible)


def _slide_xml_containing(
    archive: zipfile.ZipFile,
    *tokens: str,
) -> bytes:
    expected = [_normalize_office_text(token) for token in tokens]
    for name in sorted(
        (
            item
            for item in archive.namelist()
            if item.startswith("ppt/slides/slide")
            and item.endswith(".xml")
            and "/_rels/" not in item
        ),
        key=lambda item: int(Path(item).stem.removeprefix("slide")),
    ):
        payload = archive.read(name)
        visible = _normalized_slide_text(payload)
        if all(token in visible for token in expected):
            return payload
    raise RevisionExportUnavailable(
        "PPTX revisado sem slide esperado: " + " / ".join(tokens)
    )


def _validate_no_blocked_audience_copy(archive: zipfile.ZipFile) -> None:
    """Reject production placeholders and stale instructions in slides or notes."""

    blocked = tuple(_normalize_office_text(value) for value in BLOCKED_PPTX_AUDIENCE_COPY)
    audience_parts = sorted(
        name
        for name in archive.namelist()
        if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
        or re.fullmatch(r"ppt/notesSlides/notesSlide\d+\.xml", name)
    )
    for name in audience_parts:
        text = _normalized_slide_text(archive.read(name))
        for phrase in blocked:
            if phrase in text:
                raise RevisionExportUnavailable(
                    "PPTX revisado contém placeholder ou instrução antiga "
                    f"em {name}: {phrase}"
                )


_PML = "http://schemas.openxmlformats.org/presentationml/2006/main"
_DML = "http://schemas.openxmlformats.org/drawingml/2006/main"
_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def _ordered_slide_parts(archive: zipfile.ZipFile) -> list[str]:
    presentation = ElementTree.fromstring(archive.read("ppt/presentation.xml"))
    relationships = ElementTree.fromstring(
        archive.read("ppt/_rels/presentation.xml.rels")
    )
    targets = {
        node.attrib.get("Id", ""): node.attrib.get("Target", "")
        for node in relationships.findall(f"{{{_PKG_REL}}}Relationship")
        if node.attrib.get("Type", "").endswith("/slide")
    }
    ordered: list[str] = []
    slide_ids = presentation.findall(f".//{{{_PML}}}sldId")
    if len({node.attrib.get("id") for node in slide_ids}) != len(slide_ids):
        raise RevisionExportUnavailable("PPTX revisado contém sldId duplicado")
    for node in slide_ids:
        relation_id = node.attrib.get(f"{{{_DOC_REL}}}id", "")
        target = targets.get(relation_id, "")
        if not target:
            raise RevisionExportUnavailable(
                "PPTX revisado contém relação de slide ausente"
            )
        part = (
            target.lstrip("/")
            if target.startswith("/")
            else posixpath.normpath(posixpath.join("ppt", target))
        )
        slide_root = ElementTree.fromstring(archive.read(part))
        if str(slide_root.attrib.get("show", "1")).casefold() in {"0", "false"}:
            raise RevisionExportUnavailable("PPTX revisado contém slide oculto")
        ordered.append(part)
    if len(set(ordered)) != len(ordered):
        raise RevisionExportUnavailable("PPTX revisado contém relação de slide duplicada")
    physical = {
        name
        for name in archive.namelist()
        if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
    }
    if set(ordered) != physical:
        raise RevisionExportUnavailable(
            "PPTX revisado contém slide órfão ou fora da sequência"
        )
    return ordered


def _xfrm_bbox(node: ElementTree.Element | None) -> tuple[int, int, int, int] | None:
    if node is None:
        return None
    offset = node.find(f"{{{_DML}}}off")
    extent = node.find(f"{{{_DML}}}ext")
    if offset is None or extent is None:
        return None
    return (
        int(offset.attrib["x"]),
        int(offset.attrib["y"]),
        int(extent.attrib["cx"]),
        int(extent.attrib["cy"]),
    )


def _overlap(
    left: tuple[int, int, int, int],
    right: tuple[int, int, int, int],
) -> bool:
    left_x, left_y, left_w, left_h = left
    right_x, right_y, right_w, right_h = right
    return (
        max(left_x, right_x) < min(left_x + left_w, right_x + right_w)
        and max(left_y, right_y) < min(left_y + left_h, right_y + right_h)
    )


def _validate_native_table_slide(
    archive: zipfile.ZipFile,
    slide_number: int,
    *,
    expected_dimensions: tuple[tuple[int, int], ...],
    canvas: tuple[int, int],
) -> None:
    root = ElementTree.fromstring(
        archive.read(f"ppt/slides/slide{slide_number}.xml")
    )
    tables: list[tuple[tuple[int, int, int, int], tuple[int, int]]] = []
    for frame in root.findall(f".//{{{_PML}}}graphicFrame"):
        table = frame.find(f".//{{{_DML}}}tbl")
        if table is None:
            continue
        bbox = _xfrm_bbox(frame.find(f"{{{_PML}}}xfrm"))
        if bbox is None:
            raise RevisionExportUnavailable(
                f"slide {slide_number} contém tabela nativa sem posição"
            )
        rows = len(table.findall(f"{{{_DML}}}tr"))
        columns = len(table.findall(f"{{{_DML}}}tblGrid/{{{_DML}}}gridCol"))
        tables.append((bbox, (rows, columns)))
    if tuple(dimensions for _, dimensions in tables) != expected_dimensions:
        raise RevisionExportUnavailable(
            f"slide {slide_number} não contém as tabelas Office esperadas"
        )
    canvas_width, canvas_height = canvas
    for bbox, _ in tables:
        left, top, width, height = bbox
        if (
            left < 0
            or top < 0
            or width <= 0
            or height <= 0
            or left + width > canvas_width
            or top + height > canvas_height
        ):
            raise RevisionExportUnavailable(
                f"slide {slide_number} contém tabela fora do canvas"
            )
    shape_bboxes: list[tuple[int, int, int, int]] = []
    for shape in root.findall(f".//{{{_PML}}}sp"):
        bbox = _xfrm_bbox(shape.find(f"{{{_PML}}}spPr/{{{_DML}}}xfrm"))
        if bbox is not None:
            shape_bboxes.append(bbox)
    if any(
        _overlap(shape_bbox, table_bbox)
        for shape_bbox in shape_bboxes
        for table_bbox, _ in tables
    ):
        raise RevisionExportUnavailable(
            f"slide {slide_number} contém shape sobreposto à tabela nativa"
        )


def _native_table_text_rows(
    archive: zipfile.ZipFile,
    slide_number: int,
) -> tuple[tuple[str, ...], ...]:
    """Return the first native Office table as normalized cell text rows."""

    root = ElementTree.fromstring(
        archive.read(f"ppt/slides/slide{slide_number}.xml")
    )
    table = root.find(f".//{{{_DML}}}tbl")
    if table is None:
        return ()
    rows: list[tuple[str, ...]] = []
    for row in table.findall(f"{{{_DML}}}tr"):
        cells: list[str] = []
        for cell in row.findall(f"{{{_DML}}}tc"):
            text = " ".join(
                part.text or ""
                for part in cell.findall(f".//{{{_DML}}}t")
            )
            cells.append(" ".join(text.split()))
        rows.append(tuple(cells))
    return tuple(rows)


def _contains_blocked_rgb_color(
    xml_parts: Iterable[bytes],
    blocked_color: str,
) -> bool:
    """Match a blocked RGB value only in DrawingML color elements."""

    blocked = str(blocked_color).strip().removeprefix("#").upper()
    for xml in xml_parts:
        root = ElementTree.fromstring(xml)
        for element in root.iter():
            local_name = str(element.tag).rsplit("}", 1)[-1]
            if local_name not in {"srgbClr", "sysClr"}:
                continue
            values = (
                str(element.attrib.get("val") or "").removeprefix("#").upper(),
                str(element.attrib.get("lastClr") or "").removeprefix("#").upper(),
            )
            if blocked in values:
                return True
    return False


def validate_revision_pptx(payload: bytes) -> None:
    """Validate the visual contract directly in the exported OOXML."""

    if not _valid_zip(payload, "ppt/presentation.xml"):
        raise RevisionExportUnavailable("PPTX revisado inválido ou corrompido")
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        _validate_no_mojibake_office_archive(archive, "PPTX revisado")
        slides = [
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide")
            and name.endswith(".xml")
            and "/_rels/" not in name
        ]
        if len(slides) != EXPECTED_SLIDES:
            raise RevisionExportUnavailable(
                f"PPTX revisado deveria conter {EXPECTED_SLIDES} slides; contém {len(slides)}"
            )
        ordered_slides = _ordered_slide_parts(archive)
        if len(ordered_slides) != EXPECTED_SLIDES:
            raise RevisionExportUnavailable(
                f"sequência do PPTX deveria conter {EXPECTED_SLIDES} slides; contém {len(ordered_slides)}"
            )
        _validate_no_blocked_audience_copy(archive)
        for slide_number, (slide_path, expected_tokens) in enumerate(
            zip(ordered_slides, EXPECTED_SLIDE_SEQUENCE, strict=True),
            start=1,
        ):
            slide_text = _normalized_slide_text(archive.read(slide_path))
            missing_tokens = [
                token for token in expected_tokens if token not in slide_text
            ]
            if missing_tokens:
                raise RevisionExportUnavailable(
                    f"slide {slide_number} viola o contrato ordinal: "
                    + ", ".join(missing_tokens)
                )
        scale_slide = _slide_xml_containing(archive, "ESCALA DA INDÚSTRIA")
        if scale_slide.count(b"<c:chart") < 2:
            raise RevisionExportUnavailable(
                "slide de escala deve conter dois gráficos nativos do Office"
            )
        if "saldo fic" in _normalized_slide_text(scale_slide):
            raise RevisionExportUnavailable(
                "slide de escala voltou a exibir FIC no gráfico de PL"
            )
        scale_slide_root = ElementTree.fromstring(scale_slide)
        if str(scale_slide_root.attrib.get("show", "1")).casefold() in {"0", "false"}:
            raise RevisionExportUnavailable("slide de escala está oculto")
        presentation = ElementTree.fromstring(archive.read("ppt/presentation.xml"))
        slide_size = presentation.find(f"{{{_PML}}}sldSz")
        if slide_size is None:
            raise RevisionExportUnavailable("PPTX revisado sem dimensão de slide")
        canvas = (int(slide_size.attrib["cx"]), int(slide_size.attrib["cy"]))
        _validate_native_table_slide(
            archive,
            3,
            expected_dimensions=((7, 3),),
            canvas=canvas,
        )
        _validate_native_table_slide(
            archive,
            5,
            expected_dimensions=ISSUANCE_TAXONOMY_TABLE_DIMENSIONS,
            canvas=canvas,
        )
        for ranking_slide_number in range(10, 18):
            _validate_native_table_slide(
                archive,
                ranking_slide_number,
                expected_dimensions=((16, 8),),
                canvas=canvas,
            )
            ranking_text = _normalized_slide_text(
                archive.read(ordered_slides[ranking_slide_number - 1])
            )
            if "remuneracao-alvo" not in ranking_text:
                raise RevisionExportUnavailable(
                    f"slide {ranking_slide_number} deve exibir Remuneração-alvo"
                )
            if "preco por cota" in ranking_text:
                raise RevisionExportUnavailable(
                    f"slide {ranking_slide_number} ainda usa Preço por cota "
                    "como rótulo da rentabilidade-alvo"
                )
        for slide_number in CURRENT_TOP15_SLIDE_NUMBERS:
            _validate_native_table_slide(
                archive,
                slide_number,
                expected_dimensions=((16, 10),),
                canvas=canvas,
            )
        for slide_number, dimensions in zip(
            HISTORICAL_TOP15_SLIDE_NUMBERS,
            HISTORICAL_TOP15_TABLE_DIMENSIONS,
            strict=True,
        ):
            _validate_native_table_slide(
                archive,
                slide_number,
                expected_dimensions=(dimensions,),
                canvas=canvas,
            )
        office_xml_parts = [
            archive.read(name)
            for name in archive.namelist()
            if name.endswith(".xml") and (
                name.startswith("ppt/slides/")
                or name.startswith("ppt/theme/")
                or "/charts/chart" in name
            )
        ]
        if _contains_blocked_rgb_color(office_xml_parts, "172A3A"):
            raise RevisionExportUnavailable("PPTX revisado contém a cor navy bloqueada")
        chart_xml = b"".join(archive.read(name) for name in _chart_members(archive))
        if b'<c:smooth val="1"' in chart_xml or b'<c:smooth val="true"' in chart_xml:
            raise RevisionExportUnavailable("PPTX revisado contém linha suavizada")
        marker_tokens = chart_xml.replace(b" />", b"/>").split(b"<c:marker>")[1:]
        for token in marker_tokens:
            marker = token.split(b"</c:marker>", 1)[0]
            if b'<c:symbol val="none"' not in marker:
                raise RevisionExportUnavailable("PPTX revisado contém marker ativo")
        ranking_slide = _slide_xml_containing(
            archive, "QI LIDERA ADMINISTRAÇÃO", "BTG LIDERA GESTÃO E CUSTÓDIA"
        )
        if ranking_slide.count(b"<a:tbl>") != 0:
            raise RevisionExportUnavailable(
                "slide combinado de prestadores deve conter apenas gráficos"
            )
        if ranking_slide.count(b"<c:chart") < 6:
            raise RevisionExportUnavailable(
                "slide combinado de prestadores deve conter ao menos seis gráficos nativos do Office"
            )
        offers_slide = _slide_xml_containing(
            archive, "22 OFERTAS CONCENTRAM 42% DE TODO O VOLUME"
        )
        if offers_slide.count(b"<c:chart") < 3:
            raise RevisionExportUnavailable(
                "slide de distribuição de ofertas deve conter três gráficos nativos do Office"
            )
        placement_slide = _slide_xml_containing(
            archive,
            "GARANTIA FIRME",
            "MELHORES ESFORÇOS REPR.",
        )
        if placement_slide.count(b"<c:chart") < 4:
            raise RevisionExportUnavailable(
                "slide de volume e regime deve conter quatro gráficos nativos do Office"
            )
        combined_market_slide = _slide_xml_containing(
            archive, "FIDCS SEGUEM GANHANDO ESCALA NAS EMISSÕES"
        )
        if combined_market_slide.count(b"<c:chart") != 2:
            raise RevisionExportUnavailable(
                "slide conjunto CVM e ANBIMA deve conter dois gráficos nativos do Office"
            )
        if combined_market_slide.count(b"<a:tbl>") != 1:
            raise RevisionExportUnavailable(
                "slide conjunto CVM e ANBIMA deve conter uma tabela nativa"
            )
        stock_and_types_slide = _slide_xml_containing(
            archive,
            "SALDO E TIPOS DE FIDCS",
            "FINANCEIROS DOMINAM SALDO E NOVAS EMISSÕES",
        )
        if stock_and_types_slide.count(b"<c:chart") != 4:
            raise RevisionExportUnavailable(
                "slide de saldo e tipos deve conter quatro gráficos nativos do Office"
            )
        if stock_and_types_slide.count(b"<a:tbl>") != 0:
            raise RevisionExportUnavailable(
                "slide de saldo e tipos não deve conter tabela nativa"
            )
        taxonomy_market_slide = _slide_xml_containing(
            archive, "EMISSÕES POR CATEGORIA ANBIMA"
        )
        if taxonomy_market_slide.count(b"<c:chart") != 2:
            raise RevisionExportUnavailable(
                "slide de emissões por categoria deve conter dois gráficos nativos do Office"
            )
        if taxonomy_market_slide.count(b"<a:tbl>") != 1:
            raise RevisionExportUnavailable(
                "slide de emissões por categoria deve conter uma tabela nativa"
            )
        for slide_number in CURRENT_TOP15_SLIDE_NUMBERS:
            top15_offers_slide = archive.read(ordered_slides[slide_number - 1])
            if top15_offers_slide.count(b"<a:tbl>") != 1:
                raise RevisionExportUnavailable(
                    f"slide {slide_number} de maiores ofertas deve conter uma tabela nativa do Office"
                )
        provider_concentration_slide = _slide_xml_containing(
            archive,
            "PRESTADORES",
            "RANKING E CONCENTRAÇÃO",
        )
        if provider_concentration_slide.count(b"<c:chart") != 2:
            raise RevisionExportUnavailable(
                "slide de concentração de prestadores deve conter "
                "dois gráficos nativos do Office"
            )
        if provider_concentration_slide.count(b"<a:tbl>") != 0:
            raise RevisionExportUnavailable(
                "slide de concentração de prestadores não deve conter tabela nativa"
            )
        _slide_xml_containing(archive, "O QUE MUDA A LEITURA DO MERCADO")


def validate_revision_xlsx(payload: bytes) -> None:
    if not _valid_zip(payload, "xl/workbook.xml"):
        raise RevisionExportUnavailable("XLSX revisado inválido ou corrompido")
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            _validate_no_mojibake_office_archive(archive, "XLSX revisado")
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    except RevisionExportUnavailable:
        raise
    except Exception as exc:
        raise RevisionExportUnavailable("XLSX revisado não pôde ser lido") from exc
    try:
        missing = sorted(REQUIRED_WORKBOOK_SHEETS.difference(workbook.sheetnames))
        if missing:
            raise RevisionExportUnavailable(
                "XLSX revisado sem abas obrigatórias: " + ", ".join(missing)
            )

        audit_sheet = workbook["Auditoria emissões"]
        audit_headers = {
            str(cell.value or "").strip()
            for cell in next(audit_sheet.iter_rows(min_row=4, max_row=4), ())
        }
        missing_headers = sorted(
            REVISION_EMISSION_AUDIT_REQUIRED_HEADERS.difference(audit_headers)
        )
        if missing_headers:
            raise RevisionExportUnavailable(
                "Auditoria emissões não separa VNU de remuneração-alvo; faltam: "
                + ", ".join(missing_headers)
            )

        coverage_sheet = workbook["Cobertura emissões"]
        coverage_labels = [
            str(row[0].value or "").strip()
            for row in coverage_sheet.iter_rows(
                min_row=5,
                min_col=4,
                max_col=4,
            )
            if str(row[0].value or "").strip()
        ]
        target_count = coverage_labels.count(
            REVISION_EMISSION_COVERAGE_TARGET_LABEL
        )
        if target_count != 8:
            raise RevisionExportUnavailable(
                "Cobertura emissões deveria conter oito linhas de "
                f"{REVISION_EMISSION_COVERAGE_TARGET_LABEL}; contém {target_count}"
            )
        if "Preço por cota" in coverage_labels:
            raise RevisionExportUnavailable(
                "Cobertura emissões ainda trata VNU como rentabilidade-alvo"
            )
    finally:
        workbook.close()


def _validated_numeric_cnpj(value: object) -> str | None:
    """Return a checksum-valid 14-digit CNPJ stored as an Excel number."""

    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        digits = str(value)
    elif isinstance(value, float) and value.is_integer():
        digits = str(int(value))
    else:
        return None
    if not digits or len(digits) > 14:
        return None
    digits = digits.zfill(14)
    if len(set(digits)) == 1:
        return None

    def check_digit(base: str, weights: tuple[int, ...]) -> str:
        remainder = sum(
            int(character) * weight
            for character, weight in zip(base, weights, strict=True)
        ) % 11
        return "0" if remainder < 2 else str(11 - remainder)

    first = check_digit(
        digits[:12],
        (5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2),
    )
    second = check_digit(
        digits[:12] + first,
        (6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2),
    )
    return digits if digits[-2:] == first + second else None


def validate_revision_portfolio_xlsx(payload: bytes) -> None:
    """Validate the standalone Carteira 101 and flagship workbook."""

    if not _valid_zip(payload, "xl/workbook.xml"):
        raise RevisionExportUnavailable(
            "XLSX de Carteira 101 e Flagships inválido ou corrompido"
        )
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            _validate_no_mojibake_office_archive(
                archive,
                "XLSX de Carteira 101 e Flagships",
            )
            workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (KeyError, OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise RevisionExportUnavailable(
            "XLSX de Carteira 101 e Flagships inválido ou corrompido"
        ) from exc
    sheet_names = {
        str(node.attrib.get("name") or "")
        for node in workbook_root.iter()
        if node.tag.endswith("}sheet")
    }
    missing = sorted(REQUIRED_PORTFOLIO_WORKBOOK_SHEETS - sheet_names)
    if missing:
        raise RevisionExportUnavailable(
            "XLSX de Carteira 101 e Flagships sem abas obrigatórias: "
            + ", ".join(missing)
        )
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            chart_parts = _chart_members(archive)
            if not chart_parts:
                raise RevisionExportUnavailable(
                    "XLSX de Carteira 101 sem gráficos nativos do Office"
                )
            chart_xml = "\n".join(
                archive.read(name).decode("utf-8", errors="ignore")
                for name in chart_parts
            )
            if "Nomes editáveis" not in chart_xml:
                raise RevisionExportUnavailable(
                    "rótulos dos gráficos não referenciam a aba Nomes editáveis"
                )
    except (OSError, zipfile.BadZipFile, KeyError) as exc:
        raise RevisionExportUnavailable(
            "gráficos do XLSX de Carteira 101 não puderam ser validados"
        ) from exc

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise RevisionExportUnavailable(
            "XLSX de Carteira 101 e Flagships não pôde ser lido"
        ) from exc

    try:
        for sheet_name, expected_rows in PORTFOLIO_WORKBOOK_EXPECTED_ROWS.items():
            sheet = workbook[sheet_name]
            header_cells = next(
                sheet.iter_rows(min_row=4, max_row=4),
                (),
            )
            headers = tuple(
                str(cell.value or "").strip() for cell in header_cells
            )
            missing_headers = sorted(
                PORTFOLIO_WORKBOOK_MINIMUM_HEADERS - set(headers)
            )
            if missing_headers:
                raise RevisionExportUnavailable(
                    f"aba {sheet_name} sem cabeçalhos obrigatórios: "
                    + ", ".join(missing_headers)
                )

            cnpj_column = headers.index("CNPJ")
            data_rows = [
                row
                for row in sheet.iter_rows(
                    min_row=5,
                    max_col=len(headers),
                )
                if any(cell.value not in (None, "") for cell in row)
            ]
            if len(data_rows) != expected_rows:
                raise RevisionExportUnavailable(
                    f"aba {sheet_name} deveria conter {expected_rows} linhas; "
                    f"contém {len(data_rows)}"
                )

            cnpjs: list[str] = []
            for row_number, row in enumerate(data_rows, start=5):
                cell = row[cnpj_column]
                cnpj = _validated_numeric_cnpj(cell.value)
                if cnpj is None:
                    raise RevisionExportUnavailable(
                        f"aba {sheet_name} contém CNPJ numérico inválido "
                        f"na linha {row_number}"
                    )
                cnpjs.append(cnpj)
            duplicated = sorted(
                {
                    cnpj
                    for cnpj in cnpjs
                    if cnpjs.count(cnpj) > 1
                }
            )
            if duplicated:
                raise RevisionExportUnavailable(
                    f"aba {sheet_name} contém CNPJ duplicado: "
                    + ", ".join(duplicated)
                )
    finally:
        workbook.close()


def validate_revision_top100_xlsx(payload: bytes) -> None:
    """Validate the standalone global Top 100 plus two 2026 issuances."""

    if not _valid_zip(payload, "xl/workbook.xml"):
        raise RevisionExportUnavailable("XLSX Top 100 inválido ou corrompido")
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            _validate_no_mojibake_office_archive(archive, "XLSX Top 100 + 2")
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except RevisionExportUnavailable:
        raise
    except Exception as exc:
        raise RevisionExportUnavailable("XLSX Top 100 não pôde ser lido") from exc
    try:
        required = {"Leia-me", "Top 100 FIDCs"}
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise RevisionExportUnavailable(
                "XLSX Top 100 sem abas obrigatórias: " + ", ".join(sorted(missing))
            )
        sheet = workbook["Top 100 FIDCs"]
        headers = tuple(
            str(cell.value or "").strip()
            for cell in next(sheet.iter_rows(min_row=4, max_row=4), ())
        )
        required_headers = {
            "Ordem do export",
            "Rank geral por PL",
            "Critério de inclusão",
            "CNPJ",
            "Nome completo do fundo (CVM)",
            "PL",
            "Sub / PL atual",
            "Mínimo de Sub Jr",
            "Mínimo estrutural",
            "Preço inicial por cota",
            "Cedente / originador",
            "Sacado / devedor",
            "Tipo de recebível",
            "Tipo ANBIMA oficial",
            "Taxonomia funcional N1",
            "Middle Market · status",
            "Fonte",
        }
        missing_headers = required_headers.difference(headers)
        if missing_headers:
            raise RevisionExportUnavailable(
                "XLSX Top 100 sem cabeçalhos obrigatórios: "
                + ", ".join(sorted(missing_headers))
            )
        rows = [
            row
            for row in sheet.iter_rows(min_row=5, max_col=len(headers))
            if any(cell.value not in (None, "") for cell in row)
        ]
        if len(rows) != 102:
            raise RevisionExportUnavailable(
                f"XLSX Top 100 + 2 deveria conter 102 linhas; contém {len(rows)}"
            )
        cnpj_column = headers.index("CNPJ")
        cnpjs = [_validated_numeric_cnpj(row[cnpj_column].value) for row in rows]
        if any(value is None for value in cnpjs) or len(set(cnpjs)) != 102:
            raise RevisionExportUnavailable(
                "XLSX Top 100 + 2 deve conter 102 CNPJs numéricos válidos e únicos"
            )
        if not TOP100_PLUS2_ADDITIONAL_CNPJS.issubset(set(cnpjs)):
            raise RevisionExportUnavailable(
                "XLSX Top 100 + 2 não contém Citi-Bayer e Lavoro"
            )
        cnpj_formats = {
            str(row[cnpj_column].number_format or "") for row in rows
        }
        if cnpj_formats != {"00000000000000"}:
            raise RevisionExportUnavailable(
                "XLSX Top 100 + 2 deve exibir CNPJ com máscara de 14 dígitos"
            )
    finally:
        workbook.close()


def validate_revision_html(payload: bytes) -> None:
    """Validate the self-contained provider-flow explorer served by the app."""

    if not payload:
        raise RevisionExportUnavailable("HTML interativo de fluxos está vazio")
    if len(payload) > 2 * 1024 * 1024:
        raise RevisionExportUnavailable("HTML interativo de fluxos excede 2 MB")
    try:
        document = payload.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise RevisionExportUnavailable(
            "HTML interativo de fluxos não está em UTF-8"
        ) from exc
    _validate_no_mojibake_text(document, "HTML interativo de fluxos")
    required_tokens = (
        "<!doctype html",
        'id="provider-flow-explorer"',
        "data-chart",
        "<script",
        "Dez/24",
        "Administração",
        "Gestão",
        "Custódia",
        "CBSF / REAG",
        "Carteira 1 · evolução pela taxonomia reclassificada",
        "Carteira 1 vs. 47 CNPJs flagship",
        "flagship_curation_compact_v2",
        "carteira_1_curation_compact_v4",
        "carteira_1_taxonomy_compact_v1",
    )
    missing = [
        token
        for token in required_tokens
        if token.casefold() not in document.casefold()
    ]
    if missing:
        raise RevisionExportUnavailable(
            "HTML interativo de fluxos incompleto: " + ", ".join(missing)
        )
    if "fetch(" in document:
        raise RevisionExportUnavailable(
            "HTML interativo de fluxos depende de carregamento externo"
        )


def _candidate_paths(
    data_dir: Path,
    *,
    materialized_name: str,
    output_name: str,
    env_name: str,
) -> tuple[Path, ...]:
    explicit = os.environ.get(env_name, "").strip()
    candidates = [
        revision_dir(data_dir) / materialized_name,
        ROOT / "outputs" / output_name,
    ]
    if explicit:
        candidates.insert(0, Path(explicit).expanduser().resolve())
    return tuple(dict.fromkeys(path.resolve() for path in candidates))


def revision_pptx_candidates(data_dir: Path = DEFAULT_DATA_DIR) -> tuple[Path, ...]:
    return _candidate_paths(
        Path(data_dir),
        materialized_name=MATERIALIZED_PPTX_NAME,
        output_name="Industria_FIDC_Executivo_202607_revisado.pptx",
        env_name="FIDC_REVISION_PPTX",
    )


def revision_xlsx_candidates(data_dir: Path = DEFAULT_DATA_DIR) -> tuple[Path, ...]:
    return _candidate_paths(
        Path(data_dir),
        materialized_name=MATERIALIZED_XLSX_NAME,
        output_name="Industria_FIDC_Dados_202607_revisado.xlsx",
        env_name="FIDC_REVISION_XLSX",
    )


def revision_portfolio_xlsx_candidates(
    data_dir: Path = DEFAULT_DATA_DIR,
) -> tuple[Path, ...]:
    return _candidate_paths(
        Path(data_dir),
        materialized_name=MATERIALIZED_PORTFOLIO_XLSX_NAME,
        output_name=MATERIALIZED_PORTFOLIO_XLSX_NAME,
        env_name="FIDC_REVISION_PORTFOLIO_XLSX",
    )


def revision_top100_xlsx_candidates(
    data_dir: Path = DEFAULT_DATA_DIR,
) -> tuple[Path, ...]:
    return _candidate_paths(
        Path(data_dir),
        materialized_name=MATERIALIZED_TOP100_XLSX_NAME,
        output_name=MATERIALIZED_TOP100_XLSX_NAME,
        env_name="FIDC_REVISION_TOP100_XLSX",
    )


def revision_html_candidates(data_dir: Path = DEFAULT_DATA_DIR) -> tuple[Path, ...]:
    return _candidate_paths(
        Path(data_dir),
        materialized_name=MATERIALIZED_HTML_NAME,
        output_name="Industria_FIDC_Fluxos_Prestadores_202607.html",
        env_name="FIDC_REVISION_HTML",
    )


def _artifact_node_modules() -> Path | None:
    candidates: list[Path] = []
    configured = os.environ.get("CODEX_NODE_MODULES", "").strip()
    if configured:
        candidates.append(Path(configured).expanduser())
    candidates.extend(
        [
            ROOT / "node_modules",
            Path.home()
            / ".cache"
            / "codex-runtimes"
            / "codex-primary-runtime"
            / "dependencies"
            / "node"
            / "node_modules",
        ]
    )
    for candidate in candidates:
        if (candidate / "@oai" / "artifact-tool" / "package.json").exists():
            return candidate.resolve()
    return None


def artifact_runtime_available() -> bool:
    """Diagnostic only; the application request path never invokes the runtime."""

    return bool(shutil.which("node") and _artifact_node_modules())


def _payload_metadata(path: Path) -> tuple[str, str]:
    if not path.exists():
        return "", ""
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return "", ""
    return str(payload.get("schema_version") or ""), str(payload.get("latest_complete") or "")


def _taxonomy_review_ledger_path(
    data_dir: Path,
    payload: dict[str, object] | None,
) -> Path:
    """Resolve the payload-declared ledger inside the selected industry data dir."""

    meta = dict((payload or {}).get("taxonomy_review_meta") or {})
    configured = str(
        meta.get("ledger_path") or "taxonomy_review_actions.csv"
    ).strip()
    raw_path = Path(configured).expanduser()
    if raw_path.is_absolute():
        return raw_path.resolve()

    parts = raw_path.parts
    if "industry_study" in parts:
        anchor = parts.index("industry_study")
        suffix = parts[anchor + 1 :]
        raw_path = Path(*suffix) if suffix else Path("taxonomy_review_actions.csv")

    resolved_data_dir = Path(data_dir).resolve()
    resolved = (resolved_data_dir / raw_path).resolve()
    try:
        resolved.relative_to(resolved_data_dir)
    except ValueError as exc:
        raise RevisionExportUnavailable(
            "caminho do ledger de taxonomia está fora do diretório de dados"
        ) from exc
    return resolved


def _taxonomy_review_audit_path(
    data_dir: Path,
    payload: dict[str, object] | None,
) -> Path:
    """Resolve the payload-declared audit inside the selected industry data dir."""

    meta = dict((payload or {}).get("taxonomy_review_meta") or {})
    configured = str(
        meta.get("audit_path") or "taxonomy_review_audit.csv"
    ).strip()
    raw_path = Path(configured).expanduser()
    if raw_path.is_absolute():
        return raw_path.resolve()
    parts = raw_path.parts
    if "industry_study" in parts:
        anchor = parts.index("industry_study")
        suffix = parts[anchor + 1 :]
        raw_path = Path(*suffix) if suffix else Path("taxonomy_review_audit.csv")
    resolved_data_dir = Path(data_dir).resolve()
    resolved = (resolved_data_dir / raw_path).resolve()
    try:
        resolved.relative_to(resolved_data_dir)
    except ValueError as exc:
        raise RevisionExportUnavailable(
            "caminho da auditoria de taxonomia está fora do diretório de dados"
        ) from exc
    return resolved


def _taxonomy_review_signature(data_dir: Path, payload: dict[str, object] | None) -> str:
    ledger_path = _taxonomy_review_ledger_path(data_dir, payload)
    audit_path = _taxonomy_review_audit_path(data_dir, payload)
    return (
        f"{taxonomy_review_ledger_digest(ledger_path)}:"
        f"{taxonomy_review_audit_digest(audit_path)}"
    )


def _matching_candidate(
    paths: Iterable[Path],
    expected: dict[str, object],
    validator: Callable[[bytes], None],
) -> tuple[Path, bytes]:
    expected_hash = str(expected.get("sha256") or "")
    expected_size = int(expected.get("bytes") or 0)
    for path in paths:
        if not path.exists():
            continue
        payload = path.read_bytes()
        if expected_size and len(payload) != expected_size:
            continue
        if not expected_hash or _sha256(payload) != expected_hash:
            continue
        validator(payload)
        return path, payload
    raise RevisionExportUnavailable("arquivo publicado não corresponde ao hash do bundle")


def _load_validated_bundle(data_dir: Path = DEFAULT_DATA_DIR) -> _ValidatedBundle:
    data_dir = Path(data_dir).resolve()
    payload_path = revision_payload_path(data_dir)
    manifest_path = revision_bundle_manifest_path(data_dir)
    if not payload_path.exists():
        raise RevisionExportUnavailable(f"payload revisado ausente: {payload_path}")
    if not manifest_path.exists():
        raise RevisionExportUnavailable(f"manifest do bundle ausente: {manifest_path}")
    payload_raw = payload_path.read_bytes()
    try:
        payload = json.loads(payload_raw)
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RevisionExportUnavailable(f"bundle revisado ilegível: {exc}") from exc
    if manifest.get("schema_version") != BUNDLE_SCHEMA:
        raise RevisionExportUnavailable("schema do bundle revisado incompatível")
    payload_hash = _sha256(payload_raw)
    if manifest.get("payload_sha256") != payload_hash:
        raise RevisionExportUnavailable("payload mudou após a publicação do bundle")
    if manifest.get("source_signature") != payload_hash:
        raise RevisionExportUnavailable("assinatura de fontes do bundle não reconcilia")
    if manifest.get("payload_schema") != payload.get("schema_version"):
        raise RevisionExportUnavailable("schema do payload diverge do bundle")
    if payload.get("schema_version") != PAYLOAD_SCHEMA:
        raise RevisionExportUnavailable("schema do payload revisado incompatível")
    if manifest.get("latest_complete") != payload.get("latest_complete"):
        raise RevisionExportUnavailable("competência do bundle diverge do payload")
    taxonomy_meta = dict(payload.get("taxonomy_review_meta") or {})
    if taxonomy_meta:
        published_taxonomy_digest = str(
            taxonomy_meta.get("ledger_sha256") or ""
        ).strip()
        published_audit_digest = str(
            taxonomy_meta.get("audit_sha256") or ""
        ).strip()
        if not published_taxonomy_digest or not published_audit_digest:
            raise RevisionExportUnavailable(
                "payload revisado não registra os hashes da curadoria de taxonomia"
            )
        ledger_path = _taxonomy_review_ledger_path(data_dir, payload)
        audit_path = _taxonomy_review_audit_path(data_dir, payload)
        try:
            assert_taxonomy_review_ledger_matches_audit(
                ledger_path,
                audit_path,
            )
        except ValueError as exc:
            raise RevisionExportUnavailable(
                "curadoria ou auditoria de Outros mudou após a publicação; "
                "ledger e auditoria são inconsistentes: "
                f"{exc}"
            ) from exc
        if (
            taxonomy_review_ledger_digest(ledger_path)
            != published_taxonomy_digest
            or taxonomy_review_audit_digest(audit_path)
            != published_audit_digest
        ):
            raise RevisionExportUnavailable(
                "curadoria ou auditoria de Outros mudou após a publicação; regenere o bundle"
            )
    pptx_path, pptx_bytes = _matching_candidate(
        revision_pptx_candidates(data_dir),
        dict(manifest.get("pptx") or {}),
        validate_revision_pptx,
    )
    xlsx_path, xlsx_bytes = _matching_candidate(
        revision_xlsx_candidates(data_dir),
        dict(manifest.get("xlsx") or {}),
        validate_revision_xlsx,
    )
    portfolio_xlsx_path, portfolio_xlsx_bytes = _matching_candidate(
        revision_portfolio_xlsx_candidates(data_dir),
        dict(manifest.get("portfolio_xlsx") or {}),
        validate_revision_portfolio_xlsx,
    )
    top100_xlsx_path, top100_xlsx_bytes = _matching_candidate(
        revision_top100_xlsx_candidates(data_dir),
        dict(manifest.get("top100_xlsx") or {}),
        validate_revision_top100_xlsx,
    )
    html_path, html_bytes = _matching_candidate(
        revision_html_candidates(data_dir),
        dict(manifest.get("html") or {}),
        validate_revision_html,
    )
    return _ValidatedBundle(
        manifest=manifest,
        pptx_path=pptx_path,
        pptx_bytes=pptx_bytes,
        xlsx_path=xlsx_path,
        xlsx_bytes=xlsx_bytes,
        portfolio_xlsx_path=portfolio_xlsx_path,
        portfolio_xlsx_bytes=portfolio_xlsx_bytes,
        top100_xlsx_path=top100_xlsx_path,
        top100_xlsx_bytes=top100_xlsx_bytes,
        html_path=html_path,
        html_bytes=html_bytes,
    )


def revision_export_signature(data_dir: Path = DEFAULT_DATA_DIR) -> str:
    """Cache key for the immutable bundle and its live taxonomy ledger."""

    data_dir = Path(data_dir).resolve()
    manifest_path = revision_bundle_manifest_path(data_dir)
    manifest_digest = (
        _sha256(manifest_path.read_bytes())
        if manifest_path.exists()
        else f"missing:{manifest_path}"
    )
    payload: dict[str, object] = {}
    payload_path = revision_payload_path(data_dir)
    if payload_path.exists():
        try:
            loaded = json.loads(payload_path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                payload = loaded
        except (OSError, json.JSONDecodeError):
            payload = {}
    taxonomy_digest = _taxonomy_review_signature(data_dir, payload)
    return f"{manifest_digest}:{taxonomy_digest}"


def get_revision_export_status(data_dir: Path = DEFAULT_DATA_DIR) -> RevisionExportStatus:
    data_dir = Path(data_dir).resolve()
    payload_path = revision_payload_path(data_dir)
    manifest_path = revision_bundle_manifest_path(data_dir)
    schema, latest = _payload_metadata(payload_path)
    bundle_id = ""
    pptx_path = revision_pptx_candidates(data_dir)[0]
    xlsx_path = revision_xlsx_candidates(data_dir)[0]
    portfolio_xlsx_path = revision_portfolio_xlsx_candidates(data_dir)[0]
    top100_xlsx_path = revision_top100_xlsx_candidates(data_dir)[0]
    html_path = revision_html_candidates(data_dir)[0]
    error = ""
    valid = False
    try:
        bundle = _load_validated_bundle(data_dir)
        bundle_id = str(bundle.manifest.get("bundle_id") or "")
        pptx_path = bundle.pptx_path
        xlsx_path = bundle.xlsx_path
        portfolio_xlsx_path = bundle.portfolio_xlsx_path
        top100_xlsx_path = bundle.top100_xlsx_path
        html_path = bundle.html_path
        valid = True
    except RevisionExportUnavailable as exc:
        error = str(exc)
        if manifest_path.exists():
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                bundle_id = str(manifest.get("bundle_id") or "")
            except (OSError, json.JSONDecodeError):
                pass
    return RevisionExportStatus(
        payload_path=str(payload_path),
        payload_exists=payload_path.exists(),
        payload_schema=schema,
        latest_complete=latest,
        bundle_manifest_path=str(manifest_path),
        bundle_exists=manifest_path.exists(),
        bundle_id=bundle_id,
        bundle_valid=valid,
        validation_error=error,
        pptx_path=str(pptx_path),
        pptx_exists=pptx_path.exists(),
        xlsx_path=str(xlsx_path),
        xlsx_exists=xlsx_path.exists(),
        portfolio_xlsx_path=str(portfolio_xlsx_path),
        portfolio_xlsx_exists=portfolio_xlsx_path.exists(),
        top100_xlsx_path=str(top100_xlsx_path),
        top100_xlsx_exists=top100_xlsx_path.exists(),
        html_path=str(html_path),
        html_exists=html_path.exists(),
        artifact_runtime_available=artifact_runtime_available(),
    )


def build_revision_pptx_bytes(data_dir: Path = DEFAULT_DATA_DIR) -> bytes:
    return _load_validated_bundle(data_dir).pptx_bytes


def build_revision_xlsx_bytes(data_dir: Path = DEFAULT_DATA_DIR) -> bytes:
    return _load_validated_bundle(data_dir).xlsx_bytes


def build_revision_portfolio_xlsx_bytes(
    data_dir: Path = DEFAULT_DATA_DIR,
) -> bytes:
    return _load_validated_bundle(data_dir).portfolio_xlsx_bytes


def build_revision_top100_xlsx_bytes(
    data_dir: Path = DEFAULT_DATA_DIR,
) -> bytes:
    return _load_validated_bundle(data_dir).top100_xlsx_bytes


def build_revision_html_bytes(data_dir: Path = DEFAULT_DATA_DIR) -> bytes:
    return _load_validated_bundle(data_dir).html_bytes


__all__ = [
    "BUNDLE_SCHEMA",
    "CURRENT_TOP15_SLIDE_SEQUENCE",
    "EXPECTED_SLIDE_SEQUENCE",
    "EXPECTED_SLIDES",
    "HISTORICAL_TOP15_SLIDE_SEQUENCE",
    "HISTORICAL_TOP15_TABLE_DIMENSIONS",
    "ISSUANCE_TAXONOMY_TABLE_DIMENSIONS",
    "MATERIALIZED_HTML_NAME",
    "MATERIALIZED_PORTFOLIO_XLSX_NAME",
    "MATERIALIZED_TOP100_XLSX_NAME",
    "REQUIRED_WORKBOOK_SHEETS",
    "REQUIRED_PORTFOLIO_WORKBOOK_SHEETS",
    "REVISION_EMISSION_AUDIT_REQUIRED_HEADERS",
    "REVISION_EMISSION_COVERAGE_TARGET_LABEL",
    "STRUCTURAL_MVP_SLIDE_SEQUENCE",
    "TYPE_RANKING_SLIDE_SEQUENCE",
    "RevisionExportStatus",
    "RevisionExportUnavailable",
    "artifact_runtime_available",
    "build_revision_pptx_bytes",
    "build_revision_portfolio_xlsx_bytes",
    "build_revision_top100_xlsx_bytes",
    "build_revision_xlsx_bytes",
    "build_revision_html_bytes",
    "get_revision_export_status",
    "revision_bundle_manifest_path",
    "revision_export_signature",
    "revision_payload_path",
    "revision_html_candidates",
    "revision_portfolio_xlsx_candidates",
    "revision_top100_xlsx_candidates",
    "validate_revision_html",
    "validate_revision_portfolio_xlsx",
    "validate_revision_top100_xlsx",
    "validate_revision_pptx",
    "validate_revision_xlsx",
]
