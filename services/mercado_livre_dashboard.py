from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
from io import BytesIO
import json
from pathlib import Path
import re
from typing import Any

import pandas as pd
from pandas.errors import EmptyDataError
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


PT_MONTH_ABBR = {
    1: "jan",
    2: "fev",
    3: "mar",
    4: "abr",
    5: "mai",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "set",
    10: "out",
    11: "nov",
    12: "dez",
}

MONEY_COLUMNS = [
    "pl_total",
    "pl_total_oficial",
    "pl_total_classes",
    "pl_reconciliacao_delta",
    "pl_senior",
    "pl_subordinada_strict",
    "pl_mezzanino",
    "pl_subordinada_mezz",
    "carteira_bruta",
    "pdd_total",
    "carteira_liquida",
    "carteira_em_dia",
    "atraso_ate30",
    "atraso_31_60",
    "atraso_61_90",
    "atraso_91_180",
    "atraso_181_360",
    "vencidos_360",
    "npl_over1",
    "npl_over30",
    "npl_over60",
    "npl_over90",
    "npl_over180",
    "npl_over360",
    "baixa_over360_carteira",
    "baixa_over360_pdd",
    "baixa_over360_pl",
    "carteira_ex360",
    "pdd_ex360",
    "carteira_liquida_ex360",
    "pl_total_ex360",
    "pl_subordinada_mezz_ex360",
    "npl_over1_ex360",
    "npl_over30_ex360",
    "npl_over60_ex360",
    "npl_over90_ex360",
    "npl_over180_ex360",
    "carteira_em_dia_mais_ate30",
    "roll_rate_base_t_minus_1",
]

PRIMITIVE_SUM_COLUMNS = [
    "pl_total",
    "pl_senior",
    "pl_subordinada_strict",
    "pl_mezzanino",
    "carteira_bruta",
    "pdd_total",
    "atraso_ate30",
    "atraso_31_60",
    "atraso_61_90",
    "atraso_91_120",
    "atraso_121_150",
    "atraso_151_180",
    "atraso_181_360",
    "atraso_361_720",
    "atraso_721_1080",
    "atraso_1080",
]

WIDE_TABLE_COLUMNS = ["Bloco", "Métrica", "Memória / fórmula"]
CALCULATION_SCHEMA_VERSION = 5
OFFICIAL_PL_PATH = "DOC_ARQ/LISTA_INFORM/PATRLIQ/VL_PATRIM_LIQ"


@dataclass(frozen=True)
class MercadoLivreOutputs:
    fund_monthly: dict[str, pd.DataFrame]
    fund_wide: dict[str, pd.DataFrame]
    consolidated_monthly: pd.DataFrame
    consolidated_wide: pd.DataFrame
    warnings_df: pd.DataFrame
    metadata: dict[str, Any]


def build_mercado_livre_outputs(
    *,
    portfolio_id: str,
    portfolio_name: str,
    dashboards_by_cnpj: dict[str, tuple[str, Any]],
    period_label: str,
    official_pl_by_cnpj: dict[str, pd.DataFrame] | None = None,
) -> MercadoLivreOutputs:
    fund_monthly: dict[str, pd.DataFrame] = {}
    fund_wide: dict[str, pd.DataFrame] = {}
    warnings: list[dict[str, object]] = []
    official_pl_by_cnpj = official_pl_by_cnpj or {}
    for cnpj, (fund_name, dashboard) in dashboards_by_cnpj.items():
        monthly = build_fund_monthly_base(
            cnpj=cnpj,
            fund_name=fund_name,
            dashboard=dashboard,
            official_pl_history_df=official_pl_by_cnpj.get(cnpj),
        )
        fund_monthly[cnpj] = monthly
        fund_wide[cnpj] = build_wide_table(monthly, scope_name=fund_name)
        warnings.extend(_warning_rows(monthly, scope_name=fund_name, cnpj=cnpj))

    consolidated = build_consolidated_monthly_base(
        portfolio_name=portfolio_name,
        fund_monthly_frames=fund_monthly,
    )
    consolidated_wide = build_wide_table(consolidated, scope_name=portfolio_name)
    warnings.extend(_warning_rows(consolidated, scope_name=portfolio_name, cnpj="CONSOLIDADO"))
    metadata = {
        "schema_version": CALCULATION_SCHEMA_VERSION,
        "generated_at": _utc_now_iso(),
        "portfolio_id": portfolio_id,
        "portfolio_name": portfolio_name,
        "period_label": period_label,
        "loaded_period_label": _loaded_period_label_from_frame(consolidated),
        "competencias_disponiveis_consolidado": _competencia_list(consolidated),
        "funds": [
            {
                "cnpj": cnpj,
                "name": fund_name,
                "competencias": _competencia_list(frame),
            }
            for cnpj, (fund_name, _dashboard) in dashboards_by_cnpj.items()
            for frame in [fund_monthly.get(cnpj, pd.DataFrame())]
        ],
        "formulas": _formula_catalog(),
        "known_limitations": [
            "A visao Ex Over 360 simula baixa integral dos vencidos acima de 360 dias para fins comparaveis a instituicoes financeiras.",
            "PDD Ex Over 360 e calculada como PDD total menos a baixa dos vencidos acima de 360 dias, limitada ao saldo de PDD disponivel.",
            "Quando o saldo Over 360 excede a PDD disponivel, a diferenca e tratada como baixa residual contra PL.",
            "PL FIDC total usa PATRLIQ/VL_PATRIM_LIQ quando disponivel; a soma das classes fica como reconciliacao.",
            "Carteira Bruta usa a base canonica de direitos crediticios do dashboard, nao o campo amplo APLIC_ATIVO/VL_CARTEIRA.",
        ],
    }
    return MercadoLivreOutputs(
        fund_monthly=fund_monthly,
        fund_wide=fund_wide,
        consolidated_monthly=consolidated,
        consolidated_wide=consolidated_wide,
        warnings_df=pd.DataFrame(warnings),
        metadata=metadata,
    )


def build_fund_monthly_base(
    *,
    cnpj: str,
    fund_name: str,
    dashboard: Any,
    official_pl_history_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    competencias = list(getattr(dashboard, "competencias", []) or [])
    rows: list[dict[str, object]] = []
    info = getattr(dashboard, "fund_info", {}) or {}
    for competencia in competencias:
        row: dict[str, object] = {
            "scope": "individual",
            "fund_name": fund_name or info.get("nome_fundo") or cnpj,
            "cnpj": _digits(cnpj or info.get("cnpj_fundo")),
            "tipo_classe": _first_non_empty(info.get("nome_classe"), info.get("fundo_ou_classe")),
            "competencia": competencia,
            "competencia_dt": _competencia_to_timestamp(competencia),
        }
        row.update(_subordination_values(getattr(dashboard, "subordination_history_df", pd.DataFrame()), competencia))
        row.update(_official_pl_values(official_pl_history_df, competencia))
        row.update(_credit_values(dashboard, competencia))
        row.update(_bucket_values(getattr(dashboard, "default_buckets_history_df", pd.DataFrame()), competencia))
        rows.append(row)
    frame = pd.DataFrame(rows)
    if frame.empty:
        return _empty_monthly_frame()
    frame = frame.sort_values("competencia_dt").reset_index(drop=True)
    return _decorate_monthly_base(frame, expected_funds=1)


def build_consolidated_monthly_base(
    *,
    portfolio_name: str,
    fund_monthly_frames: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    valid_frames = [
        frame.copy()
        for frame in fund_monthly_frames.values()
        if isinstance(frame, pd.DataFrame) and not frame.empty
    ]
    if not valid_frames:
        return _empty_monthly_frame()
    combined = pd.concat(valid_frames, ignore_index=True, sort=False)
    rows: list[dict[str, object]] = []
    expected_funds = len(valid_frames)
    for competencia, group in combined.groupby("competencia", dropna=False):
        row: dict[str, object] = {
            "scope": "consolidado",
            "fund_name": portfolio_name,
            "cnpj": "CONSOLIDADO",
            "tipo_classe": "Carteira consolidada",
            "competencia": competencia,
            "competencia_dt": _competencia_to_timestamp(competencia),
            "funds_expected_count": expected_funds,
            "funds_present_count": int(group["cnpj"].nunique()) if "cnpj" in group.columns else len(group),
        }
        for column in PRIMITIVE_SUM_COLUMNS:
            row[column] = _sum_numeric(group.get(column))
        row["pl_total_classes"] = _sum_numeric(group.get("pl_total_classes"))
        row["pl_total_oficial"] = _sum_numeric(group.get("pl_total_oficial"))
        rows.append(row)
    frame = pd.DataFrame(rows).sort_values("competencia_dt").reset_index(drop=True)
    return _decorate_monthly_base(frame, expected_funds=expected_funds)


def build_wide_table(monthly_df: pd.DataFrame, *, scope_name: str) -> pd.DataFrame:
    if monthly_df.empty:
        return pd.DataFrame(columns=WIDE_TABLE_COLUMNS)
    sorted_df = _sort_monthly_by_competencia(monthly_df, descending=True)
    competencias = sorted_df["competencia"].astype(str).tolist()
    display_competencias = [_format_competencia_short(value) for value in competencias]
    metric_scales = _money_scales_by_metric(sorted_df)
    metric_specs = _wide_metric_specs()
    rows: list[dict[str, object]] = []
    for spec in metric_specs:
        values = []
        for _, item in sorted_df.iterrows():
            values.append(_format_wide_value(item.get(spec["column"]), unit=str(spec["unit"]), scale=metric_scales.get(str(spec["column"]))))
        period_values = dict(zip(display_competencias, values, strict=False))
        rows.append(
            {
                "Bloco": spec["block"],
                "Métrica": spec["metric"],
                **period_values,
                "Memória / fórmula": spec["formula"],
            }
        )
    output = pd.DataFrame(rows)
    output.attrs["scope_name"] = scope_name
    output.attrs["competencias"] = competencias
    return output


def build_validation_table(outputs: MercadoLivreOutputs) -> pd.DataFrame:
    frames = list(outputs.fund_monthly.values())
    if not outputs.consolidated_monthly.empty:
        frames.append(outputs.consolidated_monthly)
    if not frames:
        return pd.DataFrame()
    columns = [
        "fund_name",
        "cnpj",
        "competencia",
        "pl_total",
        "pl_total_oficial",
        "pl_total_classes",
        "pl_reconciliacao_delta",
        "pl_senior",
        "pl_subordinada_mezz",
        "subordinacao_total_pct",
        "carteira_bruta",
        "pdd_total",
        "npl_over90",
        "npl_over360",
        "baixa_over360_carteira",
        "baixa_over360_pdd",
        "baixa_over360_pl",
        "npl_over90_ex360",
        "carteira_ex360",
        "pdd_ex360",
        "pl_total_ex360",
        "pl_subordinada_mezz_ex360",
        "subordinacao_total_ex360_pct",
        "pdd_npl_over90_pct",
        "pdd_npl_over90_ex360_pct",
        "warnings",
    ]
    combined = pd.concat(frames, ignore_index=True, sort=False)
    available = [column for column in columns if column in combined.columns]
    return combined[available].sort_values(["fund_name", "competencia"]).reset_index(drop=True)


def order_period_columns_desc(columns: list[object] | pd.Index) -> list[str]:
    """Return only period-like columns ordered from newest to oldest."""
    period_columns = [str(column) for column in columns if _period_label_to_timestamp(column) is not None]
    return sorted(period_columns, key=lambda column: _period_label_to_timestamp(column) or pd.Timestamp.min, reverse=True)


def save_outputs_to_cache(
    outputs: MercadoLivreOutputs,
    *,
    portfolio_id: str,
    period_key: str,
    portfolio_funds: list[dict[str, str]] | tuple[Any, ...] | None = None,
    base_dir: Path | str = ".cache/mercado-livre",
) -> Path:
    identity_key = portfolio_identity_key(portfolio_funds, fallback=portfolio_id)
    root = Path(base_dir) / identity_key / _safe_path_token(period_key)
    root.mkdir(parents=True, exist_ok=True)
    for cnpj, frame in outputs.fund_monthly.items():
        frame.to_csv(root / f"monthly_{_safe_path_token(cnpj)}.csv", index=False)
    for cnpj, frame in outputs.fund_wide.items():
        frame.to_csv(root / f"wide_{_safe_path_token(cnpj)}.csv", index=False)
    outputs.consolidated_monthly.to_csv(root / "monthly_consolidado.csv", index=False)
    outputs.consolidated_wide.to_csv(root / "wide_consolidado.csv", index=False)
    outputs.warnings_df.to_csv(root / "warnings.csv", index=False)
    metadata = dict(outputs.metadata)
    metadata.update(
        {
            "storage_identity_key": identity_key,
            "period_key": period_key,
            "cache_dir": str(root),
        }
    )
    (root / "metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    return root


def portfolio_identity_key(portfolio_funds: list[dict[str, str]] | tuple[Any, ...] | None, *, fallback: str) -> str:
    cnpjs: list[str] = []
    for fund in portfolio_funds or []:
        if isinstance(fund, dict):
            cnpj = fund.get("cnpj")
        else:
            cnpj = getattr(fund, "cnpj", None)
        digits = _digits(cnpj)
        if digits:
            cnpjs.append(digits)
    if not cnpjs:
        return _safe_path_token(fallback)
    payload = {
        "schema_version": CALCULATION_SCHEMA_VERSION,
        "view": "mercado_livre",
        "cnpjs": sorted(set(cnpjs)),
    }
    digest = hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()[:20]
    return f"ml_{digest}"


def cache_dir_for_outputs(
    *,
    portfolio_id: str,
    period_key: str,
    portfolio_funds: list[dict[str, str]] | tuple[Any, ...] | None = None,
    base_dir: Path | str = ".cache/mercado-livre",
) -> Path:
    identity_key = portfolio_identity_key(portfolio_funds, fallback=portfolio_id)
    return Path(base_dir) / identity_key / _safe_path_token(period_key)


def load_outputs_from_cache(
    *,
    portfolio_id: str,
    period_key: str,
    portfolio_funds: list[dict[str, str]] | tuple[Any, ...] | None = None,
    base_dir: Path | str = ".cache/mercado-livre",
) -> MercadoLivreOutputs | None:
    root = cache_dir_for_outputs(
        portfolio_id=portfolio_id,
        period_key=period_key,
        portfolio_funds=portfolio_funds,
        base_dir=base_dir,
    )
    metadata_path = root / "metadata.json"
    if not metadata_path.exists():
        return None
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    if int(metadata.get("schema_version") or 0) != CALCULATION_SCHEMA_VERSION:
        return None
    fund_monthly: dict[str, pd.DataFrame] = {}
    fund_wide: dict[str, pd.DataFrame] = {}
    for fund in metadata.get("funds") or []:
        cnpj = _digits(fund.get("cnpj"))
        if not cnpj:
            continue
        monthly_path = root / f"monthly_{_safe_path_token(cnpj)}.csv"
        wide_path = root / f"wide_{_safe_path_token(cnpj)}.csv"
        if monthly_path.exists():
            monthly_df = _read_cache_csv(monthly_path)
            if monthly_df is None:
                return None
            fund_monthly[cnpj] = monthly_df
        if wide_path.exists():
            wide_df = _read_cache_csv(wide_path)
            if wide_df is None:
                return None
            fund_wide[cnpj] = wide_df
    consolidated_monthly_path = root / "monthly_consolidado.csv"
    consolidated_wide_path = root / "wide_consolidado.csv"
    warnings_path = root / "warnings.csv"
    if not consolidated_monthly_path.exists() or not consolidated_wide_path.exists():
        return None
    consolidated_monthly = _read_cache_csv(consolidated_monthly_path)
    consolidated_wide = _read_cache_csv(consolidated_wide_path)
    if consolidated_monthly is None or consolidated_wide is None:
        return None
    metadata["cache_dir"] = str(root)
    return MercadoLivreOutputs(
        fund_monthly=fund_monthly,
        fund_wide=fund_wide,
        consolidated_monthly=consolidated_monthly,
        consolidated_wide=consolidated_wide,
        warnings_df=_read_optional_cache_csv(warnings_path),
        metadata=metadata,
    )


def _read_cache_csv(path: Path) -> pd.DataFrame | None:
    try:
        return pd.read_csv(path)
    except EmptyDataError:
        return None


def _read_optional_cache_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except EmptyDataError:
        return pd.DataFrame()


def extract_official_pl_history_from_wide_csv(wide_csv_path: str | Path) -> pd.DataFrame:
    path = Path(wide_csv_path)
    if not path.exists():
        return pd.DataFrame(columns=["competencia", "pl_total_oficial", "pl_total_oficial_source_status"])
    wide_df = pd.read_csv(path, dtype=str, keep_default_na=False)
    competencias = [column for column in wide_df.columns if re.fullmatch(r"\d{2}/\d{4}", str(column))]
    if "tag_path" not in wide_df.columns or not competencias:
        return pd.DataFrame(columns=["competencia", "pl_total_oficial", "pl_total_oficial_source_status"])
    match = wide_df[wide_df["tag_path"].astype(str) == OFFICIAL_PL_PATH]
    if match.empty:
        return pd.DataFrame(
            {
                "competencia": competencias,
                "pl_total_oficial": [pd.NA] * len(competencias),
                "pl_total_oficial_source_status": ["missing_field"] * len(competencias),
            }
        )
    row = match.iloc[0]
    rows = []
    for competencia in competencias:
        raw = row.get(competencia)
        value = _num(raw)
        if value is None:
            status = "not_reported" if str(raw or "").strip() == "" else "not_numeric"
        elif value == 0:
            status = "reported_zero"
        else:
            status = "reported_value"
        rows.append(
            {
                "competencia": competencia,
                "pl_total_oficial": value,
                "pl_total_oficial_source_status": status,
            }
        )
    return pd.DataFrame(rows)


def build_excel_export_bytes(outputs: MercadoLivreOutputs) -> bytes:
    buffer = BytesIO()
    workbook = Workbook()
    used_sheet_names: set[str] = set()

    consolidated_ws = workbook.active
    consolidated_ws.title = "Consolidado"
    used_sheet_names.add(consolidated_ws.title)
    _write_numeric_wide_sheet(
        consolidated_ws,
        outputs.consolidated_monthly,
        scope_name=str(outputs.metadata.get("portfolio_name") or "Consolidado"),
    )

    for cnpj, monthly_df in outputs.fund_monthly.items():
        fallback_name = _fund_sheet_name(monthly_df, fallback=cnpj)
        sheet_name = _unique_excel_sheet_name(fallback_name, used_sheet_names)
        used_sheet_names.add(sheet_name)
        ws = workbook.create_sheet(sheet_name)
        _write_numeric_wide_sheet(ws, monthly_df, scope_name=fallback_name)

    _write_dataframe_sheet(workbook.create_sheet("Auditoria"), build_validation_table(outputs))
    _write_dataframe_sheet(workbook.create_sheet("Warnings"), outputs.warnings_df)
    metadata_df = pd.DataFrame(
        [
            {
                "chave": key,
                "valor": json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value,
            }
            for key, value in outputs.metadata.items()
        ]
    )
    _write_dataframe_sheet(workbook.create_sheet("Metadados"), metadata_df)
    workbook.save(buffer)
    return buffer.getvalue()


def build_consolidated_snapshot_excel_bytes(outputs: MercadoLivreOutputs) -> bytes:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Resumo 6m"
    data_ws = workbook.create_sheet("Dados gráficos")
    charts_ws = workbook.create_sheet("Gráficos")

    monthly = outputs.consolidated_monthly.copy()
    if monthly.empty:
        summary_ws.append(["Sem dados consolidados"])
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    chart_monthly = monthly.sort_values("competencia_dt").tail(6).reset_index(drop=True)
    summary_monthly = _sort_monthly_by_competencia(chart_monthly, descending=True)
    summary_period_labels = [_format_competencia_short(value) for value in summary_monthly["competencia"].astype(str).tolist()]
    chart_period_labels = [_format_competencia_short(value) for value in chart_monthly["competencia"].astype(str).tolist()]

    _write_snapshot_summary(summary_ws, summary_monthly, summary_period_labels)
    _write_snapshot_chart_data(data_ws, chart_monthly, chart_period_labels)
    _write_snapshot_charts(charts_ws, data_ws, len(chart_monthly))
    _style_plain_sheet(summary_ws)
    _style_plain_sheet(data_ws)
    charts_ws.sheet_view.showGridLines = False
    charts_ws.column_dimensions["A"].width = 2

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_snapshot_summary(ws, monthly: pd.DataFrame, period_labels: list[str]) -> None:
    ws.append(["Métrica", *period_labels])
    metrics = [
        ("PL FIDC total", "pl_total", "money"),
        ("PL Sênior", "pl_senior", "money"),
        ("Subordinada + Mezanino ex-360", "pl_subordinada_mezz_ex360", "money"),
        ("% Subordinação Total ex-360", "subordinacao_total_ex360_pct", "percent"),
        ("Carteira Bruta total", "carteira_bruta", "money"),
        ("Carteira Ex Over 360d", "carteira_ex360", "money"),
        ("PDD Ex Over 360d", "pdd_ex360", "money"),
        ("NPL Over 90d Ex 360", "npl_over90_ex360", "money"),
        ("NPL Over 90d Ex 360 / Carteira Ex 360", "npl_over90_ex360_pct", "percent"),
        ("PDD Ex / NPL Over 90d Ex 360", "pdd_npl_over90_ex360_pct", "percent"),
        ("Roll Rate 31-60d", "roll_rate_31_60_pct", "percent"),
    ]
    for label, column, unit in metrics:
        values = [_excel_snapshot_value(row.get(column), unit=unit) for _, row in monthly.iterrows()]
        ws.append([label, *values])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2):
        for cell in row:
            cell.number_format = '#,##0.0%' if isinstance(cell.value, float) and abs(cell.value) <= 20 and ws.cell(row=cell.row, column=1).value and "%" in str(ws.cell(row=cell.row, column=1).value) else '#,##0'
    ws.freeze_panes = "B2"


def _write_snapshot_chart_data(ws, monthly: pd.DataFrame, period_labels: list[str]) -> None:
    ws.append(
        [
            "Competência",
            "PL Sênior",
            "Subordinada + Mezanino ex-360",
            "% Subordinação Total ex-360",
            "NPL Over 90d Ex 360 / Carteira Ex 360",
            "PDD Ex / NPL Over 90d Ex 360",
        ]
    )
    for label, (_, row) in zip(period_labels, monthly.iterrows(), strict=False):
        ws.append(
            [
                label,
                _excel_snapshot_value(row.get("pl_senior"), unit="money"),
                _excel_snapshot_value(row.get("pl_subordinada_mezz_ex360"), unit="money"),
                _excel_snapshot_value(row.get("subordinacao_total_ex360_pct"), unit="percent"),
                _excel_snapshot_value(row.get("npl_over90_ex360_pct"), unit="percent"),
                _excel_snapshot_value(row.get("pdd_npl_over90_ex360_pct"), unit="percent"),
            ]
        )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.0%'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'
    ws.freeze_panes = "B2"


def _write_snapshot_charts(ws, data_ws, month_count: int) -> None:
    if month_count <= 0:
        return
    max_row = month_count + 1
    categories = Reference(data_ws, min_col=1, min_row=2, max_row=max_row)

    pl_chart = BarChart()
    pl_chart.type = "col"
    pl_chart.style = 10
    pl_chart.title = "Evolução de PL e Subordinação"
    pl_chart.y_axis.title = "R$"
    pl_chart.x_axis.title = "Competência"
    pl_chart.add_data(Reference(data_ws, min_col=2, max_col=3, min_row=1, max_row=max_row), titles_from_data=True)
    pl_chart.set_categories(categories)
    pl_chart.width = 33.0
    pl_chart.height = 15.0
    for series, color in zip(pl_chart.series, ["000000", "E47811"], strict=False):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color

    sub_line = LineChart()
    sub_line.add_data(Reference(data_ws, min_col=4, min_row=1, max_row=max_row), titles_from_data=True)
    sub_line.set_categories(categories)
    sub_line.y_axis.axId = 200
    sub_line.y_axis.title = "% Subordinação"
    sub_line.y_axis.crosses = "max"
    if sub_line.series:
        sub_line.series[0].graphicalProperties.line.solidFill = "3F3F3F"
        sub_line.series[0].graphicalProperties.line.width = 25000
        sub_line.series[0].marker.symbol = "circle"
    pl_chart += sub_line
    ws.add_chart(pl_chart, "B2")

    npl_chart = LineChart()
    npl_chart.title = "NPL e Cobertura Ex-Vencidos > 360d"
    npl_chart.y_axis.title = "NPL Over 90d Ex 360 / Carteira Ex 360"
    npl_chart.x_axis.title = "Competência"
    npl_chart.add_data(Reference(data_ws, min_col=5, min_row=1, max_row=max_row), titles_from_data=True)
    npl_chart.set_categories(categories)
    npl_chart.width = 33.0
    npl_chart.height = 15.0
    if npl_chart.series:
        npl_chart.series[0].graphicalProperties.line.solidFill = "000000"
        npl_chart.series[0].graphicalProperties.line.width = 25000
        npl_chart.series[0].marker.symbol = "circle"

    coverage_line = LineChart()
    coverage_line.add_data(Reference(data_ws, min_col=6, min_row=1, max_row=max_row), titles_from_data=True)
    coverage_line.set_categories(categories)
    coverage_line.y_axis.axId = 300
    coverage_line.y_axis.title = "PDD Ex / NPL Over 90d Ex 360"
    coverage_line.y_axis.crosses = "max"
    if coverage_line.series:
        coverage_line.series[0].graphicalProperties.line.solidFill = "E47811"
        coverage_line.series[0].graphicalProperties.line.width = 25000
        coverage_line.series[0].marker.symbol = "circle"
    npl_chart += coverage_line
    ws.add_chart(npl_chart, "B32")


def _excel_snapshot_value(value: object, *, unit: str) -> float | None:
    numeric = _num(value)
    if numeric is None:
        return None
    if unit == "percent":
        return numeric / 100.0
    return numeric


def _subordination_values(frame: pd.DataFrame, competencia: str) -> dict[str, object]:
    row = _latest_match(frame, competencia)
    return {
        "pl_total_classes": _num(row.get("pl_total")),
        "pl_senior": _num(row.get("pl_senior")),
        "pl_mezzanino": _num(row.get("pl_mezzanino")),
        "pl_subordinada_strict": _num(row.get("pl_subordinada_strict")),
        "pl_subordinada_mezz": _num(row.get("pl_subordinada")),
    }


def _official_pl_values(frame: pd.DataFrame | None, competencia: str) -> dict[str, object]:
    row = _latest_match(frame if frame is not None else pd.DataFrame(), competencia)
    return {
        "pl_total_oficial": _num(row.get("pl_total_oficial")),
        "pl_total_oficial_source_status": _first_non_empty(row.get("pl_total_oficial_source_status")),
    }


def _credit_values(dashboard: Any, competencia: str) -> dict[str, object]:
    default_row = _latest_match(getattr(dashboard, "default_history_df", pd.DataFrame()), competencia)
    dc_row = _latest_match(getattr(dashboard, "dc_canonical_history_df", pd.DataFrame()), competencia)
    return {
        "carteira_bruta": _num(default_row.get("direitos_creditorios")) or _num(dc_row.get("dc_total_canonico")),
        "carteira_bruta_origem": _first_non_empty(default_row.get("direitos_creditorios_fonte"), dc_row.get("dc_total_fonte_efetiva")),
        "pdd_total": _num(default_row.get("provisao_total")),
    }


def _bucket_values(frame: pd.DataFrame, competencia: str) -> dict[str, object]:
    subset = frame[frame["competencia"].astype(str) == str(competencia)].copy() if isinstance(frame, pd.DataFrame) and not frame.empty and "competencia" in frame.columns else pd.DataFrame()
    values = {int(row.get("ordem")): _num(row.get("valor")) for _, row in subset.iterrows() if pd.notna(row.get("ordem"))}
    return {
        "atraso_ate30": values.get(1),
        "atraso_31_60": values.get(2),
        "atraso_61_90": values.get(3),
        "atraso_91_120": values.get(4),
        "atraso_121_150": values.get(5),
        "atraso_151_180": values.get(6),
        "atraso_181_360": values.get(7),
        "atraso_361_720": values.get(8),
        "atraso_721_1080": values.get(9),
        "atraso_1080": values.get(10),
    }


def _decorate_monthly_base(frame: pd.DataFrame, *, expected_funds: int) -> pd.DataFrame:
    df = frame.copy()
    if "pl_total_classes" not in df.columns and "pl_total" in df.columns:
        df["pl_total_classes"] = df["pl_total"]
    for column in PRIMITIVE_SUM_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA
        df[column] = pd.to_numeric(df[column], errors="coerce")
    if "pl_total_classes" not in df.columns:
        df["pl_total_classes"] = pd.NA
    if "pl_total_oficial" not in df.columns:
        df["pl_total_oficial"] = pd.NA
    df["pl_total_classes"] = pd.to_numeric(df["pl_total_classes"], errors="coerce")
    df["pl_total_oficial"] = pd.to_numeric(df["pl_total_oficial"], errors="coerce")
    df["pl_total"] = df["pl_total_oficial"].where(df["pl_total_oficial"].notna(), df["pl_total_classes"])
    df["pl_total_usado_fonte"] = df["pl_total_oficial"].map(lambda value: "PATRLIQ/VL_PATRIM_LIQ" if pd.notna(value) else "soma_classes_cotas")
    df["pl_reconciliacao_delta"] = df["pl_total"] - df["pl_total_classes"]
    df["pl_reconciliacao_delta_pct"] = _safe_div_pct(df["pl_reconciliacao_delta"].abs(), df["pl_total"])
    df["pl_reconciliacao_warning"] = df["pl_reconciliacao_delta_pct"].gt(0.5).fillna(False)
    if "funds_expected_count" not in df.columns:
        df["funds_expected_count"] = expected_funds
    if "funds_present_count" not in df.columns:
        df["funds_present_count"] = 1
    df["pl_subordinada_mezz"] = _coalesce_numeric(df["pl_subordinada_mezz"] if "pl_subordinada_mezz" in df.columns else pd.Series(pd.NA, index=df.index), df["pl_subordinada_strict"].fillna(0.0) + df["pl_mezzanino"].fillna(0.0))
    df["subordinacao_total_pct"] = _safe_div_pct(df["pl_subordinada_mezz"], df["pl_total"])
    df["pdd_carteira_bruta_pct"] = _safe_div_pct(df["pdd_total"], df["carteira_bruta"])
    df["carteira_liquida"] = df["carteira_bruta"] - df["pdd_total"]
    df["atraso_91_180"] = df[["atraso_91_120", "atraso_121_150", "atraso_151_180"]].sum(axis=1, min_count=1)
    df["vencidos_360"] = df[["atraso_361_720", "atraso_721_1080", "atraso_1080"]].sum(axis=1, min_count=1)
    df["npl_over1"] = df[["atraso_ate30", "atraso_31_60", "atraso_61_90", "atraso_91_120", "atraso_121_150", "atraso_151_180", "atraso_181_360", "atraso_361_720", "atraso_721_1080", "atraso_1080"]].sum(axis=1, min_count=1)
    df["npl_over30"] = df[["atraso_31_60", "atraso_61_90", "atraso_91_120", "atraso_121_150", "atraso_151_180", "atraso_181_360", "atraso_361_720", "atraso_721_1080", "atraso_1080"]].sum(axis=1, min_count=1)
    df["npl_over60"] = df[["atraso_61_90", "atraso_91_120", "atraso_121_150", "atraso_151_180", "atraso_181_360", "atraso_361_720", "atraso_721_1080", "atraso_1080"]].sum(axis=1, min_count=1)
    df["npl_over90"] = df[["atraso_91_120", "atraso_121_150", "atraso_151_180", "atraso_181_360", "atraso_361_720", "atraso_721_1080", "atraso_1080"]].sum(axis=1, min_count=1)
    df["npl_over180"] = df[["atraso_181_360", "atraso_361_720", "atraso_721_1080", "atraso_1080"]].sum(axis=1, min_count=1)
    df["npl_over360"] = df["vencidos_360"]
    df["carteira_em_dia"] = df["carteira_bruta"] - df["npl_over1"]
    df["baixa_over360_carteira"] = df["npl_over360"]
    df["carteira_ex360"] = df["carteira_bruta"] - df["baixa_over360_carteira"]
    df["pdd_ex360_calculavel"] = df["pdd_total"].notna() & df["baixa_over360_carteira"].notna()
    df["baixa_over360_pdd"] = _row_min(df["pdd_total"], df["baixa_over360_carteira"])
    df["pdd_ex360"] = (df["pdd_total"] - df["baixa_over360_pdd"]).where(df["pdd_ex360_calculavel"])
    df["pdd_over360_insuficiente"] = (
        pd.to_numeric(df["pdd_total"], errors="coerce")
        < pd.to_numeric(df["baixa_over360_carteira"], errors="coerce")
    ).fillna(False)
    df["baixa_over360_pl"] = (
        pd.to_numeric(df["baixa_over360_carteira"], errors="coerce")
        - pd.to_numeric(df["baixa_over360_pdd"], errors="coerce")
    ).clip(lower=0.0)
    df["pl_total_ex360"] = df["pl_total"] - df["baixa_over360_pl"]
    df["pl_subordinada_mezz_ex360"] = (df["pl_subordinada_mezz"] - df["baixa_over360_pl"]).clip(lower=0.0)
    df["subordinacao_total_ex360_pct"] = _safe_div_pct(df["pl_subordinada_mezz_ex360"], df["pl_total_ex360"])
    df["pdd_ex360_carteira_ex360_pct"] = _safe_div_pct(df["pdd_ex360"], df["carteira_ex360"])
    df["carteira_liquida_ex360"] = df["carteira_ex360"] - df["pdd_ex360"]
    df["npl_over1_ex360"] = (df["npl_over1"] - df["npl_over360"]).clip(lower=0.0)
    df["npl_over30_ex360"] = (df["npl_over30"] - df["npl_over360"]).clip(lower=0.0)
    df["npl_over60_ex360"] = (df["npl_over60"] - df["npl_over360"]).clip(lower=0.0)
    df["npl_over90_ex360"] = (df["npl_over90"] - df["npl_over360"]).clip(lower=0.0)
    df["npl_over180_ex360"] = (df["npl_over180"] - df["npl_over360"]).clip(lower=0.0)
    df["npl_over1_pct"] = _safe_div_pct(df["npl_over1"], df["carteira_bruta"])
    df["npl_over30_pct"] = _safe_div_pct(df["npl_over30"], df["carteira_bruta"])
    df["npl_over60_pct"] = _safe_div_pct(df["npl_over60"], df["carteira_bruta"])
    df["npl_over90_pct"] = _safe_div_pct(df["npl_over90"], df["carteira_bruta"])
    df["npl_over180_pct"] = _safe_div_pct(df["npl_over180"], df["carteira_bruta"])
    df["npl_over360_pct"] = _safe_div_pct(df["npl_over360"], df["carteira_bruta"])
    df["npl_over1_ex360_pct"] = _safe_div_pct(df["npl_over1_ex360"], df["carteira_ex360"])
    df["npl_over30_ex360_pct"] = _safe_div_pct(df["npl_over30_ex360"], df["carteira_ex360"])
    df["npl_over60_ex360_pct"] = _safe_div_pct(df["npl_over60_ex360"], df["carteira_ex360"])
    df["npl_over90_ex360_pct"] = _safe_div_pct(df["npl_over90_ex360"], df["carteira_ex360"])
    df["npl_over180_ex360_pct"] = _safe_div_pct(df["npl_over180_ex360"], df["carteira_ex360"])
    df["pdd_npl_over1_pct"] = _safe_div_pct(df["pdd_total"], df["npl_over1"])
    df["pdd_npl_over30_pct"] = _safe_div_pct(df["pdd_total"], df["npl_over30"])
    df["pdd_npl_over60_pct"] = _safe_div_pct(df["pdd_total"], df["npl_over60"])
    df["pdd_npl_over90_pct"] = _safe_div_pct(df["pdd_total"], df["npl_over90"])
    df["pdd_npl_over180_pct"] = _safe_div_pct(df["pdd_total"], df["npl_over180"])
    df["pdd_npl_over360_pct"] = _safe_div_pct(df["pdd_total"], df["npl_over360"])
    df["pdd_npl_over90_ex360_pct"] = _safe_div_pct(df["pdd_ex360"], df["npl_over90_ex360"])
    df["carteira_em_dia_mais_ate30"] = df["carteira_em_dia"] + df["atraso_ate30"]
    df["roll_rate_base_t_minus_1"] = pd.to_numeric(df["carteira_em_dia_mais_ate30"], errors="coerce").shift(1)
    df["roll_rate_31_60_pct"] = _safe_div_pct(df["atraso_31_60"], df["roll_rate_base_t_minus_1"])
    df["missing_data_flag"] = df[["pl_total", "pl_senior", "pl_subordinada_mezz", "carteira_bruta", "pdd_total", "npl_over90"]].isna().any(axis=1)
    df["division_by_zero_flag"] = (
        (pd.to_numeric(df["pl_total"], errors="coerce").fillna(0) <= 0)
        | (pd.to_numeric(df["carteira_bruta"], errors="coerce").fillna(0) <= 0)
    )
    df["negative_value_flag"] = df[[column for column in MONEY_COLUMNS if column in df.columns]].lt(0).any(axis=1)
    df["not_calculable_flag"] = df["pdd_ex360_calculavel"].eq(False)
    df["warnings"] = df.apply(_row_warnings, axis=1)
    df["serie_inicio"] = _format_competencia_short(df["competencia"].iloc[0]) if not df.empty else ""
    df["periodo_inicial"] = _format_competencia_short(df["competencia"].iloc[0]) if not df.empty else ""
    df["periodo_final"] = _format_competencia_short(df["competencia"].iloc[-1]) if not df.empty else ""
    df["competencias_disponiveis"] = len(df)
    return df


def _wide_metric_specs() -> list[dict[str, str]]:
    return [
        {"block": "1. Identificação do FIDC", "metric": "Nome do fundo", "column": "fund_name", "unit": "text", "formula": "Nome resolvido do documento/carteira salva."},
        {"block": "1. Identificação do FIDC", "metric": "CNPJ", "column": "cnpj", "unit": "text", "formula": "CNPJ normalizado."},
        {"block": "1. Identificação do FIDC", "metric": "Início da série", "column": "serie_inicio", "unit": "text", "formula": "Primeira competência disponível."},
        {"block": "1. Identificação do FIDC", "metric": "Tipo/classe", "column": "tipo_classe", "unit": "text", "formula": "Classe reportada no cabeçalho, quando existir."},
        {"block": "1. Identificação do FIDC", "metric": "Período inicial", "column": "periodo_inicial", "unit": "text", "formula": "Primeira competência carregada."},
        {"block": "1. Identificação do FIDC", "metric": "Período final", "column": "periodo_final", "unit": "text", "formula": "Última competência carregada."},
        {"block": "1. Identificação do FIDC", "metric": "Quantidade de competências disponíveis", "column": "competencias_disponiveis", "unit": "count", "formula": "Contagem de competências na base mensal."},
        {"block": "1. Identificação do FIDC", "metric": "Warnings de dados faltantes", "column": "warnings", "unit": "text", "formula": "Alertas gerados pela base canônica."},
        {"block": "2. PL FIDC", "metric": "PL FIDC total", "column": "pl_total", "unit": "money", "formula": "PATRLIQ/VL_PATRIM_LIQ quando disponível; caso contrário soma das classes."},
        {"block": "2. PL FIDC", "metric": "PL FIDC oficial", "column": "pl_total_oficial", "unit": "money", "formula": "PATRLIQ/VL_PATRIM_LIQ da base oficial."},
        {"block": "2. PL FIDC", "metric": "PL soma das classes", "column": "pl_total_classes", "unit": "money", "formula": "Soma do PL das classes por quantidade de cotas × valor da cota."},
        {"block": "2. PL FIDC", "metric": "Reconciliação PL oficial vs classes", "column": "pl_reconciliacao_delta", "unit": "money", "formula": "PL FIDC total usado - PL soma das classes."},
        {"block": "2. PL FIDC", "metric": "Reconciliação PL oficial vs classes (%)", "column": "pl_reconciliacao_delta_pct", "unit": "percent", "formula": "|PL FIDC total usado - PL soma das classes| / PL FIDC total usado."},
        {"block": "2. PL FIDC", "metric": "Fonte do PL FIDC total", "column": "pl_total_usado_fonte", "unit": "text", "formula": "Indica se foi usado PL oficial ou soma das classes."},
        {"block": "2. PL FIDC", "metric": "PL Sênior", "column": "pl_senior", "unit": "money", "formula": "PL da classe sênior."},
        {"block": "2. PL FIDC", "metric": "PL Subordinada", "column": "pl_subordinada_strict", "unit": "money", "formula": "PL subordinado estrito, sem mezanino."},
        {"block": "2. PL FIDC", "metric": "PL Mezanino", "column": "pl_mezzanino", "unit": "money", "formula": "PL classificado como mezanino quando aplicável."},
        {"block": "2. PL FIDC", "metric": "Subordinada + Mezanino", "column": "pl_subordinada_mezz", "unit": "money", "formula": "PL subordinada + PL mezanino."},
        {"block": "2. PL FIDC", "metric": "% Subordinação Total", "column": "subordinacao_total_pct", "unit": "percent", "formula": "(Subordinada + Mezanino) / PL FIDC total."},
        {"block": "3. Carteira Bruta", "metric": "Carteira Bruta total", "column": "carteira_bruta", "unit": "money", "formula": "Base canônica de direitos creditórios."},
        {"block": "3. Carteira Bruta", "metric": "PDD total", "column": "pdd_total", "unit": "money", "formula": "Provisão total reportada."},
        {"block": "3. Carteira Bruta", "metric": "PDD / Carteira Bruta", "column": "pdd_carteira_bruta_pct", "unit": "percent", "formula": "PDD total / Carteira Bruta total."},
        {"block": "3. Carteira Bruta", "metric": "Carteira Líquida", "column": "carteira_liquida", "unit": "money", "formula": "Carteira Bruta total - PDD total."},
        {"block": "4. Aging / faixas de atraso", "metric": "Carteira em dia", "column": "carteira_em_dia", "unit": "money", "formula": "Carteira Bruta - NPL Over 1d."},
        {"block": "4. Aging / faixas de atraso", "metric": "Atrasada até 30 dias", "column": "atraso_ate30", "unit": "money", "formula": "Bucket VL_INAD_VENC_30."},
        {"block": "4. Aging / faixas de atraso", "metric": "Atrasos 31-60 dias", "column": "atraso_31_60", "unit": "money", "formula": "Bucket VL_INAD_VENC_31_60."},
        {"block": "4. Aging / faixas de atraso", "metric": "Atrasos 61-90 dias", "column": "atraso_61_90", "unit": "money", "formula": "Bucket VL_INAD_VENC_61_90."},
        {"block": "4. Aging / faixas de atraso", "metric": "Atrasos 91-180 dias", "column": "atraso_91_180", "unit": "money", "formula": "Soma dos buckets 91-120, 121-150 e 151-180."},
        {"block": "4. Aging / faixas de atraso", "metric": "Atrasos 181-360 dias", "column": "atraso_181_360", "unit": "money", "formula": "Bucket VL_INAD_VENC_181_360."},
        {"block": "4. Aging / faixas de atraso", "metric": "Vencidos acima de 360 dias", "column": "vencidos_360", "unit": "money", "formula": "Soma dos buckets 361-720, 721-1080 e acima de 1080 dias."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 1d", "column": "npl_over1", "unit": "money", "formula": "Soma dos buckets de atraso >= 1 dia."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 1d / Carteira", "column": "npl_over1_pct", "unit": "percent", "formula": "NPL Over 1d / Carteira Bruta."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 30d", "column": "npl_over30", "unit": "money", "formula": "Soma dos buckets >= 31 dias."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 30d / Carteira", "column": "npl_over30_pct", "unit": "percent", "formula": "NPL Over 30d / Carteira Bruta."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 60d", "column": "npl_over60", "unit": "money", "formula": "Soma dos buckets >= 61 dias."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 60d / Carteira", "column": "npl_over60_pct", "unit": "percent", "formula": "NPL Over 60d / Carteira Bruta."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 90d", "column": "npl_over90", "unit": "money", "formula": "Soma dos buckets >= 91 dias."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 90d / Carteira", "column": "npl_over90_pct", "unit": "percent", "formula": "NPL Over 90d / Carteira Bruta."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 180d", "column": "npl_over180", "unit": "money", "formula": "Soma dos buckets >= 181 dias."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 180d / Carteira", "column": "npl_over180_pct", "unit": "percent", "formula": "NPL Over 180d / Carteira Bruta."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 360d", "column": "npl_over360", "unit": "money", "formula": "Vencidos acima de 360 dias."},
        {"block": "5. NPL Over acumulado", "metric": "NPL Over 360d / Carteira", "column": "npl_over360_pct", "unit": "percent", "formula": "NPL Over 360d / Carteira Bruta."},
        {"block": "6. Cobertura", "metric": "PDD / NPL Over 1d", "column": "pdd_npl_over1_pct", "unit": "percent", "formula": "PDD total / NPL Over 1d."},
        {"block": "6. Cobertura", "metric": "PDD / NPL Over 30d", "column": "pdd_npl_over30_pct", "unit": "percent", "formula": "PDD total / NPL Over 30d."},
        {"block": "6. Cobertura", "metric": "PDD / NPL Over 60d", "column": "pdd_npl_over60_pct", "unit": "percent", "formula": "PDD total / NPL Over 60d."},
        {"block": "6. Cobertura", "metric": "PDD / NPL Over 90d", "column": "pdd_npl_over90_pct", "unit": "percent", "formula": "PDD total / NPL Over 90d."},
        {"block": "6. Cobertura", "metric": "PDD / NPL Over 180d", "column": "pdd_npl_over180_pct", "unit": "percent", "formula": "PDD total / NPL Over 180d."},
        {"block": "6. Cobertura", "metric": "PDD / NPL Over 360d", "column": "pdd_npl_over360_pct", "unit": "percent", "formula": "PDD total / NPL Over 360d, quando denominador > 0."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "Carteira Ex Over 360d", "column": "carteira_ex360", "unit": "money", "formula": "Carteira Bruta - NPL Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "Baixa Over 360d na carteira", "column": "baixa_over360_carteira", "unit": "money", "formula": "Saldo vencido acima de 360 dias baixado da carteira."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "Baixa Over 360d na PDD", "column": "baixa_over360_pdd", "unit": "money", "formula": "Menor valor entre PDD total e saldo Over 360d baixado."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "Baixa Over 360d no PL", "column": "baixa_over360_pl", "unit": "money", "formula": "Parcela da baixa Over 360d não coberta por PDD."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "PL Ex Over 360d", "column": "pl_total_ex360", "unit": "money", "formula": "PL FIDC total - baixa Over 360d no PL."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "Subordinada + Mezanino Ex Over 360d", "column": "pl_subordinada_mezz_ex360", "unit": "money", "formula": "Subordinada + Mezanino após baixa Over 360d no PL."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "% Subordinação Total Ex Over 360d", "column": "subordinacao_total_ex360_pct", "unit": "percent", "formula": "(Subordinada + Mezanino Ex Over 360d) / PL Ex Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "PDD Ex Over 360d", "column": "pdd_ex360", "unit": "money", "formula": "PDD total - baixa Over 360d na PDD."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "PDD Ex Over 360d / Carteira Ex Over 360d", "column": "pdd_ex360_carteira_ex360_pct", "unit": "percent", "formula": "PDD Ex Over 360d / Carteira Ex Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "Carteira Líquida Ex Over 360d", "column": "carteira_liquida_ex360", "unit": "money", "formula": "Carteira Ex Over 360d - PDD Ex Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 1d Ex 360", "column": "npl_over1_ex360", "unit": "money", "formula": "NPL Over 1d - NPL Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 1d Ex 360 / Carteira Ex 360", "column": "npl_over1_ex360_pct", "unit": "percent", "formula": "NPL Over 1d Ex 360 / Carteira Ex 360."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 30d Ex 360", "column": "npl_over30_ex360", "unit": "money", "formula": "NPL Over 30d - NPL Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 30d Ex 360 / Carteira Ex 360", "column": "npl_over30_ex360_pct", "unit": "percent", "formula": "NPL Over 30d Ex 360 / Carteira Ex 360."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 60d Ex 360", "column": "npl_over60_ex360", "unit": "money", "formula": "NPL Over 60d - NPL Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 60d Ex 360 / Carteira Ex 360", "column": "npl_over60_ex360_pct", "unit": "percent", "formula": "NPL Over 60d Ex 360 / Carteira Ex 360."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 90d Ex 360", "column": "npl_over90_ex360", "unit": "money", "formula": "NPL Over 90d - NPL Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 90d Ex 360 / Carteira Ex 360", "column": "npl_over90_ex360_pct", "unit": "percent", "formula": "NPL Over 90d Ex 360 / Carteira Ex 360."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 180d Ex 360", "column": "npl_over180_ex360", "unit": "money", "formula": "NPL Over 180d - NPL Over 360d."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "NPL Over 180d Ex 360 / Carteira Ex 360", "column": "npl_over180_ex360_pct", "unit": "percent", "formula": "NPL Over 180d Ex 360 / Carteira Ex 360."},
        {"block": "7. Visão Ex-Vencidos > 360d", "metric": "PDD / NPL Over 90d Ex 360", "column": "pdd_npl_over90_ex360_pct", "unit": "percent", "formula": "PDD Ex Over 360d / NPL Over 90d Ex 360."},
        {"block": "8. Roll Rate / fluxo de deterioração", "metric": "Carteira em dia + atrasada até 30d (t-1)", "column": "roll_rate_base_t_minus_1", "unit": "money", "formula": "Base do mês anterior alinhada à coluna do mês t."},
        {"block": "8. Roll Rate / fluxo de deterioração", "metric": "Atrasos 31-60d (t)", "column": "atraso_31_60", "unit": "money", "formula": "Bucket 31-60 dias do mês t."},
        {"block": "8. Roll Rate / fluxo de deterioração", "metric": "Roll Rate", "column": "roll_rate_31_60_pct", "unit": "percent", "formula": "Atrasos 31-60d_t / (Carteira em dia + atrasada até 30d)_{t-1}."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Numerador % Subordinação Total", "column": "pl_subordinada_mezz", "unit": "money", "formula": "Subordinada + Mezanino."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Denominador % Subordinação Total", "column": "pl_total", "unit": "money", "formula": "PL FIDC total."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Numerador NPL Over 90d / Carteira", "column": "npl_over90", "unit": "money", "formula": "NPL Over 90d."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Denominador NPL Over 90d / Carteira", "column": "carteira_bruta", "unit": "money", "formula": "Carteira Bruta total."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Fórmula aplicada NPL Over 90d", "column": "formula_npl_over90", "unit": "text", "formula": "Texto fixo da fórmula aplicada."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Origem da coluna Carteira Bruta", "column": "carteira_bruta_origem", "unit": "text", "formula": "Fonte efetiva selecionada pela base canônica."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Flag de dado ausente", "column": "missing_data_flag", "unit": "bool", "formula": "Verdadeiro quando campo crítico está ausente."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Flag de divisão por zero", "column": "division_by_zero_flag", "unit": "bool", "formula": "Verdadeiro quando denominador crítico é zero/ausente."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Flag de valor negativo", "column": "negative_value_flag", "unit": "bool", "formula": "Verdadeiro quando valor monetário crítico é negativo."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Flag de métrica não calculável", "column": "not_calculable_flag", "unit": "bool", "formula": "Verdadeiro quando faltam dados para calcular a baixa/PDD ex-360."},
        {"block": "9. Campos auxiliares de auditoria", "metric": "Observação técnica", "column": "warnings", "unit": "text", "formula": "Resumo dos warnings da competência."},
    ]


def _money_scales_by_block(df: pd.DataFrame) -> dict[str, tuple[float, str]]:
    specs = _wide_metric_specs()
    result: dict[str, tuple[float, str]] = {}
    for block in sorted({spec["block"] for spec in specs}):
        cols = [spec["column"] for spec in specs if spec["block"] == block and spec["unit"] == "money"]
        values = []
        for col in cols:
            if col in df.columns:
                values.extend(pd.to_numeric(df[col], errors="coerce").dropna().abs().tolist())
        result[block] = _money_scale(values)
    return result


def _money_scales_by_metric(df: pd.DataFrame) -> dict[str, tuple[float, str]]:
    specs = _wide_metric_specs()
    result: dict[str, tuple[float, str]] = {}
    for spec in specs:
        if spec["unit"] != "money":
            continue
        column = str(spec["column"])
        if column not in df.columns:
            continue
        values = pd.to_numeric(df[column], errors="coerce").dropna().abs().tolist()
        result[column] = _money_scale(values)
    return result


def _money_scale(values: list[float]) -> tuple[float, str]:
    max_value = max(values) if values else 0.0
    if max_value >= 1_000_000_000_000:
        return 1_000_000_000.0, "R$ bi"
    if max_value >= 1_000_000:
        return 1_000_000.0, "R$ mm"
    if max_value >= 1_000:
        return 1_000.0, "R$ mil"
    return 1.0, "R$"


def _format_wide_value(value: object, *, unit: str, scale: tuple[float, str] | None = None) -> str:
    if unit == "money":
        divisor, label = scale or (1.0, "R$")
        numeric = _num(value)
        if numeric is None:
            return "N/D"
        if label == "R$":
            return f"R$ {_format_decimal(numeric, 2)}"
        return f"{label} {_format_decimal(numeric / divisor, 1)}"
    if unit == "percent":
        numeric = _num(value)
        return "N/D" if numeric is None else f"{_format_decimal(numeric, 2)}%"
    if unit == "count":
        numeric = _num(value)
        return "N/D" if numeric is None else _format_decimal(numeric, 0)
    if unit == "bool":
        if pd.isna(value):
            return "N/D"
        return "Sim" if bool(value) else "Não"
    if pd.isna(value):
        return "N/D"
    return str(value)


def _format_decimal(value: object, decimals: int = 2) -> str:
    numeric = float(value)
    formatted = f"{numeric:,.{decimals}f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")


def _safe_div_pct(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    num = pd.to_numeric(numerator, errors="coerce")
    den = pd.to_numeric(denominator, errors="coerce")
    return (num / den).where(den > 0).mul(100.0)


def _row_min(left: pd.Series, right: pd.Series) -> pd.Series:
    left_num = pd.to_numeric(left, errors="coerce")
    right_num = pd.to_numeric(right, errors="coerce")
    return pd.concat([left_num, right_num], axis=1).min(axis=1, skipna=False)


def _coalesce_numeric(primary: pd.Series, fallback: pd.Series) -> pd.Series:
    primary_num = pd.to_numeric(primary, errors="coerce")
    fallback_num = pd.to_numeric(fallback, errors="coerce")
    return primary_num.where(primary_num.notna(), fallback_num)


def _latest_match(frame: pd.DataFrame, competencia: str) -> pd.Series:
    if not isinstance(frame, pd.DataFrame) or frame.empty or "competencia" not in frame.columns:
        return pd.Series(dtype="object")
    match = frame[frame["competencia"].astype(str) == str(competencia)]
    if match.empty:
        return pd.Series(dtype="object")
    return match.iloc[-1]


def _num(value: object) -> float | None:
    try:
        parsed = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    except Exception:  # noqa: BLE001
        return None
    if pd.isna(parsed):
        return None
    return float(parsed)


def _sum_numeric(series: pd.Series | None) -> float | None:
    if series is None:
        return None
    values = pd.to_numeric(series, errors="coerce")
    if values.notna().sum() == 0:
        return None
    return float(values.sum(skipna=True))


def _digits(value: object) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _first_non_empty(*values: object) -> str:
    for value in values:
        text = str(value or "").strip()
        if text and text.lower() not in {"nan", "none", "n/d", "<na>"}:
            return text
    return "N/D"


def _competencia_to_timestamp(value: object) -> pd.Timestamp:
    raw = str(value or "").strip()
    if re.fullmatch(r"\d{1,2}/\d{4}", raw):
        month, year = raw.split("/", 1)
        return pd.Timestamp(year=int(year), month=int(month), day=1)
    parsed = pd.to_datetime(raw, errors="coerce")
    return parsed if pd.notna(parsed) else pd.NaT


def _period_label_to_timestamp(value: object) -> pd.Timestamp | None:
    raw = str(value or "").strip().lower()
    if not raw:
        return None
    match = re.fullmatch(r"([a-zç]{3})/(\d{2}|\d{4})", raw)
    if match:
        month_label, year_text = match.groups()
        month_lookup = {label: month for month, label in PT_MONTH_ABBR.items()}
        month = month_lookup.get(month_label)
        if month is not None:
            year = int(year_text)
            if year < 100:
                year += 2000
            return pd.Timestamp(year=year, month=month, day=1)
    ts = _competencia_to_timestamp(raw)
    return ts if pd.notna(ts) else None


def _sort_monthly_by_competencia(monthly_df: pd.DataFrame, *, descending: bool) -> pd.DataFrame:
    if monthly_df.empty:
        return monthly_df.copy()
    df = monthly_df.copy()
    if "competencia_dt" not in df.columns:
        df["competencia_dt"] = df.get("competencia", pd.Series(index=df.index, dtype="object")).map(_competencia_to_timestamp)
    else:
        df["competencia_dt"] = pd.to_datetime(df["competencia_dt"], errors="coerce")
    return df.sort_values("competencia_dt", ascending=not descending).reset_index(drop=True)


def _format_competencia_short(value: object) -> str:
    ts = _competencia_to_timestamp(value)
    if pd.isna(ts):
        return str(value or "N/D")
    return f"{PT_MONTH_ABBR[int(ts.month)]}/{str(int(ts.year))[-2:]}"


def _competencia_list(frame: pd.DataFrame) -> list[str]:
    if frame.empty or "competencia" not in frame.columns:
        return []
    return frame["competencia"].astype(str).tolist()


def _loaded_period_label_from_frame(frame: pd.DataFrame) -> str:
    competencias = _competencia_list(frame)
    if not competencias:
        return "N/D"
    return f"{_format_competencia_short(competencias[0])} a {_format_competencia_short(competencias[-1])}"


def _row_warnings(row: pd.Series) -> str:
    warnings: list[str] = []
    if bool(row.get("missing_data_flag")):
        warnings.append("dado crítico ausente")
    if bool(row.get("division_by_zero_flag")):
        warnings.append("denominador zero/ausente")
    if bool(row.get("negative_value_flag")):
        warnings.append("valor monetário negativo")
    if str(row.get("pl_total_usado_fonte") or "") == "soma_classes_cotas":
        warnings.append("PL oficial indisponível; usando soma das classes")
    if bool(row.get("pl_reconciliacao_warning")):
        warnings.append("PL oficial diverge da soma das classes acima de 0,5%")
    if bool(row.get("pdd_over360_insuficiente")):
        warnings.append("PDD menor que Over 360; baixa residual afetou PL")
    if not bool(row.get("pdd_ex360_calculavel")):
        warnings.append("PDD ex-360 não calculável por ausência de PDD total ou Over 360")
    if int(row.get("funds_present_count") or 0) < int(row.get("funds_expected_count") or 0):
        warnings.append("consolidado parcial na competência")
    return "; ".join(warnings)


def _warning_rows(monthly_df: pd.DataFrame, *, scope_name: str, cnpj: str) -> list[dict[str, object]]:
    if monthly_df.empty:
        return [{"scope_name": scope_name, "cnpj": cnpj, "competencia": "", "warning": "base vazia"}]
    rows: list[dict[str, object]] = []
    for _, row in monthly_df.iterrows():
        warning = str(row.get("warnings") or "").strip()
        if warning:
            rows.append(
                {
                    "scope_name": scope_name,
                    "cnpj": cnpj,
                    "competencia": row.get("competencia"),
                    "warning": warning,
                }
            )
    return rows


def _formula_catalog() -> dict[str, str]:
    return {
        "subordinacao_total_pct": "(PL Subordinada + PL Mezanino) / PL FIDC total",
        "npl_over90_pct": "NPL Over 90d / Carteira Bruta",
        "pdd_npl_over90_pct": "PDD total / NPL Over 90d",
        "consolidado": "Somar valores absolutos por competência e recalcular percentuais a partir das somas.",
    }


def _empty_monthly_frame() -> pd.DataFrame:
    return pd.DataFrame(columns=["fund_name", "cnpj", "competencia", "competencia_dt"])


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _safe_path_token(value: object) -> str:
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "").strip())
    return text.strip("._") or "sem_id"


def _sheet_name_from_table(table: pd.DataFrame, *, fallback: str) -> str:
    if hasattr(table, "attrs"):
        name = str(table.attrs.get("scope_name") or "").strip()
        if name:
            return name
    return fallback


def _excel_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", value).strip() or "Sheet"
    return cleaned[:31]


def _unique_excel_sheet_name(value: str, used_names: set[str]) -> str:
    base = _excel_sheet_name(value)
    if base not in used_names:
        return base
    for suffix in range(2, 100):
        tail = f" {suffix}"
        candidate = _excel_sheet_name(f"{base[:31 - len(tail)]}{tail}")
        if candidate not in used_names:
            return candidate
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:6]
    return _excel_sheet_name(f"{base[:24]} {digest}")


def _fund_sheet_name(monthly_df: pd.DataFrame, *, fallback: str) -> str:
    if isinstance(monthly_df, pd.DataFrame) and not monthly_df.empty and "fund_name" in monthly_df.columns:
        for value in monthly_df["fund_name"].dropna().tolist():
            name = str(value).strip()
            if name and name.lower() not in {"nan", "none", "n/d", "<na>"}:
                return name
    return fallback


def _write_numeric_wide_sheet(worksheet, monthly_df: pd.DataFrame, *, scope_name: str) -> None:  # noqa: ANN001
    worksheet.sheet_view.showGridLines = False
    if monthly_df.empty:
        worksheet.append(["Métrica", "Memória / fórmula"])
        worksheet.append(["Sem dados", ""])
        _style_plain_sheet(worksheet)
        return

    sorted_df = _sort_monthly_by_competencia(monthly_df, descending=True)
    competencias = sorted_df["competencia"].astype(str).tolist()
    period_labels = [_format_competencia_short(value) for value in competencias]
    metric_scales = _money_scales_by_metric(sorted_df)
    columns = ["Métrica", *period_labels, "Memória / fórmula"]
    worksheet.append(columns)

    current_block = None
    metric_specs = _wide_metric_specs()
    for spec in metric_specs:
        block = str(spec["block"])
        if block != current_block:
            current_block = block
            worksheet.append([_section_label(block), *([""] * (len(columns) - 1))])
            _style_excel_section_row(worksheet, worksheet.max_row)

        row_values: list[object] = [_excel_text(spec["metric"])]
        for _, item in sorted_df.iterrows():
            row_values.append(_excel_metric_value(item.get(spec["column"]), unit=str(spec["unit"])))
        row_values.append(_excel_text(spec["formula"]))
        worksheet.append(row_values)
        _style_excel_metric_row(
            worksheet,
            worksheet.max_row,
            unit=str(spec["unit"]),
            scale=metric_scales.get(str(spec["column"])),
            period_start_col=2,
            period_end_col=1 + len(period_labels),
        )

    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _style_numeric_wide_sheet(worksheet, title=scope_name)


def _write_dataframe_sheet(worksheet, df: pd.DataFrame) -> None:  # noqa: ANN001
    if df is None or df.empty:
        worksheet.append(["Sem dados"])
        _style_plain_sheet(worksheet)
        return
    columns = [str(column) for column in df.columns]
    worksheet.append(columns)
    for _, row in df.iterrows():
        values = []
        for column in columns:
            value = row.get(column)
            if column.endswith("_pct") and _num(value) is not None:
                values.append(float(_num(value) or 0.0) / 100.0)
            elif isinstance(value, (dict, list, tuple)):
                values.append(_excel_text(json.dumps(value, ensure_ascii=False)))
            elif pd.isna(value):
                values.append(None)
            elif isinstance(value, str):
                values.append(_excel_text(value))
            else:
                values.append(value)
        worksheet.append(values)
    for col_idx, column in enumerate(columns, start=1):
        if column.endswith("_pct"):
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=worksheet.max_row):
                for item in cell:
                    item.number_format = "0.00%"
        elif column in MONEY_COLUMNS or column.startswith(("pl_", "carteira_", "pdd_", "npl_", "atraso_", "baixa_")):
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=worksheet.max_row):
                for item in cell:
                    if isinstance(item.value, (int, float)):
                        item.number_format = '#,##0.00'
    _style_plain_sheet(worksheet)


def _excel_metric_value(value: object, *, unit: str) -> object:
    if unit in {"money", "count"}:
        return _num(value)
    if unit == "percent":
        numeric = _num(value)
        return None if numeric is None else numeric / 100.0
    if unit == "bool":
        if pd.isna(value):
            return None
        return "Sim" if bool(value) else "Não"
    if pd.isna(value):
        return None
    text = str(value)
    return None if text.strip().upper() in {"", "N/D", "NAN", "NONE", "<NA>"} else _excel_text(text)


def _excel_text(value: object) -> str:
    text = str(value or "")
    return ILLEGAL_CHARACTERS_RE.sub("", text)


def _excel_number_format_for_unit(unit: str, scale: tuple[float, str] | None) -> str:
    if unit == "percent":
        return "0.00%"
    if unit == "count":
        return "#,##0"
    if unit == "money":
        _divisor, label = scale or (1.0, "R$")
        if label == "R$ bi":
            return '"R$" #,##0.00,,, "bi"'
        if label == "R$ mm":
            return '"R$" #,##0.00,, "mm"'
        if label == "R$ mil":
            return '"R$" #,##0.00, "mil"'
        return '"R$" #,##0.00'
    return "General"


def _excel_metric_is_highlight(metric: str) -> bool:
    normalized = metric.lower()
    destaque_terms = (
        "pl fidc total",
        "carteira bruta total",
        "carteira líquida",
        "npl over 90d / carteira",
        "npl over 90d ex 360 / carteira ex 360",
        "pdd / npl over 90d",
        "pdd / npl over 90d ex 360",
        "% subordinação total",
        "% subordinação total ex over 360d",
        "carteira ex over 360d",
        "carteira líquida ex over 360d",
        "pl ex over 360d",
        "pdd ex over 360d",
    )
    return any(term in normalized for term in destaque_terms)


def _style_excel_metric_row(
    worksheet,
    row_idx: int,
    *,
    unit: str,
    scale: tuple[float, str] | None,
    period_start_col: int,
    period_end_col: int,
) -> None:  # noqa: ANN001
    if _excel_metric_is_highlight(str(worksheet.cell(row=row_idx, column=1).value or "")):
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)
    number_format = _excel_number_format_for_unit(unit, scale)
    for col_idx in range(period_start_col, period_end_col + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        if isinstance(cell.value, (int, float)):
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right")
    worksheet.cell(row=row_idx, column=1).alignment = Alignment(indent=1)
    worksheet.cell(row=row_idx, column=worksheet.max_column).alignment = Alignment(wrap_text=True, vertical="top")


def _style_excel_section_row(worksheet, row_idx: int) -> None:  # noqa: ANN001
    fill = PatternFill("solid", fgColor="000000")
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left")


def _style_numeric_wide_sheet(worksheet, *, title: str) -> None:  # noqa: ANN001
    header_fill = PatternFill("solid", fgColor="000000")
    thin = Side(style="thin", color="E5E5E5")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = Border(bottom=thin)
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions[get_column_letter(worksheet.max_column)].width = 58
    for col_idx in range(2, worksheet.max_column):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 14
    worksheet.auto_filter.ref = worksheet.dimensions


def _section_label(block: str) -> str:
    parts = str(block).split(".", 1)
    if len(parts) == 2 and parts[0].strip().isdigit():
        return parts[1].strip()
    return str(block)


def _style_wide_sheet(worksheet, table: pd.DataFrame) -> None:  # noqa: ANN001
    header_fill = PatternFill("solid", fgColor="1F77B4")
    block_fill = PatternFill("solid", fgColor="EAF2FA")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    block_col = 1
    previous_block = None
    for row_idx in range(2, worksheet.max_row + 1):
        block = worksheet.cell(row=row_idx, column=block_col).value
        if block != previous_block:
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = block_fill
                cell.border = border
            previous_block = block
        worksheet.cell(row=row_idx, column=2).alignment = Alignment(indent=1)
    _auto_width(worksheet)
    worksheet.freeze_panes = "D2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _style_plain_sheet(worksheet) -> None:  # noqa: ANN001
    header_fill = PatternFill("solid", fgColor="1F77B4")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    _auto_width(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _auto_width(worksheet) -> None:  # noqa: ANN001
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 48)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width
