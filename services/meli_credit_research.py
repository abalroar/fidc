from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

import pandas as pd

from services.meli_credit_monitor import PT_MONTH_ABBR_TITLE


RESEARCH_SCHEMA_VERSION = 1

ROLL_RESEARCH_METRICS: tuple[dict[str, str], ...] = (
    {
        "metric_id": "roll_61_90_m3",
        "metric_name": "Roll 61-90 M-3",
        "definition": "Carteira que chegou a 61-90 dias de atraso dividida pela carteira a vencer três meses antes.",
        "numerator_col": "atraso_61_90",
        "denominator_col": "roll_61_90_m3_den",
        "formula": "atraso_61_90_t / carteira_a_vencer_t-3",
        "source_columns": "atraso_61_90; carteira_a_vencer",
    },
    {
        "metric_id": "roll_151_180_m6",
        "metric_name": "Roll 151-180 M-6",
        "definition": "Carteira que chegou a 151-180 dias de atraso dividida pela carteira a vencer seis meses antes.",
        "numerator_col": "atraso_151_180",
        "denominator_col": "roll_151_180_m6_den",
        "formula": "atraso_151_180_t / carteira_a_vencer_t-6",
        "source_columns": "atraso_151_180; carteira_a_vencer",
    },
)

NPL_RESEARCH_ROWS: tuple[dict[str, str], ...] = (
    {
        "block": "NPL Breakdown (R$)",
        "metric_id": "npl_1_360",
        "metric_name": "NPL 1-360d",
        "unit": "R$",
        "formula": "npl_1_90 + npl_91_360",
        "source_columns": "atraso_ate30; atraso_31_60; atraso_61_90; atraso_91_120; atraso_121_150; atraso_151_180; atraso_181_360",
    },
    {
        "block": "NPL Breakdown (R$)",
        "metric_id": "npl_1_30",
        "metric_name": "NPL 1-30d",
        "unit": "R$",
        "formula": "atraso_ate30",
        "source_columns": "atraso_ate30",
    },
    {
        "block": "NPL Breakdown (R$)",
        "metric_id": "npl_31_60",
        "metric_name": "NPL 31-60d",
        "unit": "R$",
        "formula": "atraso_31_60",
        "source_columns": "atraso_31_60",
    },
    {
        "block": "NPL Breakdown (R$)",
        "metric_id": "npl_61_90",
        "metric_name": "NPL 61-90d",
        "unit": "R$",
        "formula": "atraso_61_90",
        "source_columns": "atraso_61_90",
    },
    {
        "block": "NPL Breakdown (R$)",
        "metric_id": "npl_91_180",
        "metric_name": "NPL 91-180d",
        "unit": "R$",
        "formula": "atraso_91_120 + atraso_121_150 + atraso_151_180",
        "source_columns": "atraso_91_120; atraso_121_150; atraso_151_180",
    },
    {
        "block": "NPL Breakdown (R$)",
        "metric_id": "npl_181_360",
        "metric_name": "NPL 181-360d",
        "unit": "R$",
        "formula": "atraso_181_360",
        "source_columns": "atraso_181_360",
    },
    {
        "block": "% NPL of Portfolio ex-360",
        "metric_id": "npl_1_360_pct",
        "metric_name": "NPL 1-360d / carteira ex-360",
        "unit": "%",
        "formula": "(npl_1_90 + npl_91_360) / carteira_ex360",
        "source_columns": "npl_1_90; npl_91_360; carteira_ex360",
    },
    {
        "block": "% NPL of Portfolio ex-360",
        "metric_id": "npl_1_30_pct",
        "metric_name": "NPL 1-30d / carteira ex-360",
        "unit": "%",
        "formula": "atraso_ate30 / carteira_ex360",
        "source_columns": "atraso_ate30; carteira_ex360",
    },
    {
        "block": "% NPL of Portfolio ex-360",
        "metric_id": "npl_31_60_pct",
        "metric_name": "NPL 31-60d / carteira ex-360",
        "unit": "%",
        "formula": "atraso_31_60 / carteira_ex360",
        "source_columns": "atraso_31_60; carteira_ex360",
    },
    {
        "block": "% NPL of Portfolio ex-360",
        "metric_id": "npl_61_90_pct",
        "metric_name": "NPL 61-90d / carteira ex-360",
        "unit": "%",
        "formula": "atraso_61_90 / carteira_ex360",
        "source_columns": "atraso_61_90; carteira_ex360",
    },
    {
        "block": "% NPL of Portfolio ex-360",
        "metric_id": "npl_91_180_pct",
        "metric_name": "NPL 91-180d / carteira ex-360",
        "unit": "%",
        "formula": "(atraso_91_120 + atraso_121_150 + atraso_151_180) / carteira_ex360",
        "source_columns": "atraso_91_120; atraso_121_150; atraso_151_180; carteira_ex360",
    },
    {
        "block": "% NPL of Portfolio ex-360",
        "metric_id": "npl_181_360_pct",
        "metric_name": "NPL 181-360d / carteira ex-360",
        "unit": "%",
        "formula": "atraso_181_360 / carteira_ex360",
        "source_columns": "atraso_181_360; carteira_ex360",
    },
    {
        "block": "Total Credit Portfolio (<360)",
        "metric_id": "carteira_ex360",
        "metric_name": "Total Credit Portfolio (<360)",
        "unit": "R$",
        "formula": "carteira_bruta - npl_over360",
        "source_columns": "carteira_bruta; npl_over360",
    },
)


@dataclass(frozen=True)
class MeliResearchOutputs:
    roll_seasonality: pd.DataFrame
    cohort_research: pd.DataFrame
    npl_research_table: pd.DataFrame
    portfolio_duration_table: pd.DataFrame
    methodology: pd.DataFrame
    warnings: list[str]


def build_meli_research_outputs(monitor_outputs: Any) -> MeliResearchOutputs:
    """Build research-style derived datasets from already cached monitor outputs."""
    warnings: list[str] = []
    roll_frames: list[pd.DataFrame] = []
    cohort_frames: list[pd.DataFrame] = []
    npl_frames: list[pd.DataFrame] = []
    portfolio_duration_frames: list[pd.DataFrame] = []

    for scope, fund_name, cnpj, monitor, cohorts in _scope_frames(monitor_outputs):
        roll = build_roll_seasonality_frame(monitor, scope=scope, fund_name=fund_name, cnpj=cnpj)
        cohorts_research, cohort_warnings = build_cohort_research_frame(cohorts, scope=scope, fund_name=fund_name, cnpj=cnpj)
        npl_table = build_npl_research_table(monitor, scope=scope, fund_name=fund_name, cnpj=cnpj)
        portfolio_duration = build_portfolio_duration_table(monitor, scope=scope, fund_name=fund_name, cnpj=cnpj)
        roll_frames.append(roll)
        cohort_frames.append(cohorts_research)
        npl_frames.append(npl_table)
        portfolio_duration_frames.append(portfolio_duration)
        warnings.extend(cohort_warnings)
        warnings.extend(_research_warnings(monitor, scope=scope, fund_name=fund_name))

    return MeliResearchOutputs(
        roll_seasonality=_concat_frames(roll_frames),
        cohort_research=_concat_frames(cohort_frames),
        npl_research_table=_concat_frames(npl_frames),
        portfolio_duration_table=_concat_frames(portfolio_duration_frames),
        methodology=build_research_methodology_table(),
        warnings=warnings,
    )


def build_roll_seasonality_frame(
    monitor_df: pd.DataFrame,
    *,
    scope: str,
    fund_name: str,
    cnpj: str,
) -> pd.DataFrame:
    df = _prepare_monthly(monitor_df)
    rows: list[dict[str, object]] = []
    if df.empty:
        return pd.DataFrame()
    for metric in ROLL_RESEARCH_METRICS:
        numerator_col = metric["numerator_col"]
        denominator_col = metric["denominator_col"]
        for _, row in df.iterrows():
            numerator = _num(row.get(numerator_col))
            denominator = _num(row.get(denominator_col))
            value = _safe_div_pct_scalar(numerator, denominator)
            dt = pd.to_datetime(row.get("competencia_dt"), errors="coerce")
            missing = numerator is None or denominator is None
            div_zero = denominator is not None and denominator <= 0
            rows.append(
                {
                    "schema_version": RESEARCH_SCHEMA_VERSION,
                    "scope": scope,
                    "fund_name": fund_name,
                    "cnpj": cnpj,
                    "metric_id": metric["metric_id"],
                    "metric_name": metric["metric_name"],
                    "definition": metric["definition"],
                    "competencia": row.get("competencia"),
                    "competencia_dt": dt,
                    "year": int(dt.year) if pd.notna(dt) else pd.NA,
                    "month": int(dt.month) if pd.notna(dt) else pd.NA,
                    "month_label": _month_label(dt),
                    "series_name": str(int(dt.year)) if pd.notna(dt) else "N/D",
                    "value_pct": value,
                    "numerator": numerator,
                    "denominator": denominator,
                    "formula": metric["formula"],
                    "unit": "%",
                    "source_columns": metric["source_columns"],
                    "missing_data_flag": bool(missing),
                    "division_by_zero_flag": bool(div_zero),
                }
            )
    return pd.DataFrame(rows)


def build_cohort_research_frame(
    cohort_df: pd.DataFrame,
    *,
    scope: str,
    fund_name: str,
    cnpj: str,
    max_recent: int = 6,
) -> tuple[pd.DataFrame, list[str]]:
    if cohort_df is None or cohort_df.empty:
        return pd.DataFrame(), [f"{fund_name}: cohorts research não calculáveis porque a matriz de cohorts está vazia."]
    df = cohort_df.copy()
    df["cohort_dt"] = pd.to_datetime(df["cohort_dt"], errors="coerce")
    df["value_pct"] = pd.to_numeric(df.get("valor_pct"), errors="coerce")
    df["numerator"] = pd.to_numeric(df.get("numerador"), errors="coerce")
    df["denominator"] = pd.to_numeric(df.get("denominador"), errors="coerce")
    df = df.sort_values(["cohort_dt", "ordem"]).reset_index(drop=True)
    warnings: list[str] = []
    rows: list[dict[str, object]] = []

    recent_cohorts = (
        df[["cohort", "cohort_dt"]]
        .dropna(subset=["cohort_dt"])
        .drop_duplicates()
        .sort_values("cohort_dt")
        .tail(max_recent)
    )
    for _, cohort_row in recent_cohorts.iterrows():
        cohort_name = str(cohort_row["cohort"])
        cohort_rows = df[df["cohort"].eq(cohort_name)].copy()
        for _, row in cohort_rows.iterrows():
            rows.append(
                _cohort_output_row(
                    row,
                    scope=scope,
                    fund_name=fund_name,
                    cnpj=cnpj,
                    series_name=cohort_name,
                    series_type="Safra recente",
                    line_rank=int(cohort_row.name) if str(cohort_row.name).isdigit() else 0,
                )
            )

    years = sorted({int(value) for value in df["cohort_dt"].dt.year.dropna().unique().tolist()})
    latest_year = int(df["cohort_dt"].dt.year.max()) if df["cohort_dt"].notna().any() else None
    benchmark_years = [year for year in years if latest_year is None or year < latest_year]
    for year in benchmark_years[-3:]:
        rows.extend(
            _weighted_cohort_benchmark_rows(
                df[df["cohort_dt"].dt.year.eq(year)],
                scope=scope,
                fund_name=fund_name,
                cnpj=cnpj,
                series_name=f"{year} Avg",
                series_type="Média anual",
            )
        )

    cohort_lookup = (
        df[["cohort", "cohort_dt"]]
        .dropna(subset=["cohort_dt"])
        .drop_duplicates()
        .sort_values("cohort_dt")
        .tail(12)["cohort"]
        .tolist()
    )
    if cohort_lookup:
        rows.extend(
            _weighted_cohort_benchmark_rows(
                df[df["cohort"].isin(cohort_lookup)],
                scope=scope,
                fund_name=fund_name,
                cnpj=cnpj,
                series_name="LTM Avg",
                series_type="Média LTM",
            )
        )
    else:
        warnings.append(f"{fund_name}: LTM Avg de cohorts não calculável por ausência de cohorts válidos.")

    out = pd.DataFrame(rows)
    if out.empty:
        warnings.append(f"{fund_name}: nenhuma linha de cohort research foi calculada.")
    return out, warnings


def build_npl_research_table(
    monitor_df: pd.DataFrame,
    *,
    scope: str,
    fund_name: str,
    cnpj: str,
) -> pd.DataFrame:
    df = _prepare_monthly(monitor_df)
    if df.empty:
        return pd.DataFrame()
    numeric = _npl_research_values(df)
    rows: list[dict[str, object]] = []
    for spec in NPL_RESEARCH_ROWS:
        metric_id = spec["metric_id"]
        values = numeric.get(metric_id, pd.Series(index=df.index, dtype="float64"))
        unit = spec["unit"]
        mom = _variation(values, unit=unit, periods=1)
        yoy = _variation(values, unit=unit, periods=12)
        for idx, row in df.iterrows():
            value = _num(values.iloc[idx])
            numerator, denominator = _npl_numerator_denominator(metric_id, numeric, idx)
            rows.append(
                {
                    "schema_version": RESEARCH_SCHEMA_VERSION,
                    "scope": scope,
                    "fund_name": fund_name,
                    "cnpj": cnpj,
                    "table_name": "NPL and credit portfolio ex-360",
                    "block": spec["block"],
                    "metric_id": metric_id,
                    "metric_name": spec["metric_name"],
                    "competencia": row.get("competencia"),
                    "competencia_dt": row.get("competencia_dt"),
                    "value": value,
                    "unit": unit,
                    "mom_value": _num(mom.iloc[idx]),
                    "yoy_value": _num(yoy.iloc[idx]),
                    "variation_unit": "p.p." if unit == "%" else "%",
                    "numerator": numerator,
                    "denominator": denominator,
                    "formula": spec["formula"],
                    "source_columns": spec["source_columns"],
                    "missing_data_flag": bool(value is None),
                    "division_by_zero_flag": bool(unit == "%" and (_num(denominator) is None or _num(denominator) <= 0)),
                }
            )
    return pd.DataFrame(rows)


def build_portfolio_duration_table(
    monitor_df: pd.DataFrame,
    *,
    scope: str,
    fund_name: str,
    cnpj: str,
) -> pd.DataFrame:
    df = _prepare_monthly(monitor_df)
    if df.empty:
        return pd.DataFrame()
    specs = (
        {
            "block": "Total Credit Portfolio (<360)",
            "metric_id": "carteira_ex360",
            "metric_name": fund_name,
            "unit": "R$",
            "column": "carteira_ex360",
            "formula": "carteira_bruta - npl_over360",
            "source_columns": "carteira_bruta; npl_over360",
        },
        {
            "block": "Duration (meses)",
            "metric_id": "duration_months",
            "metric_name": fund_name,
            "unit": "meses",
            "column": "duration_months",
            "formula": "(Σ saldo_bucket × prazo_proxy_bucket / Σ saldo_bucket) / 30,4375",
            "source_columns": "duration_weighted_days; duration_total_saldo; malha de vencimentos",
        },
    )
    rows: list[dict[str, object]] = []
    for spec in specs:
        values = pd.to_numeric(df.get(spec["column"]), errors="coerce")
        mom = _variation(values, unit=str(spec["unit"]), periods=1)
        yoy = _variation(values, unit=str(spec["unit"]), periods=12)
        for idx, row in df.iterrows():
            value = _num(values.iloc[idx])
            denominator = _num(row.get("duration_total_saldo")) if spec["metric_id"] == "duration_months" else None
            numerator = _num(row.get("duration_weighted_days")) if spec["metric_id"] == "duration_months" else value
            rows.append(
                {
                    "schema_version": RESEARCH_SCHEMA_VERSION,
                    "scope": scope,
                    "fund_name": fund_name,
                    "cnpj": cnpj,
                    "table_name": "Portfolio and duration ex-360",
                    "block": spec["block"],
                    "metric_id": spec["metric_id"],
                    "metric_name": spec["metric_name"],
                    "competencia": row.get("competencia"),
                    "competencia_dt": row.get("competencia_dt"),
                    "value": value,
                    "unit": spec["unit"],
                    "mom_value": _num(mom.iloc[idx]),
                    "yoy_value": _num(yoy.iloc[idx]),
                    "variation_unit": "%",
                    "numerator": numerator,
                    "denominator": denominator,
                    "formula": spec["formula"],
                    "source_columns": spec["source_columns"],
                    "missing_data_flag": bool(value is None),
                    "division_by_zero_flag": False,
                }
            )
    return pd.DataFrame(rows)


def build_research_methodology_table() -> pd.DataFrame:
    rows: list[dict[str, str]] = [
        {
            "Gráfico / indicador": "Roll rate por ano - 61-90",
            "Definição": "Mede sazonalidade anual da migração para atraso 61-90 dias.",
            "Numerador": "atraso_61_90_t",
            "Denominador": "carteira_a_vencer_t-3",
            "Fórmula": "atraso_61_90_t / carteira_a_vencer_t-3",
            "Unidade": "%",
            "Fonte / coluna": "atraso_61_90; carteira_a_vencer",
            "Observação": "A comparação por mês do ano segue o padrão research; percentuais são recalculados por numerador e denominador.",
        },
        {
            "Gráfico / indicador": "Roll rate por ano - 151-180",
            "Definição": "Mede sazonalidade anual da migração para atraso 151-180 dias.",
            "Numerador": "atraso_151_180_t",
            "Denominador": "carteira_a_vencer_t-6",
            "Fórmula": "atraso_151_180_t / carteira_a_vencer_t-6",
            "Unidade": "%",
            "Fonte / coluna": "atraso_151_180; carteira_a_vencer",
            "Observação": "A defasagem de seis meses aproxima o tempo necessário para a safra migrar para 151-180 dias.",
        },
        {
            "Gráfico / indicador": "Cohorts com médias",
            "Definição": "Compara safras recentes contra médias históricas e LTM usando a mesma base inicial da safra.",
            "Numerador": "M1=até 30d no mês seguinte; M2=31-60d dois meses depois; M3=61-90d três meses depois; M4=91-120d quatro meses depois; M5=121-150d cinco meses depois; M6=151-180d seis meses depois.",
            "Denominador": "prazo_venc_30 da competência-base da safra.",
            "Fórmula": "cohort_m = atraso_bucket_t+m / prazo_venc_30_t",
            "Unidade": "%",
            "Fonte / coluna": "prazo_venc_30; atraso_ate30; atraso_31_60; atraso_61_90; atraso_91_120; atraso_121_150; atraso_151_180",
            "Observação": "Médias anuais e LTM são ponderadas por denominador: soma dos numeradores dividida pela soma dos denominadores.",
        },
        {
            "Gráfico / indicador": "Tabela NPL ex-360",
            "Definição": "Decompõe o NPL até 360 dias em valores absolutos e percentual da carteira ex-360.",
            "Numerador": "Bucket de atraso ou soma de buckets.",
            "Denominador": "carteira_ex360 nas linhas percentuais.",
            "Fórmula": "bucket / carteira_ex360; carteira_ex360 = carteira_bruta - npl_over360",
            "Unidade": "R$ e %",
            "Fonte / coluna": "Balanço de carteira e faixas de atraso do Informe Mensal Estruturado.",
            "Observação": "A visão exclui vencidos acima de 360 dias, coerente com acompanhamento de carteira limpa.",
        },
        {
            "Gráfico / indicador": "Portfolio e duration",
            "Definição": "Mostra carteira ex-360 e prazo médio ponderado por saldo.",
            "Numerador": "Carteira ex-360; ou Σ saldo_bucket × prazo_proxy_bucket para duration.",
            "Denominador": "Não aplicável para carteira; Σ saldo_bucket para duration.",
            "Fórmula": "duration_meses = (Σ saldo_bucket × prazo_proxy_bucket / Σ saldo_bucket) / 30,4375",
            "Unidade": "R$ e meses",
            "Fonte / coluna": "carteira_bruta; npl_over360; malha de vencimentos.",
            "Observação": "30,4375 = 365,25 / 12. No consolidado, duration vem da base consolidada ponderada por saldo, não de média simples entre fundos.",
        },
    ]
    return pd.DataFrame(rows)


def build_research_excel_bytes(research_outputs: MeliResearchOutputs, verification_report: pd.DataFrame | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = [
        ("Roll seasonality", research_outputs.roll_seasonality),
        ("Cohorts", research_outputs.cohort_research),
        ("NPL research", research_outputs.npl_research_table),
        ("Portfolio duration", research_outputs.portfolio_duration_table),
        ("Methodology", research_outputs.methodology),
        ("Warnings", pd.DataFrame({"warning": research_outputs.warnings})),
    ]
    if verification_report is not None:
        sheets.insert(4, ("Verification", verification_report))

    for name, frame in sheets:
        ws = wb.create_sheet(name[:31])
        _write_dataframe_numeric(ws, frame, Font=Font, PatternFill=PatternFill, get_column_letter=get_column_letter)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_dataframe_numeric(ws, frame: pd.DataFrame, *, Font, PatternFill, get_column_letter) -> None:  # noqa: ANN001
    if frame is None or frame.empty:
        ws.cell(row=1, column=1, value="Sem dados")
        return
    df = frame.copy()
    headers = list(df.columns)
    header_fill = PatternFill("solid", fgColor="000000")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        unit = str(row.get("unit") or row.get("Unidade") or "")
        variation_unit = str(row.get("variation_unit") or "")
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(value):
                cell.value = None
                continue
            if isinstance(value, pd.Timestamp):
                cell.value = value.to_pydatetime()
                cell.number_format = "mmm-yy"
            elif _is_number(value):
                number = float(value)
                if header in {"value", "value_pct", "calculated_value", "verified_value"} and unit == "%":
                    cell.value = number / 100.0
                    cell.number_format = "0.00%"
                elif header in {"mom_value", "yoy_value"} and variation_unit == "%":
                    cell.value = number / 100.0
                    cell.number_format = "0.00%"
                elif header in {"mom_value", "yoy_value"} and variation_unit == "p.p.":
                    cell.value = number
                    cell.number_format = '0.00 "p.p."'
                elif header in {"rel_diff_pct"}:
                    cell.value = number / 100.0
                    cell.number_format = "0.00%"
                elif unit == "R$" and header in {"value", "numerator", "denominator"}:
                    cell.value = number
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = number
                    cell.number_format = '#,##0.0000'
            else:
                cell.value = str(value)
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 12), 34)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def _scope_frames(monitor_outputs: Any) -> list[tuple[str, str, str, pd.DataFrame, pd.DataFrame]]:
    frames: list[tuple[str, str, str, pd.DataFrame, pd.DataFrame]] = []
    consolidated = getattr(monitor_outputs, "consolidated_monitor", pd.DataFrame())
    consolidated_cohorts = getattr(monitor_outputs, "consolidated_cohorts", pd.DataFrame())
    frames.append(("consolidado", "Consolidado", "", consolidated, consolidated_cohorts))
    for cnpj, frame in getattr(monitor_outputs, "fund_monitor", {}).items():
        name = _fund_name(frame, fallback=str(cnpj))
        cohorts = getattr(monitor_outputs, "fund_cohorts", {}).get(cnpj, pd.DataFrame())
        frames.append(("fundo", name, str(cnpj), frame, cohorts))
    return frames


def _prepare_monthly(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out["competencia_dt"] = pd.to_datetime(out.get("competencia_dt", out.get("competencia")), errors="coerce")
    return out.sort_values("competencia_dt").reset_index(drop=True)


def _npl_research_values(df: pd.DataFrame) -> dict[str, pd.Series]:
    values: dict[str, pd.Series] = {}
    values["npl_1_30"] = pd.to_numeric(df.get("atraso_ate30"), errors="coerce")
    values["npl_31_60"] = pd.to_numeric(df.get("atraso_31_60"), errors="coerce")
    values["npl_61_90"] = pd.to_numeric(df.get("atraso_61_90"), errors="coerce")
    values["npl_91_180"] = _sum_columns(df, ["atraso_91_120", "atraso_121_150", "atraso_151_180"])
    values["npl_181_360"] = pd.to_numeric(df.get("atraso_181_360"), errors="coerce")
    values["npl_1_360"] = values["npl_1_30"] + values["npl_31_60"] + values["npl_61_90"] + values["npl_91_180"] + values["npl_181_360"]
    values["carteira_ex360"] = pd.to_numeric(df.get("carteira_ex360"), errors="coerce")
    for key in ("npl_1_360", "npl_1_30", "npl_31_60", "npl_61_90", "npl_91_180", "npl_181_360"):
        values[f"{key}_pct"] = _safe_div_pct_series(values[key], values["carteira_ex360"])
    return values


def _npl_numerator_denominator(metric_id: str, values: dict[str, pd.Series], idx: int) -> tuple[float | None, float | None]:
    if metric_id.endswith("_pct"):
        base_metric = metric_id.removesuffix("_pct")
        return _num(values.get(base_metric, pd.Series(dtype="float64")).iloc[idx]), _num(values["carteira_ex360"].iloc[idx])
    if metric_id == "carteira_ex360":
        value = _num(values["carteira_ex360"].iloc[idx])
        return value, None
    value = _num(values.get(metric_id, pd.Series(dtype="float64")).iloc[idx])
    return value, None


def _sum_columns(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    existing = [pd.to_numeric(df.get(column), errors="coerce") for column in columns]
    return pd.concat(existing, axis=1).sum(axis=1, min_count=1)


def _variation(values: pd.Series, *, unit: str, periods: int) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce")
    if unit == "%":
        return numeric - numeric.shift(periods)
    base = numeric.shift(periods)
    return (numeric / base - 1.0).where(base > 0).mul(100.0)


def _cohort_output_row(
    row: pd.Series,
    *,
    scope: str,
    fund_name: str,
    cnpj: str,
    series_name: str,
    series_type: str,
    line_rank: int,
) -> dict[str, object]:
    denominator = _num(row.get("denominator"))
    numerator = _num(row.get("numerator"))
    return {
        "schema_version": RESEARCH_SCHEMA_VERSION,
        "scope": scope,
        "fund_name": fund_name,
        "cnpj": cnpj,
        "series_name": series_name,
        "series_type": series_type,
        "line_rank": line_rank,
        "cohort": row.get("cohort"),
        "cohort_dt": row.get("cohort_dt"),
        "mes_ciclo": row.get("mes_ciclo"),
        "ordem": row.get("ordem"),
        "value_pct": _safe_div_pct_scalar(numerator, denominator),
        "numerator": numerator,
        "denominator": denominator,
        "formula": "atraso_bucket_t+m / prazo_venc_30_t",
        "unit": "%",
        "source_columns": "prazo_venc_30; buckets de atraso",
        "missing_data_flag": bool(numerator is None or denominator is None),
        "division_by_zero_flag": bool(denominator is not None and denominator <= 0),
    }


def _weighted_cohort_benchmark_rows(
    df: pd.DataFrame,
    *,
    scope: str,
    fund_name: str,
    cnpj: str,
    series_name: str,
    series_type: str,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if df.empty:
        return rows
    for order, cycle in enumerate(["M1", "M2", "M3", "M4", "M5", "M6"], start=1):
        group = df[df["mes_ciclo"].eq(cycle)]
        numerator = _num(group["numerator"].sum(min_count=1))
        denominator = _num(group["denominator"].sum(min_count=1))
        value = _safe_div_pct_scalar(numerator, denominator)
        rows.append(
            {
                "schema_version": RESEARCH_SCHEMA_VERSION,
                "scope": scope,
                "fund_name": fund_name,
                "cnpj": cnpj,
                "series_name": series_name,
                "series_type": series_type,
                "line_rank": -1,
                "cohort": pd.NA,
                "cohort_dt": pd.NaT,
                "mes_ciclo": cycle,
                "ordem": order,
                "value_pct": value,
                "numerator": numerator,
                "denominator": denominator,
                "formula": "Σ atraso_bucket_t+m / Σ prazo_venc_30_t",
                "unit": "%",
                "source_columns": "prazo_venc_30; buckets de atraso",
                "missing_data_flag": bool(numerator is None or denominator is None),
                "division_by_zero_flag": bool(denominator is not None and denominator <= 0),
            }
        )
    return rows


def _research_warnings(df: pd.DataFrame, *, scope: str, fund_name: str) -> list[str]:
    warnings: list[str] = []
    if df is None or df.empty:
        return [f"{fund_name}: base vazia para visão research."]
    if "pdd_ex360" not in df.columns:
        warnings.append(f"{fund_name}: PDD ex-360 não está disponível; indicadores de cobertura research não são expandidos nesta visão.")
    if "carteira_ex360" not in df.columns:
        warnings.append(f"{fund_name}: carteira ex-360 ausente; tabelas research ficam incompletas.")
    if scope == "consolidado" and len(df) < 13:
        warnings.append("Consolidado: janela menor que 13 competências limita cálculos YoY.")
    return warnings


def _concat_frames(frames: list[pd.DataFrame]) -> pd.DataFrame:
    valid = [frame for frame in frames if frame is not None and not frame.empty]
    return pd.concat(valid, ignore_index=True, sort=False) if valid else pd.DataFrame()


def _safe_div_pct_series(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    num = pd.to_numeric(numerator, errors="coerce")
    den = pd.to_numeric(denominator, errors="coerce")
    return (num / den).where(den > 0).mul(100.0)


def _safe_div_pct_scalar(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator <= 0:
        return None
    return float(numerator / denominator * 100.0)


def _num(value: object) -> float | None:
    parsed = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(parsed):
        return None
    return float(parsed)


def _is_number(value: object) -> bool:
    if isinstance(value, bool):
        return False
    parsed = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return pd.notna(parsed)


def _month_label(dt: object) -> str:
    ts = pd.to_datetime(dt, errors="coerce")
    if pd.isna(ts):
        return "N/D"
    return PT_MONTH_ABBR_TITLE[int(ts.month)]


def _fund_name(frame: pd.DataFrame, *, fallback: str) -> str:
    if frame is not None and not frame.empty and "fund_name" in frame.columns and frame["fund_name"].notna().any():
        return str(frame["fund_name"].dropna().iloc[0])
    return fallback
