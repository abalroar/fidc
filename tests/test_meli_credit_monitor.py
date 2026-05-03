from __future__ import annotations

from io import BytesIO
import json
import unittest
import zipfile

import pandas as pd

from services.meli_credit_monitor import (
    MeliMonitorOutputs,
    build_ex360_memory_table,
    build_cohort_matrix,
    build_meli_methodology_table,
    build_meli_monitor_outputs,
    build_monitor_base,
    build_pdf_reconciliation_table,
    build_somatorio_dashboard_comparison,
)
from services.meli_credit_monitor_ppt_export import build_dashboard_meli_pptx_bytes
from services.meli_credit_research import build_meli_research_outputs, build_research_excel_bytes
from services.meli_credit_research_verification import verify_meli_research_outputs
from services.meli_credit_monitor_visuals import (
    cohort_chart,
    duration_chart,
    npl_severity_chart,
    portfolio_growth_chart,
    research_cohort_chart,
    research_roll_seasonality_chart,
    roll_rates_chart,
)
from services.mercado_livre_dashboard import build_consolidated_monthly_base


class MeliCreditMonitorTest(unittest.TestCase):
    def test_roll_rates_use_lagged_current_maturity_denominators(self) -> None:
        monthly = _sample_monthly(month_count=7)
        monitor = build_monitor_base(monthly)

        self.assertAlmostEqual(2.5, monitor.loc[3, "roll_61_90_m3_pct"])
        self.assertAlmostEqual(2.0, monitor.loc[4, "roll_91_120_m4_pct"])
        self.assertAlmostEqual(2.5, monitor.loc[5, "roll_121_150_m5_pct"])
        self.assertAlmostEqual(3.0, monitor.loc[6, "roll_151_180_m6_pct"])
        self.assertAlmostEqual(200.0, monitor.loc[3, "roll_61_90_m3_den"])
        self.assertAlmostEqual(200.0, monitor.loc[4, "roll_91_120_m4_den"])
        self.assertAlmostEqual(200.0, monitor.loc[5, "roll_121_150_m5_den"])
        self.assertAlmostEqual(200.0, monitor.loc[6, "roll_151_180_m6_den"])

    def test_roll_rates_derive_current_maturity_when_missing(self) -> None:
        monthly = _sample_monthly(month_count=7).drop(columns=["carteira_a_vencer"])
        monitor = build_monitor_base(monthly)

        self.assertAlmostEqual(200.0, monitor.loc[0, "carteira_a_vencer"])
        self.assertAlmostEqual(2.5, monitor.loc[3, "roll_61_90_m3_pct"])
        self.assertAlmostEqual(2.0, monitor.loc[4, "roll_91_120_m4_pct"])
        self.assertAlmostEqual(2.5, monitor.loc[5, "roll_121_150_m5_pct"])

    def test_cohort_matrix_uses_due_30_base_and_future_buckets(self) -> None:
        monthly = _sample_monthly(month_count=7)
        cohort = build_cohort_matrix(build_monitor_base(monthly))
        jan = cohort[cohort["cohort"] == "Jan-26"].set_index("mes_ciclo")

        self.assertAlmostEqual(1.0, jan.loc["M1", "valor_pct"])
        self.assertAlmostEqual(2.0, jan.loc["M2", "valor_pct"])
        self.assertAlmostEqual(5.0, jan.loc["M3", "valor_pct"])
        self.assertAlmostEqual(6.0, jan.loc["M6", "valor_pct"])
        self.assertNotIn("01/2026", set(cohort["cohort"]))
        jul_cohort = build_cohort_matrix(build_monitor_base(_sample_monthly(month_count=2, start="2025-07-01")))
        self.assertIn("Jul-25", set(jul_cohort["cohort"]))

    def test_consolidated_duration_is_weighted_by_maturity_balance(self) -> None:
        frame_a = pd.DataFrame(
            [
                {
                    "fund_name": "A",
                    "cnpj": "1",
                    "competencia": "01/2026",
                    "competencia_dt": pd.Timestamp("2026-01-01"),
                    "duration_total_saldo": 100.0,
                    "duration_weighted_days": 600.0,
                }
            ]
        )
        frame_b = pd.DataFrame(
            [
                {
                    "fund_name": "B",
                    "cnpj": "2",
                    "competencia": "01/2026",
                    "competencia_dt": pd.Timestamp("2026-01-01"),
                    "duration_total_saldo": 300.0,
                    "duration_weighted_days": 3600.0,
                }
            ]
        )

        consolidated = build_consolidated_monthly_base(
            portfolio_name="Carteira",
            fund_monthly_frames={"1": frame_a, "2": frame_b},
        )

        self.assertAlmostEqual(10.5, consolidated.loc[0, "duration_days"])
        self.assertAlmostEqual(10.5 / 30.4375, consolidated.loc[0, "duration_months"])

    def test_pdf_reconciliation_uses_november_2025_targets(self) -> None:
        monthly = _sample_monthly(month_count=1, start="2025-11-01")
        monthly.loc[0, "carteira_ex360"] = 7_141_000_000.0
        monthly.loc[0, "atraso_ate30"] = 266_000_000.0
        monthly.loc[0, "atraso_31_60"] = 184_000_000.0
        monthly.loc[0, "atraso_61_90"] = 150_000_000.0
        monthly.loc[0, "atraso_91_120"] = 370_000_000.0
        monthly.loc[0, "atraso_121_150"] = 0.0
        monthly.loc[0, "atraso_151_180"] = 0.0
        monthly.loc[0, "atraso_181_360"] = 642_000_000.0
        monthly.loc[0, "duration_months"] = 7.9
        monitor = build_monitor_base(monthly)

        reconciliation = build_pdf_reconciliation_table(monitor)
        npl_1_90 = reconciliation[reconciliation["Métrica"].eq("NPL 1-90d")].iloc[0]
        npl_1_360_pct = reconciliation[reconciliation["Métrica"].eq("NPL 1-360d / carteira ex-360")].iloc[0]
        no_pdf_target = reconciliation[reconciliation["Métrica"].eq("PL total")].iloc[0]

        self.assertAlmostEqual(600_000_000.0, npl_1_90["Valor app"])
        self.assertAlmostEqual(22.6, npl_1_360_pct["Valor PDF"])
        self.assertEqual("Sem alvo no PDF.", no_pdf_target["Status"])

    def test_methodology_table_documents_formula_sources(self) -> None:
        methodology = build_meli_methodology_table()

        roll = methodology[methodology["Indicador"].eq("Roll 61-90 / carteira a vencer M-3")].iloc[0]
        roll_91 = methodology[methodology["Indicador"].eq("Roll 91-120 / carteira a vencer M-4")].iloc[0]
        roll_121 = methodology[methodology["Indicador"].eq("Roll 121-150 / carteira a vencer M-5")].iloc[0]

        self.assertIn("carteira_a_vencer_t-3", roll["Fórmula"])
        self.assertIn("carteira_a_vencer_t-4", roll_91["Fórmula"])
        self.assertIn("carteira_a_vencer_t-5", roll_121["Fórmula"])
        self.assertIn("Fonte / coluna", methodology.columns)

    def test_dashboard_meli_charts_include_final_labels(self) -> None:
        monitor = build_monitor_base(_sample_monthly(month_count=7))

        roll_payload = json.dumps(roll_rates_chart(monitor).to_dict(), ensure_ascii=False)
        growth_payload = json.dumps(portfolio_growth_chart(monitor).to_dict(), ensure_ascii=False)
        npl_payload = json.dumps(npl_severity_chart(monitor).to_dict(), ensure_ascii=False)
        duration_payload = json.dumps(duration_chart(monitor, {"00000000000000": monitor}).to_dict(), ensure_ascii=False)
        cohort_payload = json.dumps(cohort_chart(build_cohort_matrix(monitor)).to_dict(), ensure_ascii=False)

        self.assertIn("mark", roll_payload)
        self.assertIn("text", roll_payload)
        self.assertIn("3,0%", roll_payload)
        self.assertIn("Roll 91-120", roll_payload)
        self.assertIn("Roll 121-150", roll_payload)
        self.assertIn("Carteira ex-360", growth_payload)
        self.assertIn("Crescimento YoY", growth_payload)
        self.assertIn('"width": "container"', growth_payload)
        self.assertIn("text", growth_payload)
        self.assertIn("vconcat", growth_payload)
        self.assertIn("Total", npl_payload)
        for payload in (roll_payload, growth_payload, npl_payload, duration_payload, cohort_payload):
            self.assertNotIn('"axis": null', payload)
            self.assertIn('"labels": true', payload)
            self.assertIn('"ticks": true', payload)

    def test_portfolio_growth_chart_labels_all_periods_in_millions_without_prefix_order_bug(self) -> None:
        monthly = _sample_monthly(month_count=3)
        monthly["carteira_ex360"] = [4_669_000_000.0, 7_674_900_000.0, 7_675_100_000.0]
        monitor = build_monitor_base(monthly)

        payload = json.dumps(portfolio_growth_chart(monitor).to_dict(), ensure_ascii=False)

        self.assertIn("4.669", payload)
        self.assertIn("7.675", payload)
        self.assertNotIn("R$ mm 7.674,9", payload)

    def test_monitor_audit_table_exposes_ex360_memory_components(self) -> None:
        monitor = build_monitor_base(_sample_monthly(month_count=3))
        monitor["vencidos_360"] = 10.0
        monitor["npl_over360"] = 10.0
        monitor["baixa_over360_carteira"] = 10.0
        rebuilt = build_meli_monitor_outputs(type("Outputs", (), {"fund_monthly": {"00000000000000": monitor}, "consolidated_monthly": monitor})())

        self.assertIn("carteira_bruta", rebuilt.audit_table.columns)
        self.assertIn("npl_over360", rebuilt.audit_table.columns)
        self.assertIn("baixa_over360_carteira", rebuilt.audit_table.columns)
        self.assertIn("carteira_ex360", rebuilt.audit_table.columns)

    def test_somatorio_dashboard_comparison_reconciles_equal_metrics(self) -> None:
        somatorio = _sample_monthly(month_count=3)
        somatorio["vencidos_360"] = 20.0
        somatorio["npl_over360"] = 20.0
        somatorio["baixa_over360_carteira"] = 20.0
        somatorio["carteira_ex360"] = somatorio["carteira_bruta"] - somatorio["npl_over360"]
        somatorio["npl_over1_ex360"] = somatorio["atraso_ate30"] + somatorio["atraso_31_60"] + somatorio["atraso_61_90"] + somatorio["atraso_91_120"] + somatorio["atraso_121_150"] + somatorio["atraso_151_180"] + somatorio["atraso_181_360"]
        somatorio["npl_over90_ex360"] = somatorio["atraso_91_120"] + somatorio["atraso_121_150"] + somatorio["atraso_151_180"] + somatorio["atraso_181_360"]
        somatorio["npl_over1_ex360_pct"] = somatorio["npl_over1_ex360"] / somatorio["carteira_ex360"] * 100.0
        somatorio["npl_over90_ex360_pct"] = somatorio["npl_over90_ex360"] / somatorio["carteira_ex360"] * 100.0
        dashboard = build_monitor_base(somatorio)
        outputs = type("Outputs", (), {"fund_monthly": {"00000000000000": somatorio}, "consolidated_monthly": somatorio})()
        monitor_outputs = MeliMonitorOutputs(
            consolidated_monitor=dashboard,
            fund_monitor={"00000000000000": dashboard},
            consolidated_cohorts=build_cohort_matrix(dashboard),
            fund_cohorts={"00000000000000": build_cohort_matrix(dashboard)},
            audit_table=pd.DataFrame(),
            pdf_reconciliation=pd.DataFrame(),
            warnings=[],
        )

        comparison = build_somatorio_dashboard_comparison(outputs, monitor_outputs)
        memory = build_ex360_memory_table(outputs)

        total_npl = comparison[comparison["metrica"].eq("NPL ex-360 total / carteira ex-360")]
        self.assertFalse(total_npl.empty)
        self.assertEqual({"OK"}, set(total_npl["status"]))
        self.assertIn("carteira_ex360", memory.columns)

    def test_dashboard_meli_cohort_chart_uses_chronological_labels_and_gray_scale(self) -> None:
        monitor = build_monitor_base(_sample_monthly(month_count=9))
        cohorts = build_cohort_matrix(monitor)

        payload = json.dumps(cohort_chart(cohorts).to_dict(), ensure_ascii=False)

        self.assertIn("Jan-26", payload)
        self.assertNotIn("01/2026", payload)
        self.assertNotIn("strokeDash", payload)
        self.assertIn("#000000", payload)
        self.assertIn("#D9D9D9", payload)
        self.assertIn("\"domain\"", payload)

    def test_dashboard_meli_pptx_export_is_valid_zip(self) -> None:
        monitor = build_monitor_base(_sample_monthly(month_count=7))
        cohorts = build_cohort_matrix(monitor)
        outputs = MeliMonitorOutputs(
            consolidated_monitor=monitor,
            fund_monitor={"00000000000000": monitor},
            consolidated_cohorts=cohorts,
            fund_cohorts={"00000000000000": cohorts},
            audit_table=pd.DataFrame(),
            pdf_reconciliation=pd.DataFrame(),
            warnings=[],
        )
        research = build_meli_research_outputs(outputs)

        pptx_bytes = build_dashboard_meli_pptx_bytes(outputs, research)

        self.assertTrue(pptx_bytes.startswith(b"PK"))
        self.assertTrue(zipfile.is_zipfile(BytesIO(pptx_bytes)))
        with zipfile.ZipFile(BytesIO(pptx_bytes)) as archive:
            names = archive.namelist()
            xml_payload = "\n".join(
                archive.read(name).decode("utf-8", errors="ignore")
                for name in names
                if name.endswith(".xml")
            )
        self.assertTrue(any(name.startswith("ppt/charts/chart") for name in names))
        self.assertIn("Análise Crédito - Consolidado: carteira e risco", xml_payload)
        self.assertIn("Análise Crédito - Consolidado: duration e cohorts", xml_payload)
        self.assertIn("NPL Over 1d ex-360", xml_payload)
        self.assertIn("NPL Over 60d ex-360", xml_payload)
        self.assertIn("Roll 61-90 por mês do ano", xml_payload)
        self.assertIn("Roll 91-120 por mês do ano", xml_payload)
        self.assertIn("Roll 121-150 por mês do ano", xml_payload)
        self.assertIn("Roll 151-180 por mês do ano", xml_payload)
        self.assertIn("Calibri", xml_payload)
        self.assertTrue("<c:dLbl" in xml_payload or "<c:dLbls" in xml_payload)
        self.assertNotIn("Cohorts com médias", xml_payload)
        self.assertNotIn("Dashboard MELI - Visão research", xml_payload)

    def test_combined_somatorio_pptx_export_contains_base_and_credit_slides(self) -> None:
        from services.somatorio_fidcs_ppt_export import build_somatorio_fidcs_pptx_bytes

        monitor = build_monitor_base(_sample_monthly(month_count=7))
        cohorts = build_cohort_matrix(monitor)
        monitor_outputs = MeliMonitorOutputs(
            consolidated_monitor=monitor,
            fund_monitor={"00000000000000": monitor},
            consolidated_cohorts=cohorts,
            fund_cohorts={"00000000000000": cohorts},
            audit_table=pd.DataFrame(),
            pdf_reconciliation=pd.DataFrame(),
            warnings=[],
        )
        research = build_meli_research_outputs(monitor_outputs)
        base_monthly = monitor.copy()
        base_monthly["pl_senior"] = 700.0
        base_monthly["pl_subordinada_mezz_ex360"] = 300.0
        base_monthly["subordinacao_total_ex360_pct"] = 30.0
        base_monthly["npl_over90_ex360_pct"] = 2.5
        base_monthly["pdd_npl_over90_ex360_pct"] = 200.0
        base_outputs = type(
            "BaseOutputs",
            (),
            {
                "consolidated_monthly": base_monthly,
                "fund_monthly": {"00000000000000": base_monthly},
            },
        )()

        pptx_bytes = build_somatorio_fidcs_pptx_bytes(
            outputs=base_outputs,
            monitor_outputs=monitor_outputs,
            research_outputs=research,
        )

        self.assertTrue(pptx_bytes.startswith(b"PK"))
        self.assertTrue(zipfile.is_zipfile(BytesIO(pptx_bytes)))
        with zipfile.ZipFile(BytesIO(pptx_bytes)) as archive:
            slide1_xml = archive.read("ppt/slides/slide1.xml").decode("utf-8", errors="ignore")
            slide2_xml = archive.read("ppt/slides/slide2.xml").decode("utf-8", errors="ignore")
            xml_payload = "\n".join(
                archive.read(name).decode("utf-8", errors="ignore")
                for name in archive.namelist()
                if name.endswith(".xml")
            )
        self.assertIn("Somatório FIDCs - Base consolidada", slide1_xml)
        self.assertIn("Análise Crédito - Consolidado: carteira e risco", slide2_xml)
        self.assertIn("Somatório FIDCs - Base consolidada", xml_payload)
        self.assertIn("Análise Crédito - Consolidado: carteira e risco", xml_payload)
        self.assertIn("NPL Over 1d ex-360", xml_payload)
        self.assertIn("Roll 91-120 por mês do ano", xml_payload)
        self.assertIn("Roll 121-150 por mês do ano", xml_payload)

    def test_research_layer_builds_auditable_metrics_and_verification(self) -> None:
        monitor = build_monitor_base(_sample_monthly(month_count=14))
        cohorts = build_cohort_matrix(monitor)
        outputs = MeliMonitorOutputs(
            consolidated_monitor=monitor,
            fund_monitor={"00000000000000": monitor},
            consolidated_cohorts=cohorts,
            fund_cohorts={"00000000000000": cohorts},
            audit_table=pd.DataFrame(),
            pdf_reconciliation=pd.DataFrame(),
            warnings=[],
        )

        research = build_meli_research_outputs(outputs)
        verification = verify_meli_research_outputs(outputs, research)
        roll = research.roll_seasonality[
            research.roll_seasonality["metric_id"].eq("roll_61_90_m3")
            & research.roll_seasonality["competencia"].eq("04/2026")
            & research.roll_seasonality["scope"].eq("consolidado")
        ].iloc[0]
        roll_91 = research.roll_seasonality[
            research.roll_seasonality["metric_id"].eq("roll_91_120_m4")
            & research.roll_seasonality["competencia"].eq("05/2026")
            & research.roll_seasonality["scope"].eq("consolidado")
        ].iloc[0]
        roll_121 = research.roll_seasonality[
            research.roll_seasonality["metric_id"].eq("roll_121_150_m5")
            & research.roll_seasonality["competencia"].eq("06/2026")
            & research.roll_seasonality["scope"].eq("consolidado")
        ].iloc[0]
        npl = research.npl_research_table[
            research.npl_research_table["metric_id"].eq("npl_1_360_pct")
            & research.npl_research_table["competencia"].eq("01/2026")
            & research.npl_research_table["scope"].eq("consolidado")
        ].iloc[0]

        self.assertAlmostEqual(2.5, roll["value_pct"])
        self.assertAlmostEqual(2.0, roll_91["value_pct"])
        self.assertAlmostEqual(2.5, roll_121["value_pct"])
        self.assertAlmostEqual(2.8, npl["value"])
        self.assertIn("numerator", research.cohort_research.columns)
        self.assertIn("Fórmula", research.methodology.columns)
        self.assertFalse(verification.empty)
        self.assertNotIn("ERRO", set(verification["status"]))

    def test_research_charts_keep_axes_visible_and_labels(self) -> None:
        monitor = build_monitor_base(_sample_monthly(month_count=14))
        cohorts = build_cohort_matrix(monitor)
        outputs = MeliMonitorOutputs(
            consolidated_monitor=monitor,
            fund_monitor={},
            consolidated_cohorts=cohorts,
            fund_cohorts={},
            audit_table=pd.DataFrame(),
            pdf_reconciliation=pd.DataFrame(),
            warnings=[],
        )
        research = build_meli_research_outputs(outputs)

        roll_payload = json.dumps(research_roll_seasonality_chart(research.roll_seasonality, metric_id="roll_61_90_m3").to_dict(), ensure_ascii=False)
        cohort_payload = json.dumps(research_cohort_chart(research.cohort_research).to_dict(), ensure_ascii=False)

        self.assertIn("Roll rate", roll_payload)
        self.assertIn("text", roll_payload)
        self.assertIn("strokeDash", cohort_payload)
        for payload in (roll_payload, cohort_payload):
            self.assertNotIn('"axis": null', payload)
            self.assertIn('"labels": true', payload)
            self.assertIn('"ticks": true', payload)

    def test_research_excel_export_opens_without_repair_and_keeps_numeric_cells(self) -> None:
        from openpyxl import load_workbook

        monitor = build_monitor_base(_sample_monthly(month_count=14))
        cohorts = build_cohort_matrix(monitor)
        outputs = MeliMonitorOutputs(
            consolidated_monitor=monitor,
            fund_monitor={},
            consolidated_cohorts=cohorts,
            fund_cohorts={},
            audit_table=pd.DataFrame(),
            pdf_reconciliation=pd.DataFrame(),
            warnings=[],
        )
        research = build_meli_research_outputs(outputs)
        verification = verify_meli_research_outputs(outputs, research)

        workbook_bytes = build_research_excel_bytes(research, verification)

        self.assertTrue(workbook_bytes.startswith(b"PK"))
        wb = load_workbook(BytesIO(workbook_bytes), data_only=False)
        self.assertIn("NPL research", wb.sheetnames)
        ws = wb["NPL research"]
        header = [cell.value for cell in ws[1]]
        value_col = header.index("value") + 1
        self.assertIsInstance(ws.cell(row=2, column=value_col).value, (int, float))

    def test_research_display_table_formats_without_numeric_dtype_upcast_error(self) -> None:
        from tabs.tab_dashboard_meli import _format_research_table

        monitor = build_monitor_base(_sample_monthly(month_count=14))
        cohorts = build_cohort_matrix(monitor)
        outputs = MeliMonitorOutputs(
            consolidated_monitor=monitor,
            fund_monitor={},
            consolidated_cohorts=cohorts,
            fund_cohorts={},
            audit_table=pd.DataFrame(),
            pdf_reconciliation=pd.DataFrame(),
            warnings=[],
        )
        research = build_meli_research_outputs(outputs)

        display = _format_research_table(research.npl_research_table)

        self.assertIn("Valor", display.columns)
        self.assertIsInstance(display.iloc[0]["Valor"], str)

    def test_dashboard_meli_kpi_cards_use_npl_over_metrics_instead_of_roll_rates(self) -> None:
        from tabs.tab_dashboard_meli import _dashboard_kpi_cards

        cards = _dashboard_kpi_cards(
            pd.Series(
                {
                    "carteira_ex360": 7_674_900_000.0,
                    "npl_over1_ex360_pct": 19.2,
                    "npl_over30_ex360_pct": 15.1,
                    "npl_over60_ex360_pct": 12.8,
                    "npl_over90_ex360_pct": 11.8,
                    "roll_61_90_m3_pct": 2.4,
                    "roll_151_180_m6_pct": 2.1,
                    "duration_months": 12.5,
                }
            )
        )
        labels = [label for label, _ in cards]

        self.assertIn("NPL Over 1d ex-360", labels)
        self.assertIn("NPL Over 30d ex-360", labels)
        self.assertIn("NPL Over 60d ex-360", labels)
        self.assertIn("NPL Over 90d ex-360", labels)
        self.assertNotIn("Roll 61-90 M-3", labels)
        self.assertNotIn("NPL 1-90 / carteira", labels)


def _sample_monthly(*, month_count: int, start: str = "2026-01-01") -> pd.DataFrame:
    rows = []
    start_ts = pd.Timestamp(start)
    for idx in range(month_count):
        ts = start_ts + pd.DateOffset(months=idx)
        rows.append(
            {
                "fund_name": "Mercado Crédito",
                "cnpj": "00000000000000",
                "competencia": f"{int(ts.month):02d}/{int(ts.year)}",
                "competencia_dt": ts,
                "carteira_ex360": 1_000.0,
                "carteira_bruta": 1_000.0,
                "carteira_em_dia": 100.0,
                "carteira_a_vencer": 200.0,
                "atraso_ate30": 1.0,
                "atraso_31_60": 2.0,
                "atraso_61_90": 5.0 if idx == 3 else 3.0,
                "atraso_91_120": 4.0,
                "atraso_121_150": 5.0,
                "atraso_151_180": 6.0,
                "atraso_181_360": 7.0,
                "prazo_venc_30": 100.0,
                "prazo_venc_31_60": 100.0,
                "prazo_venc_61_90": 0.0,
                "prazo_venc_91_120": 0.0,
                "prazo_venc_121_150": 0.0,
                "prazo_venc_151_180": 0.0,
                "prazo_venc_181_360": 0.0,
                "prazo_venc_361_720": 0.0,
                "prazo_venc_721_1080": 0.0,
                "prazo_venc_1080": 0.0,
                "pdd_ex360": 50.0,
                "npl_over90_ex360": 25.0,
                "duration_months": 6.0,
            }
        )
    return pd.DataFrame(rows)


if __name__ == "__main__":
    unittest.main()
