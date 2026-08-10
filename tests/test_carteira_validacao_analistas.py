from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "industry_study"


def test_a_base_preserva_os_101_cnpjs_e_isola_os_nove_casos() -> None:
    from services.carteira_validacao_analistas import (
        build_validation_frame,
        target_validation_frame,
    )

    carteira = build_validation_frame(DATA)
    alvos = target_validation_frame(DATA)

    assert len(carteira) == 101
    assert carteira["cnpj"].nunique() == 101
    assert len(alvos) == 9
    assert alvos["cnpj"].nunique() == 9
    assert alvos["folga_pos_pp"].lt(0).all()
    assert alvos["aporte_brl"].gt(0).all()


def test_as_tranches_reconciliam_com_o_total_usado_no_stress() -> None:
    import pandas as pd

    tranches = pd.read_csv(DATA / "carteira_cotas_tranches.csv")

    assert len(tranches) >= 101
    assert tranches["cnpj"].astype(str).str.zfill(14).nunique() == len(tranches)
    assert tranches["valor_nao_classificado_brl"].fillna(0).abs().max() < 0.01
    assert tranches["diferenca_total_vs_mensal_brl"].dropna().abs().max() < 0.01
    assert tranches["diferenca_submaismez_vs_mensal_brl"].dropna().abs().max() < 0.01


def test_o_excel_traz_formulas_validacoes_e_contraprova_documental() -> None:
    from services.middle_market_exports import build_validacao_analistas_xlsx_bytes

    workbook = load_workbook(BytesIO(build_validacao_analistas_xlsx_bytes(DATA)), data_only=False)
    assert workbook.sheetnames == ["Validação Analistas", "Carteira 101"]
    alvo = workbook["Validação Analistas"]
    carteira = workbook["Carteira 101"]
    assert alvo.max_row == 10
    assert carteira.max_row == 102

    headers = {cell.value: cell.column for cell in alvo[1]}
    obrigatorias = {
        "Segmento",
        "Analista",
        "Gerente",
        "Sub Mínima (regulamento)",
        "Subordinada / PL",
        "Sub + Mez / PL",
        "Inadimplência (R$ mm)",
        "PDD (R$ mm)",
        "Sub não provisionada — Δ (R$ mm)",
        "PL Total (R$ mm)",
        "PL Sênior (R$ mm)",
        "PL Subordinada (R$ mm)",
        "Folga pós-estresse (p.p.)",
        "Aporte necessário (R$ mm)",
        "Aporte / PL atual",
        "Comentário",
    }
    assert obrigatorias <= set(headers)
    assert all(alvo.cell(2, headers[name]).value is None for name in ("Segmento", "Analista", "Gerente", "Comentário"))

    folga = alvo.cell(2, headers["Folga pós-estresse (p.p.)"])
    aporte = alvo.cell(2, headers["Aporte necessário (R$ mm)"])
    delta = alvo.cell(2, headers["Sub não provisionada — Δ (R$ mm)"])
    inadimplencia = get_column_letter(headers["Inadimplência (R$ mm)"])
    pdd = get_column_letter(headers["PDD (R$ mm)"])
    assert delta.value == (
        f'=IF({inadimplencia}2="","",MAX(0,'
        f'{inadimplencia}2-IF({pdd}2="",0,{pdd}2)))'
    )
    assert str(folga.value).startswith("=IF(")
    assert str(aporte.value).startswith("=IF(") and "MAX(0" in aporte.value
    assert alvo.cell(1, headers["Sub não provisionada — Δ (R$ mm)"]).comment is not None
    assert alvo.cell(1, headers["Folga pós-estresse (p.p.)"]).comment is not None
    assert alvo.cell(1, headers["Aporte necessário (R$ mm)"]).comment is not None

    # Workcap, a primeira linha, não tem mezanino: valor e validação ficam vazios.
    assert alvo.cell(2, headers["Sub + Mez / PL"]).value is None
    assert alvo.cell(2, headers["Validação — Sub + Mez / PL"]).value is None
    # Sotreq tem mezanino e conserva a fórmula editável.
    sotreq = next(row for row in range(2, alvo.max_row + 1) if "SOTREQ" in str(alvo.cell(row, headers["FIDC"]).value).upper())
    assert str(alvo.cell(sotreq, headers["Sub + Mez / PL"]).value).startswith("=IF(")
    assert alvo.cell(sotreq, headers["Validação — Sub + Mez / PL"]).value == "Pendente"

    # Formato sem agrupamento de milhar; o Excel pt-BR exibe a casa decimal
    # com vírgula (1759,6), sem ponto entre os milhares.
    monetarias = {
        "PL Total (R$ mm)",
        "PL Sênior (R$ mm)",
        "PL Subordinada (R$ mm)",
        "Inadimplência (R$ mm)",
        "PDD (R$ mm)",
        "Sub não provisionada — Δ (R$ mm)",
        "Aporte necessário (R$ mm)",
    }
    assert {
        alvo.cell(2, headers[name]).number_format for name in monetarias
    } == {"0.0"}

    assert len(alvo.data_validations.dataValidation) == 1
    assert "sem cláusula localizada" in {
        str(alvo.cell(row, headers["Resultado documental"]).value)
        for row in range(2, alvo.max_row + 1)
    }
