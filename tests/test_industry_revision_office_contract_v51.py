"""Acceptance contract for the compact FIDC industry revision.

This module intentionally lives beside the legacy 47-slide assertions while
the renderer, validators and generated artifacts are migrated together.  It
tests the exported OOXML rather than presentation-library abstractions so an
image or a collection of text boxes cannot satisfy a native Office contract.
"""

from __future__ import annotations

from io import BytesIO
import json
import os
import posixpath
import re
from pathlib import Path
import unicodedata
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

from services.industry_revision_export import (
    CURRENT_TOP15_SLIDE_SEQUENCE,
    EXPECTED_SLIDE_SEQUENCE,
    EXPECTED_SLIDES,
    HISTORICAL_TOP15_SLIDE_SEQUENCE,
    HISTORICAL_TOP15_TABLE_DIMENSIONS,
    ISSUANCE_TAXONOMY_TABLE_DIMENSIONS,
    PORTFOLIO_WORKBOOK_MINIMUM_HEADERS,
    REQUIRED_WORKBOOK_SHEETS,
    REQUIRED_PORTFOLIO_WORKBOOK_SHEETS,
    REVISION_EMISSION_AUDIT_REQUIRED_HEADERS,
    REVISION_EMISSION_COVERAGE_TARGET_LABEL,
    STRUCTURAL_MVP_SLIDE_SEQUENCE,
    TYPE_RANKING_SLIDE_SEQUENCE,
    RevisionExportUnavailable,
    _contains_blocked_rgb_color,
    _validate_no_blocked_audience_copy,
    validate_revision_portfolio_xlsx,
    validate_revision_pptx,
    validate_revision_top100_xlsx,
    validate_revision_xlsx,
)


ROOT = Path(__file__).resolve().parents[1]
PPTX = Path(
    os.environ.get(
        "FIDC_TEST_PPTX",
        ROOT
        / "data"
        / "industry_study"
        / "generated_revision"
        / "industry_executive_revised.pptx",
    )
)
XLSX = Path(
    os.environ.get(
        "FIDC_TEST_XLSX",
        ROOT
        / "data"
        / "industry_study"
        / "generated_revision"
        / "industry_data_revised.xlsx",
    )
)


def test_blocked_palette_detector_only_reads_color_elements() -> None:
    metadata = b'<p:cNvPr xmlns:p="p" id="1" descr="random token 172A3A"/>'
    blocked = b'<a:srgbClr xmlns:a="a" val="172a3a"/>'
    system_fallback = b'<a:sysClr xmlns:a="a" val="window" lastClr="172A3A"/>'

    assert not _contains_blocked_rgb_color([metadata], "172A3A")
    assert _contains_blocked_rgb_color([blocked], "172A3A")
    assert _contains_blocked_rgb_color([system_fallback], "172A3A")


def test_current_top15_contract_and_renderer_distinguish_1s26_from_2025() -> None:
    assert CURRENT_TOP15_SLIDE_SEQUENCE == (
        (
            "top 15",
            "ibba esteve em 8 das 15 maiores ofertas do semestre",
            "liderou 5 delas",
            "jan–jun/26",
        ),
        (
            "top 15",
            "as 15 maiores ofertas de 2025 mantem a base anual de comparacao",
            "2025fy",
        ),
    )
    renderer = (
        ROOT / "scripts" / "build_fidc_revision_artifacts.mjs"
    ).read_text(encoding="utf-8")
    assert 'titleStartsWith: "IBBA participou de"' in renderer
    assert 'titleStartsWith: "As 15 maiores ofertas de 2025"' in renderer
    assert (
        'title: "As 15 maiores ofertas de 2025 mantêm a base anual de comparação"'
        in renderer
    )


def test_portfolio_workbook_validator_accepts_the_full_101_and_47_contract() -> None:
    validate_revision_portfolio_xlsx(_portfolio_workbook_bytes())


def test_revision_workbook_validator_separates_vnu_from_target_remuneration() -> None:
    validate_revision_xlsx(_revision_workbook_bytes())


@pytest.mark.parametrize(
    ("kwargs", "message"),
    [
        (
            {"missing_audit_header": "Fonte remuneração"},
            "não separa VNU de remuneração-alvo",
        ),
        (
            {"coverage_target_count": 7},
            "deveria conter oito linhas de Remuneração-alvo",
        ),
        (
            {"stale_price_coverage": True},
            "ainda trata VNU como rentabilidade-alvo",
        ),
    ],
)
def test_revision_workbook_validator_rejects_conflated_remuneration_contract(
    kwargs: dict[str, object],
    message: str,
) -> None:
    with pytest.raises(RevisionExportUnavailable, match=message):
        validate_revision_xlsx(_revision_workbook_bytes(**kwargs))


@pytest.mark.parametrize(
    ("kwargs", "message"),
    [
        ({"carteira_rows": 100}, "deveria conter 101 linhas"),
        ({"duplicate_cnpj": True}, "CNPJ duplicado"),
        ({"invalid_cnpj": True}, "CNPJ numérico inválido"),
        ({"text_cnpj": True}, "CNPJ numérico inválido"),
        (
            {"missing_header": "Preço por cota · leitura"},
            "sem cabeçalhos obrigatórios",
        ),
    ],
)
def test_portfolio_workbook_validator_rejects_incomplete_or_invalid_data(
    kwargs: dict[str, object],
    message: str,
) -> None:
    with pytest.raises(RevisionExportUnavailable, match=message):
        validate_revision_portfolio_xlsx(_portfolio_workbook_bytes(**kwargs))


def test_top100_workbook_validator_accepts_global_ranking_contract() -> None:
    validate_revision_top100_xlsx(_top100_workbook_bytes())


@pytest.mark.parametrize(
    "corrupt_text",
    [
        "Cr√©dito",
        "M√°quinas",
        "Adquir√™ncia",
        "CrÃ©dito",
        "Crédito � corporativo",
    ],
)
def test_top100_workbook_validator_rejects_mojibake_and_replacement_character(
    corrupt_text: str,
) -> None:
    with pytest.raises(RevisionExportUnavailable, match="texto corrompido"):
        validate_revision_top100_xlsx(
            _top100_workbook_bytes(corrupt_text=corrupt_text)
        )


@pytest.mark.parametrize(
    ("kwargs", "message"),
    [
        ({"row_count": 101}, "deveria conter 102 linhas"),
        ({"duplicate_cnpj": True}, "102 CNPJs numéricos válidos e únicos"),
        ({"invalid_cnpj": True}, "102 CNPJs numéricos válidos e únicos"),
        (
            {"missing_additional_cnpj": True},
            "não contém Citi-Bayer e Lavoro",
        ),
        (
            {"invalid_cnpj_format": True},
            "deve exibir CNPJ com máscara de 14 dígitos",
        ),
        (
            {"missing_header": "Middle Market · status"},
            "sem cabeçalhos obrigatórios",
        ),
    ],
)
def test_top100_workbook_validator_rejects_incomplete_or_invalid_data(
    kwargs: dict[str, object],
    message: str,
) -> None:
    with pytest.raises(RevisionExportUnavailable, match=message):
        validate_revision_top100_xlsx(_top100_workbook_bytes(**kwargs))


PAYLOAD = (
    ROOT
    / "data"
    / "industry_study"
    / "generated_revision"
    / "artifact_payload.json"
)

TARGET_SLIDES = EXPECTED_SLIDES

DML = "http://schemas.openxmlformats.org/drawingml/2006/main"
CHART = "http://schemas.openxmlformats.org/drawingml/2006/chart"
PML = "http://schemas.openxmlformats.org/presentationml/2006/main"
SHEET = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def _test_cnpj(index: int) -> str:
    base = f"42{index:010d}"

    def digit(value: str, weights: tuple[int, ...]) -> str:
        remainder = sum(
            int(character) * weight
            for character, weight in zip(value, weights, strict=True)
        ) % 11
        return "0" if remainder < 2 else str(11 - remainder)

    first = digit(base, (5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2))
    second = digit(
        base + first,
        (6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2),
    )
    return base + first + second


def _portfolio_workbook_bytes(
    *,
    carteira_rows: int = 101,
    flagship_rows: int = 47,
    missing_header: str | None = None,
    duplicate_cnpj: bool = False,
    invalid_cnpj: bool = False,
    text_cnpj: bool = False,
) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Leia-me"
    for sheet_name in sorted(REQUIRED_PORTFOLIO_WORKBOOK_SHEETS - {"Leia-me"}):
        workbook.create_sheet(sheet_name)

    headers = sorted(PORTFOLIO_WORKBOOK_MINIMUM_HEADERS)
    if missing_header is not None:
        headers.remove(missing_header)
    next_cnpj = 1
    for sheet_name, row_count in (
        ("Carteira 101", carteira_rows),
        ("Casos 99", 99),
        ("Flagships", flagship_rows),
    ):
        sheet = workbook[sheet_name]
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=4, column=column, value=header)
        cnpj_column = headers.index("CNPJ") + 1
        first_cnpj: int | None = None
        for offset in range(row_count):
            digits = _test_cnpj(next_cnpj)
            next_cnpj += 1
            value: object = int(digits)
            if first_cnpj is None:
                first_cnpj = int(digits)
            if duplicate_cnpj and sheet_name == "Carteira 101" and offset == 1:
                value = first_cnpj
            if invalid_cnpj and sheet_name == "Carteira 101" and offset == 0:
                value = 123
            if text_cnpj and sheet_name == "Carteira 101" and offset == 0:
                value = digits
            for column in range(1, len(headers) + 1):
                sheet.cell(row=5 + offset, column=column, value="N/D")
            cnpj_cell = sheet.cell(row=5 + offset, column=cnpj_column, value=value)
            cnpj_cell.number_format = "00000000000000"

    editable_names = workbook["Nomes editáveis"]
    editable_names.append(["Nome editável", "PL"])
    editable_names.append(["FIDC de teste", 1.0])
    chart = BarChart()
    chart.add_data(
        Reference(editable_names, min_col=2, min_row=1, max_row=2),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(editable_names, min_col=1, min_row=2, max_row=2)
    )
    workbook["Leia-me"].add_chart(chart, "A3")

    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _revision_workbook_bytes(
    *,
    missing_audit_header: str = "",
    coverage_target_count: int = 8,
    stale_price_coverage: bool = False,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in sorted(REQUIRED_WORKBOOK_SHEETS):
        workbook.create_sheet(sheet_name)

    audit = workbook["Auditoria emissões"]
    headers = sorted(REVISION_EMISSION_AUDIT_REQUIRED_HEADERS)
    if missing_audit_header:
        headers.remove(missing_audit_header)
    for column, header in enumerate(headers, start=1):
        audit.cell(row=4, column=column, value=header)

    coverage = workbook["Cobertura emissões"]
    for offset in range(coverage_target_count):
        coverage.cell(
            row=5 + offset,
            column=4,
            value=REVISION_EMISSION_COVERAGE_TARGET_LABEL,
        )
    if stale_price_coverage:
        coverage.cell(
            row=5 + coverage_target_count,
            column=4,
            value="Preço por cota",
        )

    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _top100_workbook_bytes(
    *,
    row_count: int = 102,
    missing_header: str | None = None,
    duplicate_cnpj: bool = False,
    invalid_cnpj: bool = False,
    missing_additional_cnpj: bool = False,
    invalid_cnpj_format: bool = False,
    corrupt_text: str | None = None,
) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Leia-me"
    sheet = workbook.create_sheet("Top 100 FIDCs")
    headers = [
        "Ordem do export",
        "Rank geral por PL",
        "Critério de inclusão",
        "CNPJ",
        "Nome completo do fundo (CVM)",
        "PL",
        "Sub / PL atual",
        "Mínimo de Sub Jr",
        "Mínimo estrutural",
        "Preço inicial por cota",
        "Cedente / originador",
        "Sacado / devedor",
        "Tipo de recebível",
        "Tipo ANBIMA oficial",
        "Taxonomia funcional N1",
        "Middle Market · status",
        "Fonte",
    ]
    if missing_header is not None:
        headers.remove(missing_header)
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=header)
    cnpj_column = headers.index("CNPJ") + 1
    for offset in range(row_count):
        if offset == 100:
            digits = "44302112000172"
        elif offset == 101 and not missing_additional_cnpj:
            digits = "61669748000176"
        else:
            digits = _test_cnpj(20_000 + offset)
        value: object = int(digits)
        if duplicate_cnpj and offset == 1:
            value = int(_test_cnpj(20_000))
        if invalid_cnpj and offset == 0:
            value = 123
        for column in range(1, len(headers) + 1):
            sheet.cell(row=5 + offset, column=column, value="N/D")
        sheet.cell(
            row=5 + offset,
            column=headers.index("Ordem do export") + 1,
            value=offset + 1,
        )
        sheet.cell(
            row=5 + offset,
            column=headers.index("Rank geral por PL") + 1,
            value=offset + 1,
        )
        sheet.cell(
            row=5 + offset,
            column=headers.index("Critério de inclusão") + 1,
            value=(
                "Top 100 por PL ex-FIC"
                if offset < 100
                else "Inclusão 2026 documentada"
            ),
        )
        cnpj_cell = sheet.cell(row=5 + offset, column=cnpj_column, value=value)
        cnpj_cell.number_format = (
            "General" if invalid_cnpj_format and offset == 0 else "00000000000000"
        )
        if corrupt_text is not None and offset == 0:
            sheet.cell(
                row=5 + offset,
                column=headers.index("Tipo de recebível") + 1,
                value=corrupt_text,
            )

    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _fold(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join(
        "".join(char for char in normalized if not unicodedata.combining(char))
        .lower()
        .split()
    )


def _contract_slide_numbers(*needles: str) -> tuple[int, ...]:
    folded_needles = tuple(_fold(needle) for needle in needles)
    return tuple(
        index
        for index, tokens in enumerate(EXPECTED_SLIDE_SEQUENCE, start=1)
        if all(
            any(needle in _fold(token) for token in tokens)
            for needle in folded_needles
        )
    )


def _contract_slide_number(*needles: str) -> int:
    matches = _contract_slide_numbers(*needles)
    assert len(matches) == 1, (needles, matches)
    return matches[0]


def _sequence_slide_numbers(
    sequence: tuple[tuple[str, ...], ...],
) -> tuple[int, ...]:
    return tuple(
        EXPECTED_SLIDE_SEQUENCE.index(tokens) + 1 for tokens in sequence
    )


STRUCTURAL_MVP_SLIDES = _sequence_slide_numbers(STRUCTURAL_MVP_SLIDE_SEQUENCE)
SLIDE_OFFERS_VOLUME = _contract_slide_number("emissoes crescem 15%")
SLIDE_OFFER_TICKETS = _contract_slide_number("22 ofertas concentram")
SLIDE_OFFER_REGIME = _contract_slide_number("garantia firme", "yoy ytd")
SLIDES_TOP15_CURRENT = _sequence_slide_numbers(CURRENT_TOP15_SLIDE_SEQUENCE)
SLIDES_TOP15_HISTORY = _sequence_slide_numbers(HISTORICAL_TOP15_SLIDE_SEQUENCE)
SLIDE_PROVIDER_HISTORY = _contract_slide_number("qi lidera administracao")
SLIDE_PROVIDER_RANKING = _contract_slide_number("prestadores", "ranking e concentracao")
SLIDE_HOLDER_DISTRIBUTION = _contract_slide_number("distribuicao por numero")
SLIDES_TYPE_RANKING = _sequence_slide_numbers(TYPE_RANKING_SLIDE_SEQUENCE)

SLIDE_TOKENS = {
    1: ("Indústria de FIDCs — ago-26", "Dados de referência: jun-26"),
    2: (
        "ESCALA DA INDÚSTRIA",
        "R$ 821,0 BI",
        "R$ 13,780 TRI",
        "CAGR 2015–18",
        "2020/19",
        "2022/21",
        "2026 YTD",
    ),
    3: (
        "Emissões | FIDCs seguem ganhando escala nas emissões",
        "CVM/SRE:",
        "ANBIMA:",
        "snapshot jun/26",
        "FIDCs e demais instrumentos elegíveis · R$ bi",
        "Valor encerrado por instrumento · R$ bi",
        "Emissões por instrumento",
        "2025 YoY %",
        "1S26 YTD YoY",
    ),
    4: (
        "SALDO E TIPOS DE FIDCS",
        "Financeiros dominam saldo e novas emissões",
        "Saldo ex-FIC · R$ bi",
        "Participação no saldo",
        "Novas emissões por setor · R$ bi",
        "Novas emissões por setor · %",
    ),
    5: ("EMISSÕES POR CATEGORIA ANBIMA", "Emissões por setor", "Total emitido"),
    6: (
        'Abrir "Outros" revela que 63% do mercado é crédito financeiro',
        "PRECATÓRIOS E/OU AÇÕES JUDICIAIS",
        "MULTICEDENTE/MULTISACADO",
        "RECUPERAÇÃO / FIDCS NP",
    ),
    7: (
        "Adquirência é R$ 99 bi que a taxonomia oficial não mostra",
        "33 CNPJs reclassificados, 12,1% do PL",
    ),
    8: ("Financeiro explicou 70% do crescimento da carteira",),
    9: ("RANKING · TOP 20 FIDCs",),
    SLIDE_OFFERS_VOLUME: ("Emissões crescem 15% no semestre", "jan–dez", "R$ 65,5 bi em 771 ofertas no jan–jun/26"),
    SLIDE_OFFER_TICKETS: ("22 ofertas concentram 42% de todo o volume", "> R$ 100 mi"),
    SLIDE_OFFER_REGIME: (
        "OFERTAS · VOLUME E REGIME",
        "Emissões | Garantia firme",
        "YoY YTD",
        "Melhores esforços repr. 70% do volume em 2026",
        "Número de ofertas",
        "Regime de colocação · volume",
    ),
    _contract_slide_number("o que muda"): (
        "O que muda a leitura do mercado",
        "RCVM 175",
        "771 OFERTAS",
        "R$ 65,5 BI",
        "R$ 32,4 BI",
        "DOIS FIDCS CIELO",
    ),
    SLIDE_PROVIDER_HISTORY: ("QI lidera administração; BTG lidera gestão e custódia",),
    SLIDE_PROVIDER_RANKING: ("PRESTADORES · RANKING E CONCENTRAÇÃO",),
    _contract_slide_number("quase todo o volume"): ("Quase todo o volume vai para o investidor profissional",),
    SLIDE_HOLDER_DISTRIBUTION: ("DISTRIBUIÇÃO POR NÚMERO DE COTISTAS",),
}

for slide_number, contract_tokens in enumerate(
    EXPECTED_SLIDE_SEQUENCE,
    start=1,
):
    SLIDE_TOKENS.setdefault(slide_number, contract_tokens)

REQUIRED_WORKBOOK_SHEETS_V51 = {
    "QA Inadimplência",
    "Base competência-CNPJ",
    "Base por fundo-CNPJ",
    "Concentração de monoestruturas",
    "Market share por subtipo",
    "Top 20 FIDCs",
    "Top 20 Outros",
    "Curadoria Top 20",
    "Reconciliação Tabelas I-II",
    "Comparativos históricos",
    "Ranking prestadores",
    "Taxonomia adquirência",
    "Curadoria Cartão",
    "Top 20 por Tipo ANBIMA",
    "Auditoria Top 20 Tipo",
    "Taxonomia por nível",
    "Curadoria flagship",
    "Carteira 1 curadoria",
    "Carteira 1 vs flagships",
    "Risco estrutural ativos",
    "Risco estrutural taxonomia",
    "Carteira 1 evolução",
    "Curadoria Outros Top 100",
    "Dispersão inadimplência",
    "Atribuição prestadores",
    "Fluxos prestadores",
    "Migração CBSF",
    "Checks revisão",
    "Inadimplência por recebível",
    "Histórico inad. coorte",
    "Ranking independentes",
    "FIDCs por banco",
    "Detalhe coorte bancos",
    "Ofertas encerradas",
    "Comparativo renda fixa",
    "Regime de colocação",
    "Histograma ofertas",
    "Originadores 2026",
    "Top 15 ofertas",
    "Auditoria emissões",
    "Cobertura emissões",
    "Curadoria perfis",
    "Validação emissões",
    "Emissões por categoria",
    "Público-alvo ofertas",
    "Principais conclusões",
    "Cedentes · Leia-me",
    "Cedentes · Top 437",
    "Cedentes · Cobertura",
    "Taxonomia · de-para",
    "Taxonomia · Outros",
    "Taxonomia · impacto",
}


def _require(path: Path) -> None:
    if not path.exists():
        pytest.skip(f"artefato ainda não gerado: {path}")


def _slide_text(archive: ZipFile, slide_number: int) -> str:
    root = ET.fromstring(archive.read(f"ppt/slides/slide{slide_number}.xml"))
    return " ".join(node.text or "" for node in root.iter(f"{{{DML}}}t"))


def _slide_chart_paths(archive: ZipFile, slide_number: int) -> list[str]:
    rels_path = f"ppt/slides/_rels/slide{slide_number}.xml.rels"
    rels = ET.fromstring(archive.read(rels_path))
    paths: list[str] = []
    for rel in rels.findall(f"{{{PACKAGE_REL}}}Relationship"):
        if not rel.attrib.get("Type", "").endswith("/chart"):
            continue
        target = rel.attrib["Target"]
        paths.append(
            target.lstrip("/")
            if target.startswith("/")
            else posixpath.normpath(posixpath.join("ppt/slides", target))
        )
    return paths


def _native_table_count(archive: ZipFile, slide_number: int) -> int:
    slide = ET.fromstring(archive.read(f"ppt/slides/slide{slide_number}.xml"))
    return len(slide.findall(f".//{{{DML}}}tbl"))


def _cell_text(cell: ET.Element) -> str:
    return "".join(node.text or "" for node in cell.iter(f"{{{DML}}}t")).strip()


def _xfrm_bbox(xfrm: ET.Element | None) -> tuple[int, int, int, int] | None:
    if xfrm is None:
        return None
    offset = xfrm.find(f"{{{DML}}}off")
    extent = xfrm.find(f"{{{DML}}}ext")
    if offset is None or extent is None:
        return None
    return (
        int(offset.attrib["x"]),
        int(offset.attrib["y"]),
        int(extent.attrib["cx"]),
        int(extent.attrib["cy"]),
    )


def _native_table_bboxes(
    archive: ZipFile, slide_number: int
) -> list[tuple[int, int, int, int]]:
    slide = ET.fromstring(archive.read(f"ppt/slides/slide{slide_number}.xml"))
    bboxes: list[tuple[int, int, int, int]] = []
    for frame in slide.findall(f".//{{{PML}}}graphicFrame"):
        if frame.find(f".//{{{DML}}}tbl") is None:
            continue
        bbox = _xfrm_bbox(frame.find(f"{{{PML}}}xfrm"))
        assert bbox is not None
        bboxes.append(bbox)
    return bboxes


def _shape_bboxes(
    archive: ZipFile, slide_number: int
) -> list[tuple[tuple[int, int, int, int], str]]:
    slide = ET.fromstring(archive.read(f"ppt/slides/slide{slide_number}.xml"))
    bboxes: list[tuple[tuple[int, int, int, int], str]] = []
    for shape in slide.findall(f".//{{{PML}}}sp"):
        bbox = _xfrm_bbox(shape.find(f"{{{PML}}}spPr/{{{DML}}}xfrm"))
        if bbox is None:
            continue
        text = " ".join(node.text or "" for node in shape.iter(f"{{{DML}}}t"))
        bboxes.append((bbox, text))
    return bboxes


def _overlap(
    left: tuple[int, int, int, int], right: tuple[int, int, int, int]
) -> bool:
    left_x, left_y, left_w, left_h = left
    right_x, right_y, right_w, right_h = right
    return (
        max(left_x, right_x) < min(left_x + left_w, right_x + right_w)
        and max(left_y, right_y) < min(left_y + left_h, right_y + right_h)
    )


def _sheet_names(archive: ZipFile) -> set[str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    return {
        sheet.attrib["name"]
        for sheet in workbook.findall(f".//{{{SHEET}}}sheet")
    }


def test_export_and_renderer_declare_dynamic_slide_contract() -> None:
    export_source = (ROOT / "services" / "industry_revision_export.py").read_text(
        encoding="utf-8"
    )
    renderer_source = (
        ROOT / "scripts" / "build_fidc_revision_artifacts.mjs"
    ).read_text(encoding="utf-8")

    assert EXPECTED_SLIDES == len(EXPECTED_SLIDE_SEQUENCE)
    assert EXPECTED_SLIDES == 37
    assert len(TYPE_RANKING_SLIDE_SEQUENCE) == 8
    assert STRUCTURAL_MVP_SLIDE_SEQUENCE == (
        ("risco estrutural", "financeiro", "carteira i"),
        ("risco estrutural", "adquirencia", "carteira i"),
        ("risco estrutural", "agro / revenda", "carteira i"),
        ("risco estrutural", "risco corporativo", "carteira i"),
        ("risco estrutural", "consignado inss e fgts", "carteira i"),
        ("risco estrutural", "factoring", "carteira i"),
    )
    assert len(STRUCTURAL_MVP_SLIDE_SEQUENCE) == len(STRUCTURAL_MVP_SLIDES) == 6
    assert len(CURRENT_TOP15_SLIDE_SEQUENCE) == 2
    assert len(HISTORICAL_TOP15_SLIDE_SEQUENCE) == 4
    assert HISTORICAL_TOP15_SLIDE_SEQUENCE == (
        (
            "top 15",
            "historico",
            "maiores ofertas de 2024",
            "1/2",
            "agencia",
            "rating",
        ),
        (
            "top 15",
            "historico",
            "maiores ofertas de 2024",
            "2/2",
            "agencia",
            "rating",
        ),
        (
            "top 15",
            "historico",
            "maiores ofertas de 2023",
            "1/2",
            "agencia",
            "rating",
        ),
        (
            "top 15",
            "historico",
            "maiores ofertas de 2023",
            "2/2",
            "agencia",
            "rating",
        ),
    )
    assert HISTORICAL_TOP15_TABLE_DIMENSIONS == (
        (9, 12),
        (8, 12),
        (9, 12),
        (8, 12),
    )
    assert sum(rows - 1 for rows, _ in HISTORICAL_TOP15_TABLE_DIMENSIONS[:2]) == 15
    assert sum(rows - 1 for rows, _ in HISTORICAL_TOP15_TABLE_DIMENSIONS[2:]) == 15
    assert "EXPECTED_SLIDES = len(EXPECTED_SLIDE_SEQUENCE)" in export_source
    assert "const SLIDE_CONTRACT_V1 = Object.freeze([" in renderer_source
    assert "const STRUCTURAL_MVP_SLIDE_SEQUENCE = Object.freeze([" in renderer_source
    assert "...STRUCTURAL_MVP_SLIDE_SEQUENCE.map((entry) => entry.id)" in renderer_source
    assert "const EXPECTED_SLIDES = SLIDE_CONTRACT_V1.length;" in renderer_source
    assert re.search(r"EXPECTED_SLIDES\s*!==\s*\d+", renderer_source) is None
    for slide_id in (
        "top20_fomento_2026",
        "top20_fomento_2025",
        "top20_agro_2026",
        "top20_agro_2025",
        "top20_financeiro_2026",
        "top20_financeiro_2025",
        "top20_outros_2026",
        "top20_outros_2025",
        "top15_current_2026",
        "top15_current_2025",
        "top15_history_2024_1_2",
        "top15_history_2024_2_2",
        "top15_history_2023_1_2",
        "top15_history_2023_2_2",
    ):
        assert f'"{slide_id}"' in renderer_source
    assert "function addTop20ByAnbimaTypeSlide(" in renderer_source
    assert "function addCurrentTop15Slide(" in renderer_source
    assert "function addHistoricalTop15Slide(" in renderer_source
    for sheet_name in REQUIRED_WORKBOOK_SHEETS_V51:
        assert f'"{sheet_name}"' in export_source


def test_deck_matches_dynamic_reviewed_narrative_order() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        slide_members = {
            name
            for name in archive.namelist()
            if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
        }
        assert len(slide_members) == TARGET_SLIDES

        for slide_number, tokens in SLIDE_TOKENS.items():
            text = _slide_text(archive, slide_number)
            folded_text = _fold(text)
            for token in tokens:
                assert _fold(token) in folded_text, (
                    f"slide {slide_number} deveria conter {token!r}; "
                    f"texto observado: {text[:240]!r}"
                )

        full_text = "\n".join(
            _slide_text(archive, number) for number in range(1, TARGET_SLIDES + 1)
        )

    assert "APÊNDICE · CURADORIA TOP 20" not in full_text
    assert "APÊNDICE · CASO ATLÂNTICO" not in full_text
    assert "OBSERVABILIDADE DA INADIMPLÊNCIA" not in full_text


def test_validator_rejects_a_hidden_slide_outside_slide_three() -> None:
    _require(PPTX)
    mutated = BytesIO()
    with ZipFile(PPTX) as source, ZipFile(mutated, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            content = source.read(member.filename)
            if member.filename == "ppt/slides/slide4.xml":
                root = ET.fromstring(content)
                root.set("show", "0")
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target.writestr(member, content)

    with pytest.raises(RevisionExportUnavailable, match="slide oculto"):
        validate_revision_pptx(mutated.getvalue())


@pytest.mark.parametrize(
    ("part_prefix", "blocked_text"),
    [
        ("ppt/slides/", "Clique para inserir o subtítulo"),
        ("ppt/slides/", "Click to add title"),
        ("ppt/notesSlides/", "Atualizar para junho"),
        ("ppt/notesSlides/", "Copilot"),
        ("ppt/notesSlides/", "Claude Code"),
        ("ppt/notesSlides/", "Prompt antigo"),
    ],
)
def test_validator_rejects_placeholders_and_old_instructions_in_slides_or_notes(
    part_prefix: str,
    blocked_text: str,
) -> None:
    _require(PPTX)
    mutated = BytesIO()
    with ZipFile(PPTX) as source, ZipFile(mutated, "w", ZIP_DEFLATED) as target:
        candidates = sorted(
            member.filename
            for member in source.infolist()
            if member.filename.startswith(part_prefix)
            and member.filename.endswith(".xml")
            and "/_rels/" not in member.filename
        )
        target_part = next(
            name
            for name in candidates
            if next(
                ET.fromstring(source.read(name)).iter(f"{{{DML}}}t"),
                None,
            )
            is not None
        )
        for member in source.infolist():
            content = source.read(member.filename)
            if member.filename == target_part:
                root = ET.fromstring(content)
                text_node = next(root.iter(f"{{{DML}}}t"))
                text_node.text = f"{text_node.text or ''} {blocked_text}"
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target.writestr(member, content)

    with pytest.raises(
        RevisionExportUnavailable,
        match="placeholder ou instrução antiga",
    ):
        validate_revision_pptx(mutated.getvalue())


def test_validator_allows_legitimate_methodological_use_of_prompt() -> None:
    package = BytesIO()
    with ZipFile(package, "w", ZIP_DEFLATED) as archive:
        archive.writestr(
            "ppt/notesSlides/notesSlide1.xml",
            (
                '<p:notes xmlns:p="http://schemas.openxmlformats.org/'
                'presentationml/2006/main" '
                'xmlns:a="http://schemas.openxmlformats.org/'
                'drawingml/2006/main">'
                "<a:t>Prompt usado para atualizar este artefato</a:t>"
                "</p:notes>"
            ),
        )
    with ZipFile(BytesIO(package.getvalue())) as archive:
        _validate_no_blocked_audience_copy(archive)


def test_standard_deck_omits_transition_slides_and_local_file_references() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        all_text = "\n".join(
            _slide_text(archive, number) for number in range(1, TARGET_SLIDES + 1)
        )

    for removed_title in (
        "OUTROS · FOCO ANALÍTICO",
        "OUTROS · TABELA II ANALÍTICA",
        "OUTROS · TAXONOMIA FUNCIONAL N1",
        "OUTROS · TAXONOMIA FUNCIONAL N2",
        "RANKING · TOP 20 OUTROS",
        "CBSF / REAG · DESTINO DOS FUNDOS",
        "PRESTADORES · MIGRAÇÃO EM ADMINISTRAÇÃO",
        "PRESTADORES · MIGRAÇÃO EM GESTÃO",
        "PRESTADORES · MIGRAÇÃO EM CUSTÓDIA",
        "CONCENTRAÇÃO DAS MONOESTRUTURAS",
        "MARKET SHARE · ADMINISTRAÇÃO",
        "MARKET SHARE · GESTÃO",
        "MARKET SHARE · CUSTÓDIA",
        "ADMINISTRAÇÃO POR SUBTIPO",
        "GESTÃO POR SUBTIPO",
        "CUSTÓDIA POR SUBTIPO",
    ):
        assert removed_title not in all_text
    assert ".csv" not in all_text.lower()
    assert ".xml" not in all_text.lower()
    assert "/users/" not in all_text.lower()

def test_scale_slide_keeps_two_native_bar_charts() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        chart_paths = _slide_chart_paths(archive, 2)
        charts = [ET.fromstring(archive.read(path)) for path in chart_paths]

    bar_charts = [
        chart for chart in charts if chart.find(f".//{{{CHART}}}barChart") is not None
    ]
    assert len(bar_charts) == 2

    groupings = [
        chart.find(f".//{{{CHART}}}barChart/{{{CHART}}}grouping")
        for chart in bar_charts
    ]
    assert {group.attrib.get("val") for group in groupings if group is not None} == {
        "clustered",
        "stacked",
    }

    total_series = []
    for chart in charts:
        for series in chart.findall(f".//{{{CHART}}}lineChart/{{{CHART}}}ser"):
            name = " ".join(
                node.text or "" for node in series.findall(f".//{{{CHART}}}v")
            )
            if "Total" in name:
                total_series.append(series)
    assert len(total_series) == 1
    for series in total_series:
        title = series.find(f"{{{CHART}}}tx")
        assert title is not None
        assert title.find(f"{{{CHART}}}v") is not None
        assert title.find(f"{{{CHART}}}strLit") is None
        labels = series.find(f"{{{CHART}}}dLbls")
        assert labels is not None
        children = [node.tag for node in series]
        assert children.index(f"{{{CHART}}}dLbls") < children.index(
            f"{{{CHART}}}cat"
        ) < children.index(f"{{{CHART}}}val")
        position = labels.find(f"{{{CHART}}}dLblPos")
        assert position is not None
        assert position.attrib.get("val") == "t"
        number_format = labels.find(f"{{{CHART}}}numFmt")
        assert number_format is not None
        assert number_format.attrib.get("formatCode") == "[>=1000]#\\.##0;0"


def test_scale_slide_uses_only_ex_fic_pl_and_explicit_brazilian_labels() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        text = _slide_text(archive, 2)

    assert "FIDCs ex-FIC" in text
    assert "R$ 821,0 bi" in text
    assert "crédito privado ampliado totaliza R$ 13,780 tri" in text
    assert "saldo FIC" not in text


def test_ex_fic_pl_chart_has_native_value_labels_for_every_data_point() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        bar_charts = []
        for path in _slide_chart_paths(archive, 2):
            root = ET.fromstring(archive.read(path))
            bar = root.find(f".//{{{CHART}}}barChart")
            if bar is not None:
                bar_charts.append(bar)

    ex_fic = next(
        bar
        for bar in bar_charts
        if len(bar.findall(f"{{{CHART}}}ser")) == 1
    )
    labels = ex_fic.find(f"{{{CHART}}}dLbls")
    assert labels is not None
    show_value = labels.find(f"{{{CHART}}}showVal")
    assert show_value is not None and show_value.attrib.get("val") == "1"
    series = ex_fic.find(f"{{{CHART}}}ser")
    assert series is not None
    points = series.findall(
        f"{{{CHART}}}val/{{{CHART}}}numLit/{{{CHART}}}pt"
    )
    assert len(points) == 12


def test_analytical_taxonomy_uses_only_bba_colors_and_labels_all_periods() -> None:
    _require(PPTX)
    allowed = {
        "151515",
        "30353A",
        "73787D",
        "8D9399",
        "D7DADD",
        "E7E9EB",
        "EC7000",
        "FFFFFF",
    }
    with ZipFile(PPTX) as archive:
        charts = []
        for path in _slide_chart_paths(archive, 6):
            root = ET.fromstring(archive.read(path))
            bar = root.find(f".//{{{CHART}}}barChart")
            if bar is not None:
                charts.append(bar)

    assert len(charts) == 2
    for bar in charts:
        labels = bar.find(f"{{{CHART}}}dLbls")
        assert labels is not None
        show_value = labels.find(f"{{{CHART}}}showVal")
        assert show_value is not None and show_value.attrib.get("val") == "1"
        series = bar.findall(f"{{{CHART}}}ser")
        assert len(series) == 7
        for item in series:
            colors = {
                str(node.attrib.get("val") or "").upper()
                for node in item.iter(f"{{{DML}}}srgbClr")
            }
            assert colors <= allowed
            values = item.findall(
                f"{{{CHART}}}val/{{{CHART}}}numLit/{{{CHART}}}pt"
            )
            item_labels = item.find(f"{{{CHART}}}dLbls")
            assert item_labels is not None
            assert len(item_labels.findall(f"{{{CHART}}}dLbl")) == len(values) == 4


def test_all_native_chart_data_labels_respect_ten_point_floor() -> None:
    _require(PPTX)
    violations: list[tuple[str, int]] = []
    with ZipFile(PPTX) as archive:
        chart_paths = [
            name
            for name in archive.namelist()
            if name.endswith(".xml")
            and (
                name.startswith("ppt/charts/chart")
                or name.startswith("ppt/slides/charts/chart")
            )
        ]
        for path in chart_paths:
            root = ET.fromstring(archive.read(path))
            for labels in root.findall(f".//{{{CHART}}}dLbls"):
                for properties in labels.findall(f".//{{{DML}}}defRPr"):
                    size = properties.attrib.get("sz")
                    if size is not None and int(size) < 1000:
                        violations.append((path, int(size)))

    assert not violations, violations


def test_annual_issuance_slide_contains_the_consolidated_anbima_taxonomy_table() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        slide = ET.fromstring(archive.read("ppt/slides/slide5.xml"))
        text = _slide_text(archive, 5)
    tables = slide.findall(f".//{{{DML}}}tbl")
    assert len(tables) == 1
    assert (
        len(tables[0].findall(f"{{{DML}}}tr")),
        len(tables[0].findall(f"{{{DML}}}tblGrid/{{{DML}}}gridCol")),
    ) == ISSUANCE_TAXONOMY_TABLE_DIMENSIONS[0]
    for token in (
        "EMISSÕES POR CATEGORIA ANBIMA",
        "Fomento Mercantil",
        "Agro, Indústria e Comércio",
        "Financeiro",
        "Outros",
        "Total emitido",
    ):
        assert token in text


def test_structural_chapter_uses_six_mvp_slides_and_keeps_workbook_audit() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        for slide_number, contract_tokens in zip(
            STRUCTURAL_MVP_SLIDES,
            STRUCTURAL_MVP_SLIDE_SEQUENCE,
            strict=True,
        ):
            text = _slide_text(archive, slide_number)
            folded_text = _fold(text)
            assert all(_fold(token) in folded_text for token in contract_tokens)

    assert {"Risco estrutural ativos", "Risco estrutural taxonomia"}.issubset(
        REQUIRED_WORKBOOK_SHEETS_V51
    )
    renderer_source = (
        ROOT / "scripts" / "build_fidc_revision_artifacts.mjs"
    ).read_text(encoding="utf-8")
    assert "STRUCTURAL_MVP_SLIDE_SEQUENCE" in renderer_source
    assert "payload.portfolio_export_carteira_101 || []" in renderer_source
    assert "payload.portfolio_export_flagships || []" in renderer_source


def test_native_chart_series_titles_use_schema_supported_forms() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        chart_paths = [
            name
            for name in archive.namelist()
            if "/charts/chart" in name and name.endswith(".xml")
        ]
        assert chart_paths
        for chart_path in chart_paths:
            chart = ET.fromstring(archive.read(chart_path))
            invalid_titles = chart.findall(
                f".//{{{CHART}}}ser/{{{CHART}}}tx/{{{CHART}}}strLit"
            )
            assert invalid_titles == []


def test_native_charts_are_bound_to_ptbr_without_currency_locale_prefix() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        chart_paths = [
            name
            for name in archive.namelist()
            if (
                name.endswith(".xml")
                and (
                    name.startswith("ppt/charts/chart")
                    or name.startswith("ppt/slides/charts/chart")
                )
            )
        ]
        assert chart_paths
        for chart_path in chart_paths:
            root = ET.fromstring(archive.read(chart_path))
            language = root.find(f"{{{CHART}}}lang")
            assert language is not None
            assert language.attrib.get("val") == "pt-BR"
            codes = [
                str(node.attrib.get("formatCode") or "")
                for node in root.iter(f"{{{CHART}}}numFmt")
            ]
            codes.extend(
                str(node.text or "")
                for node in root.iter(f"{{{CHART}}}formatCode")
            )
            for code in codes:
                assert "[$-416]" not in code


def test_visible_slide_text_uses_brazilian_decimal_separators() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        for slide_number in range(1, TARGET_SLIDES + 1):
            raw = archive.read(f"ppt/slides/slide{slide_number}.xml")
            text = _slide_text(archive, slide_number)
            assert b'lang="en-US"' not in raw
            assert re.search(r"R\$\s*\d+\.\d{1,2}(?:\s|$)", text) is None
            assert re.search(r"\d+\.\d+%", text) is None


def test_combined_provider_ranking_uses_six_native_charts_and_no_tables() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        assert len(_slide_chart_paths(archive, SLIDE_PROVIDER_HISTORY)) >= 6
        assert _native_table_count(archive, SLIDE_PROVIDER_HISTORY) == 0


@pytest.mark.parametrize(
    ("slide_number", "minimum_charts", "minimum_tables"),
    [
        (3, 2, 1),  # séries CVM e ANBIMA e tabela de crescimento
        (4, 4, 0),  # saldos, participações e emissões em R$ bi e %
        (5, 2, 1),  # taxonomia em R$ bi, % e tabela
        (9, 0, 2),  # Top 20 FIDCs em tabelas nativas
        (SLIDE_OFFERS_VOLUME, 2, 1),  # volume/ticket FY/YTD e acumulado mensal
    ]
    + [
        (slide_number, 0, 1)
        for slide_number in SLIDES_TYPE_RANKING
    ],
)
def test_new_analytical_slides_use_native_office_structures(
    slide_number: int, minimum_charts: int, minimum_tables: int
) -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        assert len(_slide_chart_paths(archive, slide_number)) >= minimum_charts
        assert _native_table_count(archive, slide_number) >= minimum_tables


def test_top_type_slides_split_each_type_and_period_with_originator_column() -> None:
    _require(PPTX)
    _require(PAYLOAD)
    with ZipFile(PPTX) as archive:
        slide_texts = [
            _slide_text(archive, slide_number)
            for slide_number in SLIDES_TYPE_RANKING
        ]
        assert all(
            _native_table_count(archive, slide_number) == 1
            for slide_number in SLIDES_TYPE_RANKING
        )
    assert len(slide_texts) == 8
    for text in slide_texts:
        assert "Top 15" in text
        assert any(period in text for period in ("jun/26", "dez/25"))
        assert "Originador" in text
        assert "Remuneração-alvo" in text
        assert "Preço por cota" not in text
    for type_name in (
        "Fomento Mercantil",
        "Agro, Indústria e Comércio",
        "Financeiro",
        "Outros",
    ):
        assert sum(type_name in text for text in slide_texts) == 2
    for current_slide, comparison_slide in zip(
        slide_texts[::2],
        slide_texts[1::2],
        strict=True,
    ):
        assert "jun/26 · Top 15" in current_slide
        assert "dez/25 · Top 15" in comparison_slide


def test_offer_ticket_distribution_uses_three_native_clustered_charts() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        chart_paths = _slide_chart_paths(archive, SLIDE_OFFER_TICKETS)
        assert _native_table_count(archive, SLIDE_OFFER_TICKETS) == 0
        charts = [ET.fromstring(archive.read(path)) for path in chart_paths]
        bar_charts = [
            chart for chart in charts if chart.find(f".//{{{CHART}}}barChart") is not None
        ]
        assert len(bar_charts) == 3
        for chart in bar_charts:
            bar_chart = chart.find(f".//{{{CHART}}}barChart")
            assert bar_chart is not None
            grouping = bar_chart.find(f"{{{CHART}}}grouping")
            assert grouping is not None
            assert grouping.attrib.get("val") == "clustered"
            assert len(bar_chart.findall(f"{{{CHART}}}ser")) == 3


def test_offer_placement_slide_uses_four_native_bar_charts() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        chart_paths = _slide_chart_paths(archive, SLIDE_OFFER_REGIME)
        assert _native_table_count(archive, SLIDE_OFFER_REGIME) == 0
        charts = [ET.fromstring(archive.read(path)) for path in chart_paths]
        assert sum(
            chart.find(f".//{{{CHART}}}barChart") is not None for chart in charts
        ) == 4


def test_provider_concentration_has_two_native_charts() -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        assert len(_slide_chart_paths(archive, SLIDE_PROVIDER_RANKING)) == 2
        assert _native_table_count(archive, SLIDE_PROVIDER_RANKING) == 0


@pytest.mark.parametrize(
    ("slide_number", "expected_tables"),
    [
        (slide_number, 1)
        for slide_number in (*SLIDES_TOP15_CURRENT, *SLIDES_TOP15_HISTORY)
    ],
)
def test_top15_offer_slides_use_only_on_canvas_native_tables(
    slide_number: int, expected_tables: int
) -> None:
    _require(PPTX)
    with ZipFile(PPTX) as archive:
        presentation = ET.fromstring(archive.read("ppt/presentation.xml"))
        slide_size = presentation.find(f"{{{PML}}}sldSz")
        assert slide_size is not None
        canvas_width = int(slide_size.attrib["cx"])
        canvas_height = int(slide_size.attrib["cy"])

        tables = _native_table_bboxes(archive, slide_number)
        assert len(tables) == expected_tables
        assert _native_table_count(archive, slide_number) == expected_tables
        assert _slide_chart_paths(archive, slide_number) == []
        for left, top, width, height in tables:
            assert left >= 0
            assert top >= 0
            assert width > 0
            assert height > 0
            assert left + width <= canvas_width
            assert top + height <= canvas_height

        overlapping_shapes = [
            text or "<shape sem texto>"
            for shape_bbox, text in _shape_bboxes(archive, slide_number)
            if any(_overlap(shape_bbox, table_bbox) for table_bbox in tables)
        ]
        assert overlapping_shapes == []


def test_june_offer_slide_uses_straight_markerless_native_line_chart() -> None:
    _require(PPTX)
    payload = json.loads(PAYLOAD.read_text(encoding="utf-8"))
    with ZipFile(PPTX) as archive:
        charts = [
            ET.fromstring(archive.read(path))
            for path in _slide_chart_paths(archive, SLIDE_OFFERS_VOLUME)
        ]
    line_charts = [
        chart
        for chart in charts
        if any(
            len(series.findall(f".//{{{CHART}}}pt")) > 2
            for series in chart.findall(
                f".//{{{CHART}}}lineChart/{{{CHART}}}ser"
            )
        )
    ]
    assert len(line_charts) == 1
    line_chart = line_charts[0]
    blanks = line_chart.find(f".//{{{CHART}}}dispBlanksAs")
    assert blanks is not None and blanks.attrib.get("val") == "gap"
    series = line_chart.findall(f".//{{{CHART}}}lineChart/{{{CHART}}}ser")
    assert len(series) == 3
    series_by_name = {
        str(item.findtext(f"{{{CHART}}}tx/{{{CHART}}}v") or ""): item
        for item in series
    }
    assert set(series_by_name) == {"2024", "2025", "2026"}
    for item in series:
        symbol = item.find(f".//{{{CHART}}}marker/{{{CHART}}}symbol")
        assert symbol is not None and symbol.attrib.get("val") == "none"
        smooth = item.find(f"{{{CHART}}}smooth")
        assert smooth is None or smooth.attrib.get("val") in {"0", "false"}

    point_indices: dict[str, list[int]] = {}
    point_values: dict[str, dict[int, float]] = {}
    for name, item in series_by_name.items():
        points = item.findall(
            f"{{{CHART}}}val/{{{CHART}}}numLit/{{{CHART}}}pt"
        )
        point_indices[name] = [int(point.attrib["idx"]) for point in points]
        point_values[name] = {
            int(point.attrib["idx"]): float(
                point.findtext(f"{{{CHART}}}v") or "nan"
            )
            for point in points
        }

    assert point_indices["2024"] == list(range(12))
    assert point_indices["2025"] == list(range(12))
    assert point_indices["2026"] == list(range(6))
    expected_june_2026 = sum(
        float(row["registered_volume_brl"])
        for row in payload["closed_offers_monthly"]
        if int(row["year"]) == 2026 and int(row["month"]) <= 6
    ) / 1e9
    assert point_values["2026"][5] == pytest.approx(expected_june_2026)


def test_workbook_exposes_the_v63_analysis_tabs() -> None:
    _require(XLSX)
    with ZipFile(XLSX) as archive:
        sheet_names = _sheet_names(archive)

    assert REQUIRED_WORKBOOK_SHEETS_V51.issubset(sheet_names), (
        "abas ausentes: "
        + ", ".join(sorted(REQUIRED_WORKBOOK_SHEETS_V51 - sheet_names))
    )


def test_emission_audit_sheet_materializes_180_sourced_rows_and_preserves_nd() -> None:
    _require(XLSX)
    workbook = load_workbook(XLSX, read_only=True, data_only=False)
    sheet = workbook["Auditoria emissões"]
    headers = [sheet.cell(4, column).value for column in range(1, 25)]
    assert headers == [
        "Bloco do deck",
        "Tabela / período",
        "CNPJ",
        "ID da emissão",
        "Fundo",
        "Originador",
        "Subordinação mínima",
        "Preço unitário por tipo de cota",
        "Remuneração-alvo por tipo de cota",
        "Cedente",
        "Sacado",
        "Cedente / Originador literal*",
        "Tipo de recebível literal*",
        "Fonte enriquecimento manual",
        "Fonte originador",
        "Fonte cedente",
        "Fonte originador / cedente",
        "Natureza do mínimo",
        "Fonte subordinação",
        "Fonte preço",
        "Fonte remuneração",
        "Fonte sacado",
        "Motivo N/D",
        "Status",
    ]
    rows = list(
        sheet.iter_rows(min_row=5, max_row=184, min_col=1, max_col=24, values_only=True)
    )
    assert len(rows) == 180
    assert sum(row[0] == "slides 10–17" for row in rows) == 120
    assert sum(row[0] == "slides 21–22" for row in rows) == 60
    assert all(value not in {None, ""} for row in rows for value in row)
    assert all(
        re.fullmatch(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", str(row[2]))
        for row in rows
    )
    assert all(str(row[3]).startswith(("E ", "N/D")) for row in rows)
    header_index = {header: index for index, header in enumerate(headers)}
    remuneration_index = header_index["Remuneração-alvo por tipo de cota"]
    remuneration_source_index = header_index["Fonte remuneração"]
    remuneration_rows = [
        row for row in rows if row[remuneration_index] != "N/D"
    ]
    assert remuneration_rows
    assert all(
        re.search(r"(?:CDI|DI|IPCA|SELIC|IGP-M).+%|%.+(?:CDI|DI|IPCA|SELIC|IGP-M)", str(row[remuneration_index]))
        for row in remuneration_rows
    )
    assert all("R$" not in str(row[remuneration_index]) for row in remuneration_rows)
    assert all(row[remuneration_source_index] != "N/D" for row in remuneration_rows)

    coverage = workbook["Cobertura emissões"]
    coverage_headers = [coverage.cell(4, column).value for column in range(1, 17)]
    assert coverage_headers == [
        "Tabela / período",
        "Tipo ANBIMA",
        "Competência",
        "Campo",
        "Linhas",
        "Antes · com dado",
        "Antes · cobertura",
        "Antes · PL coberto",
        "Antes · cobertura PL",
        "Depois · com dado",
        "Depois · cobertura",
        "Depois · PL coberto",
        "Depois · cobertura PL",
        "N/D depois",
        "Piso de publicação",
        "Piso atendido?",
    ]
    coverage_rows = list(
        coverage.iter_rows(min_row=5, max_row=44, min_col=1, max_col=16, values_only=True)
    )
    assert len(coverage_rows) == 40
    assert all(row[4] == 15 for row in coverage_rows)
    assert sum(row[3] == "Remuneração-alvo" for row in coverage_rows) == 8
    assert all(row[3] != "Preço por cota" for row in coverage_rows)

    profiles = workbook["Curadoria perfis"]
    assert profiles["A4"].value == "CNPJ"
    assert profiles["D4"].value == "Classificação do texto"
    assert profiles["E4"].value == "Aplicação como Cedente"
    assert profiles["F4"].value == "Valor aplicado como Cedente"
    assert profiles["G4"].value == "Aplicação como Originador"
    assert profiles["H4"].value == "Valor aplicado como Originador"
    assert profiles["M4"].value == "Data da consulta"


def test_workbook_preserves_taxonomy_levels_and_flagship_documentary_gaps() -> None:
    _require(XLSX)
    workbook = load_workbook(XLSX, read_only=True, data_only=False)

    taxonomy = workbook["Taxonomia por nível"]
    assert taxonomy["A4"].value == "Nível"
    taxonomy_rows = list(taxonomy.iter_rows(min_row=5, values_only=True))
    taxonomy_rows = [row for row in taxonomy_rows if row[0] not in {None, ""}]
    payload = json.loads(PAYLOAD.read_text(encoding="utf-8"))
    assert len(taxonomy_rows) == len(payload["taxonomy_level_history"])
    assert {row[0] for row in taxonomy_rows} == {
        "Foco analítico",
        "Tabela II analítica",
        "Taxonomia funcional N1",
        "Taxonomia funcional N2",
    }
    assert any(row[3] == "Judicial/Precatórios/NPL" for row in taxonomy_rows)
    assert all(isinstance(row[4], (int, float)) and row[4] >= 0 for row in taxonomy_rows)

    flagship = workbook["Curadoria flagship"]
    assert flagship["E4"].value == "CNPJ"
    assert flagship["G4"].value == "PL atual"
    assert flagship["I4"].value == "Subordinação atual / PL"
    assert flagship["N4"].value == "Mínimo júnior"
    assert flagship["R4"].value == "Preço/VNU numérico"
    assert flagship["AE4"].value == "Documento do regulamento revisto"
    assert flagship["AI4"].value == "Status da curadoria documental"
    rows = list(flagship.iter_rows(min_row=5, max_col=38, values_only=True))
    rows = [row for row in rows if row[4] not in {None, ""}]
    assert len(rows) == 47
    assert len({row[4] for row in rows}) == 47
    assert all(isinstance(row[6], (int, float)) and row[6] > 0 for row in rows)
    assert all(isinstance(row[8], (int, float)) and 0 <= row[8] <= 1 for row in rows)
    assert all(row[13] is None or row[13] > 0 for row in rows)
    assert all(row[17] is None or row[17] > 0 for row in rows)
    assert any(row[13] is None and row[14] == "N/D" for row in rows)
    assert any(row[17] is None and row[18] == "N/D" for row in rows)
    assert sum(str(row[34]).startswith("revisto") for row in rows) == 24
    assert sum(row[14] != "N/D" for row in rows) == 12

    carteira_1 = workbook["Carteira 1 curadoria"]
    assert carteira_1["C4"].value == "Raiz CNPJ · foto"
    assert carteira_1["D4"].value == "Nome · foto"
    assert carteira_1["J4"].value == "PL atual"
    assert carteira_1["L4"].value == "Subordinação atual / PL"
    assert carteira_1["O4"].value == "Mínimo júnior"
    assert carteira_1["T4"].value == "Emissão · mês/ano"
    carteira_rows = list(
        carteira_1.iter_rows(min_row=5, max_col=35, values_only=True)
    )
    carteira_rows = [row for row in carteira_rows if row[0] not in {None, ""}]
    assert len(carteira_rows) == 101
    assert len({row[7] for row in carteira_rows}) == 101
    assert sum(row[9] is not None for row in carteira_rows) == 78
    assert sum(row[11] is not None for row in carteira_rows) == 68
    assert sum(row[14] is not None for row in carteira_rows) == 83
    assert sum(row[19] != "N/D" for row in carteira_rows) == 97
    assert sum(str(row[31]).startswith("fora do perímetro FIDC") for row in carteira_rows) == 1
    assert all(row[9] is None or row[9] > 0 for row in carteira_rows)
    assert all(row[14] is None or row[14] > 0 for row in carteira_rows)

    structural = workbook["Risco estrutural ativos"]
    structural_headers = [
        value
        for value in next(
            structural.iter_rows(min_row=4, max_row=4, values_only=True)
        )
    ]
    structural_column = {
        header: index for index, header in enumerate(structural_headers)
    }
    structural_rows = list(structural.iter_rows(min_row=5, values_only=True))
    structural_rows = [row for row in structural_rows if row[0] not in {None, ""}]
    assert len(structural_rows) == 101
    junior = structural_column["Mínimo júnior documental"]
    total_support = structural_column["Suporte total/combinado"]
    headroom = structural_column["Folga"]
    assert sum(row[junior] is not None for row in structural_rows) == 83
    assert sum(
        row[junior] is not None or row[total_support] is not None
        for row in structural_rows
    ) == 99
    assert sum(row[headroom] is not None for row in structural_rows) == 23


def test_top20_type_workbook_keeps_rank_share_date_and_coverage_typed() -> None:
    _require(XLSX)
    workbook = load_workbook(XLSX, read_only=False, data_only=False)
    sheet = workbook["Top 20 por Tipo ANBIMA"]

    assert sheet["B4"].value == "Rank"
    assert isinstance(sheet["B5"].value, (int, float))
    assert sheet["F4"].value == "% do bucket"
    assert isinstance(sheet["F5"].value, (int, float))
    assert sheet["F5"].number_format == "0.0%"
    assert sheet["G4"].value == "Competência"
    assert sheet["G5"].value == "2026-06"
    assert "R$" not in sheet["G5"].number_format
    assert sheet["H4"].value == "Mai/26 disponível"
    assert isinstance(sheet["H5"].value, bool)
    assert "R$" not in sheet["H5"].number_format


def test_offer_workbook_uses_counts_billions_and_millions_consistently() -> None:
    _require(XLSX)
    workbook = load_workbook(XLSX, read_only=False, data_only=False)

    offers = workbook["Ofertas encerradas"]
    assert '"bi"' in offers["I5"].number_format
    assert '"mi"' in offers["J5"].number_format
    assert '"mi"' in offers["K5"].number_format
    assert '"bi"' in offers["L5"].number_format

    originators = workbook["Originadores 2026"]
    assert "R$" not in originators["D5"].number_format
    assert '"bi"' in originators["E5"].number_format
    assert '"mi"' in originators["F5"].number_format
    assert '"mi"' in originators["G5"].number_format
    assert '"bi"' in originators["H5"].number_format

    top15 = workbook["Top 15 ofertas"]
    assert '"bi"' in top15["I5"].number_format
    assert top15["K4"].value == "IBBA Coord-Líder?"
    assert top15["L4"].value == "IBBA Coord?"
    assert top15["S4"].value == "Garantia Firme?"
    assert top15["T4"].value == "Público"
    assert top15["U4"].value == "Nº de Inv."

    banks = workbook["FIDCs por banco"]
    assert banks["J4"].value == "Raízes de CNPJ listadas"
    assert banks["M4"].value == "Referências"
