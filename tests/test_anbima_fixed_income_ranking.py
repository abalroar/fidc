from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from services.anbima_fixed_income_ranking import (
    AnbimaRankingError,
    parse_annex_workbook,
    parse_ranking_workbook,
    summarize_participants,
    syndication_profile,
)


ANNEX_HEADER = [
    "Data do Anuncio de Encerramento",
    "Data de Registro CVM",
    "Registro CVM",
    "Classe *",
    "Regime de Colocação",
    "Companhia(s) Emissora(s)",
    "Originador",
    "PU de Emissão R$",
    "Quantidade Originada",
    "Valor por Originador (R$ Mil)",
    "Percentual Coordenado",
    "Risco (Securitização)",
    "CNPJ Companhia(s) Emissora(s)",
]

# Two FIDC operations: one syndicated 60/40 between two coordinators, one solo.
# The solo issuer books two CVM registrations on the same closing date, which
# ANBIMA counts as a single operation.
ANNEX_ROWS = [
    ["12/01/2026", "17/12/2025", "FDC/2025/001", "1.3.1", "GF", "FIDC ALFA",
     "ITAU BBA", 1000, 60000, 60000, 0.6, "CEDENTE ALFA", "11111111000191"],
    ["12/01/2026", "17/12/2025", "FDC/2025/001", "1.3.1", "GF", "FIDC ALFA",
     "BRADESCO BBI", 1000, 40000, 40000, 0.4, "CEDENTE ALFA", "11111111000191"],
    ["20/02/2026", "10/01/2026", "FDC/2026/010", "1.3.1", "ME", "FIDC BETA",
     "ITAU BBA", 1000, 10000, 10000, 1.0, "CEDENTE BETA", "22222222000172"],
    ["20/02/2026", "10/01/2026", "FDC/2026/011", "1.3.1", "ME", "FIDC BETA",
     "ITAU BBA", 1000, 5000, 5000, 1.0, "CEDENTE BETA", "22222222000172"],
    ["05/03/2026", "01/02/2026", "DEB/2026/001", "1.2.A", "GF", "EMPRESA GAMA",
     "BRADESCO BBI", 1000, 25000, 25000, 1.0, "N/A", "33333333000153"],
]

LEGEND_ROWS = [
    ["*Classes de Ativos:"],
    ["1.3.1) Emissão de Cotas Senior e Subordinada de FIDC"],
    ["1.2.A) Emissão Primária de Debêntures Simples - LP"],
]


def _write_annex(path: Path) -> None:
    workbook = openpyxl.Workbook()
    for index, sheet_name in enumerate(
        ("RF&Híbridos - Originação", "RF&Híbridos - Distribuição")
    ):
        sheet = (
            workbook.active if index == 0 else workbook.create_sheet()
        )
        sheet.title = sheet_name
        for _ in range(5):
            sheet.append([])
        sheet.append(["Tipo 1: Renda Fixa Consolidado"])
        sheet.append(["Características da Operação"])
        sheet.append(ANNEX_HEADER)
        for row in ANNEX_ROWS:
            sheet.append(row)
        for row in LEGEND_ROWS:
            sheet.append(row)
    workbook.save(path)


def _write_ranking(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Originação - Valor"
    for _ in range(5):
        sheet.append([])
    sheet.append(["Tipo 1: Renda Fixa Consolidado"])
    sheet.append(["Coordenadores", "Acumulado 2026"])
    sheet.append(["", "Ranking 2026", "Valor *", "Part."])
    sheet.append(["ITAU BBA", "1º", 75000, 75_000 / 140_000])
    sheet.append(["BRADESCO BBI", "2º", 65000, 65_000 / 140_000])
    sheet.append(["Total"])

    counts = workbook.create_sheet("Nº de Operações")
    for _ in range(5):
        counts.append([])
    counts.append(["Tipo 1: Renda Fixa Consolidado"])
    counts.append(["Coordenadores", "Acumulado 2026"])
    counts.append(["", "Ranking 2026", "Nº de Operações"])
    counts.append(["ITAU BBA", "1º", 2])
    counts.append(["BRADESCO BBI", "1º", 2])

    distribution = workbook.create_sheet("Distribuição")
    for _ in range(5):
        distribution.append([])
    distribution.append(["Tipo 1: Renda Fixa Consolidado"])
    distribution.append(["Distribuidores", "Acumulado 2026"])
    distribution.append(["", "Ranking 2026", "Valor *", "Part."])
    distribution.append(["ITAU BBA", "1º", 75000, 75_000 / 140_000])
    workbook.save(path)


def test_annex_values_convert_from_thousands_to_brl(tmp_path: Path) -> None:
    path = tmp_path / "anexo.xlsx"
    _write_annex(path)

    annex = parse_annex_workbook(path)

    originacao = annex[annex["role"].eq("originacao")]
    assert len(originacao) == len(ANNEX_ROWS)
    assert originacao["valor_brl"].sum() == pytest.approx(140_000_000.0)


def test_annex_drops_the_asset_class_legend(tmp_path: Path) -> None:
    path = tmp_path / "anexo.xlsx"
    _write_annex(path)

    annex = parse_annex_workbook(path)

    assert annex["participante"].ne("").all()
    assert annex["data_encerramento"].notna().all()
    assert not annex["emissor"].str.startswith("*").any()


def test_summary_reproduces_published_ranking(tmp_path: Path) -> None:
    annex_path = tmp_path / "anexo.xlsx"
    ranking_path = tmp_path / "ranking.xlsx"
    _write_annex(annex_path)
    _write_ranking(ranking_path)

    annex = parse_annex_workbook(annex_path)
    official = parse_ranking_workbook(ranking_path)
    summary = summarize_participants(annex, block_code="1").set_index(
        "participant"
    )
    published = official[
        official["measure"].eq("originacao_valor")
        & official["window"].eq("acumulado_ano")
    ].set_index("participant")

    for participant in ("ITAU BBA", "BRADESCO BBI"):
        assert summary.loc[participant, "volume_brl"] == pytest.approx(
            published.loc[participant, "value_brl_or_count"]
        )
        assert summary.loc[participant, "share"] == pytest.approx(
            published.loc[participant, "share"]
        )


def test_operations_group_registrations_of_the_same_issuance(
    tmp_path: Path,
) -> None:
    path = tmp_path / "anexo.xlsx"
    _write_annex(path)

    annex = parse_annex_workbook(path)
    fidc = summarize_participants(
        annex, block_code="1", classes=("1.3.1",)
    ).set_index("participant")

    # FIDC BETA closes two CVM registrations on the same date: one operation.
    assert fidc.loc["ITAU BBA", "operations"] == 2
    assert fidc.loc["ITAU BBA", "registrations"] == 3
    assert fidc.loc["ITAU BBA", "share"] == pytest.approx(75_000 / 115_000)


def test_syndication_profile_counts_coordinators_per_operation(
    tmp_path: Path,
) -> None:
    path = tmp_path / "anexo.xlsx"
    _write_annex(path)

    profile = parse_annex_workbook(path)
    result = syndication_profile(profile, block_code="1", classes=("1.3.1",))

    assert result.set_index("coordinators")["operations"].to_dict() == {
        1: 1,
        2: 1,
    }


def test_missing_sheet_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "incompleto.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Outra aba"
    workbook.save(path)

    with pytest.raises(AnbimaRankingError):
        parse_annex_workbook(path)
