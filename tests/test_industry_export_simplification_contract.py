"""Characterization contracts for the safe simplification of industry exports.

The published payload is the shared data contract.  The workbook and HTML may
become smaller, while the payload bytes and every block read by the deck or the
dashboard remain frozen until the Office bundle is deliberately republished.
"""

from __future__ import annotations

import ast
from decimal import Decimal
import hashlib
from io import BytesIO
import json
from pathlib import Path
import re

from openpyxl import load_workbook
import pandas as pd
import pytest

from scripts.publish_fidc_revision_bundle import USER_FACING_SNAPSHOT_SHEETS
from tabs import tab_industry_study


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "industry_study"
REVISION_DIR = DATA_DIR / "generated_revision"
PAYLOAD_PATH = REVISION_DIR / "artifact_payload.json"
BUNDLE_MANIFEST_PATH = REVISION_DIR / "industry_export_bundle.json"
RENDERER_PATH = ROOT / "scripts" / "build_fidc_revision_artifacts.mjs"
DASHBOARD_PATH = ROOT / "tabs" / "tab_industry_study.py"

PUBLISHED_PAYLOAD_BYTES = 19_172_365
PUBLISHED_PAYLOAD_SHA256 = (
    "99491539013d02389de422df4ab5b7bc14dc2604a7c48bf68c3cdc4e820115f2"
)
PUBLISHED_CONSUMER_DIMENSIONS_SHA256 = (
    "9e9874c7a832825d35cf0e0ff456e3b304703dc68a3194fa5562ec1b66fb22c2"
)
ANBIMA_2023_FIDC_VOLUME_BRL = 43_746_140_196.22
ANBIMA_SOURCE_WORKBOOK_SHA256 = (
    "1236172468f5aa3ddde24382bfa9c5f6372f9b35cd03993ef2482d358845b524"
)

# Optional compatibility read in the presentation renderer.  The current
# payload uses ``reag_admin_summary`` and intentionally has no legacy block.
OPTIONAL_ABSENT_PPTX_KEYS = frozenset({"reag_admin_migration"})
EXPECTED_WORKBOOK_SHEETS_TO_REMOVE = (
    "Conflitos Tab IV",
    "Warnings",
    "Ofertas anual",
    "Posição Itaú",
    "Ranking ofertas",
    "Cedentes",
    "Investidores hist",
    "Tipos investidor",
    "_Listas",
    "Cross-check taxonomia",
    "Taxonomia por CNPJ",
    "Reclass. adquirência",
    "Auditoria numérica",
    "Reclass. ANBIMA",
    "Reclass. CVM",
    "Fluxos visuais",
)
INHERITED_WORKBOOK_SHEETS_TO_REMOVE = (
    "Conflitos Tab IV",
    "Warnings",
    "Ofertas anual",
    "Posição Itaú",
    "Ranking ofertas",
    "Cedentes",
    "Investidores hist",
    "Tipos investidor",
    "_Listas",
)


def _published_payload() -> tuple[bytes, dict[str, object]]:
    payload_bytes = PAYLOAD_PATH.read_bytes()
    return payload_bytes, json.loads(payload_bytes)


def _pptx_payload_keys(renderer_source: str) -> set[str]:
    """Read direct payload accesses from the presentation half of the renderer."""

    presentation_source = renderer_source.split("function resetSheet", maxsplit=1)[0]
    keys = set(
        re.findall(
            r"\bpayload\.([A-Za-z_$][A-Za-z0-9_$]*)",
            presentation_source,
        )
    )
    keys.update(
        re.findall(
            r"""\bpayload\[\s*["']([^"']+)["']\s*\]""",
            presentation_source,
        )
    )
    return keys.difference(OPTIONAL_ABSENT_PPTX_KEYS)


def _dashboard_payload_keys(dashboard_source: str) -> set[str]:
    """Read render accesses and the versioned loader contract from Python AST."""

    tree = ast.parse(dashboard_source)
    keys: set[str] = set()
    loader: ast.FunctionDef | ast.AsyncFunctionDef | None = None

    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and (
            node.name == "_load_industry_revision_payload"
        ):
            loader = node
        if isinstance(node, ast.Call):
            if (
                isinstance(node.func, ast.Name)
                and node.func.id in {"_revision_frame", "_revision_history_frame"}
                and len(node.args) >= 2
                and isinstance(node.args[1], ast.Constant)
                and isinstance(node.args[1].value, str)
            ):
                keys.add(node.args[1].value)
            if (
                isinstance(node.func, ast.Attribute)
                and isinstance(node.func.value, ast.Name)
                and node.func.value.id == "payload"
                and node.func.attr == "get"
                and node.args
                and isinstance(node.args[0], ast.Constant)
                and isinstance(node.args[0].value, str)
            ):
                keys.add(node.args[0].value)
        if (
            isinstance(node, ast.Subscript)
            and isinstance(node.value, ast.Name)
            and node.value.id == "payload"
            and isinstance(node.slice, ast.Constant)
            and isinstance(node.slice.value, str)
        ):
            keys.add(node.slice.value)

    assert loader is not None
    for node in ast.walk(loader):
        if (
            isinstance(node, ast.Assign)
            and any(
                isinstance(target, ast.Name) and target.id == "required"
                for target in node.targets
            )
            and isinstance(node.value, ast.Set)
        ):
            keys.update(
                item.value
                for item in node.value.elts
                if isinstance(item, ast.Constant) and isinstance(item.value, str)
            )
        if (
            isinstance(node, ast.Call)
            and isinstance(node.func, ast.Attribute)
            and isinstance(node.func.value, ast.Name)
            and node.func.value.id == "required"
            and node.func.attr == "update"
            and node.args
            and isinstance(node.args[0], ast.Set)
        ):
            keys.update(
                item.value
                for item in node.args[0].elts
                if isinstance(item, ast.Constant) and isinstance(item.value, str)
            )
    return keys


def _block_dimensions(value: object) -> dict[str, object]:
    """Describe rows and columns without copying any analytical value."""

    if isinstance(value, list):
        columns = sorted(
            {
                str(column)
                for row in value
                if isinstance(row, dict)
                for column in row
            }
        )
        return {"kind": "list", "rows": len(value), "columns": columns}
    if isinstance(value, dict):
        nested_lists: dict[str, dict[str, object]] = {}
        for key, nested in value.items():
            if not isinstance(nested, list):
                continue
            nested_lists[str(key)] = {
                "rows": len(nested),
                "columns": sorted(
                    {
                        str(column)
                        for row in nested
                        if isinstance(row, dict)
                        for column in row
                    }
                ),
            }
        return {
            "kind": "dict",
            "keys": sorted(str(key) for key in value),
            "lists": nested_lists,
        }
    return {"kind": type(value).__name__}


def _dimensions_digest(
    payload: dict[str, object],
    keys: set[str],
) -> str:
    dimensions = {
        key: _block_dimensions(payload[key])
        for key in sorted(keys)
    }
    encoded = json.dumps(
        dimensions,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _truthy(values: pd.Series) -> pd.Series:
    return (
        values.fillna("")
        .astype(str)
        .str.strip()
        .str.casefold()
        .isin({"true", "1", "sim", "s", "yes", "y", "t"})
    )


def _cnpj_digits(value: object) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(14) if digits else ""


def _js_string_array_constant(source: str, name: str) -> tuple[str, ...] | None:
    declaration = re.search(rf"\bconst\s+{re.escape(name)}\s*=", source)
    if declaration is None:
        return None
    opening = source.find("[", declaration.end())
    if opening < 0:
        raise AssertionError(f"{name} deve ser declarado como lista literal")
    closing = source.find("]", opening + 1)
    if closing < 0:
        raise AssertionError(f"{name} contém lista literal não fechada")
    return tuple(
        re.findall(r"""["']([^"']+)["']""", source[opening + 1 : closing])
    )


def test_published_payload_and_static_consumer_contract_are_frozen() -> None:
    payload_bytes, payload = _published_payload()
    payload_digest = hashlib.sha256(payload_bytes).hexdigest()
    manifest = json.loads(BUNDLE_MANIFEST_PATH.read_text(encoding="utf-8"))

    assert len(payload_bytes) == PUBLISHED_PAYLOAD_BYTES
    assert payload_digest == PUBLISHED_PAYLOAD_SHA256
    assert manifest["payload"] == {
        "bytes": PUBLISHED_PAYLOAD_BYTES,
        "name": "artifact_payload.json",
        "sha256": PUBLISHED_PAYLOAD_SHA256,
    }
    assert manifest["payload_sha256"] == PUBLISHED_PAYLOAD_SHA256
    assert payload["schema_version"] == "fidc_revision_artifact_payload_v8"
    assert len(payload) == 150
    assert {
        "carteira_1_curation",
        "carteira_1_curation_ranges",
        "carteira_1_curation_summary",
        "carteira_1_flagship_comparison",
        "carteira_1_flagship_comparison_summary",
        "carteira_1_structural_assets",
        "carteira_1_structural_summary",
        "carteira_1_structural_taxonomy",
        "carteira_1_structural_watchlist",
        "carteira_1_taxonomy_history",
        "carteira_1_taxonomy_summary",
        "flagship_curation",
        "flagship_curation_summary",
        "flagship_families",
        "emission_field_audit",
        "issuance_taxonomy",
        "issuance_taxonomy_reconciliation",
        "issuance_taxonomy_table",
        "manual_cnpj_enrichment",
        "portfolio_export_carteira_101",
        "portfolio_export_coverage",
        "portfolio_export_flagships",
        "portfolio_export_gaps",
        "portfolio_export_manual_audit",
        "taxonomy_level_history",
    }.issubset(payload)

    renderer_source = RENDERER_PATH.read_text(encoding="utf-8")
    dashboard_source = DASHBOARD_PATH.read_text(encoding="utf-8")
    pptx_keys = _pptx_payload_keys(renderer_source)
    dashboard_keys = _dashboard_payload_keys(dashboard_source)
    consumer_keys = pptx_keys | dashboard_keys

    assert len(pptx_keys) == 85
    assert len(dashboard_keys) == 91
    assert len(consumer_keys) == 115
    assert {
        "carteira_1_flagship_comparison",
        "carteira_1_flagship_comparison_summary",
        "carteira_1_structural_summary",
        "carteira_1_structural_taxonomy",
        "carteira_1_structural_watchlist",
        "carteira_1_taxonomy_history",
        "carteira_1_taxonomy_summary",
        "flagship_curation_summary",
        "flagship_families",
        "issuance_taxonomy",
        "issuance_taxonomy_reconciliation",
        "issuance_taxonomy_table",
    }.issubset(pptx_keys)
    assert pptx_keys.issubset(payload)
    assert dashboard_keys.issubset(payload)
    assert _dimensions_digest(payload, consumer_keys) == (
        PUBLISHED_CONSUMER_DIMENSIONS_SHA256
    )


def test_published_bundle_has_converged_issuance_baseline(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload_bytes, disk_payload = _published_payload()
    disk_comparison = pd.DataFrame(
        disk_payload["fixed_income_offer_comparison"]
    )
    disk_fidc_2023 = disk_comparison[
        disk_comparison["period_label"].eq("2023 FY")
        & disk_comparison["series_label"].eq("FIDCs")
    ]

    assert len(disk_fidc_2023) == 2
    assert disk_fidc_2023["registered_volume_brl"].tolist() == pytest.approx(
        [ANBIMA_2023_FIDC_VOLUME_BRL] * 2,
        abs=0.01,
    )
    assert disk_fidc_2023["methodology"].astype(str).str.contains(
        "Correção 2023",
        regex=False,
    ).all()

    monkeypatch.setattr(tab_industry_study, "_DATA_DIR", DATA_DIR)
    loaded = tab_industry_study._load_industry_revision_payload.__wrapped__(
        PUBLISHED_PAYLOAD_SHA256
    )

    assert set(loaded) == set(disk_payload)
    for key in disk_payload.keys() - {"fixed_income_offer_comparison"}:
        assert loaded[key] == disk_payload[key], key
    pd.testing.assert_frame_equal(
        pd.DataFrame(loaded["fixed_income_offer_comparison"]),
        disk_comparison,
        check_dtype=False,
        check_like=True,
    )

    comparison = pd.DataFrame(loaded["fixed_income_offer_comparison"])
    rendered_fidc = comparison[
        comparison["view"].eq("FIDCs vs demais elegíveis")
        & comparison["series_label"].eq("FIDCs")
    ].set_index("period_label")
    expected_levels = {
        "2023 FY": ANBIMA_2023_FIDC_VOLUME_BRL,
        "2024 FY": 95_416_726_133.75,
        "2025 FY": 116_921_319_054.77,
        "2026 jan-jun": 65_488_118_983.56,
    }
    for period, expected in expected_levels.items():
        assert rendered_fidc.at[period, "registered_volume_brl"] == pytest.approx(
            expected,
            abs=0.01,
        )
    assert rendered_fidc.at["2024 FY", "yoy_growth"] == pytest.approx(
        95_416_726_133.75 / ANBIMA_2023_FIDC_VOLUME_BRL - 1
    )

    rendered_fidc_2023 = comparison[
        comparison["period_label"].eq("2023 FY")
        & comparison["series_label"].eq("FIDCs")
    ]
    assert rendered_fidc_2023["source_archive_sha256"].eq(
        ANBIMA_SOURCE_WORKBOOK_SHA256
    ).all()
    assert rendered_fidc_2023["source_dataset"].str.contains(
        "ANBIMA",
        regex=False,
    ).all()
    assert rendered_fidc_2023["methodology"].str.contains(
        "Correção 2023",
        regex=False,
    ).all()

    rendered_2023 = comparison[comparison["period_label"].eq("2023 FY")]
    assert rendered_2023.groupby("view")[
        "share_of_period_view_volume"
    ].sum().tolist() == pytest.approx([1.0, 1.0])
    expected_2023_universe = comparison[
        comparison["period_label"].eq("2023 FY")
        & comparison["view"].eq("FIDCs vs demais elegíveis")
    ]["registered_volume_brl"].sum()
    assert rendered_2023["universe_registered_volume_brl"].tolist() == (
        pytest.approx(
            [expected_2023_universe] * len(rendered_2023),
            abs=0.01,
        )
    )
    assert tab_industry_study._issuance_correction_applied(comparison)
    assert not tab_industry_study._bundle_predates_issuance_correction()

    # Loading is idempotent once the corrected series is present on disk.
    assert PAYLOAD_PATH.read_bytes() == payload_bytes
    assert hashlib.sha256(PAYLOAD_PATH.read_bytes()).hexdigest() == (
        PUBLISHED_PAYLOAD_SHA256
    )


def test_pl_history_closes_to_cent_for_every_published_competence() -> None:
    _, payload = _published_payload()
    history = payload["pl_history"]

    assert isinstance(history, list)
    assert len(history) == 12
    for row in history:
        assert isinstance(row, dict)
        total = Decimal(str(row["pl_total"]))
        direct = Decimal(str(row["pl_ex_fic"]))
        fic = Decimal(str(row["pl_fic_componente"]))
        assert abs(total - direct - fic) <= Decimal("0.01"), row["competencia"]


def test_published_ex_fic_products_exclude_all_current_fics() -> None:
    audit = pd.read_csv(
        DATA_DIR / "industry_fic_detection_audit.csv",
        dtype=str,
        low_memory=False,
    )
    audit = audit[audit["competencia"].eq("2026-06")]
    current_fics = {
        _cnpj_digits(value)
        for value in audit.loc[_truthy(audit["is_fic"]), "cnpj_fundo"]
    }
    assert len(current_fics) == 773

    fund_base = pd.read_csv(
        REVISION_DIR / "base_fundo_cnpj.csv.gz",
        usecols=["competencia", "cnpj_fundo", "pl", "is_fic", "is_fic_fidc"],
        dtype={"competencia": str, "cnpj_fundo": str},
        low_memory=False,
    )
    current_base = fund_base[fund_base["competencia"].eq("2026-06")].copy()
    base_fics = {
        _cnpj_digits(value)
        for value in current_base.loc[_truthy(current_base["is_fic"]), "cnpj_fundo"]
    }
    assert base_fics == current_fics

    for filename in (
        "top20_fidcs.csv",
        "top20_outros.csv",
        "monoestrutura_por_fundo.csv",
    ):
        frame = pd.read_csv(
            REVISION_DIR / filename,
            dtype=str,
            low_memory=False,
        )
        observed = {_cnpj_digits(value) for value in frame["cnpj_fundo"]}
        assert observed.isdisjoint(current_fics), filename
        for flag in ("is_fic", "is_fic_fidc"):
            assert not _truthy(frame[flag]).any(), f"{filename}:{flag}"

    expected_ex_fic_pl = pd.to_numeric(
        current_base.loc[~_truthy(current_base["is_fic"]), "pl"],
        errors="raise",
    ).sum()
    acquiring = pd.read_csv(
        REVISION_DIR / "adquirencia_mix_reclassificado.csv",
        low_memory=False,
    )
    acquiring = acquiring[acquiring["competencia"].eq("2026-06")]
    assert acquiring["denominador_pl_brl"].tolist() == pytest.approx(
        [expected_ex_fic_pl] * len(acquiring),
        abs=0.01,
    )
    moved_cnpjs: set[str] = set()
    for column in (
        "cnpjs_movidos_da_categoria",
        "cnpjs_movidos_para_adquirencia",
    ):
        for value in acquiring[column].dropna().astype(str):
            moved_cnpjs.update(
                _cnpj_digits(item)
                for item in value.split(";")
                if _cnpj_digits(item)
            )
    assert moved_cnpjs.isdisjoint(current_fics)


def test_user_facing_snapshot_sheets_are_preserved_by_the_input_pipeline() -> None:
    from services.industry_ppt_export import _build_legacy_industry_xlsx_bytes

    input_workbook_bytes = _build_legacy_industry_xlsx_bytes(DATA_DIR)
    workbook = load_workbook(
        BytesIO(input_workbook_bytes),
        read_only=True,
        data_only=True,
    )
    try:
        assert set(USER_FACING_SNAPSHOT_SHEETS).issubset(workbook.sheetnames)
        assert set(INHERITED_WORKBOOK_SHEETS_TO_REMOVE).issubset(
            workbook.sheetnames
        )
    finally:
        workbook.close()


def test_workbook_removal_allowlist_preserves_protected_sheets() -> None:
    from services.industry_revision_export import REQUIRED_WORKBOOK_SHEETS

    renderer_source = RENDERER_PATH.read_text(encoding="utf-8")
    build_workbook_source = renderer_source.split(
        "async function buildWorkbook",
        maxsplit=1,
    )[1].split("async function exportPresentation", maxsplit=1)[0]
    assert re.search(r"\bconst\s+INPUT_WORKBOOK\s*=", renderer_source)
    assert (
        "SpreadsheetFile.importXlsx(await FileBlob.load(INPUT_WORKBOOK))"
        in build_workbook_source
    )

    sheets_to_remove = _js_string_array_constant(
        renderer_source,
        "WORKBOOK_SHEETS_TO_REMOVE",
    )
    assert sheets_to_remove == EXPECTED_WORKBOOK_SHEETS_TO_REMOVE
    assert set(sheets_to_remove).isdisjoint(USER_FACING_SNAPSHOT_SHEETS)
    assert "removeWorkbookSheets(workbook);" in build_workbook_source
    assert "getItemOrNullObject(sheetName)" in renderer_source
    assert "sheet.delete();" in renderer_source
    assert "sheet.dataValidations.clear(used.address);" in renderer_source
    assert 'resetSheet(workbook, "Top 20 Outros")' in renderer_source
    assert set(sheets_to_remove).isdisjoint(REQUIRED_WORKBOOK_SHEETS)
    assert 'sheetName: "Cross-check taxonomia"' not in renderer_source
    assert '["Cross-check taxonomia", "A1:' not in renderer_source
    assert 'sheetName: "Taxonomia por CNPJ"' not in renderer_source
    assert '["Taxonomia por CNPJ", "A1:' not in renderer_source
    assert "addAcquiringAnbimaReviewSheet" not in renderer_source
    assert 'resetSheet(workbook, "Reclass. adquirência")' not in renderer_source
    assert '["Reclass. adquirência", "A1:' not in renderer_source
    assert "addNumericLocaleAuditSheet" not in renderer_source
    assert 'resetSheet(workbook, "Auditoria numérica")' not in renderer_source
    assert '["Auditoria numérica", "A1:' not in renderer_source
    assert 'payloadKey: "anbima_outros_reclassification"' not in renderer_source
    assert 'sheetName: "Reclass. ANBIMA"' not in renderer_source
    assert '["Reclass. ANBIMA", "A1:' not in renderer_source
    assert "addReclassificationSheet" not in renderer_source
    assert 'payloadKey: "cvm_outros_reclassification"' not in renderer_source
    assert 'sheetName: "Reclass. CVM"' not in renderer_source
    assert '["Reclass. CVM", "A1:' not in renderer_source
    assert "addProviderFlowVisualSheet" not in renderer_source
    assert 'resetSheet(workbook, "Fluxos visuais")' not in renderer_source
    assert '["Fluxos visuais", "A1:' not in renderer_source
    assert "flowAssets" not in renderer_source
    assert "PngDataUrl" not in renderer_source
    assert "FLOW_ASSET_DIR" not in renderer_source
    assert "generateProviderFlowHtml" in renderer_source

    flow_builder_source = (
        ROOT / "scripts" / "build_provider_flow_explorer.mjs"
    ).read_text(encoding="utf-8")
    assert 'require("sharp")' not in flow_builder_source
    assert "staticSvg" not in flow_builder_source
    assert "provider_flow_admin.png" not in flow_builder_source
    assert '"output-dir"' not in flow_builder_source
