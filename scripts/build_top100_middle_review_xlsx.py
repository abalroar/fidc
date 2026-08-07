"""Top 100 FIDCs por PL, em planilha nativa para a revisão do universo Middle.

A aba ``Top 100`` é uma **tabela do Excel** de verdade (um ListObject), não um
intervalo formatado: o filtro e a ordenação vêm com ela, e a coluna ``MIDDLE``
tem lista suspensa com Sim e Não, de modo que a revisão não admite grafia
livre.

    python scripts/build_top100_middle_review_xlsx.py
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from services.top100_middle_deck import load_review  # noqa: E402

DEFAULT_DATA_DIR = Path("data/industry_study")
DEFAULT_OUTPUT = Path("outputs/top100_middle/Top100_FIDCs_Revisao_Middle.xlsx")

ORANGE = "E36C0A"
WHITE = "FFFFFF"
GRAY_100 = "F5F6F7"

#: Ordem das colunas na planilha: identificação, materialidade, e o que
#: sustenta a decisão de Middle.  A coluna MIDDLE fecha, à direita, porque é
#: onde o revisor digita.
COLUMNS: tuple[tuple[str, str, float], ...] = (
    ("rank_pl", "#", 5),
    ("FIDC", "FIDC", 58),
    ("cnpj", "CNPJ do fundo", 18),
    ("pl_mm", "PL (R$ mm)", 13),
    ("competencia_pl", "Competência do PL", 16),
    ("Volume 2026 (R$ mi)", "Emissão 2026 (R$ mi)", 18),
    ("Ofertas", "Ofertas", 9),
    ("Coordenador líder", "Coordenador líder", 22),
    ("Tipo ANBIMA", "Tipo ANBIMA", 22),
    ("Foco ANBIMA", "Foco ANBIMA", 26),
    ("CNPJ cedente", "CNPJ cedente", 18),
    ("Razão social cedente", "Razão social cedente", 40),
    ("Natureza cedente", "Natureza cedente", 20),
    ("CNAE cedente", "CNAE cedente", 44),
    ("Seção CNAE", "Seção CNAE", 30),
    ("UF", "UF", 6),
    ("Porte", "Porte", 12),
    ("% do fundo", "% do fundo", 11),
    ("Nº cedentes", "Nº cedentes", 12),
    ("Setor cedente", "Setor cedente", 26),
    ("Segmento cedente", "Segmento cedente", 30),
    ("Setor sacado", "Setor sacado", 26),
    ("Segmento sacado", "Segmento sacado", 30),
    ("Sacado nomeado", "Sacado nomeado", 26),
    ("MIDDLE (preencher)", "MIDDLE", 12),
)
NUMERIC = {"pl_mm", "Volume 2026 (R$ mi)", "Ofertas", "% do fundo", "Nº cedentes"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def _cnpj(value: object) -> str:
    digits = "" if value is None or pd.isna(value) else str(value).strip().zfill(14)
    if len(digits) != 14:
        return ""
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"


def write_workbook(frame: pd.DataFrame, output: Path) -> Path:
    from openpyxl import Workbook

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Top 100"

    header_font = Font(name="Arial", size=10, bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=ORANGE)
    body_font = Font(name="Arial", size=10)

    for index, (_key, titulo, largura) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = largura
    sheet.row_dimensions[1].height = 28

    for linha, registro in enumerate(frame.to_dict("records"), start=2):
        for index, (key, _titulo, _largura) in enumerate(COLUMNS, start=1):
            bruto = registro.get(key)
            if key in {"cnpj", "CNPJ cedente"}:
                valor = _cnpj(bruto)
            elif key in NUMERIC:
                valor = pd.to_numeric(bruto, errors="coerce")
                valor = None if pd.isna(valor) else float(valor)
            else:
                valor = "" if bruto is None or pd.isna(bruto) else str(bruto)
            cell = sheet.cell(row=linha, column=index, value=valor)
            cell.font = body_font
            if key in NUMERIC:
                cell.number_format = "#,##0.0" if key != "Ofertas" else "0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center")
        if linha % 2 == 0:
            for index in range(1, len(COLUMNS) + 1):
                sheet.cell(row=linha, column=index).fill = PatternFill(
                    "solid", fgColor=GRAY_100
                )

    ultima_coluna = get_column_letter(len(COLUMNS))
    ultima_linha = len(frame) + 1
    referencia = f"A1:{ultima_coluna}{ultima_linha}"
    # Uma tabela do Excel de verdade: o filtro e a ordenação vêm com o objeto,
    # e inserir linha propaga formato e fórmula sem retrabalho.
    tabela = Table(displayName="Top100Middle", ref=referencia)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(tabela)
    sheet.freeze_panes = "C2"

    coluna_middle = get_column_letter(len(COLUMNS))
    validacao = DataValidation(
        type="list",
        formula1='"Sim,Não"',
        allow_blank=True,
        showDropDown=False,
        promptTitle="Universo Middle",
        prompt="Marque Sim quando o veículo pertencer ao universo Middle.",
    )
    sheet.add_data_validation(validacao)
    validacao.add(f"{coluna_middle}2:{coluna_middle}{ultima_linha}")

    notas = workbook.create_sheet("Fontes")
    for linha, texto in enumerate(
        (
            "Top 100 FIDCs para revisão do universo Middle",
            "",
            "Ordenação: patrimônio líquido, do maior para o menor.",
            "PL: CVM, Informe Mensal FIDC — competência mais recente em que cada "
            "fundo reportou patrimônio (coluna Competência do PL).",
            "Emissão 2026: CVM, Sistema de Registro de Ofertas — volume de cotas "
            "colocadas no ano, todos os ritos.",
            "Cedente, CNAE, UF e porte: conforme declarado no Informe Mensal; "
            "quando o fundo não declara, a célula fica vazia em vez de estimada.",
            "MIDDLE: campo de revisão. Lista suspensa com Sim e Não.",
            "",
            "Oito dos cem fundos não reportaram patrimônio em competência alguma "
            "e fecham a lista, sem PL atribuído.",
        ),
        start=1,
    ):
        cell = notas.cell(row=linha, column=1, value=texto)
        cell.font = Font(name="Arial", size=10, bold=linha == 1)
    notas.column_dimensions["A"].width = 110

    workbook.save(output)
    return output


def main() -> None:
    args = parse_args()
    frame = load_review(args.data_dir)
    caminho = write_workbook(frame, args.output)
    print(f"{caminho}: {len(frame)} fundos, {len(COLUMNS)} colunas")


if __name__ == "__main__":
    main()
