#!/usr/bin/env python3
"""Add native, cell-linked structural-risk scatter charts to the portfolio XLSX."""

from __future__ import annotations

import math
from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.properties import CalcProperties


CASES_SHEET = "Casos 99"
NAMES_SHEET = "Nomes editáveis"
HEADER_ROW = 4
DATA_ROW = 5
CHART_PREFIX = "Gráfico "

COLORS = {
    "abaixo": "7A1F3D",
    "estreita": "EC7000",
    "acima": "002B5C",
    "outro": "7A7F84",
}


def _truthy(value: object) -> bool:
    return str(value or "").strip().casefold() in {"true", "1", "sim", "s"}


def _numeric(value: object) -> float | None:
    if value in (None, "", "N/D"):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _cnpj(value: object) -> str:
    digits = "".join(character for character in str(value or "") if character.isdigit())
    return digits.zfill(14) if digits else ""


def _series_color(status: object) -> str:
    value = str(status or "").casefold()
    if "abaixo" in value:
        return COLORS["abaixo"]
    if "estreita" in value or "2 p.p." in value:
        return COLORS["estreita"]
    if "acima" in value:
        return COLORS["acima"]
    return COLORS["outro"]


def _headers(sheet) -> dict[str, int]:
    return {
        str(cell.value or "").strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value not in (None, "")
    }


def _required(headers: dict[str, int], *names: str) -> None:
    missing = [name for name in names if name not in headers]
    if missing:
        raise ValueError("Casos 99 sem colunas para os gráficos: " + ", ".join(missing))


def _alias_row_by_cnpj(names_sheet) -> dict[str, int]:
    headers = _headers(names_sheet)
    _required(headers, "CNPJ", "Nome editável para gráfico")
    mapping: dict[str, int] = {}
    for row in range(DATA_ROW, names_sheet.max_row + 1):
        cnpj = _cnpj(names_sheet.cell(row, headers["CNPJ"]).value)
        if cnpj:
            mapping[cnpj] = row
    return mapping


def _eligible_points(cases_sheet, headers: dict[str, int]) -> list[dict[str, object]]:
    _required(
        headers,
        "CNPJ",
        "Sub / PL atual",
        "Índice estrutural usado",
        "Comparável?",
        "Subtipo de risco diagnosticado",
        "Situação regulatória",
        "Categoria de risco proposta",
    )
    points: list[dict[str, object]] = []
    for row in range(DATA_ROW, cases_sheet.max_row + 1):
        x = _numeric(cases_sheet.cell(row, headers["Índice estrutural usado"]).value)
        y = _numeric(cases_sheet.cell(row, headers["Sub / PL atual"]).value)
        cnpj = _cnpj(cases_sheet.cell(row, headers["CNPJ"]).value)
        if not cnpj or x is None or y is None:
            continue
        if not _truthy(cases_sheet.cell(row, headers["Comparável?"]).value):
            continue
        subtype = str(
            cases_sheet.cell(row, headers["Subtipo de risco diagnosticado"]).value
            or "N/D"
        ).strip()
        category = str(
            cases_sheet.cell(row, headers["Categoria de risco proposta"]).value
            or "N/D"
        ).strip()
        if subtype.upper().startswith("N/D"):
            subtype = category
        points.append(
            {
                "row": row,
                "cnpj": cnpj,
                "x": x,
                "y": y,
                "subtype": subtype,
                "category": category,
                "status": cases_sheet.cell(row, headers["Situação regulatória"]).value,
            }
        )
    return points


def _add_chart_sheet(
    workbook,
    *,
    index: int,
    subtype: str,
    points: list[dict[str, object]],
    cases_sheet,
    names_sheet,
    headers: dict[str, int],
    alias_rows: dict[str, int],
) -> None:
    sheet = workbook.create_sheet(f"{CHART_PREFIX}{index:02d}")
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = f"Subordinação mínima vs. Subordinação atual · {subtype}"
    sheet["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="000000")
    sheet.merge_cells("A1:H1")
    sheet["A2"] = (
        "Cada ponto é um CNPJ com estruturas comparáveis. O rótulo e a legenda "
        "referenciam a aba Nomes editáveis; ausências e estruturas incomparáveis "
        "permanecem na aba Casos 99 e não entram como zero."
    )
    sheet["A2"].font = Font(name="Arial", size=10, color="666666")
    sheet.merge_cells("A2:H2")
    sheet["A3"] = "Cores"
    sheet["B3"] = "Vinho: abaixo do mínimo"
    sheet["C3"] = "Âmbar: folga estreita"
    sheet["D3"] = "Azul: acima do mínimo"
    sheet["E3"] = f"Pontos: {len(points)}"
    for cell, color in (("B3", COLORS["abaixo"]), ("C3", COLORS["estreita"]), ("D3", COLORS["acima"])):
        sheet[cell].fill = PatternFill("solid", fgColor=color)
        sheet[cell].font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    chart = ScatterChart()
    chart.title = f"{subtype} · mínimo estrutural x Sub/PL atual"
    chart.style = 13
    chart.height = 14.5
    chart.width = 25.0
    chart.x_axis.title = "Subordinação mínima / suporte estrutural"
    chart.y_axis.title = "Subordinação atual / PL"
    chart.x_axis.numFmt = "0%"
    chart.y_axis.numFmt = "0%"
    chart.x_axis.scaling.min = 0
    chart.y_axis.scaling.min = 0
    chart.legend.position = "r"
    chart.dLbls = DataLabelList()
    chart.dLbls.showSerName = True
    chart.dLbls.showVal = False
    chart.dLbls.showLegendKey = False

    max_value = max(max(float(point["x"]), float(point["y"])) for point in points)
    upper = min(1.1, max(0.10, math.ceil(max_value * 20 + 1) / 20))
    chart.x_axis.scaling.max = upper
    chart.y_axis.scaling.max = upper

    cases_title = cases_sheet.title.replace("'", "''")
    names_title = names_sheet.title.replace("'", "''")
    x_col = headers["Índice estrutural usado"]
    y_col = headers["Sub / PL atual"]
    for point in points:
        data_row = int(point["row"])
        alias_row = alias_rows.get(str(point["cnpj"]))
        if alias_row is None:
            raise ValueError(f"CNPJ sem nome editável: {point['cnpj']}")
        x_values = Reference(cases_sheet, min_col=x_col, min_row=data_row, max_row=data_row)
        y_values = Reference(cases_sheet, min_col=y_col, min_row=data_row, max_row=data_row)
        series = Series(y_values, x_values)
        series.tx = SeriesLabel(
            strRef=StrRef(f=f"'{names_title}'!$D${alias_row}")
        )
        series.marker.symbol = "circle"
        series.marker.size = 8
        color = _series_color(point["status"])
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)

    sheet.add_chart(chart, "A5")
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.column_dimensions["A"].width = 19
    for column in "BCDEFGH":
        sheet.column_dimensions[column].width = 23


def patch(path: Path) -> None:
    workbook = load_workbook(path)
    if CASES_SHEET not in workbook.sheetnames or NAMES_SHEET not in workbook.sheetnames:
        raise ValueError("workbook sem Casos 99 ou Nomes editáveis")
    for name in list(workbook.sheetnames):
        if name.startswith(CHART_PREFIX):
            del workbook[name]
    cases_sheet = workbook[CASES_SHEET]
    names_sheet = workbook[NAMES_SHEET]
    headers = _headers(cases_sheet)
    alias_rows = _alias_row_by_cnpj(names_sheet)
    points = _eligible_points(cases_sheet, headers)
    grouped: dict[str, list[dict[str, object]]] = {}
    for point in points:
        grouped.setdefault(str(point["subtype"]), []).append(point)
    if not grouped:
        raise ValueError("nenhum ponto comparável para os gráficos estruturais")
    for index, subtype in enumerate(sorted(grouped), start=1):
        _add_chart_sheet(
            workbook,
            index=index,
            subtype=subtype,
            points=sorted(grouped[subtype], key=lambda item: (float(item["x"]), str(item["cnpj"]))),
            cases_sheet=cases_sheet,
            names_sheet=names_sheet,
            headers=headers,
            alias_rows=alias_rows,
        )
    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(path)


def main(argv: list[str] | None = None) -> None:
    args = list(sys.argv[1:] if argv is None else argv)
    if len(args) != 1:
        raise SystemExit("uso: patch_portfolio_workbook_charts.py /caminho/arquivo.xlsx")
    path = Path(args[0]).expanduser().resolve()
    patch(path)
    print(f"[ok] gráficos nativos adicionados: {path}")


if __name__ == "__main__":
    main()
