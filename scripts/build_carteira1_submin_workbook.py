"""Consolida os 101 fundos da Carteira 1 num workbook de preenchimento da sub mínima.

Junta tudo o que já existe no repositório para cada CNPJ da carteira:

* identidade e rastro da transcrição das fotos (`industry_carteira_1_scope.csv`);
* curadoria documental da subordinação mínima júnior (`industry_carteira_1_document_curation.csv`);
* snapshot do estudo de indústria (`industry_fund_snapshot.csv.gz`);
* Informe Mensal Estruturado — o "Excel" da CVM (`vehicle_monthly.csv.gz`);
* taxonomias curadas e catálogo de dimensões;
* critérios extraídos de regulamento (`criteria_structured.csv.gz` e `data/regulatory_knowledge/`);
* inventário documental com links do Fundos.NET.

A aba mestre deixa em branco o bloco `USR_*`, que é onde o analista escreve
o índice de subordinação mínima lido no regulamento. As colunas de folga são
fórmulas do Excel: assim que `USR_sub_min_pct` é preenchido, a folga contra a
subordinação realizada do IME aparece sozinha.

Uso:
    python3 scripts/build_carteira1_submin_workbook.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
STUDY = ROOT / "data" / "industry_study"
KNOWLEDGE = ROOT / "data" / "regulatory_knowledge"
OUT_DIR = ROOT / "reports" / "carteira1_sub_minima"
PORTFOLIO_ID = "carteira1_fidc_industry_20260731"

FNET_DOC = "https://fnet.bmfbovespa.com.br/fnet/publico/exibirDocumento?id={doc}&cvm=true"
FNET_FUND = (
    "https://fnet.bmfbovespa.com.br/fnet/publico/"
    "abrirGerenciadorDocumentosCVM?cnpjFundo={cnpj}"
)

# Colunas que o analista preenche à mão. Ficam vazias de propósito.
USER_COLUMNS = [
    "USR_sub_min_pct",
    "USR_tipo_indice",
    "USR_numerador",
    "USR_denominador",
    "USR_documento_id",
    "USR_pagina",
    "USR_citacao_regulamento",
    "USR_status",
    "USR_observacao",
]


def digits(value: object) -> str:
    return re.sub(r"\D", "", str(value or "")).zfill(14)


def format_cnpj(value: str) -> str:
    raw = digits(value)
    return f"{raw[:2]}.{raw[2:5]}.{raw[5:8]}/{raw[8:12]}-{raw[12:]}"


def read_csv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, low_memory=False, **kwargs)


def key_column(frame: pd.DataFrame, column: str) -> pd.DataFrame:
    frame = frame.copy()
    frame["cnpj_key"] = frame[column].map(digits)
    return frame


def load_portfolio() -> pd.DataFrame:
    payload = json.loads((ROOT / "portfolios.json").read_text(encoding="utf-8"))
    portfolio = next(p for p in payload["portfolios"] if p["id"] == PORTFOLIO_ID)
    rows = [
        {
            "cnpj_key": digits(fund["cnpj"]),
            "cnpj": format_cnpj(fund["cnpj"]),
            "nome_carteira": fund["display_name"],
        }
        for fund in portfolio["funds"]
    ]
    return pd.DataFrame(rows)


def load_scope() -> pd.DataFrame:
    scope = key_column(read_csv(STUDY / "industry_carteira_1_scope.csv"), "cnpj_fundo")
    return scope[
        [
            "cnpj_key",
            "ordem",
            "imagem",
            "nome_foto",
            "status_identidade",
            "regra_identidade",
            "observacao_identidade",
        ]
    ]


def load_curation() -> pd.DataFrame:
    curation = key_column(
        read_csv(STUDY / "industry_carteira_1_document_curation.csv"), "cnpj_fundo"
    )
    curation = curation.rename(
        columns={
            "subordinacao_minima_junior_pct": "cur_sub_min_pct",
            "subordinacao_minima_junior_display": "cur_sub_min_display",
            "subordinacao_minima_texto": "cur_sub_min_texto",
            "subordinacao_minima_fonte": "cur_sub_min_fonte",
            "documento_id_regulamento": "cur_documento_id",
            "documento_data_regulamento": "cur_documento_data",
            "pagina_clausula": "cur_pagina_clausula",
            "paginas_lidas": "cur_paginas_lidas",
            "status_curadoria_documental": "cur_status",
            "observacao_documental": "cur_observacao",
            "emissao_data": "cur_emissao_data",
            "emissao_data_display": "cur_emissao_display",
            "emissao_fonte": "cur_emissao_fonte",
        }
    )
    curation["cur_url_regulamento"] = curation["cur_documento_id"].map(
        lambda doc: FNET_DOC.format(doc=doc) if pd.notna(doc) and str(doc).strip() else ""
    )
    return curation.drop(columns=["cnpj_fundo"])


def load_ime() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Retorna (última competência por fundo, série dos últimos 12 meses)."""
    ime = key_column(read_csv(STUDY / "vehicle_monthly.csv.gz"), "cnpj")
    ime = ime.sort_values(["cnpj_key", "competencia"])
    latest = ime.groupby("cnpj_key", as_index=False).tail(1)
    series = ime.groupby("cnpj_key", as_index=False).tail(12)
    drop = ["cnpj", "cnpj_fundo"]
    latest = latest.drop(columns=[c for c in drop if c in latest.columns])
    latest = latest.rename(columns={c: f"ime_{c}" for c in latest.columns if c != "cnpj_key"})
    return latest, series


def load_snapshot() -> pd.DataFrame:
    snapshot = key_column(read_csv(STUDY / "industry_fund_snapshot.csv.gz"), "cnpj_fundo")
    snapshot = snapshot.sort_values(["cnpj_key", "competencia"]).groupby(
        "cnpj_key", as_index=False
    ).tail(1)
    keep = [
        "cnpj_key",
        "nome_exibicao",
        "competencia",
        "segmento_principal",
        "segmento_estrategia",
        "subsegmento_estrategia",
        "fundo_estrategia",
        "classificacao_anbima",
        "admin_nome",
        "gestor_nome",
        "custodiante_nome",
        "condominio",
        "publico_alvo",
        "exclusivo",
        "is_fic_fidc",
        "indexadores",
        "tipo_cotas",
        "emission_cohort",
        "first_offer_year",
        "latest_regulamento_date",
        "has_regulatory_matrix",
        "criteria_rows",
        "criteria_keys",
        "criteria_subordination_rows",
        "criteria_monitorable_rows",
        "criteria_partial_rows",
        "criteria_not_monitorable_rows",
        "sub_min_pct_median",
        "sub_min_pct_min",
        "sub_min_pct_max",
        "tem_sub_minima",
        "camadas_com_evidencia",
        "snapshot_status",
        "document_rows",
        "document_local_ready",
        "document_missing_local",
        "document_classes",
        "document_latest_date",
        "cedentes_top",
        "grupos_economicos",
        "tranche_volume_brl",
        "valid_volume_2024_2026_brl",
    ]
    keep = [c for c in keep if c in snapshot.columns]
    snapshot = snapshot[keep]
    return snapshot.rename(
        columns={c: f"snap_{c}" for c in snapshot.columns if c != "cnpj_key"}
    )


def load_taxonomy_actions() -> pd.DataFrame:
    actions = key_column(read_csv(STUDY / "taxonomy_review_actions.csv"), "cnpj_fundo")
    actions = actions.sort_values(["cnpj_key", "updated_at_utc"]).groupby(
        "cnpj_key", as_index=False
    ).tail(1)
    keep = [
        "cnpj_key",
        "tipo_analitico",
        "foco_analitico",
        "tabela_ii_analitica",
        "taxonomia_funcional_n1",
        "taxonomia_funcional_n2",
        "confianca",
        "cedente_originador_expresso",
        "documento_id",
        "documento_data",
        "pagina_clausula",
        "status",
    ]
    actions = actions[[c for c in keep if c in actions.columns]]
    return actions.rename(
        columns={c: f"tax_{c}" for c in actions.columns if c != "cnpj_key"}
    )


def load_dimension_catalog() -> tuple[pd.DataFrame, pd.DataFrame]:
    catalog = key_column(read_csv(STUDY / "industry_dimension_catalog.csv.gz"), "cnpj_fundo")
    wide = (
        catalog.groupby(["cnpj_key", "dimension_id"])["dimension_value"]
        .apply(lambda values: " | ".join(sorted({v for v in values if str(v).strip()})))
        .unstack("dimension_id")
    )
    wide.columns = [f"dim_{column}" for column in wide.columns]
    return wide.reset_index(), catalog


def load_criteria() -> tuple[pd.DataFrame, pd.DataFrame]:
    criteria = key_column(read_csv(STUDY / "criteria_structured.csv.gz"), "cnpj_fundo")
    sub = criteria[
        criteria["chave"].isin(["subordination_ratio_min", "feature_subordination_minimum"])
    ]
    summary = (
        sub.groupby("cnpj_key")
        .agg(
            crit_sub_linhas=("chave", "size"),
            crit_sub_limites=(
                "limite_regra",
                lambda values: " | ".join(
                    dict.fromkeys(str(v).strip() for v in values if str(v).strip() and str(v) != "nan")
                )[:900],
            ),
            crit_sub_pct_min=("pct_min", "min"),
            crit_sub_pct_max=("pct_max", "max"),
            crit_sub_monitorabilidade=(
                "monitorabilidade_ime",
                lambda values: " | ".join(dict.fromkeys(str(v) for v in values if str(v) != "nan")),
            ),
            crit_sub_documentos=(
                "documento_id",
                lambda values: " | ".join(
                    dict.fromkeys(str(v) for v in values if str(v) not in {"nan", ""})
                ),
            ),
        )
        .reset_index()
    )
    return summary, criteria


def load_knowledge_criteria() -> pd.DataFrame:
    """Critérios de regulamento extraídos por fundo em data/regulatory_knowledge."""
    rows: list[dict[str, object]] = []
    for path in sorted(KNOWLEDGE.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for criterion in payload.get("criteria") or []:
            mapping = criterion.get("monitoring_mapping") or {}
            rows.append(
                {
                    "cnpj_key": digits(path.stem),
                    "fundo": payload.get("fund_name"),
                    "criterio": criterion.get("name"),
                    "chave": criterion.get("canonical_key"),
                    "tipo_evento": criterion.get("event_type"),
                    "limite_valor": criterion.get("threshold_value"),
                    "limite_unidade": criterion.get("threshold_unit"),
                    "limite_display": criterion.get("threshold_display"),
                    "comparacao": criterion.get("comparison"),
                    "confianca": criterion.get("confidence"),
                    "formula_texto": criterion.get("formula_text"),
                    "trecho_fonte": criterion.get("source_excerpt"),
                    "documento_origem": criterion.get("source_document"),
                    "documento_id": criterion.get("source_document_id"),
                    "monitorabilidade": mapping.get("status"),
                    "metrica_ime": mapping.get("ime_metric"),
                    "racional_monitoramento": mapping.get("rationale"),
                    "notas": criterion.get("notes"),
                }
            )
    return pd.DataFrame(rows)


def load_documents() -> pd.DataFrame:
    inventory = key_column(read_csv(STUDY / "document_inventory.csv.gz"), "cnpj_fundo")
    inventory["url_fundosnet"] = inventory["documento_id"].map(
        lambda doc: FNET_DOC.format(doc=doc) if str(doc).strip() not in {"", "nan"} else ""
    )
    return inventory


def build_master() -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    master = load_portfolio()
    ime_latest, ime_series = load_ime()
    dim_wide, dim_long = load_dimension_catalog()
    crit_summary, criteria_all = load_criteria()

    for frame in (
        load_scope(),
        load_curation(),
        load_snapshot(),
        ime_latest,
        load_taxonomy_actions(),
        dim_wide,
        crit_summary,
    ):
        master = master.merge(frame, on="cnpj_key", how="left")

    knowledge = load_knowledge_criteria()
    knowledge_sub = knowledge[knowledge["chave"] == "subordination_ratio_min"]
    if not knowledge_sub.empty:
        agg = (
            knowledge_sub.groupby("cnpj_key")["limite_display"]
            .apply(lambda values: " | ".join(dict.fromkeys(str(v) for v in values if str(v).strip())))
            .rename("rk_sub_min_limites")
            .reset_index()
        )
        master = master.merge(agg, on="cnpj_key", how="left")
    else:
        master["rk_sub_min_limites"] = ""

    master["url_fundosnet"] = master["cnpj_key"].map(lambda c: FNET_FUND.format(cnpj=c))
    for column in USER_COLUMNS + ["USR_folga_pp", "USR_divergencia_vs_curadoria_pp"]:
        master[column] = ""

    master["ordem"] = pd.to_numeric(master["ordem"], errors="coerce")
    master = master.sort_values("ordem").reset_index(drop=True)

    ordered = (
        ["ordem", "cnpj", "cnpj_key", "nome_carteira", "snap_nome_exibicao", "ime_denominacao"]
        + USER_COLUMNS
        + ["USR_folga_pp", "USR_divergencia_vs_curadoria_pp"]
        + [
            "cur_sub_min_pct",
            "cur_sub_min_display",
            "cur_status",
            "cur_sub_min_texto",
            "cur_sub_min_fonte",
            "cur_documento_id",
            "cur_documento_data",
            "cur_pagina_clausula",
            "cur_paginas_lidas",
            "cur_url_regulamento",
            "cur_observacao",
            "cur_emissao_data",
            "cur_emissao_display",
            "cur_emissao_fonte",
            "crit_sub_linhas",
            "crit_sub_limites",
            "crit_sub_pct_min",
            "crit_sub_pct_max",
            "crit_sub_monitorabilidade",
            "crit_sub_documentos",
            "rk_sub_min_limites",
            "snap_sub_min_pct_median",
            "snap_sub_min_pct_min",
            "snap_sub_min_pct_max",
            "snap_tem_sub_minima",
            "snap_criteria_subordination_rows",
            "url_fundosnet",
        ]
    )
    remaining = [c for c in master.columns if c not in ordered]
    master = master[[c for c in ordered if c in master.columns] + remaining]

    # Subordinação realizada do IME vem como razão; a curadoria vem em p.p.
    master.insert(
        master.columns.get_loc("ime_subordinacao_pct") + 1,
        "ime_subordinacao_pp",
        pd.to_numeric(master["ime_subordinacao_pct"], errors="coerce") * 100,
    )

    empty = [
        column
        for column in master.columns
        if not column.startswith("USR_")
        and not master[column].astype(str).str.strip().replace({"nan": ""}).any()
    ]
    master = master.drop(columns=empty)

    sheets = {
        "02_curadoria_sub_minima": load_curation().merge(
            master[["cnpj_key", "ordem", "cnpj", "nome_carteira"]], on="cnpj_key", how="right"
        ),
        "03_criterios_regulamento": criteria_all[
            criteria_all["cnpj_key"].isin(set(master["cnpj_key"]))
        ],
        "04_criterios_extraidos_json": knowledge[
            knowledge["cnpj_key"].isin(set(master["cnpj_key"]))
        ],
        "05_ime_ultimos_12m": ime_series[ime_series["cnpj_key"].isin(set(master["cnpj_key"]))],
        "06_taxonomias_long": dim_long[dim_long["cnpj_key"].isin(set(master["cnpj_key"]))],
        "07_documentos": load_documents().pipe(
            lambda frame: frame[frame["cnpj_key"].isin(set(master["cnpj_key"]))]
        ),
    }
    return master, sheets


def dictionary_frame(master: pd.DataFrame) -> pd.DataFrame:
    origem = {
        "ordem": "industry_carteira_1_scope.csv — ordem da transcrição das fotos",
        "cnpj": "portfolios.json — Carteira 1",
        "nome_carteira": "portfolios.json — display_name",
        "USR_": "EM BRANCO — preenchimento manual do analista",
        "cur_": "industry_carteira_1_document_curation.csv — curadoria documental da sub mínima júnior",
        "crit_": "criteria_structured.csv.gz — critérios estruturados de regulamento",
        "rk_": "data/regulatory_knowledge/<cnpj>.json — extração heurística por regulamento",
        "snap_": "industry_fund_snapshot.csv.gz — snapshot do estudo de indústria",
        "ime_": "vehicle_monthly.csv.gz — Informe Mensal Estruturado (o 'Excel' da CVM)",
        "tax_": "taxonomy_review_actions.csv — decisão de taxonomia curada",
        "dim_": "industry_dimension_catalog.csv.gz — catálogo de dimensões/taxonomias",
    }
    rows = []
    for column in master.columns:
        prefix = next((p for p in origem if p.endswith("_") and column.startswith(p)), None)
        rows.append(
            {
                "coluna": column,
                "origem": origem.get(column) or origem.get(prefix, "derivada neste script"),
                "preenchimento": "manual" if column.startswith("USR_") else "automático",
            }
        )
    return pd.DataFrame(rows)


def write_workbook(master: pd.DataFrame, sheets: dict[str, pd.DataFrame], path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="01_mestre", index=False)
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
        dictionary_frame(master).to_excel(writer, sheet_name="08_dicionario", index=False)

        book = writer.book
        sheet = book["01_mestre"]
        header_fill = PatternFill("solid", fgColor="1F3864")
        user_fill = PatternFill("solid", fgColor="FFF2CC")
        user_header_fill = PatternFill("solid", fgColor="BF8F00")
        headers = [cell.value for cell in sheet[1]]

        for index, cell in enumerate(sheet[1], start=1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = user_header_fill if str(cell.value).startswith("USR_") else header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            sheet.column_dimensions[get_column_letter(index)].width = (
                18 if str(cell.value).startswith("USR_") else 24
            )
        sheet.freeze_panes = "G2"
        sheet.auto_filter.ref = sheet.dimensions

        # Fórmulas de folga: dependem só do que o analista digitar.
        col_user_pct = headers.index("USR_sub_min_pct") + 1
        col_folga = headers.index("USR_folga_pp") + 1
        col_diverg = headers.index("USR_divergencia_vs_curadoria_pp") + 1
        col_ime_sub = headers.index("ime_subordinacao_pct") + 1
        col_cur_pct = headers.index("cur_sub_min_pct") + 1
        letters = {
            "user": get_column_letter(col_user_pct),
            "ime": get_column_letter(col_ime_sub),
            "cur": get_column_letter(col_cur_pct),
        }
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=col_folga).value = (
                f'=IF(OR({letters["user"]}{row}="",{letters["ime"]}{row}=""),"",'
                f'{letters["ime"]}{row}*100-{letters["user"]}{row})'
            )
            sheet.cell(row=row, column=col_diverg).value = (
                f'=IF(OR({letters["user"]}{row}="",{letters["cur"]}{row}=""),"",'
                f'{letters["user"]}{row}-{letters["cur"]}{row})'
            )
            for column in range(col_user_pct, col_user_pct + len(USER_COLUMNS)):
                sheet.cell(row=row, column=column).fill = user_fill

        for name in sheets:
            tab = book[name[:31]]
            for cell in tab[1]:
                cell.font = Font(bold=True)
            tab.freeze_panes = "A2"


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    master, sheets = build_master()
    workbook = OUT_DIR / "carteira1_101_fundos_sub_minima.xlsx"
    write_workbook(master, sheets, workbook)
    master.to_csv(OUT_DIR / "carteira1_mestre.csv", index=False)

    print(f"fundos: {len(master)}")
    print(f"workbook: {workbook.relative_to(ROOT)}")
    for name, frame in sheets.items():
        print(f"  {name}: {len(frame)} linhas")


if __name__ == "__main__":
    main()
