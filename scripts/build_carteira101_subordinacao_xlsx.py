"""Carteira 101 — subordination workbook with a native Excel bubble chart.

Sheet ``Dados`` carries one row per fund with the current and the contractual
subordination side by side, so an analyst can correct a value in place.  Sheet
``Gráfico`` holds a native Office bubble chart whose series point at those
cells: editing the data sheet redraws the chart, with no rebuild.

Sources
-------
* current subordination and PL — ``vehicle_monthly.csv.gz`` (CVM Informe Mensal
  FIDC), competence 2026-06;
* contractual minimum and structural support — the documentary curation in
  ``industry_carteira_1_document_curation.csv``, which records the regulation
  id, its date and the page of the clause;
* type and focus — the official ANBIMA class registry.

Units differ between sources and are normalized here: the Informe reports
subordination as a fraction, the curation as percentage points.

    python scripts/build_carteira101_subordinacao_xlsx.py
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl.chart import BubbleChart, Reference, Series
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_DATA_DIR = Path("data/industry_study")
DEFAULT_OUTPUT = Path("outputs/carteira101/Carteira101_Subordinacao_jun26.xlsx")
COMPETENCE = "2026-06"

ORANGE = "E36C0A"
WHITE = "FFFFFF"
GRAY_100 = "F5F6F7"
GRAY_300 = "D7DADD"
RED_FILL = "FCE4E4"
RED_TEXT = "A83C32"

#: One colour per ANBIMA type, so the chart legend reads as the deck taxonomy.
TYPE_COLORS: dict[str, str] = {
    "Financeiro": "14315C",
    "Agro, Indústria e Comércio": "2D6F51",
    "Fomento Mercantil": "E36C0A",
    "Outros": "8D9399",
    "Não classificado": "C7CBCF",
}

COLUMNS: tuple[tuple[str, str, float], ...] = (
    ("ordem", "#", 5),
    ("cnpj", "CNPJ do fundo", 17),
    ("fidc", "FIDC", 46),
    ("tipo_anbima", "Tipo ANBIMA", 24),
    ("foco_anbima", "Foco ANBIMA", 26),
    ("pl_mm", "PL jun-26 (R$ mm)", 16),
    ("sub_atual_pct", "Subord. atual (%)", 16),
    ("sub_minima_pct", "Subord. mínima (%)", 17),
    ("sub_estrutural_pct", "Subord. estrutural (%)", 19),
    ("folga_pp", "Folga vs. mínima (p.p.)", 20),
    ("folga_estrutural_pp", "Folga vs. estrutural (p.p.)", 22),
    ("doc_regulamento", "Regulamento (id)", 16),
    ("doc_data", "Data do regulamento", 17),
    ("pagina", "Página da cláusula", 16),
    ("sub_min_fonte", "Fonte da mínima", 22),
    ("status_curadoria", "Status da curadoria", 40),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def build_frame(data_dir: Path) -> pd.DataFrame:
    scope = pd.read_csv(data_dir / "industry_carteira_1_scope.csv", dtype=str)
    curation = pd.read_csv(
        data_dir / "industry_carteira_1_document_curation.csv", dtype=str
    ).set_index("cnpj_fundo")
    monthly = pd.read_csv(
        data_dir / "vehicle_monthly.csv.gz", dtype={"cnpj": str}, low_memory=False
    )
    monthly = monthly[monthly["competencia"].astype(str).eq(COMPETENCE)]
    monthly = monthly.drop_duplicates("cnpj").set_index("cnpj")
    registry = pd.read_csv(data_dir / "industry_anbima_classification.csv.gz", dtype=str)
    by_fund = registry.drop_duplicates("cnpj_fundo").set_index("cnpj_fundo")
    by_class = registry.drop_duplicates("cnpj_classe").set_index("cnpj_classe")

    def number(value: object) -> float | None:
        return float(value) if value is not None and pd.notna(value) else None

    rows: list[dict[str, object]] = []
    for record in scope.itertuples():
        cnpj = str(record.cnpj_fundo)
        current = monthly.loc[cnpj] if cnpj in monthly.index else None
        curated = curation.loc[cnpj] if cnpj in curation.index else None

        # The Informe reports subordination as a fraction; the curation records
        # percentage points.  Everything below is percentage points.
        atual = number(current["subordinacao_pct"]) if current is not None else None
        minima = (
            number(curated["subordinacao_minima_junior_pct"])
            if curated is not None
            else None
        )
        estrutural = (
            number(curated["suporte_estrutural_minimo_pct"])
            if curated is not None
            else None
        )
        patrimonio = number(current["pl"]) if current is not None else None

        rows.append(
            {
                "ordem": int(record.ordem),
                "cnpj": cnpj,
                "fidc": (
                    str(current["denominacao"])
                    if current is not None and pd.notna(current["denominacao"])
                    else str(record.nome_foto or "")
                ),
                "tipo_anbima": (
                    by_fund["tipo_anbima"].get(cnpj)
                    or by_class["tipo_anbima"].get(cnpj)
                    or "Não classificado"
                ),
                "foco_anbima": (
                    by_fund["foco_anbima"].get(cnpj)
                    or by_class["foco_anbima"].get(cnpj)
                    or ""
                ),
                "pl_mm": round(patrimonio / 1e6, 1) if patrimonio is not None else None,
                "sub_atual_pct": round(atual * 100, 2) if atual is not None else None,
                "sub_minima_pct": round(minima, 2) if minima is not None else None,
                "sub_estrutural_pct": (
                    round(estrutural, 2) if estrutural is not None else None
                ),
                "doc_regulamento": _text(curated, "documento_id_regulamento"),
                "doc_data": _text(curated, "documento_data_regulamento"),
                "pagina": _text(curated, "pagina_clausula"),
                "sub_min_fonte": _text(curated, "subordinacao_minima_fonte"),
                "status_curadoria": _text(curated, "status_curadoria_documental"),
            }
        )

    frame = pd.DataFrame(rows)
    frame["plotavel"] = (
        frame["pl_mm"].notna()
        & frame["sub_atual_pct"].notna()
        & frame["sub_minima_pct"].notna()
    )
    # Plottable rows first, grouped by type: an Excel series needs a contiguous
    # range, so each type must occupy one uninterrupted block.
    return frame.sort_values(
        ["plotavel", "tipo_anbima", "pl_mm"], ascending=[False, True, False]
    ).reset_index(drop=True)


def _text(row: object, column: str) -> str:
    if row is None:
        return ""
    value = row[column]
    return "" if pd.isna(value) else str(value)


def write_workbook(frame: pd.DataFrame, output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook

    workbook = Workbook()
    data = workbook.active
    data.title = "Dados"

    header_font = Font(name="Arial", size=10, bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=ORANGE)
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color=GRAY_300)
    border = Border(bottom=thin)

    for index, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = data.cell(row=1, column=index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data.column_dimensions[get_column_letter(index)].width = width
    data.row_dimensions[1].height = 30

    folga_index = [key for key, _, _ in COLUMNS].index("folga_pp") + 1
    folga_estrutural_index = (
        [key for key, _, _ in COLUMNS].index("folga_estrutural_pp") + 1
    )
    atual_letter = get_column_letter([k for k, _, _ in COLUMNS].index("sub_atual_pct") + 1)
    minima_letter = get_column_letter([k for k, _, _ in COLUMNS].index("sub_minima_pct") + 1)
    estrutural_letter = get_column_letter(
        [k for k, _, _ in COLUMNS].index("sub_estrutural_pct") + 1
    )

    for offset, record in enumerate(frame.itertuples(index=False), start=2):
        values = record._asdict()
        for index, (key, _, _) in enumerate(COLUMNS, start=1):
            if key == "folga_pp":
                # A formula, not a constant: correcting a value on this sheet
                # has to update the gap and the chart without a rebuild.
                value = (
                    f"=IF(OR({atual_letter}{offset}=\"\",{minima_letter}{offset}=\"\"),"
                    f"\"\",{atual_letter}{offset}-{minima_letter}{offset})"
                )
            elif key == "folga_estrutural_pp":
                value = (
                    f"=IF(OR({atual_letter}{offset}=\"\",{estrutural_letter}{offset}=\"\"),"
                    f"\"\",{atual_letter}{offset}-{estrutural_letter}{offset})"
                )
            else:
                value = values.get(key)
                if pd.isna(value):
                    value = None
            cell = data.cell(row=offset, column=index, value=value)
            cell.font = body_font
            cell.border = border
            if key in {"pl_mm"}:
                cell.number_format = "#,##0.0"
            elif key in {
                "sub_atual_pct",
                "sub_minima_pct",
                "sub_estrutural_pct",
            }:
                cell.number_format = '0.00"%"'
            elif key in {"folga_pp", "folga_estrutural_pp"}:
                cell.number_format = '+0.00;-0.00;0.00'
            elif key == "cnpj":
                cell.number_format = "@"
            if key in {"fidc", "tipo_anbima", "foco_anbima", "status_curadoria", "sub_min_fonte"}:
                cell.alignment = Alignment(vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = len(frame) + 1
    data.freeze_panes = "D2"
    data.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    negative = CellIsRule(
        operator="lessThan",
        formula=["0"],
        fill=PatternFill("solid", fgColor=RED_FILL),
        font=Font(name="Arial", size=10, bold=True, color=RED_TEXT),
    )
    for column_index in (folga_index, folga_estrutural_index):
        letter = get_column_letter(column_index)
        data.conditional_formatting.add(f"{letter}2:{letter}{last_row}", negative)

    _write_chart_sheet(workbook, data, frame)
    workbook.save(output)
    return output


def _write_chart_sheet(workbook, data, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet("Gráfico")
    sheet.sheet_view.showGridLines = False
    title = sheet.cell(row=1, column=1, value="Carteira 101 — subordinação mínima × atual")
    title.font = Font(name="Arial", size=14, bold=True, color="151515")
    subtitle = sheet.cell(
        row=2,
        column=1,
        value=(
            "Cada bolha é um FIDC; o tamanho é o PL de jun-26 em R$ mm. "
            "Acima da diagonal, a subordinação atual supera a mínima contratual. "
            "Os pontos vêm da aba Dados — editar lá redesenha o gráfico."
        ),
    )
    subtitle.font = Font(name="Arial", size=9, color="5D6369")

    keys = [key for key, _, _ in COLUMNS]
    col_x = keys.index("sub_minima_pct") + 1
    col_y = keys.index("sub_atual_pct") + 1
    col_z = keys.index("pl_mm") + 1
    col_name = keys.index("fidc") + 1

    chart = BubbleChart()
    chart.style = 18
    chart.title = "Subordinação mínima (x) × atual (y) — bolha = PL jun-26"
    chart.x_axis.title = "Subordinação mínima contratual (%)"
    chart.y_axis.title = "Subordinação atual, jun-26 (%)"
    chart.x_axis.majorGridlines = None
    chart.height = 15.5
    chart.width = 30.0

    plottable = frame[frame["plotavel"]]
    start = 2
    for anbima_type, block in plottable.groupby("tipo_anbima", sort=False):
        first = start
        last = start + len(block) - 1
        series = Series(
            values=Reference(data, min_col=col_y, min_row=first, max_row=last),
            xvalues=Reference(data, min_col=col_x, min_row=first, max_row=last),
            zvalues=Reference(data, min_col=col_z, min_row=first, max_row=last),
            title=str(anbima_type),
        )
        series.bubble3D = False
        colour = TYPE_COLORS.get(str(anbima_type), "8D9399")
        series.graphicalProperties.solidFill = colour
        series.graphicalProperties.line.solidFill = colour
        chart.series.append(series)
        start = last + 1

    sheet.add_chart(chart, "A4")

    note_row = 36
    for offset, line in enumerate(
        (
            "Como validar:",
            "1. Corrija os valores na aba Dados — as colunas de folga são fórmulas e o gráfico segue as células.",
            "2. Folga negativa fica destacada em vermelho: a subordinação atual está abaixo do mínimo contratual.",
            "3. As linhas estão agrupadas por Tipo ANBIMA porque cada série do gráfico lê um intervalo contínuo. "
            "Para mover um fundo de tipo, reordene o bloco e ajuste o intervalo da série.",
            "4. Fundos sem PL, sem subordinação atual ou sem mínima curada ficam no fim da aba Dados e fora do gráfico.",
        ),
        start=note_row,
    ):
        cell = sheet.cell(row=offset, column=1, value=line)
        cell.font = Font(name="Arial", size=9, bold=offset == note_row, color="30353A")


def main() -> None:
    args = parse_args()
    frame = build_frame(args.data_dir)
    path = write_workbook(frame, args.output)
    plottable = int(frame["plotavel"].sum())
    print(f"workbook: {path}")
    print(f"  fundos: {len(frame)} | no gráfico: {plottable}")
    for column in ("pl_mm", "sub_atual_pct", "sub_minima_pct", "sub_estrutural_pct"):
        print(f"  {column}: {int(frame[column].notna().sum())} preenchidos")


if __name__ == "__main__":
    main()
