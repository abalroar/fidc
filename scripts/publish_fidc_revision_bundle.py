"""Build and atomically publish the audited FIDC revision Office bundle.

This command is intentionally an offline publishing step.  It rebuilds the
revision analysis and editorial payload in a staging directory, invokes the
JavaScript artifact renderer there, validates every output, and only then
replaces the published files.  ``industry_export_bundle.json`` is always the
last file replaced, so the application either sees the previous valid bundle
or fails closed while a new bundle is being published.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import gzip
import hashlib
from io import BytesIO
import json
import math
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import tempfile
from typing import Callable, Iterable, Mapping
import zipfile

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.build_fidc_revision_analysis import main as build_revision_analysis
from scripts.build_fidc_revision_artifact_payload import build_payload
from scripts.build_fidc_provider_history import main as build_provider_history
from services.industry_revision_export import (
    BUNDLE_MANIFEST_NAME,
    BUNDLE_SCHEMA,
    EXPECTED_SLIDES,
    MATERIALIZED_HTML_NAME,
    MATERIALIZED_PORTFOLIO_XLSX_NAME,
    MATERIALIZED_PPTX_NAME,
    MATERIALIZED_XLSX_NAME,
    validate_revision_html,
    validate_revision_portfolio_xlsx,
    validate_revision_pptx,
    validate_revision_xlsx,
)


ROOT = Path(__file__).resolve().parents[1]
ARTIFACT_SCRIPT = ROOT / "scripts" / "build_fidc_revision_artifacts.mjs"
NATIVE_CHART_PATCHER = ROOT / "scripts" / "patch_pptx_native_market_charts.py"
PROVIDER_FLOW_BUILDER = ROOT / "scripts" / "build_provider_flow_explorer.mjs"
PAYLOAD_NAME = "artifact_payload.json"
ANALYSIS_MANIFEST_NAME = "revision_manifest.json"
PAYLOAD_SCHEMA = "fidc_revision_artifact_payload_v8"
DEFAULT_CURATION = ROOT / "outputs" / "analysis" / "top20_fidcs_curadoria.csv"
DEFAULT_TIMEOUT_SECONDS = 30 * 60

REQUIRED_DATA_INPUTS = (
    "vehicle_monthly.csv.gz",
    "industry_competence_status.csv",
    "industry_monthly.csv",
    "cotistas_tipo_monthly.csv",
    "segments_monthly.csv",
    "prestadores_latest.csv",
    "industry_offers.csv.gz",
    "industry_originators_annual.csv",
    "industry_closed_offers_annual.csv",
    "industry_closed_offers_monthly.csv",
    "industry_closed_offer_originators_2026.csv",
    "industry_closed_offer_ticket_distribution.csv",
    "industry_closed_offer_ticket_cohort.csv.gz",
    "industry_closed_offer_placement_regime.csv",
    "industry_fixed_income_offer_comparison.csv",
    "industry_anbima_market_offers.csv",
    "industry_market_offer_reconciliation.csv",
    "industry_issuance_taxonomy_delta.csv",
    "industry_bcb_expanded_credit.csv",
    "industry_offer_document_curation.csv",
    "industry_offer_rating_review.csv",
    "industry_offer_rating_curation.csv",
    "industry_offer_rating_by_offer.csv",
    "industry_top20_outros_regulation_review.csv",
    "top20_outros_regulation_curation.csv",
    "document_inventory.csv.gz",
    "taxonomy_review_actions.csv",
    "taxonomy_review_audit.csv",
    "taxonomy_user_comment_overrides.csv",
    "industry_taxonomy_document_review.csv",
    "industry_top20_taxonomy_document_review.csv",
    "industry_top20_taxonomy_document_conclusions.csv",
    "provider_ownership_curation.csv",
    "bank_fidc_curation.csv",
    "acquiring_reclassification_curation.csv",
    "card_receivables_curation.csv",
    "atlantico_curadoria.json",
    "acquiring_taxonomy_curation.json",
    "industry_flagship_scope.csv",
    "industry_flagship_document_curation.csv",
    "industry_carteira_1_scope.csv",
    "industry_carteira_1_document_curation.csv",
    "industry_cnpj_manual_enrichment.csv",
    "emission_field_audit.csv",
)
OPTIONAL_DATA_INPUTS = (
    "industry_anbima_classification.csv.gz",
    "industry_large_fund_classification.csv",
    "anbima_documentary_overrides.csv",
    "fic_perimeter_overrides.csv",
    # Auditoria da exclusão de FICs e cross-check da taxonomia: viajam com o
    # bundle para que o Excel exportado reproduza o que os gráficos mostram e
    # diga por que cada fundo saiu do universo.
    "industry_fic_detection_audit.csv",
    "industry_taxonomy_crosscheck.csv",
    "top20_profile_curation_overrides.csv",
    "industry_intelligence_manifest.json",
)
WORKBOOK_AUDIT_DATA_INPUTS = (
    "industry_fic_detection_audit.csv",
    "industry_taxonomy_crosscheck.csv",
    "taxonomy_review_actions.csv",
)
BUILDER_SOURCES = (
    ROOT / "scripts" / "build_fidc_revision_analysis.py",
    ROOT / "scripts" / "build_fidc_revision_artifact_payload.py",
    ROOT / "scripts" / "build_fidc_top20_taxonomy_document_conclusions.py",
    ROOT / "scripts" / "apply_consolidated_taxonomy_decisions.py",
    ROOT / "scripts" / "build_fidc_revision_artifacts.mjs",
    ROOT / "scripts" / "build_fidc_offer_ticket_distribution.py",
    ROOT / "scripts" / "build_fidc_closed_offers.py",
    ROOT / "scripts" / "build_fidc_closed_offer_placement_regime.py",
    ROOT / "scripts" / "build_fidc_fixed_income_offer_comparison.py",
    ROOT / "scripts" / "build_fidc_market_offer_reconciliation.py",
    ROOT / "scripts" / "build_fidc_issuance_taxonomy_delta.py",
    ROOT / "scripts" / "build_fidc_bcb_expanded_credit.py",
    ROOT / "scripts" / "build_fidc_offer_document_curation.py",
    ROOT / "scripts" / "build_fidc_offer_rating_review.py",
    ROOT / "scripts" / "build_fidc_top20_outros_regulations.py",
    ROOT / "scripts" / "build_provider_flow_explorer.mjs",
    ROOT / "scripts" / "build_fidc_provider_history.py",
    ROOT / "scripts" / "patch_pptx_native_market_charts.py",
    ROOT / "services" / "industry_revision_analysis.py",
    ROOT / "services" / "industry_revision_additions.py",
    ROOT / "services" / "industry_closed_offers.py",
    ROOT / "services" / "industry_closed_offers_source.py",
    ROOT / "services" / "industry_executive_pack.py",
    ROOT / "services" / "industry_ppt_export.py",
    ROOT / "services" / "industry_revision_export.py",
    ROOT / "services" / "industry_portfolio_export.py",
    ROOT / "services" / "industry_taxonomy_review.py",
    ROOT / "services" / "industry_provider_history.py",
    ROOT / "services" / "industry_offer_ticket_distribution.py",
    ROOT / "services" / "industry_closed_offer_rankings.py",
    ROOT / "services" / "industry_closed_offer_placement_regime.py",
    ROOT / "services" / "industry_fixed_income_offer_comparison.py",
    ROOT / "services" / "industry_market_offer_reconciliation.py",
    ROOT / "services" / "industry_issuance_taxonomy.py",
    ROOT / "services" / "industry_flagship_curation.py",
    ROOT / "services" / "industry_public_offers.py",
    ROOT / "services" / "industry_bcb_expanded_credit.py",
    ROOT / "services" / "industry_offer_document_curation.py",
    ROOT / "services" / "industry_top20_outros_regulations.py",
)
REQUIRED_ANALYSIS_FILES = {
    "base_competencia_cnpj.csv.gz",
    "base_fundo_cnpj.csv.gz",
    "source_presence_overlay.csv.gz",
    "qa_inadimplencia_competencia.csv",
    "reconciliacao_tabelas_i_ii_resumo.csv",
    "reconciliacao_tabelas_i_ii_detalhe.csv",
    "top20_fidcs.csv",
    "top20_outros.csv",
    "monoestrutura_por_fundo.csv",
    "monoestrutura_concentracao.csv",
    "market_share_por_subtipo.csv",
    "market_share_top10_fixo.csv",
    "market_share_escopo_resumo.csv",
    "prestadores_ranking_historico.csv",
    "prestadores_independentes_ranking.csv",
    "bancos_fidcs_evolucao.csv",
    "adquirencia_mix_reclassificado.csv",
    "inadimplencia_tipo_recebivel_unico.csv",
    "inadimplencia_tipo_recebivel_unico_resumo.csv",
    "inadimplencia_coorte_atual_membros.csv.gz",
    "inadimplencia_coorte_atual_historico.csv",
    "inadimplencia_coorte_atual_resumo.csv",
    "inadimplencia_coorte_revisao_resumo.csv",
    "inadimplencia_coorte_revisao_transicoes.csv",
    "inadimplencia_coorte_revisao_sensibilidade.csv",
    "inadimplencia_dispersao_subcategoria.csv",
    "inadimplencia_dispersao_resumo.csv",
    "prestadores_transicoes_resumo.csv",
    "prestadores_transicoes_links.csv",
    "prestadores_transicoes_detalhe.csv",
    "prestadores_transicoes_disponibilidade.csv",
    "reag_cbsf_coorte_resumo.csv",
    "reag_cbsf_coorte_links.csv",
    "reag_cbsf_coorte_detalhe.csv",
    "prestadores_lideranca_atribuicao.csv",
    "bancos_fidcs_detalhe.csv",
    "btg_prestadores_ex_controlados.csv",
    "btg_fidcs_controlados_reconciliacao.csv",
    "qi_atribuicao_cnpjs_legados.csv",
}
REQUIRED_PROVIDER_HISTORY_FILES = {
    "prestadores_historico_cvm_cobertura.csv",
    "prestadores_historico_cvm_manifest.json",
    "prestadores_historico_cvm_snapshot.csv.gz",
    "prestadores_historico_cvm_transicoes_detalhe.csv.gz",
    "prestadores_historico_cvm_transicoes_links.csv",
}


class RevisionBundlePublishError(RuntimeError):
    """Raised before publication when a staged revision is not trustworthy."""


@dataclass(frozen=True)
class PublishedRevisionBundle:
    bundle_id: str
    latest_complete: str
    payload_path: Path
    pptx_path: Path
    xlsx_path: Path
    portfolio_xlsx_path: Path
    html_path: Path
    manifest_path: Path


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha256_semantic_file(path: Path) -> str:
    """Hash decompressed CSV content so gzip timestamps do not change identity."""

    if path.suffix == ".gz":
        digest = hashlib.sha256()
        with gzip.open(path, "rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
    return _sha256_file(path)


def _canonical_json_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def discover_latest_complete(data_dir: Path) -> str:
    """Return the newest competence explicitly marked ``completa``."""

    status_path = Path(data_dir) / "industry_competence_status.csv"
    if not status_path.exists():
        raise RevisionBundlePublishError(f"status de competências ausente: {status_path}")
    with status_path.open(encoding="utf-8", newline="") as handle:
        rows = csv.DictReader(handle)
        complete = sorted(
            {
                str(row.get("competencia") or "").strip()
                for row in rows
                if str(row.get("publication_status") or "").strip() == "completa"
                and re.fullmatch(r"\d{4}-\d{2}", str(row.get("competencia") or "").strip())
            }
        )
    if not complete:
        raise RevisionBundlePublishError(
            f"nenhuma competência completa encontrada em {status_path}"
        )
    return complete[-1]


def discover_artifact_node_modules(
    explicit: Path | None = None,
    *,
    root: Path = ROOT,
    home: Path | None = None,
) -> Path:
    """Locate an already-installed artifact runtime without network access."""

    candidates: list[Path] = []
    if explicit is not None:
        candidates.append(Path(explicit).expanduser())
    else:
        configured = os.environ.get("CODEX_NODE_MODULES", "").strip()
        if configured:
            candidates.append(Path(configured).expanduser())
        home = Path.home() if home is None else Path(home)
        candidates.extend(
            [
                Path(root) / "node_modules",
                home
                / ".cache"
                / "codex-runtimes"
                / "codex-primary-runtime"
                / "dependencies"
                / "node"
                / "node_modules",
            ]
        )
    for candidate in candidates:
        package = candidate / "@oai" / "artifact-tool" / "package.json"
        if package.exists():
            return candidate.resolve()
    searched = ", ".join(str(path) for path in candidates) or "nenhum caminho"
    raise RevisionBundlePublishError(
        "runtime offline do @oai/artifact-tool não localizado; caminhos: " + searched
    )


def _generated_at(explicit: str = "") -> str:
    value = str(explicit or "").strip()
    if value:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc).isoformat(timespec="seconds")
    epoch = os.environ.get("SOURCE_DATE_EPOCH", "").strip()
    if epoch:
        return datetime.fromtimestamp(int(epoch), tz=timezone.utc).isoformat(
            timespec="seconds"
        )
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def validate_analysis_manifest(
    manifest: Mapping[str, object],
    *,
    revision_dir: Path,
    latest_complete: str,
) -> None:
    if str(manifest.get("latest_complete") or "") != latest_complete:
        raise RevisionBundlePublishError(
            "competência do manifest analítico diverge da publicação"
        )
    files = dict(manifest.get("files") or {})
    missing_entries = sorted(REQUIRED_ANALYSIS_FILES.difference(files))
    if missing_entries:
        raise RevisionBundlePublishError(
            "manifest analítico sem arquivos obrigatórios: " + ", ".join(missing_entries)
        )
    missing_files = sorted(
        name for name in REQUIRED_ANALYSIS_FILES if not (revision_dir / name).exists()
    )
    if missing_files:
        raise RevisionBundlePublishError(
            "staging analítico incompleto: " + ", ".join(missing_files)
        )
    checks = dict(manifest.get("checks") or {})
    if int(checks.get("top20_fidcs_rows") or 0) != 20:
        raise RevisionBundlePublishError("Top 20 FIDCs não contém exatamente 20 linhas")
    if int(checks.get("top20_outros_rows") or 0) != 20:
        raise RevisionBundlePublishError("Top 20 Outros não contém exatamente 20 linhas")
    if int(checks.get("latest_funds") or 0) <= 0:
        raise RevisionBundlePublishError("universo de fundos vazio no manifest analítico")


def validate_source_presence_coverage(
    revision_dir: Path,
    latest_complete: str,
) -> None:
    """Block publication when the raw empty-versus-zero audit is incomplete."""

    base_path = revision_dir / "base_competencia_cnpj.csv.gz"
    overlay_path = revision_dir / "source_presence_overlay.csv.gz"
    latest_rows = 0
    exact_rows = 0
    base_rows_by_period: dict[str, int] = {}
    with gzip.open(base_path, "rt", encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            competence = str(row.get("competencia") or "")
            if not competence:
                continue
            base_rows_by_period[competence] = (
                base_rows_by_period.get(competence, 0) + 1
            )
            if competence != latest_complete:
                continue
            latest_rows += 1
            if str(row.get("field_presence_exact") or "").strip().lower() in {
                "1",
                "true",
                "sim",
            }:
                exact_rows += 1
    if latest_rows <= 0:
        raise RevisionBundlePublishError(
            "base analítica sem veículos na competência mais recente"
        )
    if exact_rows != latest_rows:
        raise RevisionBundlePublishError(
            "auditoria vazio-versus-zero incompleta na competência mais recente; "
            "publique com --refresh-source-presence e o ZIP bruto CVM disponível"
        )

    overlay_rows_by_period: dict[str, int] = {}
    with gzip.open(overlay_path, "rt", encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            competence = str(row.get("competencia") or "")
            if competence:
                overlay_rows_by_period[competence] = (
                    overlay_rows_by_period.get(competence, 0) + 1
                )
    overlay_latest_rows = overlay_rows_by_period.get(latest_complete, 0)
    if overlay_latest_rows < latest_rows:
        raise RevisionBundlePublishError(
            "overlay bruto de presença não cobre a competência mais recente"
        )
    incomplete_periods = [
        competence
        for competence, base_rows in sorted(base_rows_by_period.items())
        if overlay_rows_by_period.get(competence, 0) < base_rows
    ]
    if incomplete_periods:
        sample = ", ".join(incomplete_periods[:6])
        raise RevisionBundlePublishError(
            "overlay bruto de presença não cobre o histórico completo: "
            f"{sample}; publique com --refresh-source-presence "
            "e --presence-months all"
        )


def serialize_analysis_manifest(
    manifest: Mapping[str, object], generated_at_utc: str
) -> tuple[dict[str, object], bytes]:
    """Canonicalize the analysis manifest under the publisher's build clock."""

    normalized = dict(manifest)
    normalized["generated_at_utc"] = generated_at_utc
    payload = json.dumps(
        normalized,
        ensure_ascii=False,
        indent=2,
        sort_keys=True,
    ).encode("utf-8")
    return normalized, payload


_CARD_TAXONOMY_EXPECTED_STATUS_COUNTS = {
    "Incluído em Adquirência": 26,
    "Fora de Adquirência": 17,
    "Pendente": 1,
}
_CARD_TAXONOMY_SUMMARY_STATUS_FIELDS = {
    "Incluído em Adquirência": (
        "fundos_incluidos_adquirencia",
        "pl_incluido_adquirencia_brl",
    ),
    "Fora de Adquirência": (
        "fundos_fora_adquirencia",
        "pl_fora_adquirencia_brl",
    ),
    "Pendente": (
        "fundos_pendentes_curadoria",
        "pl_pendente_curadoria_brl",
    ),
}


def _finite_payload_number(value: object, field: str) -> float:
    if isinstance(value, bool):
        raise RevisionBundlePublishError(f"payload {field} deve ser numérico")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise RevisionBundlePublishError(
            f"payload {field} deve ser numérico"
        ) from exc
    if not math.isfinite(number):
        raise RevisionBundlePublishError(f"payload {field} deve ser finito")
    return number


def _integer_payload_number(value: object, field: str) -> int:
    number = _finite_payload_number(value, field)
    if not number.is_integer():
        raise RevisionBundlePublishError(f"payload {field} deve ser inteiro")
    return int(number)


def _require_payload_amount_close(
    actual: object,
    expected: float,
    field: str,
) -> None:
    actual_number = _finite_payload_number(actual, field)
    if not math.isclose(actual_number, expected, rel_tol=1e-12, abs_tol=0.01):
        raise RevisionBundlePublishError(
            f"payload {field} não reconcilia: {actual_number} != {expected}"
        )


def validate_fic_detection_audit_provenance(data_dir: Path) -> None:
    """Reject a FIC audit that still labels the nominal signal as cadastral."""

    path = Path(data_dir) / "industry_fic_detection_audit.csv"
    if not path.is_file():
        raise RevisionBundlePublishError(
            f"auditoria de proveniência FIC ausente: {path}"
        )
    legacy_methods: dict[str, int] = {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if "fic_detection_method" not in (reader.fieldnames or ()):
            raise RevisionBundlePublishError(
                "industry_fic_detection_audit.csv sem fic_detection_method"
            )
        for row in reader:
            method = str(row.get("fic_detection_method") or "")
            if method in {"flag_cadastral", "combinacao"}:
                legacy_methods[method] = legacy_methods.get(method, 0) + 1
    if legacy_methods:
        details = ", ".join(
            f"{method}={count}"
            for method, count in sorted(legacy_methods.items())
        )
        raise RevisionBundlePublishError(
            "auditoria FIC ainda usa rótulos de proveniência legados "
            f"({details}); execute scripts/build_fic_detection_audit.py "
            f"--data-dir {Path(data_dir)}"
        )


def _validate_card_taxonomy_contract(payload: Mapping[str, object]) -> None:
    rows = payload.get("card_taxonomy_audit")
    summary = payload.get("card_taxonomy_summary")
    if not isinstance(rows, list) or len(rows) != 44:
        raise RevisionBundlePublishError(
            "card_taxonomy_audit deve conter exatamente 44 fundos"
        )
    if not isinstance(summary, Mapping):
        raise RevisionBundlePublishError("payload editorial sem card_taxonomy_summary")

    ranks: list[int] = []
    cnpjs: list[str] = []
    status_counts = {status: 0 for status in _CARD_TAXONOMY_EXPECTED_STATUS_COUNTS}
    status_pl = {status: 0.0 for status in _CARD_TAXONOMY_EXPECTED_STATUS_COUNTS}
    total_pl = 0.0

    for index, row in enumerate(rows, start=1):
        if not isinstance(row, Mapping):
            raise RevisionBundlePublishError(
                f"card_taxonomy_audit contém linha {index} inválida"
            )
        ranks.append(
            _integer_payload_number(
                row.get("ordem_materialidade"),
                f"card_taxonomy_audit[{index}].ordem_materialidade",
            )
        )
        cnpj = re.sub(r"\D", "", str(row.get("cnpj_fundo_formatado") or ""))
        if len(cnpj) != 14:
            raise RevisionBundlePublishError(
                f"card_taxonomy_audit linha {index} contém CNPJ inválido"
            )
        if row.get("cnpj_fundo_identificado") is not True:
            raise RevisionBundlePublishError(
                f"card_taxonomy_audit linha {index} sem CNPJ identificado"
            )
        cnpjs.append(cnpj)

        status = str(row.get("status_curadoria") or "").strip()
        if status not in _CARD_TAXONOMY_EXPECTED_STATUS_COUNTS:
            raise RevisionBundlePublishError(
                f"card_taxonomy_audit linha {index} contém status_curadoria inválido"
            )
        pl = _finite_payload_number(
            row.get("pl_referencia_brl"),
            f"card_taxonomy_audit[{index}].pl_referencia_brl",
        )
        if pl < 0:
            raise RevisionBundlePublishError(
                f"card_taxonomy_audit linha {index} contém PL negativo"
            )
        status_counts[status] += 1
        status_pl[status] += pl
        total_pl += pl

        source_url = str(row.get("fonte_url") or "").strip()
        if re.match(r"^https?://", source_url, flags=re.IGNORECASE) is None:
            raise RevisionBundlePublishError(
                f"card_taxonomy_audit linha {index} contém fonte_url inválida"
            )
        if row.get("consistencia_decisao_reclassificacao") != "OK":
            raise RevisionBundlePublishError(
                f"card_taxonomy_audit linha {index} contém divergência de decisão"
            )

    expected_ranks = list(range(1, 45))
    if ranks != expected_ranks:
        raise RevisionBundlePublishError(
            "card_taxonomy_audit deve ter ordem_materialidade contínua de 1 a 44"
        )
    if len(set(cnpjs)) != 44:
        raise RevisionBundlePublishError(
            "card_taxonomy_audit deve conter 44 CNPJs únicos"
        )
    if status_counts != _CARD_TAXONOMY_EXPECTED_STATUS_COUNTS:
        raise RevisionBundlePublishError(
            "card_taxonomy_audit deve reconciliar 26 incluídos, 17 fora e 1 pendente"
        )

    required_summary_fields = {
        "fundos_total",
        "fundos_incluidos_adquirencia",
        "fundos_fora_adquirencia",
        "fundos_pendentes_curadoria",
        "pl_referencia_observado_brl",
        "pl_incluido_adquirencia_brl",
        "pl_fora_adquirencia_brl",
        "pl_pendente_curadoria_brl",
        "divergencias_decisao_reclassificacao",
    }
    if missing := sorted(required_summary_fields.difference(summary)):
        raise RevisionBundlePublishError(
            "card_taxonomy_summary sem campos obrigatórios: " + ", ".join(missing)
        )
    if _integer_payload_number(summary.get("fundos_total"), "card_taxonomy_summary.fundos_total") != 44:
        raise RevisionBundlePublishError(
            "card_taxonomy_summary deve reconciliar 44 fundos"
        )
    for status, expected_count in _CARD_TAXONOMY_EXPECTED_STATUS_COUNTS.items():
        count_field, pl_field = _CARD_TAXONOMY_SUMMARY_STATUS_FIELDS[status]
        if _integer_payload_number(
            summary.get(count_field), f"card_taxonomy_summary.{count_field}"
        ) != expected_count:
            raise RevisionBundlePublishError(
                f"card_taxonomy_summary.{count_field} não reconcilia"
            )
        _require_payload_amount_close(
            summary.get(pl_field),
            status_pl[status],
            f"card_taxonomy_summary.{pl_field}",
        )
    _require_payload_amount_close(
        summary.get("pl_referencia_observado_brl"),
        total_pl,
        "card_taxonomy_summary.pl_referencia_observado_brl",
    )
    if _integer_payload_number(
        summary.get("divergencias_decisao_reclassificacao"),
        "card_taxonomy_summary.divergencias_decisao_reclassificacao",
    ) != 0:
        raise RevisionBundlePublishError(
            "card_taxonomy_summary deve registrar zero divergências"
        )


def _validate_closed_offer_originator_order(payload: Mapping[str, object]) -> None:
    rows = payload.get("closed_offer_originators_2026")
    if not isinstance(rows, list) or not rows:
        raise RevisionBundlePublishError(
            "payload editorial sem closed_offer_originators_2026"
        )
    ranks: list[int] = []
    groups: list[str] = []
    volumes: list[float] = []
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, Mapping):
            raise RevisionBundlePublishError(
                f"closed_offer_originators_2026 contém linha {index} inválida"
            )
        ranks.append(
            _integer_payload_number(
                row.get("rank"), f"closed_offer_originators_2026[{index}].rank"
            )
        )
        group = str(row.get("originator_group") or "").strip()
        if not group:
            raise RevisionBundlePublishError(
                f"closed_offer_originators_2026 linha {index} sem originator_group"
            )
        groups.append(group.casefold())
        volume = _finite_payload_number(
            row.get("registered_volume_brl"),
            f"closed_offer_originators_2026[{index}].registered_volume_brl",
        )
        if volume < 0:
            raise RevisionBundlePublishError(
                f"closed_offer_originators_2026 linha {index} contém volume negativo"
            )
        volumes.append(volume)

    if ranks != list(range(1, len(rows) + 1)):
        raise RevisionBundlePublishError(
            "closed_offer_originators_2026 deve ter ranks contínuos e únicos"
        )
    if len(set(groups)) != len(groups):
        raise RevisionBundlePublishError(
            "closed_offer_originators_2026 contém originator_group duplicado"
        )
    if any(current < following for current, following in zip(volumes, volumes[1:])):
        raise RevisionBundlePublishError(
            "closed_offer_originators_2026 deve estar em volume decrescente"
        )


def _validate_closed_offer_top15(payload: Mapping[str, object]) -> None:
    rows = payload.get("closed_offer_top15")
    summaries = payload.get("closed_offer_top15_summary")
    if not isinstance(rows, list) or len(rows) != 67:
        raise RevisionBundlePublishError(
            "closed_offer_top15 deve conter sete linhas em 2022 e 15 nos demais períodos"
        )
    if not isinstance(summaries, list) or len(summaries) != 5:
        raise RevisionBundlePublishError(
            "closed_offer_top15_summary deve conter cinco períodos"
        )

    expected_periods = (
        "2022 FY parcial",
        "2023 FY",
        "2024 FY",
        "2025 FY",
        "2026 jan-jun",
    )
    summary_by_period = {
        str(row.get("period_label") or ""): row
        for row in summaries
        if isinstance(row, Mapping)
    }
    if tuple(summary_by_period) != expected_periods:
        raise RevisionBundlePublishError(
            "closed_offer_top15_summary contém períodos incompatíveis"
        )

    for period_label in expected_periods:
        period_rows = [
            row
            for row in rows
            if isinstance(row, Mapping)
            and str(row.get("period_label") or "") == period_label
        ]
        expected_rows = 7 if period_label == "2022 FY parcial" else 15
        if len(period_rows) != expected_rows:
            raise RevisionBundlePublishError(
                f"closed_offer_top15 contém contagem inválida em {period_label}"
            )
        ranks = [
            _integer_payload_number(
                row.get("rank"),
                f"closed_offer_top15[{period_label}].rank",
            )
            for row in period_rows
        ]
        if ranks != list(range(1, expected_rows + 1)):
            raise RevisionBundlePublishError(
                f"closed_offer_top15 possui ranks inválidos em {period_label}"
            )
        volumes = [
            _finite_payload_number(
                row.get("registered_volume_brl"),
                f"closed_offer_top15[{period_label}].registered_volume_brl",
            )
            for row in period_rows
        ]
        if any(value <= 0 for value in volumes):
            raise RevisionBundlePublishError(
                f"closed_offer_top15 possui volume inválido em {period_label}"
            )
        if any(
            current < following
            for current, following in zip(volumes, volumes[1:])
        ):
            raise RevisionBundlePublishError(
                f"closed_offer_top15 deve estar em volume decrescente em {period_label}"
            )

        for row in period_rows:
            if str(row.get("status") or "").casefold() != "oferta encerrada":
                raise RevisionBundlePublishError(
                    f"closed_offer_top15 contém oferta não encerrada em {period_label}"
                )
            if str(row.get("offer_type") or "").upper() != "PRIMARIA":
                raise RevisionBundlePublishError(
                    f"closed_offer_top15 contém oferta não primária em {period_label}"
                )
            if str(row.get("security") or "").casefold() != "cotas de fidc":
                raise RevisionBundlePublishError(
                    f"closed_offer_top15 contém ativo fora de Cotas de FIDC em {period_label}"
                )
            if str(row.get("originator_group") or "").strip() == "":
                raise RevisionBundlePublishError(
                    f"closed_offer_top15 contém originador vazio em {period_label}"
                )
            rating_agency = str(row.get("rating_agency") or "").strip()
            rating_assigned = str(row.get("rating_assigned") or "").strip()
            if not rating_agency or not rating_assigned:
                raise RevisionBundlePublishError(
                    f"closed_offer_top15 contém rating vazio em {period_label}"
                )
            if (rating_agency == "N/D") != (rating_assigned == "N/D"):
                raise RevisionBundlePublishError(
                    f"closed_offer_top15 contém agência/rating incompatíveis em {period_label}"
                )
            if rating_agency != "N/D":
                for evidence_field in (
                    "rating_source_type",
                    "rating_source_url",
                    "rating_match_status",
                    "rating_evidence",
                ):
                    if not str(row.get(evidence_field) or "").strip():
                        raise RevisionBundlePublishError(
                            "closed_offer_top15 contém rating verificado sem "
                            f"{evidence_field} em {period_label}"
                        )
            for boolean_field, label_field in (
                ("ibba_coord_lead", "ibba_coord_lead_label"),
                ("firm_commitment", "firm_commitment_label"),
            ):
                expected_label = "Sim" if row.get(boolean_field) is True else "Não"
                if row.get(label_field) != expected_label:
                    raise RevisionBundlePublishError(
                        f"closed_offer_top15 contém {label_field} divergente"
                    )

        summary = summary_by_period[period_label]
        top_volume = float(sum(volumes))
        _require_payload_amount_close(
            summary.get("top15_registered_volume_brl"),
            top_volume,
            f"closed_offer_top15_summary[{period_label}].top15_registered_volume_brl",
        )
        period_volume = _finite_payload_number(
            summary.get("period_registered_volume_brl"),
            f"closed_offer_top15_summary[{period_label}].period_registered_volume_brl",
        )
        if period_volume + max(1.0, abs(period_volume) * 1e-10) < top_volume:
            raise RevisionBundlePublishError(
                f"subtotal Top 15 excede o período em {period_label}"
            )
        expected_share = top_volume / period_volume if period_volume else 0.0
        observed_share = _finite_payload_number(
            summary.get("top15_share_of_period_volume"),
            f"closed_offer_top15_summary[{period_label}].top15_share_of_period_volume",
        )
        if abs(observed_share - expected_share) > 1e-10:
            raise RevisionBundlePublishError(
                f"share do Top 15 não reconcilia em {period_label}"
            )


def _validate_fixed_income_offer_comparison(
    payload: Mapping[str, object],
) -> None:
    rows = payload.get("fixed_income_offer_comparison")
    if not isinstance(rows, list) or len(rows) != 28:
        raise RevisionBundlePublishError(
            "fixed_income_offer_comparison deve conter 28 linhas"
        )
    expected_periods = (
        "2023 FY",
        "2024 FY",
        "2025 FY",
        "2026 jan-jun",
    )
    view_a = [
        row
        for row in rows
        if isinstance(row, Mapping)
        and row.get("view") == "FIDCs vs demais elegíveis"
    ]
    view_b = [
        row
        for row in rows
        if isinstance(row, Mapping)
        and row.get("view") == "FIDCs vs instrumentos materiais de 2025"
    ]
    if len(view_a) != 8 or len(view_b) != 20:
        raise RevisionBundlePublishError(
            "comparativo de renda fixa contém cardinalidade incompatível"
        )
    for period_label in expected_periods:
        period_a = [
            row for row in view_a if row.get("period_label") == period_label
        ]
        period_b = [
            row for row in view_b if row.get("period_label") == period_label
        ]
        if len(period_a) != 2 or len(period_b) != 5:
            raise RevisionBundlePublishError(
                f"comparativo de renda fixa incompleto em {period_label}"
            )
        universe = _finite_payload_number(
            period_a[0].get("universe_registered_volume_brl"),
            f"fixed_income_offer_comparison[{period_label}].universe",
        )
        observed = sum(
            _finite_payload_number(
                row.get("registered_volume_brl"),
                f"fixed_income_offer_comparison[{period_label}].volume",
            )
            for row in period_a
        )
        if abs(observed - universe) > max(1.0, universe * 1e-10):
            raise RevisionBundlePublishError(
                f"comparativo de renda fixa não fecha o universo em {period_label}"
            )
    series_2025 = [
        str(row.get("series_label") or "")
        for row in view_b
        if row.get("period_label") == "2025 FY"
    ]
    if series_2025 != [
        "FIDCs",
        "Debêntures",
        "CRI",
        "Notas comerciais",
        "CRA",
    ]:
        raise RevisionBundlePublishError(
            "instrumentos materiais de 2025 divergentes no comparativo"
        )
    if any(
        row.get("yoy_growth") is not None
        for row in rows
        if isinstance(row, Mapping) and row.get("period_label") == "2023 FY"
    ):
        raise RevisionBundlePublishError(
            "2023 deve permanecer sem YoY no comparativo de renda fixa"
        )
    fidc_by_period = {
        str(row.get("period_label") or ""): _finite_payload_number(
            row.get("registered_volume_brl"),
            "fixed_income_offer_comparison.FIDCs",
        )
        for row in view_a
        if row.get("series_label") == "FIDCs"
    }
    official_fidc = {
        str(row.get("period_label") or ""): _finite_payload_number(
            row.get("registered_volume_brl"),
            "closed_offers_annual.registered_volume_brl",
        )
        for row in payload.get("closed_offers_annual") or []
        if isinstance(row, Mapping)
        and str(row.get("period_label") or "") in expected_periods[1:3]
    }
    anbima_2023_rows = [
        row
        for row in payload.get("market_offer_reconciliation") or []
        if isinstance(row, Mapping)
        and str(row.get("period_label") or "") == "2023 FY"
        and str(row.get("instrument_label") or "") == "FIDCs"
    ]
    if len(anbima_2023_rows) != 1:
        raise RevisionBundlePublishError(
            "reconciliação ANBIMA de FIDCs em 2023 FY deve ser única"
        )
    official_fidc["2023 FY"] = _finite_payload_number(
        anbima_2023_rows[0].get("anbima_closed_volume_brl"),
        "market_offer_reconciliation[2023 FY/FIDCs].anbima_closed_volume_brl",
    )
    official_fidc.update(
        {
            "2026 jan-jun": _finite_payload_number(
                row.get("registered_volume_brl"),
                "closed_offers_jan_june.registered_volume_brl",
            )
            for row in payload.get("closed_offers_jan_june") or []
            if isinstance(row, Mapping) and int(row.get("year") or 0) == 2026
        }
    )
    for period_label, expected in official_fidc.items():
        observed = fidc_by_period.get(period_label)
        _require_payload_amount_close(
            observed,
            expected,
            f"fixed_income_offer_comparison.FIDCs[{period_label}]",
        )


def _validate_closed_offer_placement_regime(
    payload: Mapping[str, object],
) -> None:
    rows = payload.get("closed_offer_placement_regime")
    if not isinstance(rows, list) or len(rows) != 12:
        raise RevisionBundlePublishError(
            "closed_offer_placement_regime deve conter 12 linhas"
        )
    expected_periods = ("2024 FY", "2025 FY", "2026 jan-jun")
    expected_regimes = (
        "Melhores esforços",
        "Garantia firme",
        "Misto",
        "Não informado",
    )
    official_totals = {
        str(row.get("period_label") or ""): (
            _finite_payload_number(
                row.get("closed_offers"),
                "closed_offers_annual.closed_offers",
            ),
            _finite_payload_number(
                row.get("registered_volume_brl"),
                "closed_offers_annual.registered_volume_brl",
            ),
        )
        for row in payload.get("closed_offers_annual") or []
        if isinstance(row, Mapping)
        and str(row.get("period_label") or "") in expected_periods[:2]
    }
    for row in payload.get("closed_offers_jan_june") or []:
        if isinstance(row, Mapping) and int(row.get("year") or 0) == 2026:
            official_totals["2026 jan-jun"] = (
                _finite_payload_number(
                    row.get("closed_offers"),
                    "closed_offers_jan_june.closed_offers",
                ),
                _finite_payload_number(
                    row.get("registered_volume_brl"),
                    "closed_offers_jan_june.registered_volume_brl",
                ),
            )
    for period_label in expected_periods:
        period = [
            row
            for row in rows
            if isinstance(row, Mapping)
            and row.get("period_label") == period_label
        ]
        period.sort(key=lambda row: int(row.get("regime_order") or 0))
        if tuple(row.get("placement_regime") for row in period) != expected_regimes:
            raise RevisionBundlePublishError(
                f"regimes de colocação divergentes em {period_label}"
            )
        offers = sum(
            _finite_payload_number(
                row.get("closed_offers"),
                f"closed_offer_placement_regime[{period_label}].closed_offers",
            )
            for row in period
        )
        volume = sum(
            _finite_payload_number(
                row.get("registered_volume_brl"),
                f"closed_offer_placement_regime[{period_label}].registered_volume_brl",
            )
            for row in period
        )
        expected_offers, expected_volume = official_totals[period_label]
        if offers != expected_offers:
            raise RevisionBundlePublishError(
                f"quantidade por regime não reconcilia em {period_label}"
            )
        if abs(volume - expected_volume) > max(1.0, expected_volume * 1e-10):
            raise RevisionBundlePublishError(
                f"volume por regime não reconcilia em {period_label}"
            )


def _validate_market_offer_reconciliation(
    payload: Mapping[str, object],
) -> None:
    rows = payload.get("market_offer_reconciliation")
    if not isinstance(rows, list) or len(rows) != 20:
        raise RevisionBundlePublishError(
            "market_offer_reconciliation deve conter 20 linhas"
        )
    expected_periods = (
        "2023 FY",
        "2024 FY",
        "2025 FY",
        "2026 jan-mai",
    )
    expected_instruments = (
        "Debêntures",
        "FIDCs",
        "CRI",
        "Notas comerciais",
        "CRA",
    )
    for period_label in expected_periods:
        period = [
            row
            for row in rows
            if isinstance(row, Mapping)
            and row.get("period_label") == period_label
        ]
        period.sort(key=lambda row: int(row.get("instrument_order") or 0))
        if tuple(row.get("instrument_label") for row in period) != expected_instruments:
            raise RevisionBundlePublishError(
                f"reconciliação de mercado incompleta em {period_label}"
            )
        for row in period:
            cvm = _finite_payload_number(
                row.get("cvm_registered_volume_brl"),
                "market_offer_reconciliation.cvm_registered_volume_brl",
            )
            bridge = _finite_payload_number(
                row.get("cvm_harmonization_volume_brl"),
                "market_offer_reconciliation.cvm_harmonization_volume_brl",
            )
            harmonized = _finite_payload_number(
                row.get("cvm_harmonized_volume_brl"),
                "market_offer_reconciliation.cvm_harmonized_volume_brl",
            )
            if abs(harmonized - cvm - bridge) > max(1.0, abs(harmonized) * 1e-10):
                raise RevisionBundlePublishError(
                    f"harmonização CVM não reconcilia em {period_label}"
                )
def _validate_portfolio_export_payload(payload: Mapping[str, object]) -> None:
    cohorts = (
        ("portfolio_export_carteira_101", 101, "Carteira 101"),
        ("portfolio_export_flagships", 47, "Flagships"),
    )
    for key, expected_count, label in cohorts:
        rows = payload.get(key)
        if not isinstance(rows, list) or len(rows) != expected_count:
            raise RevisionBundlePublishError(
                f"{key} deve conter {expected_count} linhas"
            )
        cnpjs: list[str] = []
        for row in rows:
            if not isinstance(row, Mapping):
                raise RevisionBundlePublishError(f"{key} contém linha inválida")
            cnpj = str(row.get("cnpj") or "").strip()
            if not re.fullmatch(r"\d{14}", cnpj):
                raise RevisionBundlePublishError(
                    f"{label} contém CNPJ que não tem 14 dígitos"
                )
            cnpjs.append(cnpj)
        if len(set(cnpjs)) != expected_count:
            raise RevisionBundlePublishError(
                f"{label} deve conter {expected_count} CNPJs únicos"
            )

    for key in ("portfolio_export_coverage", "portfolio_export_gaps"):
        rows = payload.get(key)
        if not isinstance(rows, list) or not rows:
            raise RevisionBundlePublishError(f"payload editorial sem {key}")

    manual = payload.get("portfolio_export_manual_audit")
    if not isinstance(manual, list) or not manual:
        raise RevisionBundlePublishError(
            "payload editorial sem portfolio_export_manual_audit"
        )
    roots: list[str] = []
    resolved_cnpjs: list[str] = []
    allowed_resolution = {
        "correspondencia_unica",
        "sem_correspondencia",
        "cnpj_informado",
    }
    for row in manual:
        if not isinstance(row, Mapping):
            raise RevisionBundlePublishError(
                "portfolio_export_manual_audit contém linha inválida"
            )
        root = re.sub(
            r"\D",
            "",
            str(row.get("raiz_cnpj_foto") or row.get("cnpj") or ""),
        )
        if not root:
            raise RevisionBundlePublishError(
                "auditoria manual contém raiz/CNPJ vazio"
            )
        roots.append(root)
        status = str(row.get("status_resolucao_cnpj") or "").strip()
        if status not in allowed_resolution:
            raise RevisionBundlePublishError(
                "auditoria manual contém resolução de CNPJ ambígua ou desconhecida"
            )
        candidate_count = int(row.get("quantidade_candidatos_cnpj") or 0)
        if candidate_count > 1:
            raise RevisionBundlePublishError(
                "auditoria manual contém raiz de CNPJ ambígua"
            )
        resolved = str(row.get("cnpj") or "").strip()
        if resolved:
            if not re.fullmatch(r"\d{14}", resolved):
                raise RevisionBundlePublishError(
                    "auditoria manual contém CNPJ resolvido inválido"
                )
            resolved_cnpjs.append(resolved)
        elif status != "sem_correspondencia":
            raise RevisionBundlePublishError(
                "auditoria manual perdeu o CNPJ de uma correspondência resolvida"
            )
    if len(set(roots)) != len(roots):
        raise RevisionBundlePublishError(
            "auditoria manual contém raízes/CNPJs duplicados"
        )
    if len(set(resolved_cnpjs)) != len(resolved_cnpjs):
        raise RevisionBundlePublishError(
            "auditoria manual contém CNPJs resolvidos duplicados"
        )


def validate_artifact_payload(payload: Mapping[str, object], latest_complete: str) -> None:
    if payload.get("schema_version") != PAYLOAD_SCHEMA:
        raise RevisionBundlePublishError("schema do payload editorial incompatível")
    if payload.get("latest_complete") != latest_complete:
        raise RevisionBundlePublishError("competência do payload editorial diverge")
    _validate_portfolio_export_payload(payload)
    for key in ("top20_fidcs", "top20_outros", "profiles"):
        rows = payload.get(key)
        if not isinstance(rows, list) or len(rows) != 20:
            raise RevisionBundlePublishError(f"payload {key} deve conter 20 linhas")
    emission_audit = payload.get("emission_field_audit")
    if not isinstance(emission_audit, list) or len(emission_audit) != 180:
        raise RevisionBundlePublishError(
            "emission_field_audit deve conter 180 linhas auditadas"
        )
    audit_counts: dict[str, int] = {}
    for row in emission_audit:
        if not isinstance(row, Mapping):
            raise RevisionBundlePublishError("emission_field_audit contém linha inválida")
        block = str(row.get("bloco") or "")
        audit_counts[block] = audit_counts.get(block, 0) + 1
    if audit_counts != {"slides 10–13": 120, "slides 21–22": 60}:
        raise RevisionBundlePublishError("emission_field_audit não fecha 120 + 60 linhas")
    if not payload.get("offers_as_of"):
        raise RevisionBundlePublishError("payload editorial sem data-base de ofertas")
    type_mix = payload.get("type_mix_history")
    if not isinstance(type_mix, list) or len(type_mix) != 16:
        raise RevisionBundlePublishError(
            "type_mix_history deve conter quatro categorias em quatro competências"
        )
    categories = {
        str(row.get("anbima_tipo") or "")
        for row in type_mix
        if isinstance(row, Mapping)
    }
    expected_categories = {
        "Fomento Mercantil",
        "Agro, Indústria e Comércio",
        "Financeiro",
        "Outros",
    }
    if categories != expected_categories:
        raise RevisionBundlePublishError(
            "type_mix_history deve incorporar N/D em Outros na visão editorial"
        )
    expected_periods = {"2023-12", "2024-12", "2025-12", latest_complete}
    periods = {
        str(row.get("competencia") or "")
        for row in type_mix
        if isinstance(row, Mapping)
    }
    if periods != expected_periods:
        raise RevisionBundlePublishError(
            "mix ANBIMA sem as quatro competências editoriais"
        )
    shares_by_period: dict[str, float] = {}
    for row in type_mix:
        if not isinstance(row, Mapping):
            continue
        period = str(row.get("competencia") or "")
        shares_by_period[period] = shares_by_period.get(period, 0.0) + float(
            row.get("share") or 0.0
        )
    if any(abs(total - 1.0) > 1e-8 for total in shares_by_period.values()):
        raise RevisionBundlePublishError(
            "participações do mix ANBIMA não fecham 100% por competência"
        )
    carteira_history = payload.get("carteira_1_taxonomy_history")
    if not isinstance(carteira_history, list) or len(carteira_history) != 16:
        raise RevisionBundlePublishError(
            "carteira_1_taxonomy_history deve conter quatro categorias em quatro competências"
        )
    if {
        str(row.get("competencia") or "") for row in carteira_history
        if isinstance(row, Mapping)
    } != expected_periods:
        raise RevisionBundlePublishError("Carteira 1 sem as quatro competências editoriais")
    if {
        str(row.get("anbima_tipo") or "") for row in carteira_history
        if isinstance(row, Mapping)
    } != expected_categories:
        raise RevisionBundlePublishError("Carteira 1 diverge das quatro categorias analíticas")
    for period in expected_periods:
        period_rows = [
            row for row in carteira_history
            if isinstance(row, Mapping) and str(row.get("competencia") or "") == period
        ]
        if not math.isclose(
            sum(float(row.get("portfolio_share") or 0.0) for row in period_rows),
            1.0,
            abs_tol=1e-8,
        ):
            raise RevisionBundlePublishError(f"mix da Carteira 1 não fecha 100% em {period}")
        if not math.isclose(
            sum(float(row.get("market_share") or 0.0) for row in period_rows),
            1.0,
            abs_tol=1e-8,
        ):
            raise RevisionBundlePublishError(f"mix de mercado da Carteira 1 não fecha 100% em {period}")
    if not isinstance(payload.get("carteira_1_taxonomy_summary"), Mapping):
        raise RevisionBundlePublishError("payload sem carteira_1_taxonomy_summary")
    for key in (
        "holder_distribution_history",
        "type_mix_history",
        "carteira_1_taxonomy_history",
        "receivables_history",
        "provider_concentration_history",
        "provider_historical_ranking",
        "market_share_scope_summary",
        "atlantico_history",
        "delinquency_single_receivable",
        "delinquency_frozen_cohort_history",
        "delinquency_frozen_cohort_summary",
        "delinquency_cohort_revision_transitions",
        "delinquency_cohort_revision_sensitivity",
        "delinquency_dispersion",
        "acquiring_curation_detail",
        "card_taxonomy_audit",
        "acquiring_anbima_review",
        "taxonomy_top15",
        "top20_by_anbima_type",
        "top20_by_anbima_type_coverage",
        "top20_taxonomy_review",
        "top100_outros_review",
        "numeric_locale_audit",
        "provider_independent_ranking",
        "bank_fidc_evolution",
        "bank_fidc_detail",
        "btg_provider_ex_controlled_scenario",
        "acquiring_reclassified_mix",
        "closed_offers_annual",
        "closed_offers_monthly",
        "closed_offers_jan_june",
        "closed_offers_jan_may",
        "closed_offer_ticket_distribution",
        "closed_offer_placement_regime",
        "closed_offer_originators_2026",
        "closed_offer_top15",
        "closed_offer_top15_summary",
        "top20_outros_regulation_review",
        "fixed_income_offer_comparison",
        "market_offer_reconciliation",
        "issuance_taxonomy",
        "issuance_taxonomy_table",
        "issuance_taxonomy_reconciliation",
        "bcb_expanded_credit",
        "provider_history_cvm_coverage",
        "provider_history_cvm_links",
        "provider_history_cvm_detail",
    ):
        rows = payload.get(key)
        if not isinstance(rows, list) or not rows:
            raise RevisionBundlePublishError(f"payload editorial sem {key}")
    required_columns = {
        "delinquency_single_receivable": {
            "tipo_recebivel_tabela_ii",
            "fundos_incluidos",
            "pl_incluido_brl",
            "inadimplencia_sobre_pl",
        },
        "delinquency_frozen_cohort_history": {
            "competencia",
            "tipo_recebivel_tabela_ii",
            "fundos_incluidos",
            "pl_incluido_brl",
            "inadimplencia_sobre_carteira",
            "fundos_coorte",
            "pl_coorte_referencia_brl",
        },
        "delinquency_frozen_cohort_summary": {
            "competencia",
            "fundos_incluidos",
            "pl_incluido_brl",
            "inadimplencia_sobre_carteira",
            "fundos_coorte",
            "pl_coorte_referencia_brl",
            "regra",
            "fonte",
        },
        "delinquency_cohort_revision_transitions": {
            "subtipo_anterior",
            "subtipo_atual",
            "fundos",
            "pl_atual_brl",
            "principais_fundos",
            "competencia_anterior",
            "competencia_atual",
        },
        "delinquency_cohort_revision_sensitivity": {
            "competencia",
            "tipo_recebivel_tabela_ii",
            "inadimplencia_sobre_carteira_coorte_anterior",
            "inadimplencia_sobre_carteira_coorte_atual",
            "delta_inadimplencia_pp",
            "competencia_coorte_anterior",
            "competencia_coorte_atual",
        },
        "delinquency_dispersion": {
            "competencia",
            "tipo_recebivel_tabela_ii",
            "fundos_reportantes_inadimplencia",
            "inadimplencia_total_subcategoria_brl",
            "top1_inadimplencia_brl",
            "top1_share",
            "top3_inadimplencia_brl",
            "top3_share",
            "top5_inadimplencia_brl",
            "top5_share",
            "hhi",
            "gini",
            "leitura_concentracao",
            "fonte",
        },
        "card_taxonomy_audit": {
            "ordem_materialidade",
            "cnpj_fundo_formatado",
            "cnpj_fundo_identificado",
            "denominacao",
            "criterio_inclusao",
            "categoria_tabela_ii",
            "valor_cartao_tabela_ii_brl",
            "pl_jun25_brl",
            "pl_jun25_observavel",
            "pl_referencia_brl",
            "pl_referencia_competencia",
            "status_curadoria",
            "decisao_curadoria",
            "cedente_originador",
            "devedor_sacado",
            "instrumento",
            "natureza_economica",
            "evidencia_curta",
            "fonte_url",
            "anbima_tipo",
            "anbima_foco",
            "anbima_cartao_explicito",
            "ja_curado_como_adquirencia",
            "consistencia_decisao_reclassificacao",
        },
        "acquiring_anbima_review": {
            "cnpj_fundo_formatado",
            "denominacao",
            "tipo_anbima_atual",
            "foco_anbima_atual",
            "categoria_referencia_sugerida",
            "base_alterada",
            "criterio_sugestao",
        },
        "taxonomy_top15": {
            "visao",
            "rank",
            "cnpj_fundo",
            "denominacao",
            "taxonomia_atual",
            "pl_brl",
            "competencia",
            "fonte",
            "metodologia",
        },
        "top20_by_anbima_type": {
            "tipo_exibicao",
            "rank_tipo",
            "cnpj_fundo",
            "denominacao",
            "pl",
            "competencia_pl",
            "pl_anterior_positivo",
            "administrador",
            "gestor",
            "custodiante",
            "cedente_originador",
            "cedente_status",
            "regulamento_id",
            "regulamento_data",
            "regulamento_url",
            "pagina_clausula",
            "evidencia_cedente",
            "limitacao_cedente",
        },
        "top20_by_anbima_type_coverage": {
            "tipo_exibicao",
            "fundos",
            "administrador_preenchido",
            "gestor_preenchido",
            "custodiante_preenchido",
            "cedente_curadoria_concluida",
            "regulamento_local_sem_curadoria",
            "sem_regulamento_local",
            "competencia_pl",
            "competencia_anterior_verificada",
            "fundos_pl_anterior_positivo",
        },
        "top100_outros_review": {
            "competencia_pl",
            "rank_outros_slide",
            "cnpj_fundo",
            "denominacao",
            "pl",
            "bucket_slide_atual",
            "anbima_tipo_oficial",
            "anbima_foco_oficial",
            "tabela_ii_reportada",
            "tabela_ii_dominante",
            "tabela_ii_multisegmento",
            "documento_id_base",
            "documento_data_base",
            "documento_url_base",
            "evidencia_documental",
            "cedente_originador_expresso",
            "tipo_anbima_sugerido",
            "foco_anbima_sugerido",
            "tabela_ii_sugerida",
            "perimeter_proposal",
            "is_fic_fidc_sugerido",
            "pl_correcao_perimetro_candidata_brl",
            "confianca_base",
            "status_revisao_base",
            "motivo_validacao_manual_base",
            "acao_status",
            "anbima_tipo_curado",
            "anbima_foco_curado",
            "tabela_ii_curada",
            "taxonomy_review_applied",
        },
        "acquiring_curation_detail": {
            "ordem_materialidade",
            "cnpj_fundo_formatado",
            "denominacao",
            "pl_referencia_brl",
            "pl_referencia_competencia",
            "natureza_economica",
            "categoria_tabela_ii",
            "anbima_tipo",
            "anbima_foco",
            "fonte_url",
        },
        "provider_independent_ranking": {
            "competencia",
            "papel",
            "participante",
            "rank_independente",
            "rank_geral",
            "pl_brl",
            "selected_latest_top_n",
        },
        "bank_fidc_evolution": {
            "competencia",
            "grupo_bancario",
            "pl_bruto_brl",
            "pl_brl_raw",
            "pl_recovered_official",
            "pl_display_suffix",
            "pl_source_references",
            "is_total_5_banks",
            "observado",
        },
        "bank_fidc_detail": {
            "competencia",
            "grupo_bancario",
            "cnpj_fundo",
            "denominacao",
            "pl_brl",
            "pl_brl_raw",
            "pl_recovered_official",
            "pl_display_suffix",
            "pl_source_reference",
            "observado",
        },
        "btg_provider_ex_controlled_scenario": {
            "competencia",
            "papel",
            "btg_pl_brl",
            "btg_rank",
            "fidcs_controlados_excluidos",
            "pl_controlado_excluido_brl",
            "btg_pl_ex_controlados_brl",
            "btg_rank_ex_controlados",
            "regra",
            "fonte",
        },
        "acquiring_reclassified_mix": {
            "competencia",
            "categoria_analitica",
            "pl_brl",
            "share_pl",
        },
        "closed_offers_annual": {
            "year",
            "closed_offers",
            "registered_volume_brl",
            "mean_registered_ticket_brl",
            "median_registered_ticket_brl",
            "natural_person_placed_volume_share",
            "placed_quantity_registered_volume_coverage",
            "professional_target_registered_volume_share",
        },
        "closed_offers_jan_may": {
            "year",
            "closed_offers",
            "registered_volume_brl",
            "mean_registered_ticket_brl",
        },
        "closed_offers_jan_june": {
            "year",
            "closed_offers",
            "registered_volume_brl",
            "mean_registered_ticket_brl",
        },
        "closed_offers_monthly": {
            "year",
            "month",
            "registered_volume_brl",
        },
        "closed_offer_ticket_distribution": {
            "period_label",
            "period_start",
            "period_end",
            "ticket_bucket",
            "closed_offers",
            "offer_share",
            "registered_volume_brl",
            "registered_volume_share",
            "period_mean_ticket_brl",
            "period_median_ticket_brl",
        },
        "closed_offer_originators_2026": {
            "rank",
            "originator_group",
            "closed_offers",
            "registered_volume_brl",
            "mean_registered_ticket_brl",
            "identified_registered_volume_coverage",
            "identified_registered_volume_brl",
            "confidence",
            "share_of_total_registered_volume",
        },
        "closed_offer_top15": {
            "period_label",
            "rank",
            "offer_id",
            "data_encerramento",
            "cnpj_emissor",
            "nome_emissor",
            "fund_name_short",
            "originator_group",
            "registered_volume_brl",
            "leader_name",
            "ibba_coord_lead",
            "ibba_coord_lead_label",
            "ibba_participant",
            "ibba_participant_label",
            "ibba_participant_entities",
            "ibba_participant_roles",
            "ibba_participation_source",
            "participants_source_url",
            "closing_document_url",
            "distribution_regime",
            "firm_commitment",
            "firm_commitment_label",
            "publico",
            "investor_count",
            "investor_categories",
            "coordinator_entities",
            "firm_commitment_coordinators",
            "firm_commitment_amount_by_coordinator",
            "firm_commitment_source_limitation",
            "rating_agency",
            "rating_assigned",
            "rating_scope",
            "rating_source_type",
            "rating_source_url",
            "rating_match_status",
            "rating_evidence",
            "rating_availability_status",
            "rating_limitation",
            "metadata_matched",
            "status",
            "offer_type",
            "security",
            "source_url",
            "scope",
        },
        "closed_offer_top15_summary": {
            "period_label",
            "period_closed_offers",
            "period_registered_volume_brl",
            "top15_offers",
            "top15_registered_volume_brl",
            "top15_share_of_period_volume",
            "ibba_lead_offers_top15",
            "ibba_lead_volume_top15_brl",
            "ibba_lead_share_top15_volume",
            "ibba_participation_offers_top15",
            "ibba_participation_volume_top15_brl",
            "ibba_participation_share_top15_volume",
            "firm_commitment_offers_top15",
            "firm_commitment_volume_top15_brl",
            "ibba_firm_commitment_offers_top15",
            "ibba_firm_commitment_volume_top15_brl",
            "investor_count_methodology",
            "ranking_methodology",
            "automatic_rite_registered_volume_share",
            "comparability_status",
            "coverage_note",
        },
        "fixed_income_offer_comparison": {
            "view",
            "series_label",
            "period_label",
            "registered_volume_brl",
            "previous_registered_volume_brl",
            "yoy_growth",
            "yoy_comparable",
            "universe_registered_volume_brl",
            "source_url",
            "scope",
            "excluded_instruments",
        },
        "market_offer_reconciliation": {
            "period_label",
            "instrument_label",
            "cvm_registered_volume_brl",
            "cvm_harmonization_volume_brl",
            "cvm_harmonized_volume_brl",
            "anbima_closed_volume_brl",
            "raw_gap_pct",
            "harmonized_gap_pct",
            "primary_explanation",
            "cvm_source_url",
            "anbima_source_url",
            "anbima_source_snapshot",
            "limitation",
        },
        "issuance_taxonomy": {
            "period_key",
            "period_label",
            "categoria",
            "volume_brl",
            "share",
        },
        "issuance_taxonomy_reconciliation": {
            "period_key",
            "period_label",
            "total_brl",
            "fic_excluded_brl",
            "emitted_volume_brl",
        },
        "bcb_expanded_credit": {
            "competencia",
            "period_label",
            "expanded_credit_total_brl",
            "private_expanded_credit_total_brl",
            "loans_brl",
            "public_debt_brl",
            "private_debt_brl",
            "fidc_receivables_brl",
            "other_securitization_brl",
            "external_debt_brl",
            "source_bcb",
            "source_cvm",
            "methodology",
        },
        "carteira_1_taxonomy_history": {
            "competencia",
            "period_label",
            "anbima_tipo",
            "portfolio_pl_brl",
            "portfolio_share",
            "portfolio_funds",
            "portfolio_total_brl",
            "scope_cnpjs",
            "observed_cnpjs",
            "coverage_scope_share",
            "market_pl_brl",
            "market_share",
            "market_total_brl",
            "portfolio_growth_since_start",
            "market_growth_since_start",
        },
        "closed_offer_placement_regime": {
            "period_label",
            "placement_regime",
            "closed_offers",
            "closed_offers_share",
            "registered_volume_brl",
            "registered_volume_share",
            "period_closed_offers",
            "period_registered_volume_brl",
            "source_url",
            "scope",
            "methodology",
        },
        "provider_history_cvm_coverage": {
            "papel",
            "data_referencia",
            "fundos_coorte",
            "pl_coorte_mai26_brl",
            "fundos_resolvidos_unicos",
            "pl_resolvido_unico_brl",
            "cobertura_fundos_resolvida",
            "cobertura_pl_resolvida",
            "escopo_fonte",
        },
        "provider_history_cvm_links": {
            "papel",
            "data_origem",
            "data_destino",
            "origem_prestador_grupo",
            "destino_prestador_grupo",
            "fundos",
            "pl_mai26_brl",
            "share_pl_comparavel",
            "escopo_fonte",
        },
        "provider_history_cvm_detail": {
            "papel",
            "data_origem",
            "data_destino",
            "cnpj_fundo",
            "denominacao",
            "pl_mai26_brl",
            "origem_prestador_grupo",
            "destino_prestador_grupo",
        },
    }
    for key, columns in required_columns.items():
        rows = payload.get(key)
        if not isinstance(rows, list):
            raise RevisionBundlePublishError(f"payload editorial sem {key}")
        for index, row in enumerate(rows, start=1):
            if not isinstance(row, Mapping):
                raise RevisionBundlePublishError(
                    f"payload {key} contém linha {index} inválida"
                )
            missing_columns = sorted(columns.difference(row))
            if missing_columns:
                raise RevisionBundlePublishError(
                    f"payload {key} linha {index} sem colunas obrigatórias: "
                    + ", ".join(missing_columns)
                )
    issuance_rows = list(payload.get("issuance_taxonomy") or [])
    issuance_reconciliation = {
        str(row.get("period_key") or ""): row
        for row in list(payload.get("issuance_taxonomy_reconciliation") or [])
        if isinstance(row, Mapping)
    }
    expected_issuance_periods = {"2023", "2024", "2025", "jun25", "jun26"}
    expected_issuance_categories = {
        "Fomento Mercantil",
        "Agro, Indústria e Comércio",
        "Financeiro",
        "Outros",
    }
    if set(issuance_reconciliation) != expected_issuance_periods:
        raise RevisionBundlePublishError(
            "emissões por taxonomia devem conter cinco períodos reconciliados"
        )
    issuance_by_period: dict[str, list[Mapping[str, object]]] = {}
    for row in issuance_rows:
        issuance_by_period.setdefault(str(row.get("period_key") or ""), []).append(row)
    for period_key in expected_issuance_periods:
        period_rows = issuance_by_period.get(period_key, [])
        if len(period_rows) != 4:
            raise RevisionBundlePublishError(
                f"emissões por taxonomia em {period_key} devem conter quatro tipos"
            )
        if {str(row.get("categoria") or "") for row in period_rows} != expected_issuance_categories:
            raise RevisionBundlePublishError(
                f"emissões por taxonomia em {period_key} divergem das categorias ANBIMA"
            )
        share_total = sum(float(row.get("share") or 0.0) for row in period_rows)
        if not math.isclose(share_total, 1.0, abs_tol=1e-8):
            raise RevisionBundlePublishError(
                f"shares de emissões por taxonomia em {period_key} não fecham 100%"
            )
        four_types = sum(float(row.get("volume_brl") or 0.0) for row in period_rows)
        reconciliation = issuance_reconciliation[period_key]
        if not math.isclose(
            four_types,
            float(reconciliation.get("total_brl") or 0.0),
            abs_tol=0.01,
        ):
            raise RevisionBundlePublishError(
                f"quatro tipos ANBIMA não reconciliam em {period_key}"
            )
        emitted = four_types + float(reconciliation.get("fic_excluded_brl") or 0.0)
        if not math.isclose(
            emitted,
            float(reconciliation.get("emitted_volume_brl") or 0.0),
            abs_tol=0.01,
        ):
            raise RevisionBundlePublishError(
                f"quatro tipos ANBIMA + FIC-FIDC não reconciliam com o volume emitido em {period_key}"
            )
    top20_by_type = list(payload.get("top20_by_anbima_type") or [])
    if len(top20_by_type) != 80:
        raise RevisionBundlePublishError(
            "top20_by_anbima_type deve conter exatamente 80 linhas"
        )
    expected_types = {
        "Fomento Mercantil",
        "Agro, Indústria e Comércio",
        "Financeiro",
        "Outros",
    }
    observed_types = {str(row.get("tipo_exibicao") or "") for row in top20_by_type}
    if observed_types != expected_types:
        raise RevisionBundlePublishError(
            "top20_by_anbima_type deve conter as quatro categorias do slide 8"
        )
    for type_name in expected_types:
        group = sorted(
            (row for row in top20_by_type if row.get("tipo_exibicao") == type_name),
            key=lambda row: int(row.get("rank_tipo") or 0),
        )
        if len(group) != 20 or [int(row.get("rank_tipo") or 0) for row in group] != list(range(1, 21)):
            raise RevisionBundlePublishError(
                f"ranking Top 20 inválido em {type_name}"
            )
        if any(str(row.get("competencia_pl") or "") != latest_complete for row in group):
            raise RevisionBundlePublishError(
                f"competência do Top 20 diverge em {type_name}"
            )
        if any(float(row.get("pl") or 0) <= 0 for row in group):
            raise RevisionBundlePublishError(
                f"Top 20 contém PL ausente ou não positivo em {type_name}"
            )
    if len({str(row.get("cnpj_fundo") or "") for row in top20_by_type}) != 80:
        raise RevisionBundlePublishError("Top 20 por Tipo contém CNPJ duplicado")

    historical_top20 = list(payload.get("top20_taxonomy_review") or [])
    if len(historical_top20) != 320:
        raise RevisionBundlePublishError(
            "top20_taxonomy_review deve conter 320 linhas"
        )
    historical_periods = {
        str(row.get("competencia") or "") for row in historical_top20
    }
    if historical_periods != set(expected_periods):
        raise RevisionBundlePublishError(
            "top20_taxonomy_review diverge das quatro competências editoriais"
        )
    historical_groups: dict[tuple[str, str], int] = {}
    review_ids: set[str] = set()
    for row in historical_top20:
        group = (
            str(row.get("competencia") or ""),
            str(row.get("tipo_exibicao") or ""),
        )
        historical_groups[group] = historical_groups.get(group, 0) + 1
        review_id = str(row.get("review_id") or "")
        if not re.fullmatch(r"[0-9]{14}", review_id):
            raise RevisionBundlePublishError(
                "top20_taxonomy_review contém review_id inválido"
            )
        if review_id != str(row.get("cnpj_fundo") or ""):
            raise RevisionBundlePublishError(
                "top20_taxonomy_review diverge entre review_id e CNPJ"
            )
        review_ids.add(review_id)
    if len(historical_groups) != 16 or set(historical_groups.values()) != {20}:
        raise RevisionBundlePublishError(
            "top20_taxonomy_review deve conter 20 fundos por Tipo e competência"
        )
    historical_cnpjs = {
        str(row.get("cnpj_fundo") or "") for row in historical_top20
    }
    if review_ids != historical_cnpjs:
        raise RevisionBundlePublishError(
            "top20_taxonomy_review deve reutilizar um review_id por CNPJ"
        )

    top100_outros = list(payload.get("top100_outros_review") or [])
    if len(top100_outros) != 100:
        raise RevisionBundlePublishError(
            "top100_outros_review deve conter exatamente 100 linhas"
        )
    if [int(row.get("rank_outros_slide") or 0) for row in top100_outros] != list(range(1, 101)):
        raise RevisionBundlePublishError(
            "top100_outros_review deve usar ranks sequenciais de 1 a 100"
        )
    if any(str(row.get("bucket_slide_atual") or "") != "Outros" for row in top100_outros):
        raise RevisionBundlePublishError(
            "top100_outros_review contém fundo fora do bucket exibido no slide 8"
        )
    summary_outros = payload.get("top100_outros_summary")
    if not isinstance(summary_outros, Mapping):
        raise RevisionBundlePublishError("payload sem top100_outros_summary")
    required_outros_summary = {
        "outros_oficial_brl",
        "outros_curado_brl",
        "reducao_aprovada_brl",
        "top100_outros_brl",
        "candidatos_documentais_brl",
        "candidatos_reclassificacao_tipo_brl",
        "candidatos_correcao_perimetro_brl",
        "outros_pos_candidatos_brl",
        "residual_minimo_top100_brl",
        "gap_meta_minimo_top100_brl",
        "meta_atingivel_top100",
    }
    if missing := sorted(required_outros_summary.difference(summary_outros)):
        raise RevisionBundlePublishError(
            "top100_outros_summary sem campos obrigatórios: " + ", ".join(missing)
        )
    review_meta = payload.get("taxonomy_review_meta")
    if (
        not isinstance(review_meta, Mapping)
        or not re.fullmatch(
            r"[0-9a-f]{64}", str(review_meta.get("ledger_sha256") or "")
        )
        or not re.fullmatch(
            r"[0-9a-f]{64}", str(review_meta.get("audit_sha256") or "")
        )
        or not str(review_meta.get("ledger_path") or "").strip()
        or not str(review_meta.get("audit_path") or "").strip()
    ):
        raise RevisionBundlePublishError("metadata do ledger de taxonomia inválida")
    exclusions = payload.get("market_share_exclusions")
    if not isinstance(exclusions, list) or len(exclusions) != 2:
        raise RevisionBundlePublishError(
            "payload editorial sem market_share_exclusions (duas exclusões esperadas)"
        )
    acquiring = payload.get("acquiring_taxonomy")
    if not isinstance(acquiring, Mapping) or not isinstance(acquiring.get("funds"), list):
        raise RevisionBundlePublishError("payload editorial sem acquiring_taxonomy")
    if not isinstance(payload.get("atlantico_profile"), Mapping):
        raise RevisionBundlePublishError("payload editorial sem atlantico_profile")
    if not isinstance(payload.get("delinquency_single_receivable_summary"), Mapping):
        raise RevisionBundlePublishError(
            "payload editorial sem delinquency_single_receivable_summary"
        )
    if not isinstance(payload.get("delinquency_dispersion_summary"), Mapping):
        raise RevisionBundlePublishError(
            "payload editorial sem delinquency_dispersion_summary"
        )
    if not isinstance(payload.get("acquiring_anbima_review_summary"), Mapping):
        raise RevisionBundlePublishError(
            "payload editorial sem acquiring_anbima_review_summary"
        )
    summary = payload["delinquency_single_receivable_summary"]
    required_summary = {
        "fundos_universo_ex_fic_pl_positivo",
        "pl_universo_ex_fic_positivo_brl",
        "fundos_incluidos",
        "pl_incluido_brl",
        "cobertura_pl",
        "fundos_multitipo_excluidos",
        "pl_multitipo_excluido_brl",
        "fundos_sem_tipo_excluidos",
        "pl_sem_tipo_excluido_brl",
        "fundos_inad_supera_carteira_excluidos",
        "pl_inad_supera_carteira_excluido_brl",
        "fundos_fic_excluidos",
        "pl_fic_excluido_brl",
    }
    missing_summary = sorted(required_summary.difference(summary))
    if missing_summary:
        raise RevisionBundlePublishError(
            "payload delinquency_single_receivable_summary sem campos obrigatórios: "
            + ", ".join(missing_summary)
        )
    if len(payload.get("closed_offers_annual") or []) != 5:
        raise RevisionBundlePublishError(
            "payload editorial deve conter ofertas anuais de 2022 a 2026"
        )
    for key in (
        "provider_transition_summary",
        "reag_admin_summary",
        "provider_leadership_attribution",
    ):
        if not isinstance(payload.get(key), Mapping):
            raise RevisionBundlePublishError(f"payload editorial sem {key}")
    if not isinstance(payload.get("conclusion_metrics"), Mapping):
        raise RevisionBundlePublishError("payload editorial sem conclusion_metrics")
    cohort_revision = payload.get("delinquency_cohort_revision_summary")
    if not isinstance(cohort_revision, Mapping):
        raise RevisionBundlePublishError(
            "payload editorial sem delinquency_cohort_revision_summary"
        )
    required_cohort_revision = {
        "competencia_anterior",
        "competencia_atual",
        "fundos_coorte_anterior",
        "fundos_coorte_atual",
        "fundos_reclassificados",
        "fundos_entraram",
        "fundos_sairam",
    }
    if missing := sorted(required_cohort_revision.difference(cohort_revision)):
        raise RevisionBundlePublishError(
            "payload delinquency_cohort_revision_summary sem campos obrigatórios: "
            + ", ".join(missing)
        )
    _validate_card_taxonomy_contract(payload)
    _validate_closed_offer_originator_order(payload)
    _validate_closed_offer_top15(payload)
    _validate_fixed_income_offer_comparison(payload)
    _validate_market_offer_reconciliation(payload)
    _validate_closed_offer_placement_regime(payload)
    conclusion_metrics = payload["conclusion_metrics"]
    required_btg_metrics = {
        "btg_bank_cohort_listed_roots",
        "btg_bank_cohort_observed_funds",
        "btg_bank_cohort_pl_brl",
        "btg_bank_cohort_combo_funds",
        "btg_bank_cohort_combo_pl_brl",
    }
    if missing := sorted(required_btg_metrics.difference(conclusion_metrics)):
        raise RevisionBundlePublishError(
            "payload conclusion_metrics sem coorte bancária BTG: "
            + ", ".join(missing)
        )
    for key in (
        "provider_transition_links",
        "provider_transition_detail",
        "provider_transition_role_availability",
        "reag_admin_links",
        "reag_admin_detail",
        "btg_controlled_reconciliation",
        "qi_legacy_attribution",
    ):
        rows = payload.get(key)
        if not isinstance(rows, list) or not rows:
            raise RevisionBundlePublishError(f"payload editorial sem {key}")


_MONTH_ABBR = (
    "jan",
    "fev",
    "mar",
    "abr",
    "mai",
    "jun",
    "jul",
    "ago",
    "set",
    "out",
    "nov",
    "dez",
)


def _competence_label(competence: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{2})", competence)
    if not match:
        raise RevisionBundlePublishError(f"competência inválida: {competence}")
    year, month = int(match.group(1)), int(match.group(2))
    if month not in range(1, 13):
        raise RevisionBundlePublishError(f"competência inválida: {competence}")
    return f"{_MONTH_ABBR[month - 1]}/{str(year)[-2:]}"


def validate_deck_snapshot(payload: bytes, latest_complete: str) -> None:
    """Guard against publishing a deck whose visible snapshot is hardcoded."""

    expected = _competence_label(latest_complete).casefold()
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        slide_xml = b"".join(
            archive.read(name)
            for name in archive.namelist()
            if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
        )
    visible = slide_xml.decode("utf-8", errors="ignore").casefold()
    if expected not in visible:
        raise RevisionBundlePublishError(
            f"PPTX não contém a competência publicada ({expected})"
        )


def collect_input_hashes(
    *,
    data_dir: Path,
    curation_path: Path,
    input_workbook: Path,
    artifact_script: Path = ARTIFACT_SCRIPT,
) -> dict[str, str]:
    """Hash every external input that can change the published bundle."""

    data_dir = Path(data_dir)
    paths: list[tuple[str, Path]] = []
    for name in REQUIRED_DATA_INPUTS:
        path = data_dir / name
        if not path.exists():
            raise RevisionBundlePublishError(f"input obrigatório ausente: {path}")
        paths.append((f"data/{name}", path))
    for name in OPTIONAL_DATA_INPUTS:
        path = data_dir / name
        if path.exists():
            paths.append((f"data/{name}", path))
    if not Path(curation_path).exists():
        raise RevisionBundlePublishError(f"curadoria Top 20 ausente: {curation_path}")
    if not Path(input_workbook).exists():
        raise RevisionBundlePublishError(f"workbook-base ausente: {input_workbook}")
    if not Path(artifact_script).exists():
        raise RevisionBundlePublishError(f"renderer ausente: {artifact_script}")
    paths.extend(
        [
            ("curation/top20.csv", Path(curation_path)),
            ("workbook/input.xlsx", Path(input_workbook)),
        ]
    )
    for path in BUILDER_SOURCES:
        paths.append((f"builder/{path.name}", path))
    if Path(artifact_script).resolve() not in {path.resolve() for _, path in paths}:
        paths.append((f"builder/{Path(artifact_script).name}", Path(artifact_script)))
    return {label: _sha256_semantic_file(path) for label, path in sorted(paths)}


def _artifact_runtime_metadata(
    node: Path,
    node_modules: Path,
    *,
    artifact_script: Path = ARTIFACT_SCRIPT,
) -> dict[str, str]:
    package_path = node_modules / "@oai" / "artifact-tool" / "package.json"
    package = json.loads(package_path.read_text(encoding="utf-8"))
    completed = subprocess.run(
        [str(node), "--version"],
        check=True,
        capture_output=True,
        text=True,
        timeout=15,
    )
    return {
        "node_version": completed.stdout.strip(),
        "artifact_tool_version": str(package.get("version") or "desconhecida"),
        "renderer_sha256": _sha256_file(artifact_script),
    }


def build_bundle_manifest(
    *,
    payload_bytes: bytes,
    payload: Mapping[str, object],
    analysis_manifest_bytes: bytes,
    pptx_bytes: bytes,
    xlsx_bytes: bytes,
    input_hashes: Mapping[str, str],
    renderer: Mapping[str, str],
    generated_at_utc: str,
    html_bytes: bytes = b"",
    portfolio_xlsx_bytes: bytes = b"",
) -> dict[str, object]:
    """Build the content-addressed manifest consumed by the application."""

    payload_hash = _sha256_bytes(payload_bytes)
    pptx_hash = _sha256_bytes(pptx_bytes)
    xlsx_hash = _sha256_bytes(xlsx_bytes)
    portfolio_xlsx_hash = _sha256_bytes(portfolio_xlsx_bytes)
    html_hash = _sha256_bytes(html_bytes)
    input_signature = _sha256_bytes(_canonical_json_bytes(dict(input_hashes)))
    bundle_id = (
        str(payload.get("latest_complete") or "unknown").replace("-", "")
        + "_"
        + payload_hash[:16]
    )
    return {
        "schema_version": BUNDLE_SCHEMA,
        "bundle_id": bundle_id,
        "generated_at_utc": generated_at_utc,
        "latest_complete": str(payload.get("latest_complete") or ""),
        "offers_as_of": str(payload.get("offers_as_of") or ""),
        "payload_schema": str(payload.get("schema_version") or ""),
        # Kept for the read-path contract; input_signature is the true source hash.
        "source_signature": payload_hash,
        "input_signature": input_signature,
        "inputs": dict(input_hashes),
        "renderer": dict(renderer),
        "renderer_version": str(renderer.get("renderer_version") or ""),
        "renderer_sha256": str(renderer.get("renderer_sha256") or ""),
        "payload_sha256": payload_hash,
        "payload": {
            "name": PAYLOAD_NAME,
            "sha256": payload_hash,
            "bytes": len(payload_bytes),
        },
        "analysis_manifest": {
            "name": ANALYSIS_MANIFEST_NAME,
            "sha256": _sha256_bytes(analysis_manifest_bytes),
            "bytes": len(analysis_manifest_bytes),
        },
        "pptx": {
            "name": MATERIALIZED_PPTX_NAME,
            "sha256": pptx_hash,
            "bytes": len(pptx_bytes),
        },
        "xlsx": {
            "name": MATERIALIZED_XLSX_NAME,
            "sha256": xlsx_hash,
            "bytes": len(xlsx_bytes),
        },
        "portfolio_xlsx": {
            "name": MATERIALIZED_PORTFOLIO_XLSX_NAME,
            "sha256": portfolio_xlsx_hash,
            "bytes": len(portfolio_xlsx_bytes),
        },
        "html": {
            "name": MATERIALIZED_HTML_NAME,
            "sha256": html_hash,
            "bytes": len(html_bytes),
        },
        "checks": {
            "slides": EXPECTED_SLIDES,
            "top20_fidcs": len(list(payload.get("top20_fidcs") or [])),
            "top20_outros": len(list(payload.get("top20_outros") or [])),
            "profiles": len(list(payload.get("profiles") or [])),
            "top20_by_anbima_type": len(
                list(payload.get("top20_by_anbima_type") or [])
            ),
            "top20_taxonomy_review": len(
                list(payload.get("top20_taxonomy_review") or [])
            ),
            "top100_outros_review": len(
                list(payload.get("top100_outros_review") or [])
            ),
            "portfolio_export_carteira_101": len(
                list(payload.get("portfolio_export_carteira_101") or [])
            ),
            "portfolio_export_flagships": len(
                list(payload.get("portfolio_export_flagships") or [])
            ),
            "portfolio_export_coverage": len(
                list(payload.get("portfolio_export_coverage") or [])
            ),
            "portfolio_export_gaps": len(
                list(payload.get("portfolio_export_gaps") or [])
            ),
        },
    }


def validate_bundle_manifest(
    manifest: Mapping[str, object],
    *,
    payload_bytes: bytes,
    payload: Mapping[str, object],
    analysis_manifest_bytes: bytes,
    pptx_bytes: bytes,
    xlsx_bytes: bytes,
    html_bytes: bytes = b"",
    portfolio_xlsx_bytes: bytes = b"",
) -> None:
    if manifest.get("schema_version") != BUNDLE_SCHEMA:
        raise RevisionBundlePublishError("schema do manifest de publicação incompatível")
    if manifest.get("payload_schema") != payload.get("schema_version"):
        raise RevisionBundlePublishError("schema do payload diverge do bundle")
    if manifest.get("latest_complete") != payload.get("latest_complete"):
        raise RevisionBundlePublishError("competência do payload diverge do bundle")
    expected = {
        "payload_sha256": _sha256_bytes(payload_bytes),
        "source_signature": _sha256_bytes(payload_bytes),
    }
    for key, value in expected.items():
        if manifest.get(key) != value:
            raise RevisionBundlePublishError(f"hash inválido no manifest: {key}")
    files = (
        ("payload", payload_bytes),
        ("analysis_manifest", analysis_manifest_bytes),
        ("pptx", pptx_bytes),
        ("xlsx", xlsx_bytes),
        ("portfolio_xlsx", portfolio_xlsx_bytes),
        ("html", html_bytes),
    )
    for key, content in files:
        entry = dict(manifest.get(key) or {})
        if entry.get("sha256") != _sha256_bytes(content):
            raise RevisionBundlePublishError(f"hash inválido no manifest: {key}")
        if entry.get("bytes") is None or int(entry["bytes"]) != len(content):
            raise RevisionBundlePublishError(f"tamanho inválido no manifest: {key}")
    if not re.fullmatch(r"\d{6}_[0-9a-f]{16}", str(manifest.get("bundle_id") or "")):
        raise RevisionBundlePublishError("bundle_id inválido")


def validate_renderer_manifest(
    manifest: Mapping[str, object],
    *,
    payload_bytes: bytes,
    payload: Mapping[str, object],
    pptx_bytes: bytes,
    xlsx_bytes: bytes,
    renderer_sha256: str,
    html_bytes: bytes = b"",
    portfolio_xlsx_bytes: bytes = b"",
) -> None:
    """Validate the renderer's own manifest before creating the publish manifest."""

    payload_hash = _sha256_bytes(payload_bytes)
    if manifest.get("schema_version") != BUNDLE_SCHEMA:
        raise RevisionBundlePublishError("renderer produziu manifest com schema inválido")
    if manifest.get("payload_schema") != payload.get("schema_version"):
        raise RevisionBundlePublishError("renderer usou schema de payload divergente")
    if manifest.get("latest_complete") != payload.get("latest_complete"):
        raise RevisionBundlePublishError("renderer usou competência divergente")
    if manifest.get("payload_sha256") != payload_hash:
        raise RevisionBundlePublishError("renderer não reconciliou o hash do payload")
    if manifest.get("renderer_sha256") != renderer_sha256:
        raise RevisionBundlePublishError("renderer executado diverge do snapshot publicado")
    for key, content in (
        ("pptx", pptx_bytes),
        ("xlsx", xlsx_bytes),
        ("portfolio_xlsx", portfolio_xlsx_bytes),
        ("html", html_bytes),
    ):
        entry = dict(manifest.get(key) or {})
        if entry.get("sha256") != _sha256_bytes(content):
            raise RevisionBundlePublishError(f"manifest do renderer diverge em {key}")
        if entry.get("bytes") is None or int(entry["bytes"]) != len(content):
            raise RevisionBundlePublishError(f"manifest do renderer diverge em {key}")
    checks = dict(manifest.get("checks") or {})
    if int(checks.get("slides") or 0) != EXPECTED_SLIDES:
        raise RevisionBundlePublishError(
            f"manifest do renderer não contém {EXPECTED_SLIDES} slides"
        )
    if any(int(checks.get(key) or 0) != 20 for key in ("top20_fidcs", "top20_outros", "profiles")):
        raise RevisionBundlePublishError("manifest do renderer falhou nos checks Top 20")
    if int(checks.get("top20_by_anbima_type") or 0) != 80:
        raise RevisionBundlePublishError("manifest do renderer falhou no Top 20 por Tipo")
    if int(checks.get("top20_taxonomy_review") or 0) != 320:
        raise RevisionBundlePublishError(
            "manifest do renderer falhou na fila histórica Top 20"
        )
    if int(checks.get("top100_outros_review") or 0) != 100:
        raise RevisionBundlePublishError("manifest do renderer falhou na fila Top 100 Outros")
    if portfolio_xlsx_bytes:
        if int(checks.get("portfolio_export_carteira_101") or 0) != 101:
            raise RevisionBundlePublishError(
                "manifest do renderer falhou no export Carteira 101"
            )
        if int(checks.get("portfolio_export_flagships") or 0) != 47:
            raise RevisionBundlePublishError(
                "manifest do renderer falhou no export Flagships"
            )
        if int(checks.get("portfolio_export_coverage") or 0) <= 0:
            raise RevisionBundlePublishError(
                "manifest do renderer não registrou a cobertura do export estrutural"
            )


def publish_staged_bundle(
    *,
    staged_revision_dir: Path,
    staged_pptx: Path,
    staged_xlsx: Path,
    staged_bundle_manifest: Path,
    publish_dir: Path,
    staged_html: Path | None = None,
    staged_portfolio_xlsx: Path | None = None,
    replace: Callable[[str | bytes | os.PathLike[str] | os.PathLike[bytes], str | bytes | os.PathLike[str] | os.PathLike[bytes]], None] = os.replace,
) -> tuple[Path, Path, Path]:
    """Move staged outputs into place, replacing the bundle manifest last."""

    publish_dir = Path(publish_dir)
    publish_dir.mkdir(parents=True, exist_ok=True)
    for source in sorted(Path(staged_revision_dir).iterdir(), key=lambda item: item.name):
        if source.is_file() and source.name != BUNDLE_MANIFEST_NAME:
            replace(source, publish_dir / source.name)
    target_pptx = publish_dir / MATERIALIZED_PPTX_NAME
    target_xlsx = publish_dir / MATERIALIZED_XLSX_NAME
    target_manifest = publish_dir / BUNDLE_MANIFEST_NAME
    replace(staged_pptx, target_pptx)
    replace(staged_xlsx, target_xlsx)
    if staged_portfolio_xlsx is not None:
        replace(
            staged_portfolio_xlsx,
            publish_dir / MATERIALIZED_PORTFOLIO_XLSX_NAME,
        )
    if staged_html is not None:
        replace(staged_html, publish_dir / MATERIALIZED_HTML_NAME)
    # This commit marker is deliberately last.
    replace(staged_bundle_manifest, target_manifest)
    return target_pptx, target_xlsx, target_manifest


def _run_artifact_builder(
    *,
    node: Path,
    artifact_script: Path,
    provider_flow_builder: Path,
    node_modules: Path,
    data_dir: Path,
    input_workbook: Path,
    revision_dir: Path,
    payload_path: Path,
    output_dir: Path,
    pptx_path: Path,
    xlsx_path: Path,
    portfolio_xlsx_path: Path,
    html_path: Path,
    renderer_manifest_path: Path,
    timeout_seconds: int,
) -> None:
    env = os.environ.copy()
    env.update(
        {
            "CODEX_NODE_MODULES": str(node_modules),
            "FIDC_DATA_DIR": str(data_dir),
            "FIDC_INPUT_WORKBOOK": str(input_workbook),
            "FIDC_REVISION_DIR": str(revision_dir),
            "FIDC_PAYLOAD_PATH": str(payload_path),
            "FIDC_OUTPUT_DIR": str(output_dir),
            "FIDC_QA_DIR": str(output_dir / "qa"),
            "FIDC_OUTPUT_PPTX": str(pptx_path),
            "FIDC_OUTPUT_XLSX": str(xlsx_path),
            "FIDC_OUTPUT_PORTFOLIO_XLSX": str(portfolio_xlsx_path),
            "FIDC_OUTPUT_HTML": str(html_path),
            "FIDC_PROVIDER_FLOW_BUILDER": str(provider_flow_builder),
            "FIDC_EXPORT_MANIFEST": str(renderer_manifest_path),
            "FIDC_SKIP_QA": "1",
        }
    )
    phases = (
        ("PPTX", "4096", {"FIDC_SKIP_WORKBOOK": "1"}),
        (
            "XLSX",
            "8192",
            {"FIDC_SKIP_PRESENTATION": "1", "FIDC_WRITE_MANIFEST": "1"},
        ),
    )
    for phase, heap_mb, overrides in phases:
        phase_env = env | overrides
        try:
            completed = subprocess.run(
                [str(node), f"--max-old-space-size={heap_mb}", str(artifact_script)],
                cwd=ROOT,
                env=phase_env,
                check=False,
                capture_output=True,
                text=True,
                timeout=timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise RevisionBundlePublishError(
                f"renderer {phase} excedeu {timeout_seconds}s"
            ) from exc
        if completed.returncode:
            detail = (completed.stderr or completed.stdout or "falha sem log").strip()
            raise RevisionBundlePublishError(
                f"renderer {phase} do bundle falhou: " + detail[-4000:]
            )
    if (
        not pptx_path.exists()
        or not xlsx_path.exists()
        or not portfolio_xlsx_path.exists()
        or not html_path.exists()
        or not renderer_manifest_path.exists()
    ):
        raise RevisionBundlePublishError(
            "renderer não produziu PPTX/XLSX Carteira 101/HTML/manifest"
        )


def _validate_input_workbook(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            if "xl/workbook.xml" not in archive.namelist():
                raise RevisionBundlePublishError("workbook-base não é um XLSX válido")
    except zipfile.BadZipFile as exc:
        raise RevisionBundlePublishError("workbook-base não é um XLSX válido") from exc


USER_FACING_SNAPSHOT_SHEETS = (
    "PL histórico",
    "PL anual",
    "Mix ANBIMA",
    "Fila curadoria",
    "Hist cotistas",
    "Monoestrutura",
    "Rankings ANBIMA",
    "Cobertura",
    "Competências",
    "Indústria mensal",
)


def validate_user_facing_workbook_snapshot(
    payload: bytes,
    latest_complete: str,
) -> None:
    """Require inherited analytical tabs to reach the published competence.

    The reviewed renderer imports a workbook before adding its revision tabs.
    This guard prevents a valid June bundle from silently retaining May data in
    the inherited analyst-facing sheets.
    """

    from openpyxl import load_workbook

    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:  # pragma: no cover - openpyxl has many parser errors
        raise RevisionBundlePublishError(
            "workbook revisado não pôde ser auditado por competência"
        ) from exc
    try:
        missing = sorted(
            sheet for sheet in USER_FACING_SNAPSHOT_SHEETS if sheet not in workbook.sheetnames
        )
        if missing:
            raise RevisionBundlePublishError(
                "workbook revisado sem abas herdadas auditáveis: " + ", ".join(missing)
            )
        stale: list[str] = []
        for sheet_name in USER_FACING_SNAPSHOT_SHEETS:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip().casefold() for value in next(rows, ())]
            try:
                competence_index = headers.index("competencia")
            except ValueError:
                stale.append(f"{sheet_name} (sem coluna competencia)")
                continue
            competences = {
                str(row[competence_index]).strip()
                for row in rows
                if competence_index < len(row)
                and re.fullmatch(r"\d{4}-\d{2}", str(row[competence_index] or "").strip())
            }
            observed_latest = max(competences) if competences else "ausente"
            if observed_latest != latest_complete:
                stale.append(f"{sheet_name} ({observed_latest})")
        if stale:
            raise RevisionBundlePublishError(
                "abas herdadas não reconciliadas à competência publicada "
                f"{latest_complete}: "
                + ", ".join(stale)
            )
    finally:
        workbook.close()


def materialize_current_workbook_base(
    data_dir: Path,
    output_path: Path,
    latest_complete: str,
) -> Path:
    """Build the inherited workbook tabs from the same current source snapshot."""

    from services.industry_ppt_export import _build_legacy_industry_xlsx_bytes

    payload = _build_legacy_industry_xlsx_bytes(Path(data_dir))
    validate_user_facing_workbook_snapshot(payload, latest_complete)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(payload)
    return output_path


def publish_revision_bundle(
    *,
    data_dir: Path,
    publish_dir: Path,
    curation_path: Path,
    input_workbook: Path,
    latest_complete: str = "",
    raw_dir: Path = ROOT / ".cache" / "cvm-industry-study",
    provider_history_archive: Path | None = None,
    refresh_source_presence: bool = False,
    presence_months: Iterable[str] = ("all",),
    skip_download: bool = True,
    node_modules: Path | None = None,
    timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS,
    generated_at_utc: str = "",
) -> PublishedRevisionBundle:
    """Build, validate and atomically publish a complete revision snapshot."""

    data_dir = Path(data_dir).resolve()
    publish_dir = Path(publish_dir).resolve()
    curation_path = Path(curation_path).resolve()
    input_workbook = Path(input_workbook).resolve()
    raw_dir = Path(raw_dir).resolve()
    provider_history_archive = (
        Path(provider_history_archive).expanduser().resolve()
        if provider_history_archive is not None
        else raw_dir.parent / "cvm-cadastro" / "cad_fi_hist.zip"
    )
    latest_complete = latest_complete or discover_latest_complete(data_dir)
    _validate_input_workbook(input_workbook)
    validate_fic_detection_audit_provenance(data_dir)
    # Capture the long-running renderer once.  The staged build must execute
    # the exact bytes recorded in the input signature even if the worktree is
    # edited concurrently.
    artifact_script_bytes = ARTIFACT_SCRIPT.read_bytes()
    native_chart_patcher_bytes = NATIVE_CHART_PATCHER.read_bytes()
    provider_flow_builder_bytes = PROVIDER_FLOW_BUILDER.read_bytes()
    input_hashes = collect_input_hashes(
        data_dir=data_dir,
        curation_path=curation_path,
        input_workbook=input_workbook,
    )
    input_hashes[f"builder/{ARTIFACT_SCRIPT.name}"] = _sha256_bytes(
        artifact_script_bytes
    )
    input_hashes[f"builder/{PROVIDER_FLOW_BUILDER.name}"] = _sha256_bytes(
        provider_flow_builder_bytes
    )
    node_text = shutil.which("node")
    if not node_text:
        raise RevisionBundlePublishError("Node.js não localizado para o build offline")
    node = Path(node_text).resolve()
    resolved_modules = discover_artifact_node_modules(node_modules)
    published_at = _generated_at(generated_at_utc)

    publish_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix=".fidc-revision-publish-",
        dir=publish_dir.parent,
    ) as tmp_text:
        stage = Path(tmp_text)
        stage_revision = stage / "revision"
        stage_exports = stage / "exports"
        stage_revision.mkdir(parents=True)
        stage_exports.mkdir(parents=True)
        stage_data = stage / "data_inputs"
        stage_data.mkdir(parents=True)
        for name in WORKBOOK_AUDIT_DATA_INPUTS:
            source = data_dir / name
            if not source.is_file():
                raise RevisionBundlePublishError(
                    f"input auditável do workbook ausente: {source}"
                )
            target = stage_data / name
            shutil.copy2(source, target)
            expected_hash = input_hashes.get(f"data/{name}")
            if expected_hash and _sha256_semantic_file(target) != expected_hash:
                raise RevisionBundlePublishError(
                    f"input auditável mudou durante a publicação: {name}"
                )
        staged_renderer = stage / ARTIFACT_SCRIPT.name
        staged_renderer.write_bytes(artifact_script_bytes)
        staged_native_chart_patcher = stage / NATIVE_CHART_PATCHER.name
        staged_native_chart_patcher.write_bytes(native_chart_patcher_bytes)
        staged_provider_flow_builder = stage / PROVIDER_FLOW_BUILDER.name
        staged_provider_flow_builder.write_bytes(provider_flow_builder_bytes)
        renderer = _artifact_runtime_metadata(
            node,
            resolved_modules,
            artifact_script=staged_renderer,
        )

        months = [str(value).strip() for value in presence_months if str(value).strip()]
        if not any(value.casefold() == "all" for value in months) and latest_complete not in months:
            months.append(latest_complete)
        analysis_args = [
            "--data-dir",
            str(data_dir),
            "--output-dir",
            str(stage_revision),
            "--latest-complete",
            latest_complete,
            "--raw-dir",
            str(raw_dir),
            "--presence-months",
            ",".join(months),
        ]
        published_overlay = publish_dir / "source_presence_overlay.csv.gz"
        if published_overlay.is_file():
            analysis_args.extend(
                ["--source-presence-overlay", str(published_overlay)]
            )
        if refresh_source_presence:
            analysis_args.append("--refresh-source-presence")
        if skip_download:
            analysis_args.append("--skip-download")
        build_revision_analysis(analysis_args)

        analysis_manifest_path = stage_revision / ANALYSIS_MANIFEST_NAME
        analysis_manifest = json.loads(analysis_manifest_path.read_text(encoding="utf-8"))
        if not refresh_source_presence:
            published_overlay = publish_dir / "source_presence_overlay.csv.gz"
            staged_overlay = stage_revision / "source_presence_overlay.csv.gz"
            if published_overlay.is_file():
                shutil.copy2(published_overlay, staged_overlay)
                with gzip.open(
                    staged_overlay,
                    "rt",
                    encoding="utf-8",
                    newline="",
                ) as handle:
                    reader = csv.reader(handle)
                    header = next(reader, [])
                    row_count = sum(1 for _ in reader)
                files = dict(analysis_manifest.get("files") or {})
                files["source_presence_overlay.csv.gz"] = {
                    "rows": row_count,
                    "columns": len(header),
                }
                analysis_manifest["files"] = files
        # The analysis builder records wall-clock time.  Replace it with the
        # publisher timestamp so SOURCE_DATE_EPOCH/--generated-at-utc also
        # stabilizes the staged analysis metadata.
        analysis_manifest, analysis_manifest_bytes = serialize_analysis_manifest(
            analysis_manifest, published_at
        )
        analysis_manifest_path.write_bytes(analysis_manifest_bytes)
        validate_analysis_manifest(
            analysis_manifest,
            revision_dir=stage_revision,
            latest_complete=latest_complete,
        )
        validate_source_presence_coverage(stage_revision, latest_complete)

        provider_history_args = [
            "--fund-base",
            str(stage_revision / "base_fundo_cnpj.csv.gz"),
            "--ownership-curation",
            str(data_dir / "provider_ownership_curation.csv"),
            "--output-dir",
            str(stage_revision),
            "--cache-zip",
            str(provider_history_archive),
            "--latest-competence",
            latest_complete,
        ]
        if skip_download:
            provider_history_args.append("--skip-download")
        build_provider_history(provider_history_args)
        missing_provider_history = sorted(
            name
            for name in REQUIRED_PROVIDER_HISTORY_FILES
            if not (stage_revision / name).exists()
        )
        if missing_provider_history:
            raise RevisionBundlePublishError(
                "staging do histórico de prestadores incompleto: "
                + ", ".join(missing_provider_history)
            )
        provider_history_manifest_path = (
            stage_revision / "prestadores_historico_cvm_manifest.json"
        )
        provider_history_manifest = json.loads(
            provider_history_manifest_path.read_text(encoding="utf-8")
        )
        provider_history_manifest["generated_at_utc"] = published_at
        provider_history_manifest_path.write_text(
            json.dumps(
                provider_history_manifest,
                ensure_ascii=False,
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        if provider_history_archive.exists():
            input_hashes["source/cad_fi_hist.zip"] = _sha256_file(
                provider_history_archive
            )
        for path in sorted(stage_revision.iterdir(), key=lambda item: item.name):
            if path.is_file() and path.name != ANALYSIS_MANIFEST_NAME:
                input_hashes[f"analysis/{path.name}"] = _sha256_semantic_file(path)

        payload = build_payload(
            data_dir=data_dir,
            revision_dir=stage_revision,
            curation_path=curation_path,
            latest=latest_complete,
        )
        taxonomy_ledger = data_dir / "taxonomy_review_actions.csv"
        taxonomy_audit = data_dir / "taxonomy_review_audit.csv"
        if _sha256_semantic_file(taxonomy_ledger) != input_hashes.get(
            "data/taxonomy_review_actions.csv"
        ):
            raise RevisionBundlePublishError(
                "ledger de taxonomia mudou durante a publicação; execute novamente para preservar atomicidade"
            )
        if _sha256_semantic_file(taxonomy_audit) != input_hashes.get(
            "data/taxonomy_review_audit.csv"
        ):
            raise RevisionBundlePublishError(
                "auditoria de taxonomia mudou durante a publicação; execute novamente para preservar atomicidade"
            )
        payload["generated_at"] = published_at
        validate_artifact_payload(payload, latest_complete)
        payload_bytes = json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ).encode("utf-8")
        payload_path = stage_revision / PAYLOAD_NAME
        payload_path.write_bytes(payload_bytes)

        # The workbook supplied as visual/reference input may predate the
        # published data snapshot.  Rebuild its inherited analytical tabs from
        # the same data directory before the artifact renderer adds the audited
        # revision sheets.  The PPTX renderer does not consume these tabs.
        staged_input_workbook = materialize_current_workbook_base(
            data_dir,
            stage / "workbook_current.xlsx",
            latest_complete,
        )
        input_hashes["workbook/generated_current.xlsx"] = _sha256_file(
            staged_input_workbook
        )

        staged_pptx = stage_exports / MATERIALIZED_PPTX_NAME
        staged_xlsx = stage_exports / MATERIALIZED_XLSX_NAME
        staged_portfolio_xlsx = stage_exports / MATERIALIZED_PORTFOLIO_XLSX_NAME
        staged_html = stage_exports / MATERIALIZED_HTML_NAME
        renderer_manifest_path = stage / "renderer_export_bundle.json"
        _run_artifact_builder(
            node=node,
            artifact_script=staged_renderer,
            provider_flow_builder=staged_provider_flow_builder,
            node_modules=resolved_modules,
            data_dir=stage_data,
            input_workbook=staged_input_workbook,
            revision_dir=stage_revision,
            payload_path=payload_path,
            output_dir=stage_exports,
            pptx_path=staged_pptx,
            xlsx_path=staged_xlsx,
            portfolio_xlsx_path=staged_portfolio_xlsx,
            html_path=staged_html,
            renderer_manifest_path=renderer_manifest_path,
            timeout_seconds=timeout_seconds,
        )
        pptx_bytes = staged_pptx.read_bytes()
        xlsx_bytes = staged_xlsx.read_bytes()
        portfolio_xlsx_bytes = staged_portfolio_xlsx.read_bytes()
        html_bytes = staged_html.read_bytes()
        renderer_manifest = json.loads(renderer_manifest_path.read_text(encoding="utf-8"))
        validate_renderer_manifest(
            renderer_manifest,
            payload_bytes=payload_bytes,
            payload=payload,
            pptx_bytes=pptx_bytes,
            xlsx_bytes=xlsx_bytes,
            portfolio_xlsx_bytes=portfolio_xlsx_bytes,
            html_bytes=html_bytes,
            renderer_sha256=str(renderer["renderer_sha256"]),
        )
        renderer = {
            **renderer,
            "renderer_version": str(renderer_manifest.get("renderer_version") or ""),
        }
        validate_revision_pptx(pptx_bytes)
        validate_revision_xlsx(xlsx_bytes)
        validate_revision_portfolio_xlsx(portfolio_xlsx_bytes)
        validate_user_facing_workbook_snapshot(xlsx_bytes, latest_complete)
        validate_revision_html(html_bytes)
        validate_deck_snapshot(pptx_bytes, latest_complete)

        manifest = build_bundle_manifest(
            payload_bytes=payload_bytes,
            payload=payload,
            analysis_manifest_bytes=analysis_manifest_bytes,
            pptx_bytes=pptx_bytes,
            xlsx_bytes=xlsx_bytes,
            portfolio_xlsx_bytes=portfolio_xlsx_bytes,
            html_bytes=html_bytes,
            input_hashes=input_hashes,
            renderer=renderer,
            generated_at_utc=published_at,
        )
        validate_bundle_manifest(
            manifest,
            payload_bytes=payload_bytes,
            payload=payload,
            analysis_manifest_bytes=analysis_manifest_bytes,
            pptx_bytes=pptx_bytes,
            xlsx_bytes=xlsx_bytes,
            portfolio_xlsx_bytes=portfolio_xlsx_bytes,
            html_bytes=html_bytes,
        )
        staged_manifest = stage / BUNDLE_MANIFEST_NAME
        staged_manifest.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )

        target_pptx, target_xlsx, target_manifest = publish_staged_bundle(
            staged_revision_dir=stage_revision,
            staged_pptx=staged_pptx,
            staged_xlsx=staged_xlsx,
            staged_portfolio_xlsx=staged_portfolio_xlsx,
            staged_html=staged_html,
            staged_bundle_manifest=staged_manifest,
            publish_dir=publish_dir,
        )

    # Re-read the committed files; the manifest is now the publication marker.
    committed_payload = publish_dir / PAYLOAD_NAME
    target_portfolio_xlsx = publish_dir / MATERIALIZED_PORTFOLIO_XLSX_NAME
    validate_bundle_manifest(
        json.loads(target_manifest.read_text(encoding="utf-8")),
        payload_bytes=committed_payload.read_bytes(),
        payload=json.loads(committed_payload.read_text(encoding="utf-8")),
        analysis_manifest_bytes=(publish_dir / ANALYSIS_MANIFEST_NAME).read_bytes(),
        pptx_bytes=target_pptx.read_bytes(),
        xlsx_bytes=target_xlsx.read_bytes(),
        portfolio_xlsx_bytes=target_portfolio_xlsx.read_bytes(),
        html_bytes=(publish_dir / MATERIALIZED_HTML_NAME).read_bytes(),
    )
    validate_revision_pptx(target_pptx.read_bytes())
    validate_revision_xlsx(target_xlsx.read_bytes())
    validate_revision_portfolio_xlsx(target_portfolio_xlsx.read_bytes())
    validate_revision_html((publish_dir / MATERIALIZED_HTML_NAME).read_bytes())
    return PublishedRevisionBundle(
        bundle_id=str(manifest["bundle_id"]),
        latest_complete=latest_complete,
        payload_path=committed_payload,
        pptx_path=target_pptx,
        xlsx_path=target_xlsx,
        portfolio_xlsx_path=target_portfolio_xlsx,
        html_path=publish_dir / MATERIALIZED_HTML_NAME,
        manifest_path=target_manifest,
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=ROOT / "data/industry_study")
    parser.add_argument(
        "--publish-dir",
        type=Path,
        default=ROOT / "data/industry_study/generated_revision",
    )
    parser.add_argument("--latest-complete", default="")
    parser.add_argument("--curation", type=Path, default=DEFAULT_CURATION)
    parser.add_argument(
        "--input-workbook",
        type=Path,
        default=(Path(os.environ["FIDC_INPUT_WORKBOOK"]) if os.environ.get("FIDC_INPUT_WORKBOOK") else None),
        help="workbook-base obrigatório; também pode vir de FIDC_INPUT_WORKBOOK",
    )
    parser.add_argument("--raw-dir", type=Path, default=ROOT / ".cache/cvm-industry-study")
    parser.add_argument(
        "--provider-history-archive",
        type=Path,
        default=None,
        help=(
            "cad_fi_hist.zip da CVM; vazio usa o cache irmão "
            ".cache/cvm-cadastro/cad_fi_hist.zip"
        ),
    )
    parser.add_argument("--refresh-source-presence", action="store_true")
    parser.add_argument(
        "--presence-months",
        default="all",
        help="competências separadas por vírgula; 'all' reprocessa todo o histórico",
    )
    parser.add_argument("--skip-download", action="store_true")
    parser.add_argument("--node-modules", type=Path, default=None)
    parser.add_argument("--timeout-seconds", type=int, default=DEFAULT_TIMEOUT_SECONDS)
    parser.add_argument(
        "--generated-at-utc",
        default="",
        help="timestamp ISO opcional; SOURCE_DATE_EPOCH também é respeitado",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    if args.input_workbook is None:
        raise SystemExit(
            "--input-workbook é obrigatório para publicar o bundle revisado"
        )
    result = publish_revision_bundle(
        data_dir=args.data_dir,
        publish_dir=args.publish_dir,
        curation_path=args.curation,
        input_workbook=args.input_workbook,
        latest_complete=str(args.latest_complete or "").strip(),
        raw_dir=args.raw_dir,
        provider_history_archive=args.provider_history_archive,
        refresh_source_presence=bool(args.refresh_source_presence),
        presence_months=[item.strip() for item in args.presence_months.split(",")],
        skip_download=bool(args.skip_download),
        node_modules=args.node_modules,
        timeout_seconds=max(1, int(args.timeout_seconds)),
        generated_at_utc=args.generated_at_utc,
    )
    print(
        f"[ok] bundle {result.bundle_id} publicado em {result.manifest_path.parent} "
        f"(competência {result.latest_complete})"
    )


if __name__ == "__main__":
    main()
