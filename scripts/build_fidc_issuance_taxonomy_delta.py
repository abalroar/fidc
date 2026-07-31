"""Decompose the 2023→2024 FIDC issuance jump by curated ANBIMA taxonomy.

Answers "which sectors explain the issuance growth" with the same taxonomy the
industry dashboard charts use: official ANBIMA type per fund, under the
documentary-curation overlay of ``taxonomy_review_actions.csv``.

The decomposition runs over the closed-offer cohort observed in the CVM/SRE
registry.  2023 is the one year that registry cannot carry alone — it captures
R$ 26,5 bi of the R$ 43,7 bi the ANBIMA Boletim closes — so the table carries
an explicit "não observado" bridge row for 2023 instead of allocating the gap
to categories by assumption.  With the bridge, the table totals equal the
levels charted by the panel after the 2023 ANBIMA correction.

Outputs, under ``outputs/analysis``:

- ``emissoes_fidc_2023_2024_taxonomia.csv`` — the table, machine-readable.
- ``emissoes_fidc_2023_2024_taxonomia.xlsx`` — the same table as a plain
  Excel sheet ready to paste into PPT, plus a per-focus detail sheet.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import re
import shutil
import sys
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pandas as pd

from services.industry_market_offer_reconciliation import load_anbima_market_offers
from services.industry_taxonomy_review import (
    apply_taxonomy_review_overlay,
    load_taxonomy_review_actions,
    normalize_cnpj,
)

COHORT_FILENAME = "industry_closed_offer_ticket_cohort.csv.gz"
FUND_BASE = Path("generated_revision") / "base_fundo_cnpj.csv.gz"
LEDGER_FILENAME = "taxonomy_review_actions.csv"
OUTPUT_BASENAME = "emissoes_fidc_2023_2024_taxonomia"

PERIODS = ("2023 FY", "2024 FY")
UNOBSERVED_LABEL = "Não observado na base CVM em 2023 (gap vs ANBIMA)"
ND_LABEL = "Sem classificação (N/D)"
FIC_LABEL = "FIC-FIDC (fundos que investem em cotas de outros FIDCs)"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=Path("data/industry_study"))
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/analysis"))
    return parser.parse_args(argv)


def _display_label(value: object) -> str:
    text = str(value or "").strip()
    if not text or text.casefold() in {"nan", "none", "n/d"}:
        return ND_LABEL
    if text == "FIC-FIDC":
        return FIC_LABEL
    return text


def build_issuance_taxonomy_frames(
    data_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, float]]:
    """Return the category table, the per-focus detail and the audit totals."""

    cohort = pd.read_csv(data_dir / COHORT_FILENAME, low_memory=False)
    offers = cohort[cohort["period_label"].isin(PERIODS)].copy()
    offers["cnpj_n"] = offers["cnpj_emissor"].astype(str).map(normalize_cnpj)
    offers["registered_volume_brl"] = pd.to_numeric(
        offers["registered_volume_brl"], errors="coerce"
    ).fillna(0.0)

    base = pd.read_csv(
        data_dir / FUND_BASE,
        usecols=["competencia", "cnpj_fundo", "anbima_tipo", "anbima_foco"],
        low_memory=False,
    )
    base["cnpj_fundo"] = base["cnpj_fundo"].astype(str).map(normalize_cnpj)
    photo = base.sort_values("competencia").drop_duplicates("cnpj_fundo", keep="last")
    actions = load_taxonomy_review_actions(data_dir / LEDGER_FILENAME)
    photo = apply_taxonomy_review_overlay(photo, actions)
    lookup = photo.set_index("cnpj_fundo")[["anbima_tipo_curado", "anbima_foco_curado"]]

    offers = offers.merge(lookup, left_on="cnpj_n", right_index=True, how="left")
    offers["categoria"] = offers["anbima_tipo_curado"].map(_display_label)
    offers["foco"] = offers["anbima_foco_curado"].map(_display_label)

    observed = (
        offers.groupby(["categoria", "period_label"])["registered_volume_brl"]
        .sum()
        .unstack("period_label")
        .reindex(columns=list(PERIODS))
        .fillna(0.0)
    )
    observed.columns = ["volume_2023_brl", "volume_2024_brl"]

    anbima = load_anbima_market_offers(data_dir)
    anbima_2023 = float(
        anbima.loc[
            anbima["instrument_label"].eq("FIDCs")
            & anbima["period_label"].eq("2023 FY"),
            "closed_volume_brl",
        ].iloc[0]
    )
    observed_2023 = float(observed["volume_2023_brl"].sum())
    observed_2024 = float(observed["volume_2024_brl"].sum())
    gap_2023 = anbima_2023 - observed_2023
    if gap_2023 < 0:
        raise SystemExit(
            "coorte CVM de 2023 excede o total ANBIMA; revisar fontes antes de publicar"
        )

    table = observed.sort_values("volume_2024_brl", ascending=False).reset_index()
    table = pd.concat(
        [
            table,
            pd.DataFrame(
                [
                    {
                        "categoria": UNOBSERVED_LABEL,
                        "volume_2023_brl": gap_2023,
                        "volume_2024_brl": 0.0,
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    total_2023 = float(table["volume_2023_brl"].sum())
    total_2024 = float(table["volume_2024_brl"].sum())
    table["share_2023"] = table["volume_2023_brl"] / total_2023
    table["share_2024"] = table["volume_2024_brl"] / total_2024
    table["delta_brl"] = table["volume_2024_brl"] - table["volume_2023_brl"]

    focus = (
        offers.groupby(["foco", "period_label"])["registered_volume_brl"]
        .sum()
        .unstack("period_label")
        .reindex(columns=list(PERIODS))
        .fillna(0.0)
    )
    focus.columns = ["volume_2023_brl", "volume_2024_brl"]
    focus["delta_brl"] = focus["volume_2024_brl"] - focus["volume_2023_brl"]
    focus = focus.sort_values("delta_brl", ascending=False).reset_index()

    # O delta por categoria é medido sobre a parcela observada de 2023, que
    # cobre 61% do ano.  Se o não observado tivesse a mesma composição, cada
    # 2023 subiria pelo mesmo fator e os deltas encolheriam — a pergunta é se o
    # pódio muda.  Testar isso é mais honesto do que escolher uma das hipóteses.
    ranked = table[~table["categoria"].eq(UNOBSERVED_LABEL)].copy()
    coverage_scale = anbima_2023 / observed_2023 if observed_2023 else 1.0
    ranked["delta_prorata_brl"] = ranked["volume_2024_brl"] - (
        ranked["volume_2023_brl"] * coverage_scale
    )
    top3_reported = ranked.nlargest(3, "delta_brl")["categoria"].tolist()
    top3_prorata = ranked.nlargest(3, "delta_prorata_brl")["categoria"].tolist()
    if top3_reported != top3_prorata:
        raise SystemExit(
            "o pódio do delta muda conforme a hipótese de cobertura de 2023: "
            f"{top3_reported} versus {top3_prorata}; publicar exige explicitar "
            "as duas leituras em vez de uma"
        )

    totals = {
        "observed_2023_brl": observed_2023,
        "observed_2024_brl": observed_2024,
        "anbima_2023_brl": anbima_2023,
        "gap_2023_brl": gap_2023,
        "total_2023_brl": total_2023,
        "total_2024_brl": total_2024,
        "coverage_scale_2023": coverage_scale,
        "top3_labels": top3_reported,
    }
    return table, focus, totals


_CELL_PATTERN = re.compile(
    r'<c r="(?P<ref>[A-Z]+\d+)"(?P<attrs>[^>/]*)>(?P<body>.*?)</c>', re.DOTALL
)
_FORMULA_PATTERN = re.compile(r"<f[ >].*?</f>|<f/>", re.DOTALL)


def inject_cached_values(path: Path, cached: dict[str, dict[str, float]]) -> int:
    """Write the result of every formula next to the formula itself.

    openpyxl emits formulas with no cached value, so until the file is opened
    in a spreadsheet application every formula cell reads back as empty to
    pandas, to ``load_workbook(data_only=True)`` and to previewers.  The
    canonical fix is to recalculate with LibreOffice, which is unavailable in
    this runtime, so the values computed here — the same arithmetic the
    formulas express — are written into the sheet XML as ``<v>``.

    The formulas stay in the file and remain authoritative: a spreadsheet
    recalculates them on open and overwrites these values.  Returns how many
    cells were filled, so the caller can assert none was missed.
    """

    sheet_targets = {
        f"xl/worksheets/sheet{index}.xml": values
        for index, values in enumerate(cached.values(), start=1)
    }
    staging = path.with_suffix(".tmp.xlsx")
    filled = 0
    with ZipFile(path) as source, ZipFile(
        staging, "w", compression=ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            values = sheet_targets.get(item.filename)
            if values:
                text = payload.decode("utf-8")

                def replace(match: re.Match[str]) -> str:
                    nonlocal filled
                    body = match.group("body")
                    if "<f" not in body:
                        return match.group(0)
                    value = values.get(match.group("ref"))
                    if value is None:
                        return match.group(0)
                    formula = _FORMULA_PATTERN.search(body)
                    if formula is None:
                        return match.group(0)
                    filled += 1
                    attrs = match.group("attrs")
                    return (
                        f'<c r="{match.group("ref")}"{attrs}>'
                        f"{formula.group(0)}<v>{escape(repr(float(value)))}</v></c>"
                    )

                payload = _CELL_PATTERN.sub(replace, text).encode("utf-8")
            target.writestr(item, payload)
    shutil.move(staging, path)
    return filled


def write_outputs(
    table: pd.DataFrame,
    focus: pd.DataFrame,
    totals: dict[str, float],
    output_dir: Path,
) -> tuple[Path, Path]:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"{OUTPUT_BASENAME}.csv"
    table.to_csv(csv_path, index=False)

    xlsx_path = output_dir / f"{OUTPUT_BASENAME}.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Emissões por categoria"
    normal = Font(name="Arial", size=11)
    bold = Font(name="Arial", size=11, bold=True)
    small = Font(name="Arial", size=9)

    sheet["A1"] = "Emissões de FIDCs por categoria — 2023 vs 2024 (R$ bilhões)"
    sheet["A1"].font = bold
    headers = (
        "Categoria",
        "2023 (R$ bi)",
        "2023 (%)",
        "2024 (R$ bi)",
        "2024 (%)",
        "Delta (R$ bi)",
    )
    header_row = 3
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.font = bold

    first_data = header_row + 1
    total_row = first_data + len(table)
    total_2023_bi = totals["total_2023_brl"] / 1e9
    total_2024_bi = totals["total_2024_brl"] / 1e9
    category_cache: dict[str, float] = {}
    for offset, record in enumerate(table.itertuples(index=False)):
        row = first_data + offset
        value_2023 = record.volume_2023_brl / 1e9
        value_2024 = record.volume_2024_brl / 1e9
        sheet.cell(row=row, column=1, value=record.categoria).font = normal
        sheet.cell(row=row, column=2, value=value_2023)
        sheet.cell(row=row, column=3, value=f"=B{row}/B${total_row}")
        sheet.cell(row=row, column=4, value=value_2024)
        sheet.cell(row=row, column=5, value=f"=D{row}/D${total_row}")
        sheet.cell(row=row, column=6, value=f"=D{row}-B{row}")
        category_cache[f"C{row}"] = value_2023 / total_2023_bi
        category_cache[f"E{row}"] = value_2024 / total_2024_bi
        category_cache[f"F{row}"] = value_2024 - value_2023
    label = sheet.cell(
        row=total_row,
        column=1,
        value="Total (igual aos gráficos do painel)",
    )
    label.font = bold
    sheet.cell(row=total_row, column=2, value=f"=SUM(B{first_data}:B{total_row - 1})")
    sheet.cell(row=total_row, column=3, value=f"=B{total_row}/B${total_row}")
    sheet.cell(row=total_row, column=4, value=f"=SUM(D{first_data}:D{total_row - 1})")
    sheet.cell(row=total_row, column=5, value=f"=D{total_row}/D${total_row}")
    sheet.cell(row=total_row, column=6, value=f"=D{total_row}-B{total_row}")
    category_cache[f"B{total_row}"] = total_2023_bi
    category_cache[f"C{total_row}"] = 1.0
    category_cache[f"D{total_row}"] = total_2024_bi
    category_cache[f"E{total_row}"] = 1.0
    category_cache[f"F{total_row}"] = total_2024_bi - total_2023_bi
    for row in range(first_data, total_row + 1):
        for column, mask in ((2, "#,##0.0"), (3, "0.0%"), (4, "#,##0.0"), (5, "0.0%"), (6, "#,##0.0")):
            cell = sheet.cell(row=row, column=column)
            cell.number_format = mask
            cell.font = bold if row == total_row else normal

    notes = (
        "Fonte: CVM/SRE (ofertas públicas primárias encerradas, snapshot 24/jul/26) "
        "com taxonomia ANBIMA sob curadoria documental do projeto "
        "(taxonomy_review_actions.csv).",
        "2023 foi o primeiro ano da Resolução CVM 160: a base granular da CVM "
        f"observa R$ {totals['observed_2023_brl'] / 1e9:,.1f} bi dos "
        f"R$ {totals['anbima_2023_brl'] / 1e9:,.1f} bi encerrados segundo a ANBIMA "
        "(Boletim de Mercado de Capitais, mai/26). A linha 'Não observado' carrega a "
        "diferença em vez de distribuí-la por hipótese.",
        "A decomposição por categoria usa a parcela observada na CVM; em 2024 a "
        "cobertura da CVM é integral. FIC-FIDCs destacados por serem fundos de "
        "cotas: o dinheiro captado investe em outros FIDCs.",
        "Sensibilidade: se o volume não observado de 2023 tivesse a mesma "
        f"composição do observado (fator {totals['coverage_scale_2023']:.2f}×), as "
        "três categorias que mais cresceram seriam as mesmas — "
        + ", ".join(totals["top3_labels"])
        + ". O ranking não depende da hipótese de cobertura.",
    )
    note_row = total_row + 2
    for offset, note in enumerate(notes):
        cell = sheet.cell(row=note_row + offset, column=1, value=note)
        cell.font = small
    sheet.column_dimensions["A"].width = 52
    for letter in ("B", "C", "D", "E", "F"):
        sheet.column_dimensions[letter].width = 13

    detail = book.create_sheet("Detalhe por foco")
    detail["A1"] = "Emissões por Foco ANBIMA (parcela observada na CVM) — R$ bilhões"
    detail["A1"].font = bold
    for column, header in enumerate(
        ("Foco ANBIMA", "2023 (R$ bi)", "2024 (R$ bi)", "Delta (R$ bi)"), start=1
    ):
        cell = detail.cell(row=3, column=column, value=header)
        cell.font = bold
    focus_cache: dict[str, float] = {}
    for offset, record in enumerate(focus.itertuples(index=False)):
        row = 4 + offset
        value_2023 = record.volume_2023_brl / 1e9
        value_2024 = record.volume_2024_brl / 1e9
        detail.cell(row=row, column=1, value=record.foco).font = normal
        detail.cell(row=row, column=2, value=value_2023)
        detail.cell(row=row, column=3, value=value_2024)
        detail.cell(row=row, column=4, value=f"=C{row}-B{row}")
        focus_cache[f"D{row}"] = value_2024 - value_2023
        for column in (2, 3, 4):
            cell = detail.cell(row=row, column=column)
            cell.number_format = "#,##0.0"
            cell.font = normal
    detail.column_dimensions["A"].width = 34
    for letter in ("B", "C", "D"):
        detail.column_dimensions[letter].width = 13

    book.save(xlsx_path)
    expected = len(category_cache) + len(focus_cache)
    filled = inject_cached_values(
        xlsx_path,
        {sheet.title: category_cache, detail.title: focus_cache},
    )
    if filled != expected:
        raise SystemExit(
            f"cache de fórmulas incompleto: {filled} de {expected} células "
            "preenchidas; o arquivo abriria com células vazias fora do Excel"
        )
    return csv_path, xlsx_path


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    table, focus, totals = build_issuance_taxonomy_frames(args.data_dir)
    csv_path, xlsx_path = write_outputs(table, focus, totals, args.output_dir)
    top3 = table[~table["categoria"].eq(UNOBSERVED_LABEL)].nlargest(3, "delta_brl")
    print(f"[ok] tabela materializada: {csv_path}")
    print(f"[ok] excel materializado: {xlsx_path}")
    print(
        "totais: 2023 R$ {a:.2f} bi (CVM {b:.2f} + gap {c:.2f}) | 2024 R$ {d:.2f} bi".format(
            a=totals["total_2023_brl"] / 1e9,
            b=totals["observed_2023_brl"] / 1e9,
            c=totals["gap_2023_brl"] / 1e9,
            d=totals["total_2024_brl"] / 1e9,
        )
    )
    for record in top3.itertuples(index=False):
        print(
            f"  top delta: {record.categoria}: "
            f"{record.volume_2023_brl / 1e9:.2f} → {record.volume_2024_brl / 1e9:.2f} "
            f"(+{record.delta_brl / 1e9:.2f} bi)"
        )
    print(
        "sensibilidade: o pódio se mantém com o não observado de 2023 "
        f"distribuído pro-rata (fator {totals['coverage_scale_2023']:.2f}×)"
    )


if __name__ == "__main__":
    main()
