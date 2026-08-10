"""Gera o Excel simples de validação dos analistas da Carteira 101."""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from services.carteira_validacao_analistas import (
    build_validation_frame,
    target_validation_frame,
)


DEFAULT_DATA_DIR = Path("data/industry_study")
DEFAULT_OUTPUT = Path("outputs/validacao_analistas/Validacao_Analistas.xlsx")

ORANGE = "E36C0A"
WHITE = "FFFFFF"
GRAY_300 = "D7DADD"
RED_FILL = "FCE4E4"
RED_TEXT = "A83C32"

COLUMNS: tuple[tuple[str, str, float], ...] = (
    ("ordem", "#", 5),
    ("cnpj", "CNPJ", 18),
    ("fidc", "FIDC", 48),
    ("competencia", "Competência", 12),
    ("segmento", "Segmento", 18),
    ("analista", "Analista", 18),
    ("gerente", "Gerente", 18),
    ("incluido_slide", "No slide", 11),
    ("cobertura_pct", "Cobertura PDD / Inadimplência", 19),
    ("referencia_pct", "Sub Mínima usada no stress", 19),
    ("resultado_documental", "Resultado documental", 23),
    ("valor_documental_pct", "Sub Mínima (regulamento)", 20),
    ("documento", "Documento", 31),
    ("pagina", "Página", 10),
    ("trecho", "Trecho de referência", 72),
    ("minimo_fonte", "Fonte do mínimo usado", 42),
    ("pl_total_cotas_brl", "PL Total (R$ mm)", 17),
    ("pl_senior_brl", "PL Sênior (R$ mm)", 17),
    ("pl_mezanino_brl", "PL Mezanino (R$ mm)", 18),
    ("pl_subordinada_brl", "PL Subordinada (R$ mm)", 20),
    ("subordinada_sobre_pl", "Subordinada / PL", 17),
    ("submaismez_sobre_pl", "Sub + Mez / PL", 17),
    ("dc_inadimplentes", "Inadimplência (R$ mm)", 19),
    ("pdd_brl", "PDD (R$ mm)", 16),
    ("deficit_brl", "Sub não provisionada — Δ (R$ mm)", 24),
    ("sub_pos_pct", "Sub pós-estresse", 17),
    ("folga_pos_pp", "Folga pós-estresse (p.p.)", 20),
    ("aporte_brl", "Aporte necessário (R$ mm)", 21),
    ("aporte_sobre_pl", "Aporte / PL atual", 17),
    ("validacao_sub_minima", "Validação — Sub Mínima", 19),
    ("validacao_sub_pl", "Validação — Sub / PL", 18),
    ("validacao_sub_mez_pl", "Validação — Sub + Mez / PL", 21),
    ("validacao_pl_total", "Validação — PL Total", 18),
    ("validacao_pl_senior", "Validação — PL Sênior", 19),
    ("validacao_pl_subordinada", "Validação — PL Subordinada", 22),
    ("comentario", "Comentário", 46),
)

NOTES = {
    "referencia_pct": (
        "Percentual aplicado no teste de estresse já validado. A confirmação "
        "documental aparece nas colunas seguintes; divergências e remissões a "
        "suplemento permanecem explícitas."
    ),
    "valor_documental_pct": (
        "Preenchido somente quando a apuração precision-first localizou no "
        "regulamento um percentual ancorado em cláusula de piso e patrimônio líquido."
    ),
    "pl_total_cotas_brl": "Total de cotas da Tabela X.2 do Informe Mensal, denominador do stress.",
    "submaismez_sobre_pl": "Fica em branco quando a Tabela X.2 não traz cota mezanino.",
    "deficit_brl": (
        "Fórmula em Excel: Δ = máx(Inadimplência − PDD, 0). Se a PDD estiver "
        "em branco, a fórmula aplica zero, conforme a lógica validada do stress."
    ),
    "folga_pos_pp": "Folga = Sub pós-estresse − Sub Mínima usada no stress.",
    "aporte_brl": (
        "Aporte = máx(0; [m × (Total − Δ) − (Sub + Mez − Δ)] ÷ (1 − m)), "
        "em que m é a Sub Mínima usada no stress."
    ),
    "aporte_sobre_pl": "Aporte / PL atual = Aporte necessário ÷ PL Total.",
}

VALIDATION_KEYS = (
    "validacao_sub_minima",
    "validacao_sub_pl",
    "validacao_sub_mez_pl",
    "validacao_pl_total",
    "validacao_pl_senior",
    "validacao_pl_subordinada",
)


def _is_missing(value: object) -> bool:
    return value is None or (isinstance(value, float) and pd.isna(value)) or str(value) == ""


def _columns() -> dict[str, int]:
    return {key: pos for pos, (key, _, _) in enumerate(COLUMNS, start=1)}


def _letter(key: str) -> str:
    return get_column_letter(_columns()[key])


def _value(record: dict[str, object], key: str, row: int):
    total = _letter("pl_total_cotas_brl")
    mezz = _letter("pl_mezanino_brl")
    sub = _letter("pl_subordinada_brl")
    minimo = _letter("referencia_pct")
    inadimplencia = _letter("dc_inadimplentes")
    pdd = _letter("pdd_brl")
    delta = _letter("deficit_brl")
    sub_pos = _letter("sub_pos_pct")
    aporte = _letter("aporte_brl")

    if key in {"segmento", "analista", "gerente", "comentario"}:
        return None
    if key == "incluido_slide":
        return "Sim" if bool(record.get(key)) else "Não"
    if key in {"cobertura_pct", "referencia_pct", "valor_documental_pct"}:
        value = record.get(key)
        return None if _is_missing(value) else float(value) / 100.0
    if key in {
        "pl_total_cotas_brl",
        "pl_senior_brl",
        "pl_mezanino_brl",
        "pl_subordinada_brl",
        "dc_inadimplentes",
        "pdd_brl",
    }:
        value = record.get(key)
        if _is_missing(value):
            return None
        if key == "pl_mezanino_brl" and not bool(record.get("tem_mezanino")):
            return None
        return float(value) / 1e6
    if key == "subordinada_sobre_pl":
        return f'=IF(OR({total}{row}="",{sub}{row}="",{total}{row}=0),"",{sub}{row}/{total}{row})'
    if key == "submaismez_sobre_pl":
        if not bool(record.get("tem_mezanino")):
            return None
        return f'=IF(OR({total}{row}="",{sub}{row}="",{mezz}{row}="",{total}{row}=0),"",({sub}{row}+{mezz}{row})/{total}{row})'
    if key == "deficit_brl":
        return (
            f'=IF({inadimplencia}{row}="","",MAX(0,'
            f'{inadimplencia}{row}-IF({pdd}{row}="",0,{pdd}{row})))'
        )
    if key == "sub_pos_pct":
        return (
            f'=IF(OR({total}{row}="",{sub}{row}="",{delta}{row}="",'
            f'{total}{row}={delta}{row}),"",({sub}{row}+IF({mezz}{row}="",0,'
            f'{mezz}{row})-{delta}{row})/({total}{row}-{delta}{row}))'
        )
    if key == "folga_pos_pp":
        return f'=IF(OR({sub_pos}{row}="",{minimo}{row}=""),"",{sub_pos}{row}-{minimo}{row})'
    if key == "aporte_brl":
        return (
            f'=IF(OR({minimo}{row}="",{total}{row}="",{sub}{row}="",'
            f'{delta}{row}="",{minimo}{row}=1),"",MAX(0,('
            f'{minimo}{row}*({total}{row}-{delta}{row})-('
            f'{sub}{row}+IF({mezz}{row}="",0,{mezz}{row})-{delta}{row}))/'
            f'(1-{minimo}{row})))'
        )
    if key == "aporte_sobre_pl":
        return f'=IF(OR({aporte}{row}="",{total}{row}="",{total}{row}=0),"",{aporte}{row}/{total}{row})'
    if key in VALIDATION_KEYS:
        related = {
            "validacao_sub_minima": "referencia_pct",
            "validacao_sub_pl": "pl_subordinada_brl",
            "validacao_sub_mez_pl": "pl_mezanino_brl",
            "validacao_pl_total": "pl_total_cotas_brl",
            "validacao_pl_senior": "pl_senior_brl",
            "validacao_pl_subordinada": "pl_subordinada_brl",
        }[key]
        if related == "pl_mezanino_brl" and not bool(record.get("tem_mezanino")):
            return None
        return "Pendente" if not _is_missing(record.get(related)) else None
    value = record.get(key)
    return None if _is_missing(value) else value


def _write_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title)
    header_font = Font(name="Arial", size=10, bold=True, color=WHITE)
    body_font = Font(name="Arial", size=10)
    header_fill = PatternFill("solid", fgColor=ORANGE)
    thin = Side(style="thin", color=GRAY_300)
    border = Border(bottom=thin)

    for column, (key, label, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(1, column, label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if key in NOTES:
            cell.comment = Comment(NOTES[key], "Dados da Indústria")
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.row_dimensions[1].height = 42

    for row, record in enumerate(frame.to_dict("records"), start=2):
        for column, (key, _, _) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row, column, _value(record, key, row))
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal=("left" if key in {"fidc", "resultado_documental", "documento", "trecho", "minimo_fonte", "comentario"} else "center"),
                vertical="top",
                wrap_text=key in {"fidc", "resultado_documental", "documento", "trecho", "minimo_fonte", "comentario"},
            )
            if key == "cnpj":
                cell.number_format = "@"
            elif key in {"cobertura_pct", "referencia_pct", "valor_documental_pct", "subordinada_sobre_pl", "submaismez_sobre_pl", "sub_pos_pct", "aporte_sobre_pl"}:
                cell.number_format = "0.00%"
            elif key == "folga_pos_pp":
                cell.number_format = "+0.00%;-0.00%;0.00%"
            elif key in {
                "pl_total_cotas_brl",
                "pl_senior_brl",
                "pl_mezanino_brl",
                "pl_subordinada_brl",
                "dc_inadimplentes",
                "pdd_brl",
                "deficit_brl",
                "aporte_brl",
            }:
                # Sem agrupamento de milhar: em Excel pt-BR, 1759.6 aparece
                # como 1759,6, sem o ponto de milhar.
                cell.number_format = "0.0"

    last_row = len(frame) + 1
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions.group(2, last_row, hidden=False)

    validation = DataValidation(type="list", formula1='"Sim,Não,Pendente"', allow_blank=True)
    validation.error = "Use Sim, Não ou Pendente."
    validation.errorTitle = "Valor inválido"
    validation.prompt = "Confirme a variável ou registre a contestação no comentário."
    validation.promptTitle = "Validação do analista"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    for key in VALIDATION_KEYS:
        letter = _letter(key)
        validation.add(f"{letter}2:{letter}{last_row}")

    folga_letter = _letter("folga_pos_pp")
    sheet.conditional_formatting.add(
        f"{folga_letter}2:{folga_letter}{last_row}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=RED_FILL),
            font=Font(name="Arial", size=10, bold=True, color=RED_TEXT),
        ),
    )


def write_workbook(
    target: pd.DataFrame,
    carteira101: pd.DataFrame,
    output: Path,
) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "Validação Analistas", target)
    _write_sheet(workbook, "Carteira 101", carteira101)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output)
    return output


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = write_workbook(
        target_validation_frame(args.data_dir),
        build_validation_frame(args.data_dir),
        args.output,
    )
    print(output)


if __name__ == "__main__":
    main()
