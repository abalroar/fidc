"""A munição da apuração num arquivo só: o que os documentos dizem e o que falta.

Três abas, cada uma respondendo a uma pergunta diferente:

``Apuração``
    um fundo por linha, com cada família de cláusula marcada, o trecho literal
    que a sustenta, a contraprova de terceiro e o diagnóstico consolidado;

``Subordinação``
    o mínimo em uso confrontado com o regulamento, com veredito, página e
    trecho;

``Em branco``
    só o que **falta** — uma linha por lacuna nomeada, que é a lista de trabalho
    de quem vai ler os documentos à mão.

A aba em branco existe porque é o pedido central: a varredura automática não
fecha o caso de todo fundo, e fingir que fecha seria pior do que dizer onde ela
parou.

    python scripts/build_carteira_apuracao_xlsx.py
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from services.carteira_apuracao_documental import (  # noqa: E402
    APURACAO_NAME,
    FAMILIAS,
)
from services.carteira_subordinacao import DEFAULT_DATA_DIR  # noqa: E402
from services.carteira_validacao_subordinacao import (  # noqa: E402
    DIVERGE,
    REMETE_SUPLEMENTO,
    SEM_CLAUSULA,
    SEM_DOCUMENTO,
)

VALIDACAO_NAME = "carteira_subordinacao_validacao.csv"
DEFAULT_OUTPUT = ROOT / "outputs" / "carteira_apuracao" / "Carteira_Apuracao_Documental.xlsx"

ORANGE = "E36C0A"
WHITE = "FFFFFF"
GRAY = "F5F6F7"


def _cabecalho_apuracao() -> list[tuple[str, str, int]]:
    colunas = [
        ("cnpj", "CNPJ", 20),
        ("fundo", "FIDC", 34),
        ("secao", "Seção", 22),
        ("caso", "Caso na CVM", 40),
        ("carteira_mm", "Carteira R$ mm", 15),
        ("inadimplencia_mm", "Inad. R$ mm", 13),
        ("pdd_mm", "PDD R$ mm", 13),
        ("diagnostico", "Diagnóstico documental", 60),
        ("lacunas", "O que falta apurar", 48),
        ("contraprova_fontes", "Contraprova de terceiro", 30),
        ("contraprova_trecho", "Trecho da contraprova", 70),
        ("terceiro_veredito", "Medição de terceiro", 30),
        ("terceiro_competencia", "Competência da medição", 18),
        ("terceiro_vencidos_brl", "Vencidos no terceiro (R$)", 20),
        ("terceiro_pdd_brl", "PDD no terceiro (R$)", 18),
        ("terceiro_documento", "Documento da medição", 26),
    ]
    for chave, familia in FAMILIAS.items():
        rotulo = str(familia["rotulo"])
        colunas.append((chave, rotulo, 14))
        colunas.append((f"{chave}_sinal", f"{rotulo} — sinal", 14))
        colunas.append((f"{chave}_fonte", f"{rotulo} — fonte", 26))
        colunas.append((f"{chave}_trecho", f"{rotulo} — trecho", 70))
    return colunas


CABECALHO_VALIDACAO = [
    ("cnpj", "CNPJ", 20),
    ("fundo", "FIDC", 34),
    ("minimo_em_uso_pct", "Mínimo em uso (%)", 17),
    ("valor_documental_pct", "Mínimo no documento (%)", 22),
    ("veredito", "Veredito", 22),
    ("familia", "Forma da cláusula", 18),
    ("documento", "Documento", 34),
    ("pagina", "Página", 9),
    ("fonte_no_registro", "Fonte registrada", 40),
    ("trecho", "Trecho literal", 90),
]

CABECALHO_BRANCO = [
    ("cnpj", "CNPJ", 20),
    ("fundo", "FIDC", 34),
    ("frente", "Frente", 22),
    ("lacuna", "O que falta", 62),
    ("onde_procurar", "Onde procurar", 52),
]


def _escrever(workbook, titulo: str, colunas, frame: pd.DataFrame, nome_tabela: str):
    sheet = workbook.create_sheet(titulo)
    fonte_cabecalho = Font(name="Arial", size=10, bold=True, color=WHITE)
    preenchimento = PatternFill("solid", fgColor=ORANGE)
    corpo = Font(name="Arial", size=10)

    for indice, (_chave, rotulo, largura) in enumerate(colunas, start=1):
        celula = sheet.cell(row=1, column=indice, value=rotulo)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento
        celula.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(indice)].width = largura
    sheet.row_dimensions[1].height = 30

    for linha, registro in enumerate(frame.to_dict("records"), start=2):
        for indice, (chave, _rotulo, _largura) in enumerate(colunas, start=1):
            valor = registro.get(chave, "")
            if isinstance(valor, float) and pd.isna(valor):
                valor = ""
            celula = sheet.cell(row=linha, column=indice, value=valor)
            celula.font = corpo
            celula.alignment = Alignment(vertical="top", wrap_text=len(_largura * "x") > 30)

    ultima = get_column_letter(len(colunas))
    total = max(len(frame) + 1, 2)
    tabela = Table(displayName=nome_tabela, ref=f"A1:{ultima}{total}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(tabela)
    sheet.freeze_panes = "C2"
    return sheet


def montar_em_branco(apuracao: pd.DataFrame, validacao: pd.DataFrame) -> pd.DataFrame:
    """Uma linha por lacuna — a lista de trabalho, não o relatório."""

    linhas: list[dict[str, str]] = []
    for fundo in apuracao.itertuples():
        for lacuna in str(getattr(fundo, "lacunas", "") or "").split(";"):
            lacuna = lacuna.strip()
            if not lacuna:
                continue
            linhas.append(
                {
                    "cnpj": fundo.cnpj,
                    "fundo": fundo.fundo,
                    "frente": "inadimplência / PDD",
                    "lacuna": lacuna,
                    "onde_procurar": (
                        "regulamento e anexos da classe; relatório de rating; "
                        "demonstrações financeiras; verificação de lastro"
                    ),
                }
            )

    pendentes = {DIVERGE, REMETE_SUPLEMENTO, SEM_CLAUSULA, SEM_DOCUMENTO}
    onde = {
        DIVERGE: "conferir a cláusula na página indicada e decidir qual valor vale",
        REMETE_SUPLEMENTO: "baixar o Suplemento/Anexo da Classe citado no trecho",
        SEM_CLAUSULA: "ler o regulamento inteiro; o piso pode estar em anexo",
        SEM_DOCUMENTO: "baixar o regulamento do fundo na FundosNET",
    }
    for fundo in validacao[validacao["veredito"].isin(pendentes)].itertuples():
        linhas.append(
            {
                "cnpj": fundo.cnpj,
                "fundo": fundo.fundo,
                "frente": "subordinação mínima",
                "lacuna": f"{fundo.veredito} (em uso {fundo.minimo_em_uso_pct})",
                "onde_procurar": onde[fundo.veredito],
            }
        )
    return pd.DataFrame(linhas)


def main() -> None:
    from openpyxl import Workbook

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    apuracao = pd.read_csv(args.data_dir / APURACAO_NAME, dtype={"cnpj": str}).fillna("")
    validacao = pd.read_csv(
        args.data_dir / VALIDACAO_NAME, dtype={"cnpj": str}
    ).fillna("")
    branco = montar_em_branco(apuracao, validacao)

    workbook = Workbook()
    workbook.remove(workbook.active)
    _escrever(workbook, "Apuração", _cabecalho_apuracao(), apuracao, "Apuracao")
    _escrever(workbook, "Subordinação", CABECALHO_VALIDACAO, validacao, "Subordinacao")
    _escrever(workbook, "Em branco", CABECALHO_BRANCO, branco, "EmBranco")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(
        f"{len(apuracao)} fundos apurados · {len(validacao)} mínimos conferidos · "
        f"{len(branco)} lacunas -> {args.output}"
    )


if __name__ == "__main__":
    main()
